"""Excel report builders for the programme modules.

One workbook per module (inventory / milestone shifts / variance), matching
the dcma.report_xlsx look. Each accepts an optional AI narrative which lands
on its own sheet. UI-independent: every builder returns bytes.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .inventory import ProgrammeInventory
from .milestones import MilestoneSeries, MilestoneShiftResult
from .variance import VarianceResult

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F3864")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
SLIP_FILL = PatternFill("solid", fgColor="FFC7CE")
GAIN_FILL = PatternFill("solid", fgColor="C6EFCE")


def _wb_bytes(wb: Workbook) -> bytes:
    """Serialise with build provenance — EVERY workbook exits here.

    The stamp (toolkit, commit, generation time) lands three ways: the
    print footer of every sheet, a visible line under the first sheet's
    content, and the file's description property — so the document can
    testify which code produced it (audit F-04, answered per-export).
    """
    from buildinfo import build_stamp
    stamp = build_stamp()
    for ws in wb.worksheets:
        # (H6) formula neutralisation: these workbooks never write
        # intentional formulas, so any cell openpyxl typed as a formula
        # got there from XER-derived text beginning with '=' (a hostile
        # activity name or evidence field). Re-type it as a literal
        # string — the text is preserved exactly, nothing executes.
        for row in ws.iter_rows():
            for c in row:
                if c.data_type == "f":
                    c.data_type = "s"
        ws.oddFooter.left.text = stamp
        ws.oddFooter.left.size = 7
        # every table gets an Excel autofilter: _header_row() recorded
        # where the header sits. Sheets with MULTIPLE stacked tables are
        # skipped — one filter range would swallow the lower table and
        # filtering would hide its rows. Computed before the stamp line
        # below so the stamp never lands inside a filter range.
        anchors = getattr(ws, "_hdr_anchors", [])
        if len(anchors) == 1 and not ws.auto_filter.ref:
            row, ncols = anchors[0]
            if ws.max_row > row and ncols:
                ws.auto_filter.ref = (
                    f"A{row}:{get_column_letter(ncols)}{ws.max_row}")
    ws0 = wb.worksheets[0]
    c = ws0.cell(row=ws0.max_row + 2, column=1, value=stamp)
    c.font = Font(size=8, italic=True, color="808080")
    wb.properties.description = stamp
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _title(ws, text: str, span: int) -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=2, column=1,
            value=f"Generated {datetime.now():%Y-%m-%d %H:%M}").font = Font(
        italic=True, color="6E7781")


def _header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = THIN_BORDER
    # record the table anchor so _wb_bytes can set the autofilter
    if not hasattr(ws, "_hdr_anchors"):
        ws._hdr_anchors = []
    ws._hdr_anchors.append((row, len(headers)))


def _autofit(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _narrative_sheet(wb: Workbook, narrative: str | None) -> None:
    if not narrative:
        return
    ws = wb.create_sheet("AI Narrative")
    ws["A1"] = "AI-Generated Narrative"
    ws["A1"].font = Font(size=13, bold=True, color="1F3864")
    ws.column_dimensions["A"].width = 110
    for i, para in enumerate(narrative.split("\n"), start=3):
        c = ws.cell(row=i, column=1, value=para)
        c.alignment = WRAP


def _notes_sheet(wb: Workbook, notes: list[str], title: str = "Notes") -> None:
    if not notes:
        return
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = Font(size=13, bold=True, color="1F3864")
    ws.column_dimensions["A"].width = 110
    for i, n in enumerate(notes, start=3):
        c = ws.cell(row=i, column=1, value=f"• {n}")
        c.alignment = WRAP


def _fmt(d) -> str:
    return f"{d:%Y-%m-%d}" if d else "—"


# --------------------------------------------------------------------------- #
# Module 0 — Data inventory
# --------------------------------------------------------------------------- #
def build_inventory_xlsx(
    inv: ProgrammeInventory, narrative: str | None = None
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    _title(ws, "Programme Data Inventory", 8)

    headers = ["File", "Project", "Data Date", "Role", "Activities",
               "Relationships", "Milestones", "Activity Codes"]
    _header_row(ws, 4, headers)
    for i, r in enumerate(inv.revisions, start=5):
        role = ("Baseline" if r.is_baseline
                else "Current" if r.is_current else "Update")
        values = [r.file_name, r.project_short_name or "—", _fmt(r.data_date),
                  role, r.activity_count, r.relationship_count,
                  r.milestone_count, "Yes" if r.has_activity_codes else "No"]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
    _autofit(ws, {1: 30, 2: 22, 3: 12, 4: 10, 5: 11, 6: 14, 7: 11, 8: 14})
    ws.freeze_panes = "A5"

    _notes_sheet(wb, inv.missing + inv.warnings, "Missing & Warnings")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 3 — Milestone shifts
# --------------------------------------------------------------------------- #
def build_milestone_xlsx(
    result: MilestoneShiftResult,
    series: list[MilestoneSeries],
    narrative: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Summary"
    _title(ws, "Milestone Shift Tracker", 6)

    headers = ["Activity ID", "Milestone", "First Forecast", "Latest Date",
               "Total Shift (days)", "Achieved"]
    _header_row(ws, 4, headers)
    for i, s in enumerate(series, start=5):
        shift = round(s.total_shift_days, 1) if s.total_shift_days is not None else None
        values = [s.key, s.name, _fmt(s.first_value), _fmt(s.last_value),
                  shift, "Yes" if s.is_achieved else "No"]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col == 5 and isinstance(shift, float):
                c.fill = SLIP_FILL if shift > 0 else GAIN_FILL
    _autofit(ws, {1: 14, 2: 55, 3: 14, 4: 14, 5: 17, 6: 10})
    ws.freeze_panes = "A5"

    # Full revision-by-revision detail.
    dws = wb.create_sheet("Revision Detail")
    _title(dws, "Milestone Dates per Revision", 5)
    _header_row(dws, 4, ["Activity ID", "Milestone", "Data Date",
                         "Forecast/Actual Date", "Status"])
    row = 5
    for s in series:
        for p in s.points:
            values = [s.key, s.name, _fmt(p.data_date), _fmt(p.value_date),
                      "Actual" if p.is_actual else "Forecast"]
            for col, v in enumerate(values, start=1):
                c = dws.cell(row=row, column=col, value=v)
                c.border = THIN_BORDER
            row += 1
    _autofit(dws, {1: 14, 2: 55, 3: 12, 4: 18, 5: 10})
    dws.freeze_panes = "A5"

    notes = list(result.warnings)
    notes += [
        f"Possible renamed milestone (unconfirmed): {m.task_code} "
        f"'{m.task_name}' ~ {m.matched_to_key} '{m.matched_to_name}' "
        f"({m.similarity:.0%})"
        for m in result.needs_confirmation
    ]
    _notes_sheet(wb, notes, "Warnings")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 4 — As-planned vs as-recorded
# --------------------------------------------------------------------------- #
def build_variance_xlsx(
    var: VarianceResult, narrative: str | None = None
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Variance"
    _title(ws, f"As-Planned vs As-Recorded — by {var.code_type_name}", 9)

    headers = [var.code_type_name, "Planned Start", "Planned Finish",
               "Recorded Start", "Recorded Finish", "Δ Start (days)",
               "Δ Finish (days)", "Planned Acts", "Recorded Acts"]
    _header_row(ws, 4, headers)
    for i, g in enumerate(var.groups, start=5):
        sd = round(g.start_delta_days, 1) if g.start_delta_days is not None else None
        fd = round(g.finish_delta_days, 1) if g.finish_delta_days is not None else None
        values = [g.code_value, _fmt(g.planned.start), _fmt(g.planned.finish),
                  _fmt(g.recorded.start), _fmt(g.recorded.finish), sd, fd,
                  g.planned.activity_count, g.recorded.activity_count]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col in (6, 7) and isinstance(v, float):
                c.fill = SLIP_FILL if v > 0 else GAIN_FILL
    _autofit(ws, {1: 42, 2: 13, 3: 13, 4: 13, 5: 14, 6: 13, 7: 13, 8: 12, 9: 13})
    ws.freeze_panes = "A5"

    _notes_sheet(wb, var.caveats + var.warnings, "Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 5 — Baseline planned critical path
# --------------------------------------------------------------------------- #
def build_critical_path_xlsx(cp, narrative: str | None = None) -> bytes:
    """cp: CriticalPathResult (imported lazily to avoid a cycle)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Critical Path"
    _title(ws, f"Planned Critical Path — {cp.programme_label}", 8)

    ws.cell(row=3, column=1, value=(
        f"{len(cp.critical)} critical (TF <= {cp.float_tolerance_days:.0f}d), "
        f"{len(cp.near_critical)} near-critical (TF <= "
        f"{cp.near_critical_days:.0f}d); "
        f"{'continuous path' if cp.is_continuous else f'{cp.chain_segments} broken segments'}"
    )).font = Font(italic=True)

    # Band is DATA, not decoration: the membership carried only by the
    # fill colour is unreadable to a filter, a screen reader, or a
    # printed page — the same colour-only defect purged from the gantts
    headers = ["Activity ID", "Activity Name", "Type", "Band",
               "Early Start", "Early Finish", "Duration (d)",
               "Total Float (d)"]
    _header_row(ws, 5, headers)
    row = 6
    for a in cp.activities:
        band = ("Critical" if a.band == "critical" else "Near-critical")
        values = [a.task_code, a.name,
                  "Milestone" if a.is_milestone else "Task", band,
                  _fmt(a.early_start), _fmt(a.early_finish),
                  a.duration_days, a.total_float_days]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = THIN_BORDER
            if col == 8 and isinstance(v, float):
                c.fill = SLIP_FILL if v < 0 else (
                    GAIN_FILL if a.band == "critical" else PatternFill(
                        "solid", fgColor="FFF2CC"))
        row += 1
    _autofit(ws, {1: 16, 2: 55, 3: 11, 4: 13, 5: 12, 6: 12, 7: 12,
                  8: 14})
    ws.freeze_panes = "A6"

    if cp.links:
        lws = wb.create_sheet("Driving Links")
        _title(lws, "Logic Links Between Critical Activities", 4)
        _header_row(lws, 4, ["Predecessor", "Type", "Successor", "Lag (d)"])
        for i, lk in enumerate(cp.links, start=5):
            for col, v in enumerate(
                    [lk.pred_code, lk.link_type, lk.succ_code, lk.lag_days],
                    start=1):
                c = lws.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
        _autofit(lws, {1: 18, 2: 8, 3: 18, 4: 10})
        lws.freeze_panes = "A5"

    _notes_sheet(wb, cp.caveats + cp.warnings, "Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 6 — Revision comparison / change log
# --------------------------------------------------------------------------- #
def build_comparison_xlsx(cmp, narrative: str | None = None,
                          impact=None, attribution=None,
                          provenance=None) -> bytes:
    """cmp: ComparisonResult. When the page ran the impact screening,
    the completion attribution or the provenance timeline, their
    tables ship in the same workbook — the report carries everything
    the page shows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Programme Revision Comparison", 4)
    ws.cell(row=3, column=1, value=(
        f"'{cmp.old_label}' (DD {_fmt(cmp.old_data_date)}, finish "
        f"{_fmt(cmp.old_finish)})  vs  '{cmp.new_label}' "
        f"(DD {_fmt(cmp.new_data_date)}, finish {_fmt(cmp.new_finish)})"
    )).font = Font(italic=True)

    _header_row(ws, 5, ["Change category", "Count"])
    for i, (k, v) in enumerate(cmp.category_counts.items(), start=6):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.border = b.border = THIN_BORDER
        if k.startswith("Actual dates") and v:
            a.fill = b.fill = SLIP_FILL
    _autofit(ws, {1: 42, 2: 10})

    def _acts_sheet(title, refs):
        if not refs:
            return
        s = wb.create_sheet(title)
        _header_row(s, 1, ["Activity ID", "Activity", "Type",
                           "Start", "Finish", "Duration (d)"])
        for i, a in enumerate(refs, start=2):
            vals = [a.task_code, a.name,
                    "Milestone" if a.is_milestone else "Task",
                    _fmt(a.start), _fmt(a.finish),
                    a.duration_days if a.duration_days is not None else "—"]
            for col, v in enumerate(vals, start=1):
                s.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(s, {1: 18, 2: 52, 3: 10, 4: 12, 5: 12, 6: 12})
        s.freeze_panes = "A2"

    _acts_sheet("Added", cmp.added)
    _acts_sheet("Deleted", cmp.deleted)

    def _changes_sheet(title, changes, red_all=False):
        if not changes:
            return
        s = wb.create_sheet(title)
        _header_row(s, 1, ["Activity / Link", "Name", "Was", "Now",
                           "Delta (d)"])
        for i, c in enumerate(changes, start=2):
            vals = [c.task_code, c.name, c.old_value, c.new_value,
                    c.delta_days if c.delta_days is not None else ""]
            for col, v in enumerate(vals, start=1):
                cell = s.cell(row=i, column=col, value=v)
                cell.border = THIN_BORDER
                if red_all or (col == 5 and isinstance(v, (int, float))
                               and v > 0):
                    cell.fill = SLIP_FILL
                elif col == 5 and isinstance(v, (int, float)) and v < 0:
                    cell.fill = GAIN_FILL
        _autofit(s, {1: 26, 2: 46, 3: 26, 4: 30, 5: 10})
        s.freeze_panes = "A2"

    _changes_sheet("Duration Changes", cmp.duration_changes)
    _changes_sheet("Constraint Changes", cmp.constraint_changes)
    _changes_sheet("Calendar Changes", cmp.calendar_changes)
    _changes_sheet("Calendar Definitions", cmp.calendar_def_changes,
                   red_all=True)
    _changes_sheet("Scheduling Options", cmp.sched_options_changes,
                   red_all=True)
    _changes_sheet("Renamed", cmp.renamed)
    _changes_sheet("Lag Changes", cmp.lag_changes)
    _changes_sheet("Actuals Changed", cmp.actual_date_changes, red_all=True)

    def _logic_sheet(title, links):
        if not links:
            return
        s = wb.create_sheet(title)
        _header_row(s, 1, ["Predecessor", "Pred name", "Type",
                           "Successor", "Succ name", "Lag (d)"])
        for i, lk in enumerate(links, start=2):
            vals = [lk.pred_code, lk.pred_name, lk.link_type,
                    lk.succ_code, lk.succ_name, lk.lag_days]
            for col, v in enumerate(vals, start=1):
                s.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(s, {1: 18, 2: 40, 3: 7, 4: 18, 5: 40, 6: 9})
        s.freeze_panes = "A2"

    _logic_sheet("Logic Added", cmp.logic_added)
    _logic_sheet("Logic Removed", cmp.logic_removed)

    if impact is not None and impact.ranked:
        s = wb.create_sheet("Materiality Rank")
        _header_row(s, 1, ["Score", "Band", "Category", "Activity / Link",
                           "Name", "Change", "Delta (d)", "TF now (d)",
                           "Red flag"])
        for i, rc in enumerate(impact.ranked, start=2):
            vals = [rc.score, rc.band, rc.category, rc.ref, rc.name,
                    rc.detail, rc.delta_days, rc.total_float_new,
                    "RED FLAG" if rc.red_flag else ""]
            for col, v in enumerate(vals, start=1):
                c = s.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 9 and rc.red_flag:
                    c.fill = SLIP_FILL
        _autofit(s, {1: 8, 2: 14, 3: 26, 4: 26, 5: 38, 6: 40, 7: 10,
                     8: 10, 9: 11})
        s.freeze_panes = "A2"

    if attribution is not None and attribution.changes:
        s = wb.create_sheet("Completion Attribution")
        s["A1"] = ("One change at a time: the later revision "
                   "re-scheduled with that single change reverted. "
                   "Contribution +ve = the change pushed completion "
                   "later. Kernel-vs-kernel; contributions interact "
                   "and need not sum to the total movement.")
        s["A1"].font = Font(bold=True)
        _header_row(s, 3, ["Category", "Change", "Name", "Detail",
                           "Band", "Completion WITH change",
                           "Completion WITHOUT", "Contribution (d)",
                           "Tested", "Note"])
        for i, a in enumerate(attribution.changes, start=4):
            vals = [a.category, a.ref, a.name, a.detail, a.band,
                    _fmt(a.completion_with), _fmt(a.completion_without),
                    a.contribution_days,
                    "yes" if a.tested else "no", a.note]
            for col, v in enumerate(vals, start=1):
                c = s.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 8 and (a.contribution_days or 0) > 0:
                    c.fill = SLIP_FILL
                elif col == 8 and (a.contribution_days or 0) < 0:
                    c.fill = GAIN_FILL
        _autofit(s, {1: 22, 2: 26, 3: 34, 4: 34, 5: 13, 6: 20, 7: 20,
                     8: 15, 9: 8, 10: 34})
        s.freeze_panes = "A4"

    if provenance is not None and provenance.windows:
        s = wb.create_sheet("Provenance")
        _header_row(s, 1, ["Window", "Completion moved (d)",
                           "Retro actual changes"]
                    + provenance.categories)
        for i, w in enumerate(provenance.windows, start=2):
            vals = [f"{w.old_label} -> {w.new_label}",
                    w.completion_moved_days, w.red_flag_count] + \
                   [w.counts.get(cat, 0) for cat in provenance.categories]
            for col, v in enumerate(vals, start=1):
                c = s.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 3 and w.red_flag_count:
                    c.fill = SLIP_FILL
        _autofit(s, {1: 40, 2: 18, 3: 18})
        s.freeze_panes = "A2"

    _notes_sheet(wb, cmp.warnings + cmp.caveats
                 + (list(impact.caveats) + list(impact.warnings)
                    if impact is not None else [])
                 + (list(attribution.caveats)
                    + list(attribution.warnings)
                    if attribution is not None else []),
                 "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 7 — Windows / period movement
# --------------------------------------------------------------------------- #
def build_windows_xlsx(res, narrative: str | None = None) -> bytes:
    """res: programme.windows.WindowsResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Windows"
    _title(ws, "Windows / Period Movement Analysis", 10)
    if res.total_movement_days is not None:
        ws.cell(row=3, column=1, value=(
            f"Cumulative completion movement: "
            f"{res.total_movement_days:+.0f} days across "
            f"{len(res.windows)} window(s)")).font = Font(italic=True)

    headers = ["#", "From", "To", "Window start", "Window end",
               "Window (d)", "Finish (from)", "Finish (to)",
               "Movement (d)", "Path similarity"]
    _header_row(ws, 5, headers)
    for i, w in enumerate(res.windows, start=6):
        sim = f"{w.cp_similarity:.0%}" if w.cp_similarity is not None else "—"
        vals = [w.index, w.from_label, w.to_label, _fmt(w.start),
                _fmt(w.end), w.window_days, _fmt(w.finish_old),
                _fmt(w.finish_new), w.movement_days, sim]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col == 9 and isinstance(v, (int, float)):
                c.fill = SLIP_FILL if v > 0 else GAIN_FILL
    _autofit(ws, {1: 4, 2: 26, 3: 26, 4: 13, 5: 13, 6: 11, 7: 13, 8: 13,
                  9: 13, 10: 14})
    ws.freeze_panes = "A6"

    shifts = [(w.index, s) for w in res.windows for s in w.shifts]
    if shifts:
        s2 = wb.create_sheet("Path Changes")
        _header_row(s2, 1, ["Window", "Direction", "Activity ID", "Activity"])
        for i, (widx, s) in enumerate(shifts, start=2):
            vals = [widx, s.direction, s.task_code, s.name]
            for col, v in enumerate(vals, start=1):
                c = s2.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 2:
                    c.fill = SLIP_FILL if v == "joined" else GAIN_FILL
        _autofit(s2, {1: 8, 2: 10, 3: 20, 4: 56})
        s2.freeze_panes = "A2"

    # Traceback: the rows behind each window's movement figure —
    # driving-path activities with stored finishes on both sides.
    drivers = [(w.index, d) for w in res.windows for d in w.drivers]
    if drivers:
        s3 = wb.create_sheet("Window Drivers")
        _header_row(s3, 1, ["Window", "Activity ID", "Activity",
                            "Membership", "Finish (from)", "Finish (to)",
                            "Slip (d)", "Basis (to)"])
        for i, (widx, d) in enumerate(drivers, start=2):
            vals = [widx, d.task_code, d.name, d.membership,
                    _fmt(d.finish_old), _fmt(d.finish_new),
                    d.slip_days, d.basis_new]
            for col, v in enumerate(vals, start=1):
                c = s3.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 7 and isinstance(v, (int, float)) and v:
                    c.fill = SLIP_FILL if v > 0 else GAIN_FILL
        _autofit(s3, {1: 8, 2: 20, 3: 46, 4: 11, 5: 13, 6: 13,
                      7: 9, 8: 10})
        s3.freeze_panes = "A2"

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 8 — Progress S-curve
# --------------------------------------------------------------------------- #
def build_progress_xlsx(res, narrative: str | None = None) -> bytes:
    """res: programme.progress.ProgressResult"""
    from .progress import WEIGHT_OPTIONS
    wb = Workbook()
    ws = wb.active
    ws.title = "S-Curve"
    _title(ws, "Progress S-Curve (Planned vs As-Recorded)", 4)
    ws.cell(row=3, column=1, value=(
        f"Weighting: {WEIGHT_OPTIONS.get(res.weight_scheme, res.weight_scheme)}"
        + (f" | Planned {res.planned_pct_at_dd}% vs recorded "
           f"{res.recorded_pct_at_dd}% at latest data date"
           if res.planned_pct_at_dd is not None else "")
        + (f" | time offset {res.time_offset_days:+.0f}d"
           if res.time_offset_days is not None else "")
    )).font = Font(italic=True)

    _header_row(ws, 5, ["Month end", "Planned cum %", "Recorded cum %"])
    rec = {p.date.strftime("%Y-%m"): p.cum_pct for p in res.recorded_curve}
    pla = {p.date.strftime("%Y-%m"): p.cum_pct for p in res.planned_curve}
    months = sorted(set(rec) | set(pla))
    prev_p = prev_r = None
    for i, m in enumerate(months, start=6):
        p = pla.get(m, prev_p if prev_p is not None else None)
        r = rec.get(m, prev_r if prev_r is not None else None)
        prev_p, prev_r = p, r
        vals = [m, round(p, 1) if p is not None else "",
                round(r, 1) if r is not None else ""]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(ws, {1: 12, 2: 14, 3: 15})
    ws.freeze_panes = "A6"

    if res.revision_points:
        s2 = wb.create_sheet("Revision Points")
        _header_row(s2, 1, ["Revision", "Data date", "Recorded %",
                            "Planned %", "Gap (pts)"])
        for i, rp in enumerate(res.revision_points, start=2):
            gap = (round(rp.planned_pct - rp.recorded_pct, 1)
                   if rp.planned_pct is not None
                   and rp.recorded_pct is not None else "")
            vals = [rp.label, _fmt(rp.data_date), rp.recorded_pct,
                    rp.planned_pct, gap]
            for col, v in enumerate(vals, start=1):
                c = s2.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 5 and isinstance(v, (int, float)):
                    c.fill = SLIP_FILL if v > 0 else GAIN_FILL
        _autofit(s2, {1: 30, 2: 12, 3: 12, 4: 12, 5: 10})

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 9 — Float erosion
# --------------------------------------------------------------------------- #
def build_float_erosion_xlsx(res, narrative: str | None = None) -> bytes:
    """res: programme.float_erosion.FloatErosionResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Float Profile"
    _title(ws, "Float Erosion Review", 8)
    _header_row(ws, 4, ["Revision", "Data date", "Incomplete",
                        "Median TF (d)", "Min TF (d)", "Critical (TF<=0)",
                        "Negative", f"Near (<= {res.near_days:.0f}d)"])
    for i, s in enumerate(res.snapshots, start=5):
        vals = [s.label, _fmt(s.data_date), s.incomplete_count,
                s.median_float, s.min_float, s.critical_count,
                s.negative_count, s.near_count]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col == 7 and isinstance(v, int) and v > 0:
                c.fill = SLIP_FILL
    _autofit(ws, {1: 28, 2: 12, 3: 11, 4: 13, 5: 11, 6: 15, 7: 10, 8: 13})
    ws.freeze_panes = "A5"

    deltas = [(w, d, "eroded") for w in res.windows for d in w.top_eroders] \
        + [(w, d, "gained") for w in res.windows for d in w.top_gainers]
    if deltas:
        s2 = wb.create_sheet("Top Movers")
        _header_row(s2, 1, ["Window", "Direction", "Activity ID", "Activity",
                            "TF was (d)", "TF now (d)", "Delta (d)"])
        for i, (w, d, direction) in enumerate(deltas, start=2):
            vals = [f"{w.from_label} -> {w.to_label}", direction,
                    d.task_code, d.name, d.old_tf, d.new_tf,
                    round(d.delta, 1)]
            for col, v in enumerate(vals, start=1):
                c = s2.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 7:
                    c.fill = SLIP_FILL if d.delta < 0 else GAIN_FILL
        _autofit(s2, {1: 34, 2: 10, 3: 18, 4: 46, 5: 11, 6: 11, 7: 10})
        s2.freeze_panes = "A2"

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 10 — Planned resource loading
# --------------------------------------------------------------------------- #
def build_resources_xlsx(res, narrative: str | None = None) -> bytes:
    """res: programme.resources.ResourceLoadingResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resources"
    _title(ws, "Planned Resource Loading", 5)
    ws.cell(row=3, column=1,
            value=f"Programme: {res.programme_label} — PLANNED loading, "
                  "not actual expenditure").font = Font(italic=True)
    _header_row(ws, 5, ["Resource", "Name", "Type", "Total planned qty",
                        "Assignments"])
    for i, r in enumerate(res.resources, start=6):
        vals = [r.short_name, r.name, r.rsrc_type,
                round(r.total_qty, 1), r.assignment_count]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(ws, {1: 14, 2: 34, 3: 12, 4: 17, 5: 12})
    ws.freeze_panes = "A6"

    if res.histogram:
        s2 = wb.create_sheet("Monthly Loading")
        months = sorted({p.month_end for p in res.histogram})
        names = [r.short_name for r in res.resources]
        grid = {(p.resource, p.month_end): p.qty for p in res.histogram}
        _header_row(s2, 1, ["Month"] + names)
        for i, m in enumerate(months, start=2):
            s2.cell(row=i, column=1, value=f"{m:%Y-%m}").border = THIN_BORDER
            for j, n in enumerate(names, start=2):
                q = grid.get((n, m))
                c = s2.cell(row=i, column=j,
                            value=round(q, 1) if q else None)
                c.border = THIN_BORDER
        _autofit(s2, {1: 10, **{j: 13 for j in range(2, len(names) + 2)}})
        s2.freeze_panes = "B2"

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 12 — As-built critical path
# --------------------------------------------------------------------------- #
def build_asbuilt_xlsx(trace, narrative: str | None = None,
                       roll=None, links=None,
                       gantt_png: bytes | None = None) -> bytes:
    """trace: ActualTraceResult; roll: RollupResult; links: umbrella
    links; gantt_png: the path gantt as a figure sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "As-Built Path"
    _title(ws, "As-Built Critical Path — backward trace", 6)
    ws.cell(row=3, column=1, value=(
        f"Traced back from {trace.terminal_code} · "
        f"{len(trace.activities)} activities · "
        f"{trace.asbuilt_count} as-built, {trace.in_progress_count} in "
        f"progress, {trace.forecast_count} forecast · data date "
        f"{_fmt(trace.data_date)}")).font = Font(italic=True)
    if trace.hybrid:
        c = ws.cell(row=4, column=1, value=(
            "HYBRID PATH — the milestone was not achieved; rows marked "
            "'forecast' are the programme's remaining early dates, not a "
            "record of what happened."))
        c.font = Font(bold=True, color="9B3227")

    _header_row(ws, 6, ["#", "Basis", "Activity ID", "Activity",
                        "Start", "Finish"])
    for i, a in enumerate(trace.activities, start=1):
        vals = [i, a.basis, a.task_code, a.name,
                _fmt(a.act_start), _fmt(a.act_finish)]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i + 6, column=col, value=v)
            c.border = THIN_BORDER
            if col == 2 and v == "as-built":
                c.fill = GAIN_FILL
            elif col == 2 and v == "forecast":
                c.fill = SLIP_FILL
    _autofit(ws, {1: 5, 2: 13, 3: 20, 4: 52, 5: 13, 6: 13})
    ws.freeze_panes = "A7"

    if trace.links:
        s2 = wb.create_sheet("Hand-Offs")
        _header_row(s2, 1, ["Predecessor", "Pred name", "Kind",
                            "Successor", "Succ name", "Gap (d)",
                            "Programmed logic", "Confidence"])
        for i, lk in enumerate(trace.links, start=2):
            vals = [lk.pred_code, lk.pred_name, lk.kind, lk.succ_code,
                    lk.succ_name, lk.gap_days,
                    "Yes" if lk.had_logic else "SEQUENCE ONLY", lk.score]
            for col, v in enumerate(vals, start=1):
                c = s2.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 7:
                    c.fill = GAIN_FILL if lk.had_logic else SLIP_FILL
        _autofit(s2, {1: 18, 2: 38, 3: 12, 4: 18, 5: 38, 6: 9, 7: 17,
                      8: 11})
        s2.freeze_panes = "A2"

    if roll is not None and roll.umbrellas:
        s3 = wb.create_sheet("Work Packages")
        s3["A1"] = ("Umbrella dates are measured from CRITICAL-PATH "
                    "members only — grouping is presentation and does "
                    "not move the measured delay.")
        s3["A1"].font = Font(bold=True)
        _header_row(s3, 3, ["Work package", "Members", "On CP",
                            "Measured start", "Measured finish",
                            "Driving member", "Full group runs on (d)",
                            "Measured?"])
        for i, u in enumerate(roll.umbrellas, start=4):
            vals = [u.name, u.member_count, u.on_path_count,
                    _fmt(u.actual_start), _fmt(u.actual_finish),
                    u.driving_member or "", u.presentation_only_days,
                    "yes" if u.measured else "PRESENTATION ONLY"]
            for col, v in enumerate(vals, start=1):
                c = s3.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 8:
                    c.fill = GAIN_FILL if u.measured else SLIP_FILL
        _autofit(s3, {1: 34, 2: 9, 3: 7, 4: 15, 5: 15, 6: 20, 7: 20,
                      8: 18})
        s3.freeze_panes = "A4"

        members = roll.member_rows()
        if members:
            s4 = wb.create_sheet("Work Package Members")
            _header_row(s4, 1, ["Work package", "Activity ID", "Activity",
                                "On critical path", "Actual start",
                                "Actual finish", "Drives finish"])
            for i, m in enumerate(members, start=2):
                vals = [m["umbrella"], m["task_code"], m["name"],
                        m["on_critical_path"], _fmt(m["actual_start"]),
                        _fmt(m["actual_finish"]),
                        m["drives_umbrella_finish"]]
                for col, v in enumerate(vals, start=1):
                    c = s4.cell(row=i, column=col, value=v)
                    c.border = THIN_BORDER
                    if col == 4:
                        c.fill = GAIN_FILL if v == "yes" else SLIP_FILL
            _autofit(s4, {1: 30, 2: 20, 3: 44, 4: 17, 5: 13, 6: 13,
                          7: 13})
            s4.freeze_panes = "A2"

    if links:
        s5 = wb.create_sheet("Package Links")
        _header_row(s5, 1, ["From", "To", "Basis", "Hand-offs",
                            "Logic-evidenced", "Sequence only",
                            "Underlying activity hand-offs"])
        for i, r in enumerate(links, start=2):
            vals = [r["from"], r["to"], r["basis"], r["hand_off_count"],
                    r["logic_evidenced"], r["sequence_only"],
                    "; ".join(r["hand_offs"][:8])]
            for col, v in enumerate(vals, start=1):
                c = s5.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                if col == 3 and v == "logic":
                    c.fill = GAIN_FILL
                elif col == 3 and v == "sequence only":
                    c.fill = SLIP_FILL
        _autofit(s5, {1: 30, 2: 30, 3: 14, 4: 10, 5: 16, 6: 14, 7: 60})
        s5.freeze_panes = "A2"

    if gantt_png:
        _image_sheet(wb, "Gantt — as-built critical path", gantt_png)
    _notes_sheet(wb, list(trace.warnings) + list(trace.caveats)
                 + (list(roll.caveats) + list(roll.warnings)
                    if roll is not None else []),
                 "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 13 — Sequence coding
# --------------------------------------------------------------------------- #
def build_sequence_xlsx(seq, mapping_rows, narrative: str | None = None) -> bytes:
    """seq: SequenceResult; mapping_rows: list[MappingRow] (disclosed)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sequence Bands"
    _title(ws, "Construction Sequence by Work Front (Actual Dates)", 6)
    ws.cell(row=3, column=1, value=(
        "Mapping " + ("CONFIRMED by analyst"
                      if seq.mapping_confirmed else
                      "AUTO-PROPOSED (not analyst-confirmed)")
        + f" — {seq.mapped_activities} actualised activities mapped"
    )).font = Font(italic=True)
    _header_row(ws, 5, ["Work front", "Stage", "Activities", "Complete",
                        "Actual start", "Actual finish"])
    r = 6
    for b in sorted(seq.bands,
                    key=lambda b: (b.front,
                                   seq.stage_order.index(b.stage)
                                   if b.stage in seq.stage_order else 99)):
        vals = [b.front, b.stage, b.activity_count, b.complete_count,
                _fmt(b.act_start), _fmt(b.act_finish)]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=col, value=v).border = THIN_BORDER
        r += 1
    _autofit(ws, {1: 26, 2: 30, 3: 11, 4: 10, 5: 12, 6: 13})
    ws.freeze_panes = "A6"

    s2 = wb.create_sheet("Fronts by Finish")
    _header_row(s2, 1, ["Work front", "Last recorded finish"])
    for i, (f, fin) in enumerate(seq.fronts_by_finish, start=2):
        s2.cell(row=i, column=1, value=f).border = THIN_BORDER
        s2.cell(row=i, column=2, value=_fmt(fin)).border = THIN_BORDER
    _autofit(s2, {1: 30, 2: 20})

    # The disclosed mapping — the Basis-of-Analysis artefact.
    s3 = wb.create_sheet("Mapping (Disclosed)")
    _header_row(s3, 1, ["Activity ID", "Activity", "Work front",
                        "Front evidence", "Stage", "Stage evidence"])
    for i, m in enumerate(mapping_rows, start=2):
        vals = [m.task_code, m.name, m.front, m.front_evidence,
                m.stage, m.stage_evidence]
        for col, v in enumerate(vals, start=1):
            s3.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(s3, {1: 18, 2: 46, 3: 22, 4: 20, 5: 30, 6: 22})
    s3.freeze_panes = "A2"

    _notes_sheet(wb, seq.warnings + seq.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 14 — Hierarchy rebuild (outline gantt table)
# --------------------------------------------------------------------------- #
def build_hierarchy_xlsx(h, narrative: str | None = None) -> bytes:
    """h: programme.hierarchy.HierarchyResult.

    Sheet 1 mirrors the viewer: indented groups with rollup dates and
    Excel's own collapsible +/- row outlines; activities beneath. Sheet 2
    is a flat table (one dimension column per level) for pivoting.
    """
    from .hierarchy import UNASSIGNED  # noqa: F401  (documented grouping)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchy"
    _title(ws, "Rebuilt Programme Hierarchy", 7)
    ws.cell(row=3, column=1, value=(
        f"Programme: {h.programme_label} — levels: "
        + " › ".join(h.dimension_labels)
        + f" | {h.placed_activities} activities, validation "
        + ("complete" if h.is_complete else "FAILED"))).font = Font(
        italic=True)

    headers = ["Hierarchy / Activity", "Activity ID", "Start", "Finish",
               "Activities", "Complete", "Status"]
    _header_row(ws, 5, headers)
    ws.sheet_properties.outlinePr.summaryBelow = False

    GROUP_FILLS = [PatternFill("solid", fgColor=c)
                   for c in ("D9E2F3", "E8EEF9", "F3F7FC", "FAFCFE")]
    row_idx = 6

    def walk(node, depth):
        nonlocal row_idx
        for child in sorted(node.children.values(),
                            key=lambda c: (c.start or datetime.max, c.name)):
            vals = ["    " * depth + child.name, "",
                    _fmt(child.start), _fmt(child.finish),
                    child.activity_count, child.complete_count, ""]
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=row_idx, column=col, value=v)
                c.border = THIN_BORDER
                c.font = Font(bold=True, color="1F3864")
                c.fill = GROUP_FILLS[min(depth, len(GROUP_FILLS) - 1)]
            if depth:
                ws.row_dimensions[row_idx].outline_level = min(depth, 7)
            row_idx += 1
            walk(child, depth + 1)
            for a in sorted(child.activities,
                            key=lambda a: (a.start or datetime.max,
                                           a.task_code)):
                vals = ["    " * (depth + 1)
                        + ("◆ " if a.is_milestone else "") + a.name,
                        a.task_code, _fmt(a.start), _fmt(a.finish),
                        "", "", a.status]
                for col, v in enumerate(vals, start=1):
                    c = ws.cell(row=row_idx, column=col, value=v)
                    c.border = THIN_BORDER
                    if col == 7:
                        c.fill = (GAIN_FILL if a.status == "complete"
                                  else SLIP_FILL
                                  if a.status == "in progress" else None
                                  ) or PatternFill()
                ws.row_dimensions[row_idx].outline_level = min(depth + 1, 7)
                row_idx += 1

    walk(h.root, 0)
    _autofit(ws, {1: 64, 2: 18, 3: 11, 4: 11, 5: 10, 6: 10, 7: 12})
    ws.freeze_panes = "A6"

    # --- flat sheet for pivoting ------------------------------------------
    s2 = wb.create_sheet("Flat Table")
    _header_row(s2, 1, list(h.dimension_labels)
                + ["Activity ID", "Activity", "Start", "Finish",
                   "Milestone", "Status"])
    r2 = 2

    def flat(node, path):
        nonlocal r2
        for child in node.children.values():
            flat(child, path + [child.name])
            for a in child.activities:
                vals = (path + [child.name]
                        + [a.task_code, a.name, _fmt(a.start),
                           _fmt(a.finish),
                           "Yes" if a.is_milestone else "",
                           a.status])
                for col, v in enumerate(vals, start=1):
                    s2.cell(row=r2, column=col, value=v).border = THIN_BORDER
                r2 += 1

    flat(h.root, [])
    _autofit(s2, {i: 24 for i in range(1, len(h.dimension_labels) + 1)}
             | {len(h.dimension_labels) + 1: 18,
                len(h.dimension_labels) + 2: 48,
                len(h.dimension_labels) + 3: 11,
                len(h.dimension_labels) + 4: 11,
                len(h.dimension_labels) + 5: 10,
                len(h.dimension_labels) + 6: 12})
    s2.freeze_panes = "A2"

    _notes_sheet(wb, h.warnings + h.caveats, "Validation & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Module 15 — Time Impact Analysis
# --------------------------------------------------------------------------- #
def build_tia_xlsx(res, narrative: str | None = None,
                   audit: dict | None = None,
                   run_history: list[dict] | None = None) -> bytes:
    """res: programme.tia.TIAResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact"
    _title(ws, "Time Impact Analysis", 5)
    e = res.event
    ws.cell(row=3, column=1, value=(
        f"Event {e.event_id}: {e.title} | programme "
        f"{res.programme_label} (DD {_fmt(res.data_date)}) | completion "
        f"{_fmt(res.completion_pre)} -> {_fmt(res.completion_post)} "
        + (f"({res.completion_delta_days:+.1f}d)"
           if res.completion_delta_days is not None else "")
    )).font = Font(italic=True)
    _header_row(ws, 5, ["Milestone", "Name", "Pre-impact", "Post-impact",
                        "Delta (d)"])
    for i, m in enumerate(res.milestone_impacts, start=6):
        vals = [m.code, m.name, _fmt(m.pre), _fmt(m.post), m.delta_days]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col == 5 and isinstance(v, (int, float)):
                c.fill = SLIP_FILL if v > 0 else GAIN_FILL
    _autofit(ws, {1: 16, 2: 46, 3: 12, 4: 12, 5: 10})
    ws.freeze_panes = "A6"

    s2 = wb.create_sheet("Fragnet")
    _header_row(s2, 1, ["ID", "Activity", "Duration (d)", "Calendar", "Predecessors",
                        "Successors", "Source / rationale", "Assumptions",
                        "Confidence"])
    from .tia import links_to_text
    for i, f in enumerate(res.fragnet, start=2):
        vals = [f.act_id, f.name, f.duration_days, f.calendar_id,
                links_to_text(f.predecessors), links_to_text(f.successors),
                f.rationale, f.assumptions, f.confidence]
        for col, v in enumerate(vals, start=1):
            s2.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(s2, {1: 12, 2: 38, 3: 12, 4: 16, 5: 26, 6: 26, 7: 34, 8: 30, 9: 11})
    s2.freeze_panes = "A2"

    s3 = wb.create_sheet("Event")
    s3.column_dimensions["A"].width = 18
    s3.column_dimensions["B"].width = 90
    rows = [("Event ID", e.event_id), ("Title", e.title),
            ("Description", e.description),
            ("Date raised", _fmt(e.date_raised)),
            ("Responsibility (asserted)", e.responsibility_asserted),
            ("Evidence noted", e.evidence_note), ("Area / system", e.area),
            ("Discipline", e.discipline),
            ("Project context", e.project_context),
            ("Construction work package", e.work_package)]
    for i, (k, v) in enumerate(rows, start=1):
        s3.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = s3.cell(row=i, column=2, value=v)
        c.alignment = WRAP

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    if audit:
        audit_ws = wb.create_sheet("Audit Trail")
        _header_row(audit_ws, 1, ["Item", "Value"])
        for row_no, (key, value) in enumerate(audit.items(), start=2):
            audit_ws.cell(row=row_no, column=1,
                          value=key.replace("_", " ").title()).border = THIN_BORDER
            audit_ws.cell(row=row_no, column=2,
                          value=str(value)).border = THIN_BORDER
        _autofit(audit_ws, {1: 34, 2: 90})
    if run_history:
        history_ws = wb.create_sheet("Run History")
        keys = list(dict.fromkeys(
            key for run in run_history for key in run.keys()))
        _header_row(history_ws, 1, keys)
        for row_no, run in enumerate(run_history, start=2):
            for col_no, key in enumerate(keys, start=1):
                history_ws.cell(row=row_no, column=col_no,
                                value=str(run.get(key, ""))).border = THIN_BORDER
        _autofit(history_ws, {i: 22 for i in range(1, len(keys) + 1)})
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)

# --------------------------------------------------------------------------- #
# Explain This Delay
# --------------------------------------------------------------------------- #
def build_explain_xlsx(res, narrative: str | None = None,
                       confirmed: list[dict] | None = None) -> bytes:
    """res: programme.explain.ExplainResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movement (Facts)"
    _title(ws, f"Explain This Delay — {res.target_code}", 4)
    ws.cell(row=3, column=1, value=(
        f"{res.target_name} | total movement "
        + (f"{res.total_movement_days:+.0f} days"
           if res.total_movement_days is not None else "n/a")
        + (" | ACHIEVED" if res.achieved else "")
    )).font = Font(italic=True)
    _header_row(ws, 5, ["Revision", "Data date", "Forecast / Actual",
                        "Kind"])
    for i, p in enumerate(res.points, start=6):
        vals = [p.label, _fmt(p.data_date), _fmt(p.forecast),
                "ACTUAL" if p.is_actual else "forecast"]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(ws, {1: 30, 2: 12, 3: 16, 4: 10})

    s2 = wb.create_sheet("Drivers (Inference)")
    _header_row(s2, 1, ["Window", "Movement (d)", "Path similarity",
                        "Attribution", "Direction", "Activity ID",
                        "Activity"])
    r = 2
    for w in res.windows:
        rel = "reliable" if w.attribution_reliable else "UNCERTAIN"
        sim = (f"{w.path_similarity:.0f}%"
               if w.path_similarity is not None else "n/a")
        rows = w.shifts or [None]
        for s in rows:
            vals = [f"W{w.index}: {w.from_label} -> {w.to_label}",
                    w.movement_days, sim, rel,
                    s.direction if s else "", s.task_code if s else "",
                    s.name if s else ""]
            for col, v in enumerate(vals, start=1):
                c = s2.cell(row=r, column=col, value=v)
                c.border = THIN_BORDER
                if col == 4 and not w.attribution_reliable:
                    c.fill = SLIP_FILL
            r += 1
    _autofit(s2, {1: 34, 2: 12, 3: 13, 4: 12, 5: 10, 6: 18, 7: 44})
    s2.freeze_panes = "A2"

    if confirmed:
        s3 = wb.create_sheet("Confirmed Drivers")
        s3["A1"] = ("Drivers PROMOTED from candidate to confirmed by the "
                    "analyst, with the evidence relied on. Everything "
                    "not listed here remains an inferred candidate.")
        s3["A1"].font = Font(italic=True)
        _header_row(s3, 3, ["Window", "Direction", "Activity ID",
                            "Activity", "Evidence relied on"])
        for i, row in enumerate(confirmed, start=4):
            vals = [row.get("window", ""), row.get("direction", ""),
                    row.get("task_code", ""), row.get("name", ""),
                    row.get("note", "") or "(no evidence note recorded)"]
            for col, v in enumerate(vals, start=1):
                c = s3.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
                c.alignment = WRAP
                if col == 5 and not row.get("note"):
                    c.fill = SLIP_FILL
        _autofit(s3, {1: 34, 2: 10, 3: 18, 4: 40, 5: 56})
        s3.freeze_panes = "A4"

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)

    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 6b — comparison impact & materiality
# --------------------------------------------------------------------------- #

def build_impact_xlsx(imp, narrative: str | None = None) -> bytes:
    """imp: programme.comparison_impact.ComparisonImpact"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Comparison Impact & Materiality Screening", 4)
    ws.cell(row=3, column=1, value=(
        f"'{imp.old_label}' vs '{imp.new_label}' — completion moved "
        + (f"{imp.completion_moved_days:+.0f} calendar days"
           if imp.completion_moved_days is not None else "n/a")
        + f"; trace terminals {imp.end_old or '—'} / {imp.end_new or '—'}"
    )).font = Font(italic=True)
    _header_row(ws, 5, ["Path position", "Changes"])
    for i, (band, n) in enumerate(sorted(imp.band_counts.items()), start=6):
        a = ws.cell(row=i, column=1, value=band)
        b = ws.cell(row=i, column=2, value=n)
        a.border = b.border = THIN_BORDER
        if band == "critical" and n:
            a.fill = b.fill = SLIP_FILL
    _autofit(ws, {1: 24, 2: 12})

    s = wb.create_sheet("Materiality rank")
    _header_row(s, 1, ["Score", "Path position", "Category",
                       "Activity / Link", "Name", "Change", "Delta (d)",
                       "TF now (d)", "Red flag"])
    for i, c in enumerate(imp.ranked, start=2):
        vals = [c.score, c.band, c.category, c.ref, c.name, c.detail,
                c.delta_days if c.delta_days is not None else "",
                c.total_float_new if c.total_float_new is not None else "",
                "YES" if c.red_flag else ""]
        for col, v in enumerate(vals, start=1):
            cell = s.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER
            if c.red_flag and col == 9:
                cell.fill = SLIP_FILL
    _autofit(s, {1: 8, 2: 14, 3: 34, 4: 26, 5: 40, 6: 44, 7: 10, 8: 10,
                 9: 9})
    s.freeze_panes = "A2"

    _notes_sheet(wb, imp.warnings + imp.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 17 — progress transfer
# --------------------------------------------------------------------------- #

def build_transfer_xlsx(tr, narrative: str | None = None) -> bytes:
    """tr: programme.progress_transfer.ProgressTransferResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Progress Transfer — Network vs Progress Decomposition", 4)
    ws.cell(row=3, column=1, value=(
        f"Network donor '{tr.network_label}' · progress donor "
        f"'{tr.progress_label}' · data date {_fmt(tr.data_date)}"
    )).font = Font(italic=True)
    rows = [
        ("Progress transferred — in-progress starts", tr.applied_starts),
        ("Progress transferred — completed activities",
         tr.applied_finishes),
        ("Network activities with no progress match",
         tr.not_in_progress_file),
        ("Actualised activities not in the network donor",
         len(tr.unmatched_progress)),
        ("Reference forecast (progress donor, own network)",
         _fmt(tr.completion_reference)),
        ("Shared-activity forecast (donor network)",
         _fmt(tr.completion_logic_only)),
        ("Full transferred forecast (incl. unmatched scope)",
         _fmt(tr.completion_transferred)),
        ("LOGIC/DURATION EFFECT (days, scope excluded)",
         tr.network_effect_days),
        ("SCOPE EFFECT (days, unmatched activities)",
         tr.scope_effect_days),
        ("Calibration vs P6 own forecast (days)", tr.calibration_days),
    ]
    _header_row(ws, 5, ["Measure", "Value"])
    for i, (k, v) in enumerate(rows, start=6):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2,
                    value=v if v is not None else "n/a")
        a.border = b.border = THIN_BORDER
        if k.startswith(("LOGIC", "SCOPE")):
            a.font = Font(bold=True)
    _autofit(ws, {1: 46, 2: 22})

    if tr.milestones:
        s = wb.create_sheet("Milestones")
        _header_row(s, 1, ["Milestone", "Name", "Transferred",
                           "Reference", "Delta (d)"])
        for i, m in enumerate(tr.milestones, start=2):
            vals = [m.code, m.name, _fmt(m.transferred),
                    _fmt(m.reference),
                    m.delta_days if m.delta_days is not None else ""]
            for col, v in enumerate(vals, start=1):
                s.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(s, {1: 18, 2: 46, 3: 13, 4: 13, 5: 10})
        s.freeze_panes = "A2"

    if tr.driving_chain:
        s2 = wb.create_sheet("Driving chain")
        _header_row(s2, 1, ["Activity", "Name", "Start", "Finish"])
        for i, step in enumerate(tr.driving_chain, start=2):
            vals = [step["id"], step["name"], _fmt(step["start"]),
                    _fmt(step["finish"])]
            for col, v in enumerate(vals, start=1):
                s2.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(s2, {1: 18, 2: 52, 3: 13, 4: 13})
        s2.freeze_panes = "A2"

    _notes_sheet(wb, tr.warnings + tr.caveats,
                 "Statusing & Caveats")
    _narrative_sheet(wb, narrative)
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Project library — chain-of-custody register export
# --------------------------------------------------------------------------- #

def build_custody_xlsx(records) -> bytes:
    """records: list[programme.store.FileRecord]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Custody register"
    _title(ws, "Programme File Custody Register", 7)
    _header_row(ws, 4, ["Registered (UTC)", "Project", "File",
                        "Data date", "Activities", "Size (bytes)",
                        "SHA-256"])
    for i, r in enumerate(records, start=5):
        vals = [r.added_utc, r.project, r.file_name,
                r.data_date or "—",
                r.activity_count if r.activity_count is not None else "",
                r.size_bytes, r.sha256]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(ws, {1: 22, 2: 20, 3: 28, 4: 12, 5: 10, 6: 12, 7: 66})
    ws.freeze_panes = "A5"
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 18 — out-of-sequence screening & as-built repair
# --------------------------------------------------------------------------- #

def build_oos_xlsx(label, flags, plan, report=None,
                   evolution=None) -> bytes:
    """OOS module workbook: flags, repair plan/register, evolution.

    flags: list[programme.oos.OutOfSequenceFlag]
    plan:  list[programme.oos.RepairItem]
    report: programme.oos.RepairReport | None (after an export)
    evolution: programme.oos.OOSEvolution | None
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Out-of-Sequence Screening & As-Built Repair", 4)
    concrete = sum(1 for f in flags if f.rec_link_type
                   not in ("", "review"))
    rows = [
        ("Programme", label),
        ("Out-of-sequence records", len(flags)),
        ("Concrete as-built fits", concrete),
        ("Review-class flags (never auto-applied)",
         len(flags) - concrete),
        ("Repair plan items", len(plan)),
        ("Blocked (would duplicate an existing link)",
         sum(1 for r in plan if r.blocked)),
    ]
    if report is not None:
        rows += [
            ("Repairs applied to the exported copy",
             len(report.applied)),
            ("Round-trip QA", "PASSED" if report.qa_passed else "FAILED"),
            ("Source SHA-256", report.source_sha256),
            ("Repaired-copy SHA-256", report.output_sha256),
        ]
    _header_row(ws, 3, ["Measure", "Value"])
    for i, (k, v) in enumerate(rows, start=4):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v)
        a.border = b.border = THIN_BORDER
    _autofit(ws, {1: 44, 2: 70})

    s = wb.create_sheet("Flags")
    _header_row(s, 1, ["Predecessor", "Pred name", "Link", "Successor",
                       "Succ name", "Overlap (d)", "As-built fix",
                       "Fix basis (analyst to confirm)", "Detail"])
    for i, f in enumerate(flags[:2000], start=2):
        vals = [f.pred_code, f.pred_name, f.link_type, f.succ_code,
                f.succ_name,
                f.overlap_days if f.overlap_days is not None else "",
                f.rec_link, f.rec_basis, f.detail]
        for col, v in enumerate(vals, start=1):
            cell = s.cell(row=i, column=col, value=v)
            cell.border = THIN_BORDER
            if col == 7 and f.rec_link_type not in ("", "review"):
                cell.fill = GAIN_FILL
    _autofit(s, {1: 18, 2: 30, 3: 6, 4: 18, 5: 30, 6: 10, 7: 16, 8: 58,
                 9: 52})
    s.freeze_panes = "A2"

    if plan:
        sp = wb.create_sheet("Repair register")
        _header_row(sp, 1, ["Predecessor", "Successor", "Old link",
                            "New link", "Lag (cal. d)", "Lag (hr)",
                            "Applied", "Blocked", "Basis"])
        applied_keys = ({(r.pred_code, r.succ_code)
                         for r in report.applied}
                        if report is not None else set())
        for i, r in enumerate(plan, start=2):
            applied = ("YES" if (r.pred_code, r.succ_code) in applied_keys
                       else ("" if report is not None
                             else ("planned" if r.apply and not r.blocked
                                   else "")))
            vals = [r.pred_code, r.succ_code, r.old_link,
                    f"{r.new_type.replace('PR_', '')} "
                    f"{r.new_lag_days_cal:+.0f}d",
                    r.new_lag_days_cal, r.new_lag_hr, applied,
                    r.blocked, r.basis]
            for col, v in enumerate(vals, start=1):
                cell = sp.cell(row=i, column=col, value=v)
                cell.border = THIN_BORDER
                if col == 8 and r.blocked:
                    cell.fill = SLIP_FILL
        _autofit(sp, {1: 18, 2: 18, 3: 12, 4: 12, 5: 11, 6: 9, 7: 9,
                      8: 40, 9: 58})
        sp.freeze_panes = "A2"

    if evolution is not None and evolution.windows:
        se = wb.create_sheet("Evolution")
        _header_row(se, 1, ["Window", "New OOS", "Resolved",
                            "Total after"])
        for i, w in enumerate(evolution.windows, start=2):
            vals = [f"{w.from_label} -> {w.to_label}", len(w.new_flags),
                    w.resolved_count, w.total_after]
            for col, v in enumerate(vals, start=1):
                se.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(se, {1: 52, 2: 10, 3: 10, 4: 12})

    from .oos import OOS_CAVEATS, REPAIR_CAVEATS
    notes = (report.qa_notes if report is not None else [])
    _notes_sheet(wb, notes + REPAIR_CAVEATS + OOS_CAVEATS,
                 "QA & Caveats")
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 19 — concurrency screening
# --------------------------------------------------------------------------- #

def build_concurrency_xlsx(res, narrative: str | None = None) -> bytes:
    """res: programme.concurrency.ConcurrencyResult"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Matrix"
    _title(ws, "Concurrent-Delay Screening (time-overlap, per window)", 9)
    _header_row(ws, 4, ["Window", "Period", "Movement (d)",
                        "Employer days", "Contractor days", "Both (d)",
                        "Unclassified (d)", "Concurrent candidate",
                        "Pacing flag"])
    for i, w in enumerate(res.windows, start=5):
        period = (f"{w.start:%Y-%m-%d} -> {w.end:%Y-%m-%d}"
                  if w.start and w.end else "-")
        vals = [f"W{w.index}: {w.from_label} -> {w.to_label}", period,
                w.movement_days, w.employer_days, w.contractor_days,
                w.both_days, w.unclassified_days,
                "YES" if w.concurrent_candidate else "",
                "enquire" if w.pacing_flag else ""]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN_BORDER
            if col == 8 and w.concurrent_candidate:
                c.fill = SLIP_FILL
    _autofit(ws, {1: 34, 2: 24, 3: 12, 4: 13, 5: 14, 6: 9, 7: 14,
                  8: 18, 9: 11})
    ws.freeze_panes = "A5"

    s = wb.create_sheet("Events Screened")
    _header_row(s, 1, ["Event", "Title", "Asserted responsibility",
                       "Screened party", "From", "To", "Duration (d)",
                       "Note"])
    for i, e in enumerate(res.events, start=2):
        vals = [e.event_id, e.title, e.asserted, e.party,
                f"{e.start:%Y-%m-%d}", f"{e.end:%Y-%m-%d}",
                e.duration_days,
                "no fragnet - screened as a single day"
                if e.single_day else ""]
        for col, v in enumerate(vals, start=1):
            s.cell(row=i, column=col, value=v).border = THIN_BORDER
    _autofit(s, {1: 14, 2: 40, 3: 22, 4: 14, 5: 12, 6: 12, 7: 12, 8: 34})
    s.freeze_panes = "A2"

    _notes_sheet(wb, res.warnings + res.caveats, "Warnings & Caveats")
    _narrative_sheet(wb, narrative)
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Module 20 — impacted as-planned
# --------------------------------------------------------------------------- #

def build_iap_xlsx(label: str, iap: dict,
                   narrative: str | None = None) -> bytes:
    """iap: dict from programme.impacted_asplanned.run_impacted_asplanned"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "Impacted As-Planned (baseline + event fragnets)", 4)
    rows = [
        ("Baseline programme", label),
        ("Events inserted", iap.get("events_used", 0)),
        ("Events skipped (no tie-in / no fragnet)",
         len(iap.get("skipped_events", []))),
        ("Baseline modelled completion",
         _fmt(iap.get("completion_pre"))),
        ("Impacted modelled completion",
         _fmt(iap.get("completion_final"))),
        ("TOTAL MODELLED IMPACT (days)", iap.get("total_delta_days")),
    ]
    _header_row(ws, 4, ["Measure", "Value"])
    for i, (k, v) in enumerate(rows, start=5):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=v if v is not None else "n/a")
        a.border = b.border = THIN_BORDER
        if k.startswith("TOTAL"):
            a.font = Font(bold=True)
    _autofit(ws, {1: 42, 2: 24})

    if iap.get("rows"):
        s = wb.create_sheet("Per-Event Increments")
        _header_row(s, 1, ["Event", "Title", "Date", "Incremental (d)",
                           "Completion after"])
        for i, r in enumerate(iap["rows"], start=2):
            vals = [r.get("event_id"), r.get("title"),
                    _fmt(r.get("date_raised")),
                    r.get("incremental_delta_days"),
                    _fmt(r.get("completion_after"))]
            for col, v in enumerate(vals, start=1):
                s.cell(row=i, column=col, value=v).border = THIN_BORDER
        _autofit(s, {1: 14, 2: 44, 3: 12, 4: 14, 5: 15})
        s.freeze_panes = "A2"

    notes = (iap.get("warnings", []) + iap.get("concurrency", [])
             + iap.get("caveats", []) + [iap.get("caveat", "")])
    _notes_sheet(wb, [n for n in notes if n], "Warnings & Caveats")
    _narrative_sheet(wb, narrative)
    return _wb_bytes(wb)


def _image_sheet(wb, caption: str, png: bytes) -> None:
    """One chart PNG on its own sheet (the deliverable figure)."""
    from openpyxl.drawing.image import Image as _XLImage
    ws = wb.create_sheet(caption[:31])
    ws["A1"] = caption
    ws["A1"].font = Font(bold=True, size=13)
    ws.add_image(_XLImage(io.BytesIO(png)), "A3")


def build_simple_xlsx(title: str, sheets: dict[str, list[dict]],
                      notes: list[str] | None = None,
                      images: dict[str, bytes] | None = None) -> bytes:
    """Generic workbook: one sheet per {name: rows-of-dicts} + notes.
    Used by the stepped methods (As-Planned vs As-Built, Collapsed
    As-Built) whose tables are analyst-shaped. ``images`` embeds chart
    PNGs, one sheet each — the final gantt travels WITH the workbook."""
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = name[:31]
        first = False
        if not rows:
            ws["A1"] = "(empty)"
            continue
        headers = list(rows[0].keys())
        _header_row(ws, 1, headers)
        for i, row in enumerate(rows, start=2):
            for col, h in enumerate(headers, start=1):
                v = row.get(h)
                if isinstance(v, datetime):
                    v = f"{v:%Y-%m-%d}"
                c = ws.cell(row=i, column=col, value=v)
                c.border = THIN_BORDER
        _autofit(ws, {i + 1: min(max(len(str(h)) + 4, 12), 46)
                      for i, h in enumerate(headers)})
        ws.freeze_panes = "A2"
    for cap, png in (images or {}).items():
        if png:
            _image_sheet(wb, cap, png)
    if notes:
        _notes_sheet(wb, notes, "Notes & Caveats")
    return _wb_bytes(wb)


# --------------------------------------------------------------------------- #
# Report appendix — the complete tables, as their own workbook
# --------------------------------------------------------------------------- #

_SHEET_BAD = set(r'[]:*?/\\')


def _sheet_name(idx: int, caption: str, used: set[str]) -> str:
    """Excel sheet name: <=31 chars, no reserved characters, unique."""
    clean = "".join(" " if c in _SHEET_BAD else c for c in caption)
    name = f"A{idx} {clean}"[:31].strip()
    base, n = name, 2
    while name.lower() in used:
        suffix = f" ({n})"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name.lower())
    return name


def build_appendix_xlsx(title: str,
                        tables: list[tuple[str, list[dict]]],
                        notes: list[str] | None = None) -> bytes:
    """The report appendix: one sheet per table, complete, plus an index.

    Built entirely in code — the narrative model only ever sees the top
    rows, so the full record costs no tokens and no generation time.
    """
    wb = Workbook()
    used: set[str] = set()

    idx = wb.active
    idx.title = "Index"
    _title(idx, f"{title} — appendix", 3)
    idx.cell(row=3, column=1, value=(
        "Every table in full. The narrative reports the most material "
        "rows; these sheets are the complete record."
    )).font = Font(italic=True)
    _header_row(idx, 5, ["Sheet", "Table", "Rows"])

    for i, (caption, rows) in enumerate(tables, start=1):
        name = _sheet_name(i, caption, used)
        for col, v in enumerate((name, caption, len(rows)), start=1):
            idx.cell(row=5 + i, column=col, value=v).border = THIN_BORDER
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value=f"A{i}. {caption}").font = Font(
            bold=True, size=12)
        if not rows:
            ws.cell(row=3, column=1, value="(no rows)")
            continue
        headers = list(rows[0].keys())
        _header_row(ws, 3, [str(h) for h in headers])
        for r, row in enumerate(rows, start=4):
            for c, h in enumerate(headers, start=1):
                v = row.get(h)
                if isinstance(v, datetime):
                    v = f"{v:%Y-%m-%d}"
                ws.cell(row=r, column=c, value=v).border = THIN_BORDER
        _autofit(ws, {c + 1: min(max(len(str(h)) + 4, 12), 46)
                      for c, h in enumerate(headers)})
        ws.freeze_panes = "A4"

    _autofit(idx, {1: 34, 2: 54, 3: 10})
    if notes:
        _notes_sheet(wb, notes, "Notes & Caveats")
    return _wb_bytes(wb)
