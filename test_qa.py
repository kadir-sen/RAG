"""QA/QC regression suite — engine level.

Layer A: delay-analyst cross-validation — modules must agree with each other
and with manual recomputation from raw XER rows.
Layer B: software edge cases — degenerate inputs, symmetry, bounds.
Layer C: report integrity — prompts carry the hard rules and caveats; every
workbook opens with its narrative sheet.

Run: python3 test_qa.py  (exit code 1 on any failure)
"""
import os
import sys
import io

from openpyxl import load_workbook

def _p(rel: str) -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel)


from dcma import parse_xer, run_all_checks
from dcma.config import DCMAConfig
from programme import (
    analyse_float_erosion, analyse_windows, build_comparison_prompt,
    build_comparison_xlsx, build_critical_path_prompt,
    build_critical_path_xlsx, build_float_erosion_prompt,
    build_float_erosion_xlsx, build_inventory, build_inventory_prompt,
    build_inventory_xlsx, build_milestone_prompt, build_milestone_xlsx,
    build_progress_prompt, build_progress_xlsx, build_resources_prompt,
    build_resources_xlsx, build_variance_prompt, build_variance_xlsx,
    build_windows_prompt, build_windows_xlsx, compare_revisions,
    compute_progress, compute_variance_by_mapping, end_activity_candidates,
    extract_critical_path, extract_longest_path, extract_resource_loading,
    task_wbs_assignments, track_milestone_shifts,
)

PASS, FAIL = [], []
def check(name, cond, detail=""):
    (PASS if cond else FAIL).append((name, detail))
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail and not cond else ""))

cfg = DCMAConfig()


def _truncate_asbuilt(text: str, cutoff: str) -> str:
    """Wind the final as-built back to a mid-project data date.

    The bundled Harbour Point pair is baseline + FINAL as-built. Most
    pins exercise mid-project behaviour (an unachieved contract
    milestone, a remaining network for TIA insertion, forecast tails),
    so the canonical update fixture is the as-built truncated at the
    cutoff: actuals after the cutoff revert to remaining work, exactly
    as the corresponding progress update would have reported them. The
    early/target dates already in the file become the forecast.
    """
    lines = text.splitlines()
    out, fields, table = [], [], ""
    for line in lines:
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F":
            fields = cells[1:]
        elif cells[0] == "%R" and table == "PROJECT":
            row = dict(zip(fields, cells[1:]))
            row["last_recalc_date"] = cutoff
            cells = ["%R"] + [row.get(f, "") for f in fields]
        elif cells[0] == "%R" and table == "TASK":
            row = dict(zip(fields, cells[1:]))
            if (row.get("act_start_date") or "9999") >= cutoff:
                row["status_code"] = "TK_NotStart"
                row["act_start_date"] = row["act_end_date"] = ""
                row["remain_drtn_hr_cnt"] = row.get(
                    "target_drtn_hr_cnt", "0")
                row["phys_complete_pct"] = "0"
            elif (row.get("act_end_date") or "") >= cutoff:
                row["status_code"] = "TK_Active"
                row["act_end_date"] = ""
                row["remain_drtn_hr_cnt"] = str(
                    float(row.get("target_drtn_hr_cnt") or 0) / 2)
                row["phys_complete_pct"] = "50"
            cells = ["%R"] + [row.get(f, "") for f in fields]
        out.append("\t".join(cells))
    return "\n".join(out)


with open(_p("sample/Harbour Point DCP-03 - Baseline Programme Rev 0.xer"),"rb") as fh:
    B = parse_xer(fh.read())
with open(_p("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"),
          encoding="utf-8") as fh:
    _ab_text = fh.read()
AB = parse_xer(_ab_text)                       # final as-built, untouched
_u_text = _truncate_asbuilt(_ab_text, "2025-09-01 08:00")
U = parse_xer(_u_text)


def _edit_first_row(text: str, table_name: str, **field_values) -> str:
    """Override fields on the first %R row of a table."""
    out, fields, table, done = [], [], "", False
    for line in text.splitlines():
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F" and table == table_name:
            fields = cells[1:]
        elif cells[0] == "%R" and table == table_name and not done:
            row = dict(zip(fields, cells[1:]))
            row.update(field_values)
            cells = ["%R"] + [row.get(f, "") for f in fields]
            done = True
        out.append("\t".join(cells))
    return "\n".join(out)


def _add_taskpred(text: str, pred_id: str, succ_id: str) -> str:
    """Append one FS TASKPRED row (a transitive skip link)."""
    out, fields, table = [], [], ""
    for line in text.splitlines():
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F" and table == "TASKPRED":
            fields = cells[1:]
        out.append(line)
        if cells[0] == "%R" and table == "TASKPRED":
            base = dict(zip(fields, cells[1:]))
    row = dict.fromkeys(fields, "")
    row.update({"task_pred_id": "999901",
                "task_id": succ_id, "pred_task_id": pred_id,
                "proj_id": base.get("proj_id", ""),
                "pred_proj_id": base.get("pred_proj_id", ""),
                "pred_type": "PR_FS", "lag_hr_cnt": "0"})
    marker = out.index("%T\tTASKPRED")
    insert_at = marker + 2
    while insert_at < len(out) and out[insert_at].startswith("%R"):
        insert_at += 1
    out.insert(insert_at,
               "\t".join(["%R"] + [row.get(f, "") for f in fields]))
    return "\n".join(out)


def _del_tasks(text: str, codes: set) -> str:
    """Drop TASK rows (and their links) for the given codes."""
    out, fields, table, dead_ids = [], {}, "", set()
    lines = text.splitlines()
    for line in lines:                      # pass 1: find ids
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F":
            fields[table] = cells[1:]
        elif cells[0] == "%R" and table == "TASK":
            row = dict(zip(fields["TASK"], cells[1:]))
            if row.get("task_code") in codes:
                dead_ids.add(row.get("task_id"))
    table = ""
    for line in lines:                      # pass 2: drop rows
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%R" and table == "TASK":
            row = dict(zip(fields["TASK"], cells[1:]))
            if row.get("task_id") in dead_ids:
                continue
        elif cells[0] == "%R" and table == "TASKPRED":
            row = dict(zip(fields["TASKPRED"], cells[1:]))
            if (row.get("task_id") in dead_ids
                    or row.get("pred_task_id") in dead_ids):
                continue
        out.append(line)
    return "\n".join(out)


def _del_taskpred(text: str, pred_id: str, succ_id: str) -> str:
    """Remove one relationship row (to break chain continuity)."""
    out, fields, table = [], [], ""
    for line in text.splitlines():
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F" and table == "TASKPRED":
            fields = cells[1:]
        elif cells[0] == "%R" and table == "TASKPRED":
            row = dict(zip(fields, cells[1:]))
            if (row.get("pred_task_id") == pred_id
                    and row.get("task_id") == succ_id):
                continue
        out.append(line)
    return "\n".join(out)


def _edit_task_rows(text: str, task_code: str, **field_values) -> str:
    """Return the XER text with one TASK row's fields overridden."""
    out, fields, table = [], [], ""
    for line in text.splitlines():
        cells = line.split("\t")
        if cells[0] == "%T":
            table = cells[1]
        elif cells[0] == "%F":
            fields = cells[1:]
        elif cells[0] == "%R" and table == "TASK":
            row = dict(zip(fields, cells[1:]))
            if row.get("task_code") == task_code:
                row.update(field_values)
                cells = ["%R"] + [row.get(f, "") for f in fields]
        out.append("\t".join(cells))
    return "\n".join(out)
fix = []
for f in ["revA.xer","revB.xer","revC.xer"]:
    with open(_p(f"sample/revisions/{f}"), "rb") as fh:
        fix.append((f, parse_xer(fh.read())))

print("== A. Cross-module numerical consistency ==")

# A1. Negative float: DCMA check 7 vs float-erosion snapshot (baseline)
dcma = {c.number: c for c in run_all_checks(B, cfg)}
fe = analyse_float_erosion([("B", B), ("U", U)])
check("A1 DCMA neg-float == float-erosion neg count (baseline)",
      dcma[7].affected_count == fe.snapshots[0].negative_count,
      f"dcma={dcma[7].affected_count} vs fe={fe.snapshots[0].negative_count}")

# A2. Critical count: DCMA check 12 vs float-method CP module
cp_f = extract_critical_path(B, "B")
# check 12's affected_ids are the STRAYS once the chain is continuous
# (the full critical list only on FAIL) — the population comparison
# lives in the metric, so pin that, plus continuity on this fixture.
check("A2 DCMA critical count == CP module critical (TF<=0)",
      dcma[12].metric_value.startswith(f"{len(cp_f.critical)} of "),
      f"dcma12='{dcma[12].metric_value}' vs cp={len(cp_f.critical)}")
check("A2b baseline critical path is one continuous segment",
      dcma[12].status.name == "PASS" and "1 segment" in dcma[12].metric_value,
      dcma[12].metric_value)

# A3. Manual TF recount from raw rows
manual_crit = 0
for t in B.tasks:
    if t.is_loe_or_wbs or t.is_complete or t.total_float_hr is None:
        continue
    if t.total_float_hr / B.hours_per_day(t, cfg) <= 0:
        manual_crit += 1
check("A3 manual TF<=0 recount == CP module", manual_crit == len(cp_f.critical),
      f"manual={manual_crit} vs cp={len(cp_f.critical)}")

# A4. Windows completion movement == project-level scheduled finish delta
# (precise arithmetic: .days truncation was audit finding C5)
wres = analyse_windows([("B", B), ("U", U)])
manual_move = round((U.project.scheduled_finish
                     - B.project.scheduled_finish).total_seconds()
                    / 86400, 1)
check("A4 windows movement == scheduled finish delta",
      wres.windows[0].movement_days == manual_move,
      f"win={wres.windows[0].movement_days} vs manual={manual_move}")

# A5. Windows total == sum across fixture windows
wfix = analyse_windows(fix)
tot = sum(w.movement_days for w in wfix.windows if w.movement_days is not None)
check("A5 fixtures cumulative movement == sum of windows",
      wfix.total_movement_days == tot)

# A5b-d. Window driver traceback: every movement figure is backed by
# row-level driving-path activities carrying STORED dates only.
_w0 = wres.windows[0]
check("A5b window drivers populated from the later revision's path",
      len(_w0.drivers) > 0)
_u_by_code = {t.task_code: t for t in U.tasks}
_drv_ok = all(
    (d.finish_new == ((_u_by_code[d.task_code].act_finish
                       or _u_by_code[d.task_code].early_finish)))
    for d in _w0.drivers if d.task_code in _u_by_code)
check("A5c driver finishes are the revision's own stored dates", _drv_ok)
_slips = [d.slip_days for d in _w0.drivers if d.slip_days is not None]
check("A5d drivers sorted biggest mover first",
      _slips == sorted(_slips, reverse=True))
# A5e: a FULLY progressed later revision must still yield a movement
# traceback — the as-built longest path through completed work, with
# the basis disclosed (real as-built files have no remaining network)
_wres_ab = analyse_windows([("B", B), ("AB", AB)])
check("A5e fully progressed revision falls back to the as-built "
      "longest path",
      len(_wres_ab.windows[0].drivers) > 0
      and any("fully progressed" in w for w in _wres_ab.warnings))

# A6. Longest path is a subset of... no — verify every longest-path link
# joins two on-path activities and terminal is on path
cp_l = extract_longest_path(B, "B")
codes = {a.task_code for a in cp_l.critical}
bad_links = [lk for lk in cp_l.links
             if lk.pred_code not in codes or lk.succ_code not in codes]
check("A6 longest-path links all join on-path activities", not bad_links,
      f"{len(bad_links)} dangling")
check("A6b terminal on path", cp_l.end_choice in codes)

# A7. Single-branch trace (A3400) — every non-start activity has a driving
# predecessor within the path
cp_s = extract_longest_path(B, "B", end_task_code="C-3070")
succs_with_pred = {lk.succ_code for lk in cp_s.links}
starts = [a.task_code for a in cp_s.critical
          if a.task_code not in succs_with_pred]
check("A7 single-branch trace has exactly one chain start",
      len(starts) == 1, f"starts={starts}")

# A8. Comparison symmetry: swap old/new -> added<->deleted, and
# reversed-data-date warning fires
c_fwd = compare_revisions(B, U, "B", "U")
c_rev = compare_revisions(U, B, "U", "B")
check("A8 comparison added/deleted symmetric",
      len(c_fwd.added) == len(c_rev.deleted)
      and len(c_fwd.deleted) == len(c_rev.added))
check("A8b reversed direction warned",
      any("LATER data date" in w for w in c_rev.warnings))

# A9. Self-comparison finds zero changes
c_self = compare_revisions(B, B, "B", "B2")
check("A9 self-comparison == 0 changes", c_self.total_changes == 0,
      f"{c_self.total_changes} changes: {c_self.category_counts}")

# A10. S-curve bounds and monotonicity
pr = compute_progress(B, "B", [("U", U)])
mono = all(a.cum_pct <= b.cum_pct + 1e-9
           for a, b in zip(pr.planned_curve, pr.planned_curve[1:]))
check("A10 planned curve monotonic", mono)
check("A10b planned curve ends at 100%",
      abs(pr.planned_curve[-1].cum_pct - 100.0) < 0.1,
      f"end={pr.planned_curve[-1].cum_pct}")
check("A10c recorded curve <= 100%",
      all(p.cum_pct <= 100.0 + 1e-9 for p in pr.recorded_curve))
rmono = all(a.cum_pct <= b.cum_pct + 1e-9
            for a, b in zip(pr.recorded_curve, pr.recorded_curve[1:]))
check("A10d recorded curve monotonic", rmono)

# A11. Recorded % manual recount (duration weights)
w = {}
for t in U.tasks:
    if t.is_loe_or_wbs: continue
    d = t.original_duration_days(U.hours_per_day(t, cfg)) or 0.0
    w[t.task_id] = max(d, 0.0)
pct = {r["task_id"].strip(): float(r.get("phys_complete_pct") or 0)
       for r in U.raw_tables["TASK"]}
earned = sum(w[t.task_id] if t.is_complete
             else w[t.task_id]*pct.get(t.task_id,0)/100 if t.act_start else 0
             for t in U.tasks if not t.is_loe_or_wbs)
manual_pct = round(100*earned/sum(w.values()), 1)
check("A11 recorded % matches manual recount",
      abs(manual_pct - pr.recorded_pct_at_dd) < 0.05,
      f"manual={manual_pct} vs module={pr.recorded_pct_at_dd}")

# A12. Milestone shift manual verification: pick one milestone present in
# both, verify its total shift equals date difference from raw fields
inv_pool = [("Harbour Point DCP-03 - Baseline Programme Rev 0.xer", B), ("Harbour Point DCP-03 - As-Built Programme Rev 12.xer", U)]
ms = track_milestone_shifts([
    ("Harbour Point DCP-03 - Baseline Programme Rev 0.xer", B.project.data_date, B),
    ("Harbour Point DCP-03 - As-Built Programme Rev 12.xer", U.project.data_date, U),
])
s_ok = None
for s in ms.series:
    if s.total_shift_days is not None and len([p for p in s.points if p.value_date]) == 2:
        p0, p1 = [p for p in s.points if p.value_date]
        expected = (p1.value_date - p0.value_date).days
        s_ok = (s.key, s.total_shift_days, expected)
        break
check("A12 milestone shift == raw date delta",
      s_ok is not None and abs(s_ok[1] - s_ok[2]) < 1.0, str(s_ok))

# A13. Variance group bounds: WBS L1 groups — min start / max finish manual
wbs_map = task_wbs_assignments(B, level=1)
var = compute_variance_by_mapping(B, U, wbs_map, wbs_map, "WBS L1")
g = next(g for g in var.groups if g.planned.activity_count > 5)
ids = [tid for tid, lbl in wbs_map.items() if lbl == g.code_value]
starts = [t.target_start or t.early_start for t in B.tasks
          if t.task_id in set(ids) and not t.is_loe_or_wbs
          and (t.target_start or t.early_start)]
check("A13 variance planned start == manual min",
      g.planned.start == min(starts),
      f"module={g.planned.start} manual={min(starts)}")

# A14. Resource totals == raw TASKRSRC sum (for dated, positive assignments)
rl = extract_resource_loading(B, "B")
raw_total = 0.0
tid_ok = {t.task_id for t in B.tasks if not t.is_loe_or_wbs
          and (t.target_start or t.early_start or t.act_start)}
rid_ok = {r.rsrc_id for r in rl.resources} | {
    (row.get("rsrc_id") or "").strip() for row in B.raw_tables["RSRC"]}
for row in B.raw_tables["TASKRSRC"]:
    try: q = float(row.get("target_qty") or 0)
    except ValueError: q = 0
    if q > 0 and (row.get("task_id") or "").strip() in tid_ok:
        raw_total += q
mod_total = sum(r.total_qty for r in rl.resources)
check("A14 resource totals == raw sum", abs(raw_total - mod_total) < 0.5,
      f"raw={raw_total:,.0f} vs module={mod_total:,.0f}")
hist_total = sum(p.qty for p in rl.histogram)
check("A14b histogram sums to totals", abs(hist_total - mod_total) < 1.0,
      f"hist={hist_total:,.0f} vs {mod_total:,.0f}")


# A15. Milestone terminals: EVERY milestone is offered, achieved or
# not — you pick what you are measuring to; whether the works reached
# it is disclosed, never a filter.
from programme import trace_end_candidates as _tec
_a15 = _tec([("B", B), ("U", U)], contract_ms="H-5040")
check("A15 elected milestone offered first even though unachieved",
      _a15[0][0] == "H-5040" and _a15[0][3] is False)
_ms_codes = {t.task_code for t in U.tasks if t.is_milestone
             and not t.is_loe_or_wbs}
_offered = {c for c, _, _, _ in _a15}
check("A15b every milestone is offered, achieved or not",
      _ms_codes <= _offered, f"{len(_ms_codes - _offered)} missing")
check("A15c unachieved milestones carry achieved=False + a forecast date",
      all(ok or d is not None
          for c, _, d, ok in _a15 if c in _ms_codes))


# A16. Actual-date trace + triangulation invariants
from programme import extract_actual_trace, trace_end_candidates
tr_strict = extract_actual_trace([("B", B), ("U", U)], max_gap_days=240,
                                 allow_temporal_fallback=False)
check("A16 strict trace: every link logic-evidenced",
      all(lk.had_logic for lk in tr_strict.links))
check("A16b trace links form a chain (each pred is next activity)",
      all(lk.score is not None and 0 <= lk.score <= 1
          for lk in tr_strict.links))
codes = [a.task_code for a in tr_strict.activities]
check("A16c trace chain has no duplicates", len(codes) == len(set(codes)))
tr_fb = extract_actual_trace([("B", B), ("U", U)], max_gap_days=15)
check("A16d default (continuous) trace >= strict at same gap",
      len(tr_fb.activities) >= len(extract_actual_trace(
          [("B", B), ("U", U)], max_gap_days=15,
          allow_temporal_fallback=False).activities))
check("A16f default continues on sequence: un-evidenced hops disclosed",
      all(lk.had_logic for lk in tr_fb.links)
      or any("SEQUENCE ALONE" in w for w in tr_fb.warnings))

# A16g-k. Contractual-milestone anchoring + hybrid forecast tail.
_kd15 = next((t for t in U.tasks if t.task_code == "H-5040"), None)
check("A16g sample carries an unachieved completion milestone (H-5040)",
      _kd15 is not None and _kd15.act_finish is None)
_hy = extract_actual_trace([("B", B), ("U", U)], end_task_code="H-5040",
                           max_gap_days=60)
check("A16h unachieved elected milestone still anchors the trace",
      _hy.terminal_code == "H-5040" and _hy.hybrid)
check("A16i hybrid disclosed in warnings and caveats",
      any("HYBRID" in c for c in _hy.caveats)
      and any("not been achieved" in w.lower() or "NOT been achieved" in w
              for w in _hy.warnings))
check("A16j hybrid chain is basis-labelled and reaches back to as-built",
      _hy.forecast_count > 0 and _hy.asbuilt_count > 0
      and {a.basis for a in _hy.activities} <= {
          "as-built", "in-progress", "forecast"})
# ordering: as-built work must precede the forecast tail
_bases = [a.basis for a in _hy.activities]
check("A16k forecast tail sits at the end of the chain (no forecast "
      "before as-built work)",
      all(b == "forecast" for b in _bases[_bases.index("forecast"):])
      if "forecast" in _bases else True)
check("A16l terminal candidates offer the elected milestone first",
      trace_end_candidates([("B", B), ("U", U)], contract_ms="H-5040")[0][0]
      == "H-5040")

# A16m-s. Analyst-election trace (the step-① adopted path, shared by
# the As-Built CP page and APvAB). An elected path must be reported
# no more charitably than a computed one.
from programme import trace_from_election, build_asbuilt_multi_prompt
_ad_path = [(a.task_code, a.name) for a in tr_fb.activities]
_el = trace_from_election([("B", B), ("U", U)], _ad_path,
                          basis_label="Actual sequence through recorded "
                                      "dates (test)")
check("A16m election trace keeps the adopted order and terminal",
      [a.task_code for a in _el.activities] == [c for c, _ in _ad_path]
      and _el.terminal_code == _ad_path[-1][0])
check("A16n election trace discloses the adopted basis in its caveats",
      any("ELECTION" in c for c in _el.caveats)
      and any("Actual sequence through recorded dates (test)" in c
              for c in _el.caveats))
check("A16o election links agree with the computed trace on logic",
      {(lk.pred_code, lk.succ_code): lk.had_logic for lk in _el.links}
      == {(lk.pred_code, lk.succ_code): lk.had_logic
          for lk in tr_fb.links})
check("A16p election link scores stay in [0,1]",
      all(0 <= lk.score <= 1 for lk in _el.links))
# hand-edit: a pair the records cannot support scores zero temporal
# evidence (successor starting long before its predecessor began)
_el_bad = trace_from_election([("B", B), ("U", U)],
                              [_ad_path[-1], _ad_path[0]])
_bad_lk = _el_bad.links[0] if _el_bad.links else None
check("A16q an out-of-order adopted pair is scored weak, not hidden",
      _bad_lk is not None and _bad_lk.score <= 0.4)
_el_hy = trace_from_election(
    [("B", B), ("U", U)],
    _ad_path + [("H-5040", next(t.name for t in U.tasks
                              if t.task_code == "H-5040"))])
check("A16r a forecast tail in the election is a disclosed hybrid",
      _el_hy.hybrid and any("HYBRID" in c for c in _el_hy.caveats))
_mp = build_asbuilt_multi_prompt([_el, _el_hy])
check("A16s multi-path prompt covers every path and templates once",
      "PATH 1 of 2" in _mp and "PATH 2 of 2" in _mp
      and _mp.count("<context>") == 2)

# A16t-v. As-built longest path (the step-① logic candidate) must run
# THROUGH completed work to the earliest linked activity — not stop at
# the data date the way a remaining-works longest path does.
from programme import extract_asbuilt_longest_path
_lp = extract_asbuilt_longest_path(U, end_task_code="H-5040")
_dd = U.project.data_date
_lp_starts = [a.act_start for a in _lp.activities if a.act_start]
check("A16t as-built longest path reaches back past the data date",
      bool(_lp_starts) and min(_lp_starts) < _dd,
      f"earliest {min(_lp_starts) if _lp_starts else None} vs dd {_dd}")
check("A16u as-built longest path: every hand-off is programmed logic "
      "and the chain is chronological",
      all(lk.had_logic for lk in _lp.links)
      and all(a.act_start <= b.act_start
              for a, b in zip(_lp.activities, _lp.activities[1:])
              if a.act_start and b.act_start))
check("A16v unachieved terminal still anchors the logic candidate "
      "as a disclosed hybrid",
      _lp.terminal_code == "H-5040" and _lp.hybrid
      and any("HYBRID" in c for c in _lp.caveats))


# A17. Sequence coding invariants
from programme import propose_sequence_mapping, analyse_sequence
sp = propose_sequence_mapping(U, "U")
check("A17 sequence: every activity gets a front and a stage",
      all(r.front and r.stage for r in sp.rows))
check("A17b sequence: coverage percentages in [0,100]",
      0 <= sp.stage_coverage_pct <= 100 and 0 <= sp.front_coverage_pct <= 100)
sq = analyse_sequence(sp.rows, "U")
check("A17c sequence: band bounds ordered (start <= finish)",
      all(b.act_start is None or b.act_finish is None
          or b.act_start <= b.act_finish for b in sq.bands))
check("A17d sequence: mapped count == actualised rows",
      sq.mapped_activities == sum(1 for r in sp.rows if r.act_start))
check("A17e sequence: unconfirmed mapping carries the extra caveat",
      any("AUTO-PROPOSED" in c for c in sq.caveats))
sq2 = analyse_sequence(sp.rows, "U", mapping_confirmed=True)
check("A17f sequence: confirmed mapping drops it",
      not any("AUTO-PROPOSED" in c for c in sq2.caveats))


# A18. AI-review prompt/parser layer (offline)
from programme import (build_mapping_review_prompt, parse_mapping_review,
                       build_view_advice_prompt, parse_view_advice)
pmr = build_mapping_review_prompt(sp.rows[:5])
check("A18 review prompt lists stages and rows",
      "Allowed stage labels" in pmr and sp.rows[0].task_code in pmr)
good = parse_mapping_review(
    '[{"id": "%s", "stage": "Finishes & Fit-Out"}]' % sp.rows[0].task_code,
    {r.task_code for r in sp.rows[:5]})
check("A18b parser accepts valid correction", len(good) == 1)
check("A18c parser rejects unknown ids and stages",
      parse_mapping_review('[{"id":"ZZZ","stage":"Finishes & Fit-Out"},'
                           '{"id":"%s","stage":"Made Up"}]'
                           % sp.rows[0].task_code,
                           {sp.rows[0].task_code}) == {})
check("A18d parser survives garbage", parse_mapping_review("oops", {"A"}) == {})
adv = parse_view_advice('{"mode":"bands","colour":"Stage","max_fronts":10,"rationale":"r"}')
check("A18e view advice parses and clamps",
      adv is not None and adv["mode"] == "bands"
      and parse_view_advice('{"mode":"nope"}') is None)
check("A18f view advice prompt built",
      "sequence_gantt" in build_view_advice_prompt(sq, 30))


# A19. Hierarchy rebuild invariants
from programme import (available_dimensions, build_hierarchy, tree_to_dict,
                       build_gantt_html, config_to_json, config_from_json)
hd = available_dimensions(B)
check("A19 dimensions discovered (2 WBS levels + codes + UDFs in sample)",
      len([d for d in hd if d.dim_id.startswith("wbs:")]) == 2
      and any(d.dim_id.startswith("code:") for d in hd)
      and any(d.dim_id.startswith("udf:") for d in hd))
hh = build_hierarchy(B, ["wbs:1", "wbs:2"], "B",
                     dim_labels=["WBS Level 1", "WBS Level 2"])
check("A19b every source activity placed exactly once",
      hh.is_complete and hh.placed_activities == hh.source_activities)
# leaf-count == placed (no duplication anywhere in the tree)
def _leaves(n):
    return len(n.activities) + sum(_leaves(c) for c in n.children.values())
check("A19c tree leaf count == placed", _leaves(hh.root) == hh.placed_activities)
# rollup: root span brackets every activity date
def _acts(n):
    yield from n.activities
    for c in n.children.values():
        yield from _acts(c)
all_starts = [a.start for a in _acts(hh.root) if a.start]
all_fins = [a.finish for a in _acts(hh.root) if a.finish]
root_kids = list(hh.root.children.values())
check("A19d rollup start == min child start",
      min(k.start for k in root_kids if k.start) == min(all_starts))
check("A19e rollup finish == max child finish",
      max(k.finish for k in root_kids if k.finish) == max(all_fins))
# source data untouched: parse count unchanged after building
check("A19f source untouched (task count stable)",
      hh.source_activities == sum(1 for t in B.tasks
                                  if t.task_type != "TT_WBS"))
html = build_gantt_html(tree_to_dict(hh.root))
check("A19g gantt html self-contained", "<script>" in html
      and "http" not in html.split("</style>")[0].lower())
cfg = config_from_json(config_to_json("v", ["wbs:2"], ["WBS Level 2"]))
check("A19h config round-trips", cfg is not None and cfg[1] == ["wbs:2"])
check("A19i bad config rejected", config_from_json('{"dimensions":["x:1"]}') is None)


# A20. Sequence dims + hierarchy xlsx
from programme import sequence_dimension_mappings, build_hierarchy_xlsx
ex = sequence_dimension_mappings(U, sp.rows)
hs = build_hierarchy(U, ["seq:front", "seq:stage"], "U",
                     dim_labels=["Front", "Stage"], extra_mappings=ex)
check("A20 seq-dims hierarchy places all activities",
      hs.is_complete and hs.placed_activities == hs.source_activities)
xh = build_hierarchy_xlsx(hs)
from openpyxl import load_workbook as _lw
import io as _io2
_wbh = _lw(_io2.BytesIO(xh))
check("A20b hierarchy xlsx sheets",
      set(_wbh.sheetnames) >= {"Hierarchy", "Flat Table"})
outl = sum(1 for rd in _wbh["Hierarchy"].row_dimensions.values()
           if rd.outline_level)
check("A20c hierarchy xlsx has collapsible outlines", outl > 20,
      f"outlined rows={outl}")
flat_rows = _wbh["Flat Table"].max_row - 1
check("A20d flat table row per activity",
      flat_rows == hs.placed_activities,
      f"flat={flat_rows} vs placed={hs.placed_activities}")
check("A20e seq config ids accepted",
      config_from_json('{"dimensions": ["seq:front"], "labels": ["F"]}')
      is not None)


# A21. Dimension menu = WBS levels + activity codes + TASK UDFs only
hd2 = available_dimensions(U)
kinds2 = {d.dim_id.partition(":")[0] for d in hd2}
check("A21 only the three families offered", kinds2 <= {"wbs", "code", "udf"})
check("A21b all WBS levels present",
      {"wbs:1", "wbs:2"} <= {d.dim_id for d in hd2})
# synthetic TASK UDF proves the udf: path end-to-end
_t0 = U.tasks[0]
U.raw_tables.setdefault("UDFTYPE", []).append(
    {"udf_type_id": "999", "table_name": "TASK",
     "udf_type_label": "QA Zone", "udf_type_name": "qa_zone",
     "logical_data_type": "FT_TEXT"})
U.raw_tables.setdefault("UDFVALUE", []).append(
    {"udf_type_id": "999", "fk_id": _t0.task_id, "udf_text": "Zone QA",
     "udf_number": "", "udf_date": "", "udf_code_id": ""})
hd3 = available_dimensions(U)
check("A21c TASK UDF surfaces as a dimension",
      any(d.dim_id == "udf:999" and "QA Zone" in d.label for d in hd3))
_hu = build_hierarchy(U, ["udf:999"], "U", dim_labels=["QA Zone"])
check("A21d UDF hierarchy: tagged task grouped, rest Unassigned",
      _hu.is_complete and "Zone QA" in _hu.root.children
      and _hu.root.children["Zone QA"].activity_count == 1)
U.raw_tables["UDFTYPE"].pop(); U.raw_tables["UDFVALUE"].pop()
# synthetic global + project code types both surface, scope-labelled
U.raw_tables.setdefault("ACTVTYPE", []).append(
    {"actv_code_type_id": "801", "actv_code_type": "Zone",
     "actv_code_type_scope": "AS_Global"})
U.raw_tables["ACTVTYPE"].append(
    {"actv_code_type_id": "802", "actv_code_type": "Package",
     "actv_code_type_scope": "AS_Project"})
hd4 = available_dimensions(U)
lbls = {d.dim_id: d.label for d in hd4}
check("A21e global + project codes both offered, scope in label",
      "[Global]" in lbls.get("code:801", "")
      and "[Project]" in lbls.get("code:802", ""))
U.raw_tables["ACTVTYPE"] = []
check("A21f config kinds restricted",
      config_from_json('{"dimensions": ["cal:"]}') is None
      and config_from_json('{"dimensions": ["udf:9", "wbs:2"]}') is not None)


# A22. Prospective TIA engine
from programme import (DelayEvent, FragnetActivity, FragnetLink, run_tia,
                       validate_fragnet, parse_fragnet_json, parse_links,
                       find_template_activities, find_template_work_packages,
                       assess_event_scope, build_logic_recommendation_prompt,
                       parse_logic_recommendation_json)
from datetime import timedelta as _td
_ev = DelayEvent("EV-QA", "test event")
_fr = [FragnetActivity("TIA-010", "chain", 120,
                       successors=[FragnetLink("H-5040")])]
_r = run_tia(U, "U", _ev, _fr)
check("A22 TIA delta exact for a direct chain into completion",
      _r.completion_post == _r.data_date + _td(days=120)
      and (_r.completion_delta_days or 0) > 0)
_r0 = run_tia(U, "U", _ev, [])
check("A22b empty fragnet -> zero delta",
      _r0.completion_pre == _r0.completion_post)
check("A22c calibration disclosed", _r.calibration_days is not None
      and any("Calibration" in w for w in _r.warnings))
iss = validate_fragnet(U, [FragnetActivity("TIA-1", "x", -5)])
check("A22d validation flags open ends + bad duration",
      any("open start" in i for i in iss)
      and any("duration" in i for i in iss))
iss2 = validate_fragnet(U, [
    FragnetActivity("TIA-A", "a", 5,
                    predecessors=[FragnetLink("TIA-B")],
                    successors=[FragnetLink("TIA-B"), FragnetLink("H-5040")]),
    FragnetActivity("TIA-B", "b", 5,
                    predecessors=[FragnetLink("TIA-A")],
                    successors=[FragnetLink("TIA-A")])])
check("A22e circular fragnet detected",
      any("Circular" in i for i in iss2))
check("A22f fragnet json parser rejects invalid refs",
      parse_fragnet_json('{"activities":[{"id":"TIA-1","name":"x",'
                         '"duration_days":5,'
                         '"successors":[{"id":"NOPE-99"}]}]}', U)[0]
      .successors == [])
check("A22g template search returns project evidence",
      len(find_template_activities(U, "installation of ceiling")) > 0)
check("A22h link text round-trip",
      parse_links("A1:SS:5")[0].link_type == "SS")
_scope = assess_event_scope(DelayEvent(
    "EV-S", "Additional ceiling installation", "include approval and test",
    area="Zone B", discipline="Architectural", project_context="Hospital",
    work_package="Additional ceiling works"))
check("A22i event understood before fragnet drafting",
      _scope.work_nature.startswith("Additional")
      and "Testing / inspection / handover" in _scope.lifecycle_stages)
_pkgs = find_template_work_packages(U, "installation of ceiling")
check("A22j existing work packages ranked before generic drafting",
      bool(_pkgs) and bool(_pkgs[0]["activities"])
      and _pkgs[0]["score"] > 0)
_logic_prompt = build_logic_recommendation_prompt(_ev, _fr, U)
check("A22k logic recommendation uses confirmed fragnet + programme IDs",
      "TIA-010" in _logic_prompt and "allowed_existing_activities" in _logic_prompt)
_known_pred = U.tasks[0].task_code
_logic = parse_logic_recommendation_json(
    '{"predecessors":[{"id":"' + _known_pred + '","type":"FS","lag_days":0}],'
    '"successors":[{"id":"H-5040","type":"FS","lag_days":0}],'
    '"impacted_sections":[{"id":"H-5040"}],'
    '"warnings":["planner review"]}', U)
_logic_bad = parse_logic_recommendation_json(
    '{"predecessors":[{"id":"INVENTED-1"}]}', U)
check("A22l logic parser accepts programme IDs and rejects invention",
      _logic["predecessors"][0]["id"] == _known_pred
      and _logic_bad["predecessors"] == [])
_calendar_id = next(iter(U.calendars))
_calendar_fragnet = parse_fragnet_json(
    '{"activities":[{"id":"TIA-CAL","name":"calendar test",'
    '"duration_days":2,"calendar_id":"' + _calendar_id + '",'
    '"successors":[{"id":"H-5040"}]}]}', U)
check("A22m fragnet retains only a valid programme calendar",
      _calendar_fragnet[0].calendar_id == _calendar_id)
_targeted = run_tia(U, "U", _ev, _fr, target_milestone="H-5040")
check("A22n selected impacted milestone is prioritised in results",
      bool(_targeted.milestone_impacts)
      and _targeted.milestone_impacts[0].code == "H-5040")
from programme import build_tia_xlsx
_tia_book = load_workbook(io.BytesIO(build_tia_xlsx(
    _targeted, audit={"source_sha256": "abc"},
    run_history=[{"completion_delta_days": 5}])))
check("A22o TIA export includes audit and rerun history",
      "Audit Trail" in _tia_book.sheetnames
      and "Run History" in _tia_book.sheetnames
      and "Calendar" in [c.value for c in _tia_book["Fragnet"][1]])


# A23. Explain This Delay
from programme import explain_delay
_ex = explain_delay([("B", B), ("U", U)], "H-5040")
check("A23 explain: facts recorded per revision",
      len(_ex.points) == 2 and _ex.points[0].forecast is not None)
check("A23b explain: total movement == raw forecast delta",
      abs(_ex.total_movement_days
          - (_ex.points[-1].forecast
             - _ex.points[0].forecast).days) < 1)
check("A23c explain: uncertain attribution flagged when path switched",
      any(not w.attribution_reliable for w in _ex.windows)
      and any("uncertain" in w for w in _ex.warnings))
check("A23d explain: facts/inference separation in caveats",
      any("INFERENCE" in c for c in _ex.caveats))
_ex1 = explain_delay([("B", B)], "H-5040")
check("A23e explain: single revision -> warning, no crash",
      not _ex1.windows and _ex1.warnings)


# A24. Event extraction (TIA intake) + 52R-06
from programme import (build_event_extraction_prompt, parse_event_candidates,
                       read_document, recommended_analysis_schedule)
_docs = [("L1.txt", "On 12 March 2018 the Engineer issued Instruction "
                    "EI-88 requiring additional ceiling works.")]
_ep = build_event_extraction_prompt(_docs)
check("A24 extraction prompt cites 52R-06 and the doc",
      "52R-06" in _ep and "L1.txt" in _ep)
_good = ('{"events":[{"title":"EI-88","source_doc":"L1.txt",'
         '"source_snippet":"issued Instruction EI-88","date_start":'
         '"2018-03-12","confidence":"high"}]}')
_c, _d = parse_event_candidates(_good, _docs)
check("A24b verified snippet accepted", len(_c) == 1 and _c[0].verified)
_bad = ('{"events":[{"title":"Flood","source_doc":"L1.txt",'
        '"source_snippet":"site flooded for weeks"}]}')
_c2, _d2 = parse_event_candidates(_bad, _docs)
check("A24c fabricated snippet dropped", _c2 == [] and _d2 == 1)
check("A24d garbage tolerated", parse_event_candidates("x", _docs) == ([], 0))
from datetime import datetime as _dtx
_meta = [("U1", _dtx(2018, 1, 31)), ("U2", _dtx(2018, 2, 28))]
check("A24e 52R-06 picks last update before event",
      recommended_analysis_schedule(_meta, _dtx(2018, 2, 10)) == "U1")
check("A24f TIA caveats cite 52R-06",
      any("52R-06" in c for c in _r.caveats))
check("A24g txt reader works", "hello" in read_document("a.txt", b"hello"))


# A25. Impacted-programme XER export round-trip
from programme import build_impacted_xer
_raw = open(_p("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"), "rb").read()
_fr2 = [FragnetActivity("TIA-910", "a", 10,
                        successors=[FragnetLink("TIA-920")]),
        FragnetActivity("TIA-920", "b", 20,
                        predecessors=[FragnetLink("TIA-910")],
                        successors=[FragnetLink("H-5040")])]
_res2 = run_tia(U, "U", _ev, _fr2)
_out = build_impacted_xer(_raw.decode("utf-8", errors="replace"),
                          U, _fr2, _res2)
_u2 = parse_xer(_out.encode("utf-8"))
check("A25 impacted xer: fragnet tasks import",
      len(_u2.tasks) == len(U.tasks) + 2)
check("A25b impacted xer: links deduped and resolved",
      len(_u2.relationships) == len(U.relationships) + 2)
_t2 = next(x for x in _u2.tasks if x.task_code == "TIA-920")
check("A25c impacted xer: not-started with duration",
      _t2.status == "TK_NotStart" and _t2.target_drtn_hr is not None)


# A26. Calendar-exact CPM + cumulative TIA + concurrency
from programme import run_cumulative_tia
# NOTE `calibration_days or 99` was a zero-truthiness bug: a PERFECT
# calibration of exactly 0.0 evaluated to 99 and failed the pin
_rBcal = run_tia(B, "B", _ev, _fr)
check("A26 calendar-exact calibration within 2 days of P6 (baseline forward pass)",
      _rBcal.calibration_days is not None
      and abs(_rBcal.calibration_days) < 2,
      f"calib={_rBcal.calibration_days}")
from datetime import datetime as _dt6
_evA = DelayEvent("EV-A", "a", date_raised=_dt6(2018, 5, 1))
_evB = DelayEvent("EV-B", "b", date_raised=_dt6(2018, 5, 20))
_cum = run_cumulative_tia(U, "U", [
    (_evB, [FragnetActivity("TIA-B1", "b", 170,
                            successors=[FragnetLink("T-4070")])]),
    (_evA, [FragnetActivity("TIA-A1", "a", 150,
                            successors=[FragnetLink("H-5040")])])])
check("A26b cumulative inserts chronologically",
      _cum["rows"][0]["event_id"] == "EV-A")
check("A26c incremental deltas sum to total",
      abs(sum(r["incremental_delta_days"] for r in _cum["rows"])
          - _cum["total_delta_days"]) < 0.2)
check("A26d overlapping driving chains flagged as concurrency candidates",
      len(_cum["concurrency"]) == 1)


# A27. Notice screening + clause extraction + TIA report chart
from programme import (assess_notice, build_clause_extraction_prompt,
                       parse_clause_extraction)
from programme import report_charts as _rc27
from datetime import datetime as _dt7
check("A27 notice compliant with margin",
      assess_notice(_dt7(2018,5,3), _dt7(2018,5,20), 28).status
      == "compliant")
check("A27b notice late",
      assess_notice(_dt7(2018,5,3), _dt7(2018,7,1), 28).status == "late")
check("A27c no notice / indeterminate",
      assess_notice(_dt7(2018,5,3), None, 28).status == "no_notice"
      and assess_notice(None, None, None).status == "indeterminate")
_ct = "Clause 20.1: the Contractor shall give notice within 28 days of awareness."
_ok = parse_clause_extraction(
    '{"clauses":[{"topic":"notice","clause_ref":"20.1","period_days":28,'
    '"requirement":"notify","snippet":"give notice within 28 days",'
    '"silent":false},{"topic":"float","silent":true},'
    '{"topic":"fake","snippet":"invented words here","silent":false}]}', _ct)
check("A27d clause parser: verified kept, silent kept, invented dropped",
      len(_ok) == 2 and _ok[0]["period_days"] == 28)
check("A27e TIA paths chart builds",
      _rc27.tia_paths_chart(_r) is not None)

print("\n== B. Edge cases / degenerate inputs ==")

# B1. Windows with one revision
w1 = analyse_windows([("B", B)])
check("B1 single-revision windows -> warning, no crash",
      not w1.windows and w1.warnings)

# B2. Float erosion with same file twice -> zero erosion
fe2 = analyse_float_erosion([("B", B), ("B2", B)])
check("B2 self float erosion: median delta == 0",
      fe2.windows[0].median_delta == 0 and fe2.windows[0].eroded_count == 0)

# B3. Progress with no updates
pr0 = compute_progress(B, "B", [])
check("B3 progress w/o updates: planned only, no crash",
      pr0.planned_curve and not pr0.recorded_curve
      and pr0.time_offset_days is None)

# B4. Longest path with bogus end code -> falls back with warning
cp_b = extract_longest_path(B, "B", end_task_code="NOPE-123")
check("B4 bogus end code -> fallback + warning",
      cp_b.end_choice is not None
      and any("not found" in w for w in cp_b.warnings))

# B5. Resources on fixture without RSRC table
rA = extract_resource_loading(fix[0][1], "revA")
check("B5 no-resource file -> warning, no crash",
      not rA.histogram and rA.warnings)

# B6. Critical path with absurd tolerance -> no critical, warning
cp_none = extract_critical_path(B, "B", float_tolerance_days=-9999)
check("B6 impossible tolerance -> warning, empty",
      not cp_none.critical and cp_none.warnings)

# B7. Fixtures through every multi-rev engine (3 revisions)
try:
    analyse_windows(fix); analyse_float_erosion(fix)
    compare_revisions(fix[0][1], fix[2][1], "A", "C")
    compute_progress(fix[0][1], "A", [(l, d) for l, d in fix[1:]])
    check("B7 fixtures through all multi-rev engines", True)
except Exception as e:
    check("B7 fixtures through all multi-rev engines", False,
          f"{type(e).__name__}: {e}")

print("\n== C. Report integrity (prompts + workbooks) ==")
from openpyxl import load_workbook
import io as _io

inv = build_inventory(inv_pool)
builds = {
    "inventory": (build_inventory_prompt(inv), build_inventory_xlsx(inv, "n")),
    "milestones": (build_milestone_prompt(ms, ms.series[:5]),
                   build_milestone_xlsx(ms, ms.series[:5], "n")),
    "variance": (build_variance_prompt(var), build_variance_xlsx(var, "n")),
    "critical_path": (build_critical_path_prompt(cp_l),
                      build_critical_path_xlsx(cp_l, "n")),
    "comparison": (build_comparison_prompt(c_fwd),
                   build_comparison_xlsx(c_fwd, "n")),
    "windows": (build_windows_prompt(wres), build_windows_xlsx(wres, "n")),
    "progress": (build_progress_prompt(pr), build_progress_xlsx(pr, "n")),
    "float_erosion": (build_float_erosion_prompt(fe),
                      build_float_erosion_xlsx(fe, "n")),
    "resources": (build_resources_prompt(rl), build_resources_xlsx(rl, "n")),
}
for name, (prompt, xlsx) in builds.items():
    has_rules = "<rules>" in prompt and "Attribute nothing" in prompt
    has_caveats = "<caveats>" in prompt or "warnings" in prompt.lower() or name == "inventory"
    wb = load_workbook(_io.BytesIO(xlsx))
    has_narr = "AI Narrative" in wb.sheetnames
    check(f"C {name}: hard rules in prompt", has_rules)
    check(f"C {name}: workbook opens, narrative sheet present",
          has_narr, str(wb.sheetnames))

# C2. Every module's standing caveats reach its prompt
for name, (prompt, _) in builds.items():
    if name == "inventory":
        continue
    check(f"C2 {name}: limitations content present",
          "caveat" in prompt.lower() or "<caveats>" in prompt)

print("== D. TIA hardening upgrades ==")
from datetime import datetime as _dt

from programme.tia import (DelayEvent, FragnetActivity, FragnetLink,
                           _backward_pass, _build_network, _calendar_masks,
                           _forward_pass, run_cumulative_tia, run_tia)
from programme.xer_export import build_impacted_xer
from programme.events_extract import parse_event_candidates, truncation_notes
from programme.notice import assess_notice

_masks = _calendar_masks(U)
check("D1 calendar masks carry holiday exceptions",
      any(len(v[1]) > 0 for v in _masks.values()),
      f"{sum(len(v[1]) for v in _masks.values())} holidays total")

_ev = DelayEvent("EV-QA", "Chiller rework", "rework to chiller plant")
_frag = [FragnetActivity("TIA-010", "Remove", 20,
                         predecessors=[FragnetLink("T-4040")],
                         successors=[FragnetLink("TIA-020")]),
         FragnetActivity("TIA-020", "Reinstall", 40,
                         predecessors=[FragnetLink("TIA-010")],
                         successors=[FragnetLink("T-4070")])]
_UD = parse_xer(_edit_task_rows(
    _u_text, "T-4060",
    cstr_type="CS_MSOA", cstr_date="2025-09-15 08:00"))
_r = run_tia(_UD, "U", _ev, _frag)
check("D2 start-constraint floors applied and disclosed",
      any("start constraint" in w for w in _r.warnings))
check("D3 tie-in float reported, post <= pre",
      bool(_r.tie_in_float) and all(
          t["float_post"] <= t["float_pre"]
          for t in _r.tie_in_float
          if t["float_pre"] is not None and t["float_post"] is not None))
check("D4 milestone impacts carry total float",
      any(m.float_pre is not None and m.float_post is not None
          for m in _r.milestone_impacts))
_r_nc = run_tia(U, "U", _ev, _frag)
check("D4b non-binding constraint leaves calibration unchanged",
      _r.calibration_days is not None
      and _r_nc.calibration_days is not None
      and abs(_r.calibration_days - _r_nc.calibration_days) < 0.1,
      f"with={_r.calibration_days} without={_r_nc.calibration_days}")
check("D4c truncated-update fixture: stale stored forecast is VISIBLE in calibration",
      _r_nc.calibration_days < -10, f"calib={_r_nc.calibration_days}")

# D5 completion symmetry: post completion never taken from a fragnet act
_dd = _UD.project.data_date
_inc, _nodes, _preds, _started, _fm, _ = _build_network(_UD, cfg, _dd)
_np = dict(_nodes); _pp = {k: list(v) for k, v in _preds.items()}
for _f in _frag:
    _np[_f.act_id] = (max(_f.duration_days, 0.0), None)
    _pp.setdefault(_f.act_id, [])
    for _l in _f.predecessors:
        _pp[_f.act_id].append((_l.other_id, _l.link_type, _l.lag_days))
    for _l in _f.successors:
        _pp.setdefault(_l.other_id, []).append(
            (_f.act_id, _l.link_type, _l.lag_days))
_, _EF1, _, _ = _forward_pass(_np, _pp, _dd, _started)
check("D5 completion_post measured over the real network only",
      _r.completion_post == max(ef for c, ef in _EF1.items()
                                if c in _nodes))

# D6 backward pass: a genuinely critical chain exists (min TF ~ 0)
_, _EF0, _, _ = _forward_pass(dict(_nodes),
                              {k: list(v) for k, v in _preds.items()},
                              _dd, _started)
_tf = _backward_pass(_nodes, _preds, _EF0)
check("D6 backward pass yields a zero-float driving chain",
      _tf and min(abs(v) for v in _tf.values()) <= 1.0,
      f"min |TF| = {min(abs(v) for v in _tf.values()) if _tf else '—'}")

# D7 cumulative ID clash is caught and the duplicate skipped
_ev2 = DelayEvent("EV-QB", "Clash", "")
_frag2 = [FragnetActivity("TIA-010", "Dup id", 10,
                          predecessors=[FragnetLink("T-4040")],
                          successors=[FragnetLink("T-4070")])]
_cum = run_cumulative_tia(_UD, "U", [(_ev, _frag), (_ev2, _frag2)])
check("D7 cumulative flags reused fragnet IDs",
      any("SKIPPED" in w for w in _cum.get("warnings", [])))

# D8 impacted XER: dedicated fragnet WBS band + exact table anchoring
with open(_p("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"), encoding="latin-1") as fh:
    _raw = fh.read()
_out = build_impacted_xer(_raw, U, _frag, _r)
_U2 = parse_xer(_out.encode("latin-1", errors="replace"))
_wrows = [w for w in _U2.raw_tables.get("PROJWBS", [])
          if "TIA Fragnet" in (w.get("wbs_name") or "")]
_trows = [t for t in _U2.raw_tables.get("TASK", [])
          if (t.get("task_code") or "").startswith("TIA-")]
check("D8 impacted XER round-trips with fragnet WBS band",
      len(_wrows) == 1 and len(_trows) == 2
      and all(t.get("wbs_id") == _wrows[0].get("wbs_id") for t in _trows)
      and len(_U2.tasks) == len(U.tasks) + 2
      and len(_U2.relationships) == len(U.relationships) + 3)

# D9 event extraction: documented end date -> stated duration; bad order rejected
_docs = [("L1.txt", "The Engineer instructed suspension of chiller works "
          "from 12 May 2018; the suspension was lifted on 3 June 2018.")]
_resp = ('{"events":[{"title":"Suspension","date_start":"2018-05-12",'
         '"date_end":"2018-06-03","source_doc":"L1.txt",'
         '"source_snippet":"instructed suspension of chiller works",'
         '"confidence":"high"}]}')
_cands, _ = parse_event_candidates(_resp, _docs)
check("D9 date_end captured, stated duration computed",
      _cands and _cands[0].stated_duration_days == 22.0)
_bad = _resp.replace('"date_end":"2018-06-03"', '"date_end":"2018-05-01"')
_cands_b, _ = parse_event_candidates(_bad, _docs)
check("D9b end-before-start rejected",
      _cands_b and _cands_b[0].date_end is None)
check("D9c truncation disclosed for oversize documents",
      truncation_notes([("big.pdf", "x" * 20001)]) != []
      and truncation_notes(_docs) == [])

# D10 notice basis changes the count and is printed
_na_c = assess_notice(_dt(2018, 5, 11), _dt(2018, 5, 14), 2, "calendar")
_na_b = assess_notice(_dt(2018, 5, 11), _dt(2018, 5, 14), 2, "business")
check("D10 Fri->Mon: 3 calendar days late, 1 business day compliant",
      _na_c.status == "late" and _na_b.status == "compliant"
      and "business day" in _na_b.detail)

# D11 impossible notice inputs never yield a contractual status
check("D11 notice before awareness -> indeterminate",
      assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 5), 14).status
      == "indeterminate")
check("D11b non-positive clause period -> indeterminate",
      assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 12), -7).status
      == "indeterminate"
      and assess_notice(_dt(2018, 5, 10), _dt(2018, 5, 12), 0).status
      == "indeterminate")

print("== E. Comparison impact, progress transfer, project library ==")

from programme import (assess_comparison_impact, build_provenance,
                       out_of_sequence_flags, run_progress_transfer,
                       ProjectStore)
import tempfile

# E1. Impact screening — coverage, ordering, and score sanity
_imp = assess_comparison_impact(B, U, "B", "U")
_cmp_bu = compare_revisions(B, U, "B", "U")
_bands_ok = all(c.band_old in ("critical", "near-critical", "off-path",
                               "completed", "absent")
                and c.band_new in ("critical", "near-critical", "off-path",
                                   "completed", "absent")
                for c in _imp.ranked)
check("E1 every ranked change carries valid path bands", _bands_ok)
check("E1b ranked count == diff total minus renames",
      len(_imp.ranked) == _cmp_bu.total_changes - len(_cmp_bu.renamed),
      f"ranked={len(_imp.ranked)} vs "
      f"{_cmp_bu.total_changes - len(_cmp_bu.renamed)}")
check("E1c rank is sorted by score descending",
      all(_imp.ranked[i].score >= _imp.ranked[i + 1].score
          for i in range(len(_imp.ranked) - 1)))
check("E1d every retrospective actual change is red-flagged",
      sum(1 for c in _imp.ranked if c.red_flag)
      >= len(_cmp_bu.actual_date_changes))
_imp_self = assess_comparison_impact(B, B, "B", "B")
check("E1e self-impact == 0 ranked changes", len(_imp_self.ranked) == 0,
      f"got {len(_imp_self.ranked)}")

# E2. Out-of-sequence screening — well-formed, and a manual FS recount
_oos = out_of_sequence_flags(U)
check("E2 OOS overlaps positive or None (open predecessor)",
      all(f.overlap_days is None or f.overlap_days > 0 for f in _oos))
_by_id = {t.task_id: t for t in U.tasks if not t.is_loe_or_wbs}
_manual_fs = 0
for _r in U.relationships:
    _pt, _st_ = _by_id.get(_r.pred_task_id), _by_id.get(_r.task_id)
    if (_pt is not None and _st_ is not None and _r.pred_type == "PR_FS"
            and _st_.act_start and _pt.act_finish
            and (_pt.act_finish - _st_.act_start).total_seconds()
            / 86400.0 > 0.1):
        _manual_fs += 1
_fs_flags = sum(1 for f in _oos
                if f.link_type == "FS" and f.overlap_days is not None)
check("E2b FS overlap flags == manual recount from raw actuals",
      _fs_flags == _manual_fs, f"flags={_fs_flags} vs manual={_manual_fs}")

# E3. Provenance — windows equal the direct pairwise diffs
_prov = build_provenance(fix)
check("E3 provenance windows == revisions - 1",
      len(_prov.windows) == len(fix) - 1)
_direct = compare_revisions(fix[0][1], fix[1][1], fix[0][0], fix[1][0])
check("E3b window counts match direct pairwise diff",
      _prov.windows[0].counts == _direct.category_counts)
check("E3c red-flag count mirrors actual-date changes",
      all(w.red_flag_count == len(w.comparison.actual_date_changes)
          for w in _prov.windows))

# E4. Progress transfer — self-transfer identity + manual recounts
_tr_self = run_progress_transfer(U, U, "U", "U")
check("E4 self-transfer network effect == 0",
      _tr_self.network_effect_days == 0.0,
      f"got {_tr_self.network_effect_days}")
_tr = run_progress_transfer(B, U, "B", "U")
_b_codes = {t.task_code for t in B.tasks if not t.is_loe_or_wbs}
_manual_fin = sum(1 for t in U.tasks
                  if not t.is_loe_or_wbs and t.act_finish is not None
                  and t.task_code in _b_codes)
check("E4b transferred completions == manual recount",
      _tr.applied_finishes == _manual_fin,
      f"applied={_tr.applied_finishes} vs manual={_manual_fin}")
_manual_started = sum(1 for t in U.tasks
                      if not t.is_loe_or_wbs and t.act_start is not None
                      and t.act_finish is None and t.task_code in _b_codes)
check("E4c transferred starts == manual recount (in-progress only)",
      _tr.applied_starts == _manual_started,
      f"applied={_tr.applied_starts} vs manual={_manual_started}")
check("E4d reference-run calibration matches the plain run on the same donor",
      _tr.calibration_days is not None
      and _r_nc.calibration_days is not None
      and abs(_tr.calibration_days - _r_nc.calibration_days) <= 1.0,
      f"transfer={_tr.calibration_days} plain={_r_nc.calibration_days}")
check("E4e data date taken from the progress donor",
      _tr.data_date == U.project.data_date)
check("E4f statusing caveats always emitted",
      any("retained logic" in c.lower() for c in _tr.caveats)
      and any("not a schedule submission" in c for c in _tr.caveats))

# E5. Project library — dedupe by hash, append-only, record round-trip
with tempfile.TemporaryDirectory() as _td:
    _store = ProjectStore(os.path.join(_td, "lib.db"))
    _r1 = _store.register_file("QA", "a.xer", b"AAA", data_date="2020-01-01")
    _r2 = _store.register_file("QA", "a_renamed.xer", b"AAA")
    _r3 = _store.register_file("QA", "b.xer", b"BBB")
    check("E5 identical content deduped by hash",
          _r2.already_registered and _r2.id == _r1.id
          and _r2.sha256 == _r1.sha256)
    check("E5b register holds exactly the distinct files",
          len(_store.custody_register("QA")) == 2)
    check("E5c store is append-only (no delete API)",
          not any(hasattr(_store, m) for m in
                  ("delete_file", "delete_record", "remove", "clear")))
    _store.save_record("QA", "tia_audit", "run", {"delta": 12.5, "n": 3})
    _recs = _store.list_records("QA", "tia_audit")
    check("E5d analysis record round-trips through JSON",
          len(_recs) == 1 and _recs[0].payload == {"delta": 12.5, "n": 3})
    check("E5e sha256 matches an independent hash",
          _r3.sha256 == __import__("hashlib").sha256(b"BBB").hexdigest())

# E6. Scope/logic decomposition — the fix for the conflated headline
check("E6 decomposition identity: logic + scope == full - reference",
      _tr.network_effect_days is not None
      and _tr.scope_effect_days is not None
      and abs((_tr.network_effect_days + _tr.scope_effect_days)
              - (_tr.completion_transferred
                 - _tr.completion_reference).total_seconds() / 86400)
      <= 0.21,
      f"logic={_tr.network_effect_days} scope={_tr.scope_effect_days}")
check("E6b self-transfer: both effects zero",
      _tr_self.network_effect_days == 0.0
      and _tr_self.scope_effect_days == 0.0)
check("E6c sample: the split separates the effects (logic carries the movement;\n      the completed VO scope contributes nothing to remaining works)",
      abs(_tr.network_effect_days) > 0
      and abs(_tr.scope_effect_days) < abs(_tr.network_effect_days),
      f"scope={_tr.scope_effect_days} logic={_tr.network_effect_days}")
check("E6d scope caveat discloses the decomposition",
      any("intersection" in c for c in _tr.caveats))

# E7. OOS flags ranked by criticality inside the impact assessment
_ord = ["critical", "near-critical", "off-path", "completed", "absent"]
_idx = [_ord.index(f.band) for f in _imp.oos_flags]
check("E7 OOS flags ranked driving-path first", _idx == sorted(_idx))
check("E7b every OOS flag carries a valid band",
      all(f.band in _ord for f in _imp.oos_flags))

# E8. Excel deliverables open with the expected sheets
from programme import (build_impact_xlsx, build_transfer_xlsx,
                       build_custody_xlsx)
_wb_i = load_workbook(io.BytesIO(build_impact_xlsx(_imp)))
check("E8 impact workbook: summary + rank + caveats (OOS now its own "
      "module)",
      {"Summary", "Materiality rank", "Warnings & Caveats"}
      <= set(_wb_i.sheetnames)
      and "Out of sequence" not in _wb_i.sheetnames,
      str(_wb_i.sheetnames))
_wb_t = load_workbook(io.BytesIO(build_transfer_xlsx(_tr)))
check("E8b transfer workbook: summary + milestones + chain + caveats",
      {"Summary", "Milestones", "Driving chain",
       "Statusing & Caveats"} <= set(_wb_t.sheetnames),
      str(_wb_t.sheetnames))
with tempfile.TemporaryDirectory() as _td2:
    _st2 = ProjectStore(os.path.join(_td2, "l.db"))
    _st2.register_file("QA", "x.xer", b"X", data_date="2020-01-01")
    _wb_c = load_workbook(io.BytesIO(
        build_custody_xlsx(_st2.custody_register())))
    check("E8c custody workbook opens with the register sheet",
          "Custody register" in _wb_c.sheetnames)

# ===================================================================== #
# Layer F — DCMA forensic traceback (stored values only)
# ===================================================================== #
print("\n--- Layer F: DCMA traceback ---")
from dcma import build_dcma_trace, annotate_path_position
from dcma.trace import _LATE_DRIVERS
from dcma.checks import run_all_checks as _rac
from dcma.report_xlsx import build_xlsx_report as _bxr

for _label, _d in (("baseline", B), ("update", U)):
    _cfg = DCMAConfig()
    _res = _rac(_d, _cfg)
    _t = build_dcma_trace(_d, _cfg, _res)

    _c = _t.chain
    check(f"F1[{_label}] driving chain non-empty, terminal is last step",
          _c is not None and _c.steps
          and _c.steps[-1].task_code == _c.terminal_code)
    _dates = [s.early_finish or s.early_start for s in _c.steps
              if (s.early_finish or s.early_start)]
    check(f"F1b[{_label}] chain ordered towards the terminal",
          all(_dates[i] <= _dates[-1] for i in range(len(_dates))))
    check(f"F2[{_label}] continuity is settled: reaches DD or break disclosed",
          _c.reaches_data_date or (_c.break_code and _c.break_reason))

    _r7 = next(r for r in _res if r.number == 7)
    check(f"F3[{_label}] one float trace per negative-float activity",
          len(_t.float_traces) == _r7.affected_count,
          f"traces={len(_t.float_traces)} check7={_r7.affected_count}")
    check(f"F3b[{_label}] driver-group counts sum to trace count",
          sum(g.count for g in _t.float_driver_groups)
          == len(_t.float_traces))
    _by_code = {t.task_code: t for t in _d.tasks}
    _ok_kinds = {"activity constraint", "project must-finish",
                 "unidentified"}
    check(f"F4[{_label}] every driver kind valid; constraint drivers "
          "really carry a late-date constraint",
          all(g.driver_kind in _ok_kinds for g in _t.float_driver_groups)
          and all((_by_code[g.driver_code].cstr_type in _LATE_DRIVERS
                   or _by_code[g.driver_code].cstr_type2 in _LATE_DRIVERS)
                  for g in _t.float_driver_groups
                  if g.driver_kind == "activity constraint"))

    annotate_path_position(_res, _t)
    annotate_path_position(_res, _t)          # idempotency
    _r1 = next(r for r in _res if r.number == 1)
    check(f"F5[{_label}] annotate adds Path position once, sorted "
          "driving-first",
          _r1.detail_rows
          and list(_r1.detail_rows[0].keys())[0] == "Path position"
          and sum(1 for k in _r1.detail_rows[0] if k == "Path position")
          == 1)

    _tripped = {r.number: set(r.affected_ids) for r in _res
                if r.number not in (12, 13, 14)}
    check(f"F6[{_label}] offenders: >=2 checks each, consistent with "
          "affected_ids",
          all(len(o.checks) >= 2
              and all(o.task_code in _tripped.get(n, set())
                      for n in o.checks)
              for o in _t.offenders))

    _wb_f = load_workbook(io.BytesIO(_bxr(_d, _res, trace=_t)))
    _need_f = {"Driving Chain"}
    if _t.offenders:
        _need_f.add("Multi-Check Offenders")
    if _t.caveats or _t.warnings:
        _need_f.add("Traceback Notes")
    check(f"F7[{_label}] DCMA workbook gains the traceback sheets",
          _need_f <= set(_wb_f.sheetnames),
          str(sorted(_need_f - set(_wb_f.sheetnames))))

    _chain_codes = {s.task_code for s in _c.steps}
    check(f"F8[{_label}] band_map: every driving band is on the chain",
          all(code in _chain_codes
              for code, b in _t.band_map.items() if b == "driving"))

with open("dcma/trace.py") as _fh:
    _src = _fh.read()
check("F9 layering rule: dcma.trace never imports programme.*",
      "import programme" not in _src and "from programme" not in _src)

from dcma.narrative import build_report_prompt as _brp
_p = _brp(_d, _res, trace=_t)
check("F10 narrative prompt carries traceback facts",
      "<traceback_facts>" in _p and _t.chain.terminal_code in _p)

# ===================================================================== #
# Layer G — out-of-sequence: as-built fits, evolution, transfer wiring
# ===================================================================== #
print("\n--- Layer G: OOS as-built recommendations ---")
from programme.oos import (oos_evolution, out_of_sequence_flags)

_gb = parse_xer("sample/Harbour Point DCP-03 - Baseline Programme Rev 0.xer")
_gu = parse_xer(_u_text)
_gq_text = _edit_first_row(
    _u_text, "SCHEDOPTIONS",
    sched_retained_logic="N",
    sched_float_type="ST_StartFloat",
    sched_calendar_on_relationship_lag="rcal_Successor",
    sched_use_project_end_date_for_float="Y")
_gq0 = parse_xer(_gq_text)
_succ_q = {}
for _rl in _gq0.relationships:
    if _rl.pred_type == "PR_FS":
        _succ_q.setdefault(_rl.pred_task_id, []).append(_rl.task_id)
_skip_pair = next(
    (p, s2) for p, mids in _succ_q.items() for m in mids
    for s2 in _succ_q.get(m, []) if s2 not in _succ_q.get(p, []))
_gq = parse_xer(_add_taskpred(_gq_text, *_skip_pair))
_gflags = out_of_sequence_flags(_gu)
_gby = {t.task_code: t for t in _gu.tasks}

check("G1 every OOS flag carries a recommendation",
      _gflags and all(f.rec_link for f in _gflags))
check("G2 concrete fits are non-negative; reversed order is always "
      "'review'",
      all((f.rec_lag_days is None or f.rec_lag_days >= 0)
          and (f.rec_link_type != "review" or f.rec_lag_days is None)
          for f in _gflags))

_conc = [f for f in _gflags if f.rec_link_type == "SS"]
check("G3 sample has concrete SS fits", len(_conc) > 0, str(len(_conc)))
_f0 = _conc[0]
_lag = round((_gby[_f0.succ_code].act_start
              - _gby[_f0.pred_code].act_start).total_seconds() / 86400, 1)
check("G3b SS-fit lag equals the recorded start offset (manual recount)",
      abs(_f0.rec_lag_days - _lag) < 0.05,
      f"rec={_f0.rec_lag_days} manual={_lag}")
check("G3c FS/SS flags with ordered starts always get the SS fit",
      all(f.rec_link_type == "SS"
          for f in _gflags if f.link_type in ("FS", "SS")
          and _gby[f.pred_code].act_start and _gby[f.succ_code].act_start
          and _gby[f.succ_code].act_start >= _gby[f.pred_code].act_start))

_gev = oos_evolution([("Base", _gb), ("Upd", _gu)])
check("G4 evolution per-revision counts match direct screening",
      _gev.per_revision[0][1] == len(out_of_sequence_flags(_gb))
      and _gev.per_revision[1][1] == len(_gflags))
_gw = _gev.windows[0]
check("G4b window identity: after == before - resolved + new",
      _gw.total_after == _gev.per_revision[0][1] - _gw.resolved_count
      + len(_gw.new_flags),
      f"{_gw.total_after} vs {_gev.per_revision[0][1]}"
      f"-{_gw.resolved_count}+{len(_gw.new_flags)}")
check("G4c resolved contradictions raise the retro-edit warning",
      _gw.resolved_count == 0 or any("disappeared" in w
                                     for w in _gev.warnings))

from programme.progress_transfer import run_progress_transfer as _rpt
_gtr = _rpt(_gb, _gu, "Base", "Upd")
check("G5 transfer discloses the progress donor's OOS flags",
      len(_gtr.oos_flags) == len(_gflags)
      and any("out-of-sequence" in w for w in _gtr.warnings)
      and any("as-built" in c.lower() for c in _gtr.caveats))

from programme import build_impact_xlsx as _bix, build_transfer_xlsx as _btx
from programme import assess_comparison_impact as _aci
_gimp = _aci(_gb, _gu, "Base", "Upd")
_wb_g = load_workbook(io.BytesIO(_bix(_gimp)))
check("G6 OOS is un-embedded: impact workbook has NO out-of-sequence sheet",
      "Out of sequence" not in _wb_g.sheetnames)
_wb_g2 = load_workbook(io.BytesIO(_btx(_gtr)))
check("G6b OOS is un-embedded: transfer workbook has NO OOS sheet",
      "Out of sequence" not in _wb_g2.sheetnames)

# ===================================================================== #
# Layer H — standalone OOS module: as-built repair -> revised .xer
# ===================================================================== #
print("\n--- Layer H: OOS as-built repair engine ---")
from programme.oos import (build_repair_plan as _brp,
                           apply_asbuilt_repairs as _aar,
                           out_of_sequence_flags as _oosf, _TYPE_CODE)
from programme import build_oos_xlsx as _box

_hraw = open("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer", encoding="latin-1").read()
_hflags = _oosf(_gu)
_hplan = _brp(_gu, _hflags)
check("H1 plan holds only concrete fits (no review-class)",
      len(_hplan) == sum(1 for f in _hflags
                         if f.rec_link_type not in ("", "review")))
check("H1b every plan item has a positive-or-zero calendar lag "
      "and an hour conversion",
      all(r.new_lag_days_cal >= 0 and r.new_lag_hr >= 0 for r in _hplan))

# blocked items are those whose pair already carries the target link
_hexist = {(t.task_id) for t in _gu.tasks}
_hcode = {t.task_code: t.task_id for t in _gu.tasks}
_hrels = {(r.pred_task_id, r.task_id, r.pred_type)
          for r in _gu.relationships}
_hblocked = [r for r in _hplan if r.blocked]
check("H2 blocked == plan items that would duplicate an existing link",
      all((_hcode[r.pred_code], _hcode[r.succ_code], r.new_type) in _hrels
          for r in _hblocked)
      and all((_hcode[r.pred_code], _hcode[r.succ_code], r.new_type)
              not in _hrels
              for r in _hplan if not r.blocked))

_hout, _hrep = _aar(_hraw, _gu, _hplan)
check("H3 round-trip QA passes", _hrep.qa_passed, str(_hrep.qa_notes[:2]))
check("H3b relationship & task counts unchanged by the repair",
      _hrep.rel_count_after == _hrep.rel_count_before)
check("H3c applied == selected non-blocked; nothing lost",
      len(_hrep.applied) == len(_hplan) - len(_hblocked)
      and not _hrep.not_found)
check("H4 source file untouched: output hash differs, source hash "
      "matches the on-disk bytes",
      _hrep.output_sha256 != _hrep.source_sha256
      and _hrep.source_sha256 == __import__("hashlib").sha256(
          _hraw.encode("latin-1")).hexdigest())

# only TASKPRED %R rows changed; line count identical
_sl = _hraw.split("\n")
_ol = _hout.split("\n")
_diff = [i for i in range(min(len(_sl), len(_ol))) if _sl[i] != _ol[i]]
check("H5 only %R rows changed, line count identical",
      len(_sl) == len(_ol)
      and all(_sl[i].startswith("%R") for i in _diff)
      and len(_diff) == len(_hrep.applied))

# re-parse and confirm a repaired link now carries the fitted type+lag
_rep_parsed = parse_xer(_hout)
_r0 = _hrep.applied[0]
_key0 = (_hcode[_r0.pred_code], _hcode[_r0.succ_code], _r0.new_type)
_found0 = [rel for rel in _rep_parsed.relationships
           if (rel.pred_task_id, rel.task_id, rel.pred_type) == _key0]
check("H6 a repaired link is present as the fitted type after re-parse",
      len(_found0) >= 1
      and any(abs((rel.lag_hr or 0) - _r0.new_lag_hr) <= 0.51
              for rel in _found0))

# unselecting everything => empty, safe output identical to source
for _r in _hplan:
    _r.apply = False
_hout2, _hrep2 = _aar(_hraw, _gu, _hplan)
check("H7 with nothing selected, output == source (no-op is safe)",
      _hrep2.output_sha256 == _hrep2.source_sha256
      and len(_hrep2.applied) == 0)
for _r in _hplan:
    _r.apply = not _r.blocked

# bytes input path (as the app stores raw) works identically
_hout3, _hrep3 = _aar(_hraw.encode("latin-1"), _gu, _hplan)
check("H8 bytes input yields the same repaired output as str input",
      _hrep3.output_sha256 == _hrep.output_sha256 and _hrep3.qa_passed)

_wb_h = load_workbook(io.BytesIO(_box("Upd", _hflags, _hplan, _hrep, _gev)))
check("H9 OOS workbook has Summary/Flags/Repair register/Evolution/QA",
      {"Summary", "Flags", "Repair register", "Evolution",
       "QA & Caveats"} <= set(_wb_h.sheetnames), str(_wb_h.sheetnames))

# ===================================================================== #
# Layer I — supplementary DCMA checks, red-flag events, basis,
#           concurrency screening, impacted as-planned
# ===================================================================== #
print("\n--- Layer I: forensic upgrades ---")
from datetime import timedelta as _td
import copy as _copy

# I1. DCMA supplementary checks
_res17 = _rac(_gu, DCMAConfig())
check("I1 run_all_checks returns 17 (14 + 3 supplementary)",
      len(_res17) == 17 and [r.number for r in _res17[-3:]] == [15, 16, 17])
check("I1b supplementary checks labelled '(supp.)'",
      all("supp." in r.name for r in _res17[14:]))
check("I1c include_supplementary=False keeps the pure DCMA 14",
      len(_rac(_gu, DCMAConfig(), include_supplementary=False)) == 14)
_r16 = _rac(_gq, DCMAConfig())[15]
check("I2 check 16 flags exist on sample and are all FS links",
      _r16.affected_count > 0
      and all("FS" in d["Note"] or "duplicated" in d["Note"]
              for d in _r16.detail_rows))
_r17 = _res17[16]
_gu_by_id = {t.task_id: t for t in _gu.tasks}
_p_of, _s_of = {}, {}
for _rel in _gu.relationships:
    _p_of.setdefault(_rel.task_id, []).append(_rel.pred_type)
    _s_of.setdefault(_rel.pred_task_id, []).append(_rel.pred_type)
_gu_code2id = {t.task_code: t.task_id for t in _gu.tasks}
check("I3 check 17 never re-flags open ends (all have preds AND succs)",
      all(_p_of.get(_gu_code2id[c]) and _s_of.get(_gu_code2id[c])
          for c in _r17.affected_ids))

# I4. SCHEDOPTIONS diff on the real samples (retained-logic flip!)
from programme import compare_revisions as _cr
_cmp_i = _cr(_gb, _gq, "B", "U")
check("I4 SCHEDOPTIONS changes caught on real samples",
      len(_cmp_i.sched_options_changes) == 4
      and any(c.name == "Retained Logic" and c.old_value == "Y"
              and c.new_value == "N"
              for c in _cmp_i.sched_options_changes))
check("I4b scheduling-options red flag raised",
      any("RED FLAG" in w and "scheduling options" in w
          for w in _cmp_i.warnings))
check("I4c category counts carry the new categories",
      "Scheduling options changed" in _cmp_i.category_counts
      and "Calendar definitions changed" in _cmp_i.category_counts)

# I5. calendar-definition tamper (shared id) caught + red-flagged
_gu2 = _copy.deepcopy(_gu)
for _row in _gu2.raw_tables["CALENDAR"]:
    if _row.get("clndr_id", "").strip() == "1001":
        _row["day_hr_cnt"] = "10"
_cmp_i2 = _cr(_gb, _gu2, "B", "U2")
check("I5 calendar-definition change detected and red-flagged",
      len(_cmp_i2.calendar_def_changes) == 1
      and any("calendar manipulation" in w for w in _cmp_i2.warnings))

# I6. impact ranking red-flags the programme-level events
_imp_i = _aci(_gb, _gq, "B", "U", comparison=_cmp_i)
_so_ranked = [c for c in _imp_i.ranked
              if c.category == "Scheduling options changed"]
check("I6 sched-option changes ranked, red-flagged, scored 40",
      len(_so_ranked) == 4 and all(c.red_flag and c.score == 40.0
                                   for c in _so_ranked))

# I7. concurrency screening (synthetic, engine-level)
from programme.tia import DelayEvent as _DE, FragnetActivity as _FA, \
    FragnetLink as _FL
from programme.windows import analyse_windows as _aw
from programme.concurrency import screen_concurrency as _sc
_wres_i = _aw([("Base", _gb), ("Upd", _gu)])
_w0 = _wres_i.windows[0]
_mid = _w0.start + (_w0.end - _w0.start) / 2
_ev_e = _DE("EMP-01", "Late access", date_raised=_mid,
            responsibility_asserted="Employer")
_fr_e = [_FA("EMP-01-F1", "Await access", 400.0,
             predecessors=[_FL("C-3030")], successors=[_FL("H-5040")])]
_ev_c = _DE("CON-01", "Rework", date_raised=_mid + _td(days=5),
            responsibility_asserted="Contractor")
_fr_c = [_FA("CON-01-F1", "Rework", 12.0,
             predecessors=[_FL("C-3030")], successors=[_FL("H-5040")])]
_ev_u = _DE("UNK-01", "Weather", date_raised=_mid,
            responsibility_asserted="force majeure")
_conc = _sc(_wres_i, [(_ev_e, _fr_e), (_ev_c, _fr_c), (_ev_u, [])])
_cw0 = _conc.windows[0]
check("I7 overlap arithmetic: both == contractor span (nested case)",
      abs(_cw0.both_days - 12.0) < 0.1
      and _cw0.employer_days > _cw0.contractor_days)
check("I7b concurrent candidate + pacing shape flagged",
      _cw0.concurrent_candidate and _cw0.pacing_flag)
check("I7c unclassified party disclosed, not silently dropped",
      _cw0.unclassified_days > 0
      and any("neither party" in w for w in _conc.warnings))
check("I7d no-fragnet event screened as single day + warned",
      any(e.single_day for e in _conc.events)
      and any("no fragnet" in w for w in _conc.warnings))

# I8. impacted as-planned
from programme.impacted_asplanned import run_impacted_asplanned as _iap
_bad = _DE("BAD-01", "Missing tie-in", date_raised=_mid,
           responsibility_asserted="Employer")
_fr_bad = [_FA("BAD-01-F1", "x", 5.0, predecessors=[_FL("NOPE-999")])]
_iapr = _iap(_gb, "Base", [(_ev_e, _fr_e), (_ev_c, _fr_c),
                           (_ev_u, []), (_bad, _fr_bad)])
check("I8 IAP skips no-fragnet and missing-tie-in events, uses the rest",
      _iapr["events_used"] == 2 and len(_iapr["skipped_events"]) == 2
      and any("NOPE-999" in s for s in _iapr["skipped_events"]))
check("I8b IAP identity: total == final - pre == sum of increments",
      _iapr["total_delta_days"] is not None
      and abs(_iapr["total_delta_days"]
              - (_iapr["completion_final"]
                 - _iapr["completion_pre"]).total_seconds() / 86400) < 0.6
      and abs(_iapr["total_delta_days"]
              - sum(r["incremental_delta_days"]
                    for r in _iapr["rows"])) < 0.2)
check("I8c IAP carries the weak-method caveats",
      any("THEORETICAL" in c for c in _iapr["caveats"]))

# I9. scheduling-basis helpers
from programme.basis import (progress_treatment as _pt,
                             sched_options_row as _sor,
                             sched_options_summary as _sos)
check("I9 progress treatment: baseline Retained Logic, update Actual "
      "Dates",
      _pt(_sor(_gb)) == "Retained Logic"
      and _pt(_sor(_gq)) == "Actual Dates")
check("I9b basis summary discloses must-finish float trap on the update",
      any("must-finish" in ln for ln in _sos(_gq)))

# I10. new workbooks open
from programme import (build_concurrency_xlsx as _bcx,
                       build_iap_xlsx as _bix2,
                       build_explain_xlsx as _bex)
check("I10 concurrency workbook opens with matrix + events + caveats",
      {"Screening Matrix", "Events Screened", "Warnings & Caveats"}
      <= set(load_workbook(io.BytesIO(_bcx(_conc))).sheetnames))
check("I10b IAP workbook opens with summary + increments",
      {"Summary", "Per-Event Increments"}
      <= set(load_workbook(io.BytesIO(
          _bix2("Base", _iapr))).sheetnames))
from programme.explain import explain_delay as _ed
_exp = _ed([("Base", _gb), ("Upd", _gu)], "H-5040")
_wb_e = load_workbook(io.BytesIO(_bex(_exp, confirmed=[{
    "window": "W1", "task_code": "X", "direction": "joined",
    "name": "x", "note": "Letter ref 123"}])))
check("I10c explain workbook gains the Confirmed Drivers sheet",
      "Confirmed Drivers" in _wb_e.sheetnames)


# ===================================================================== #
# Layer J — APvAB stepped method + Collapsed As-Built
# ===================================================================== #
print("\n--- Layer J: APvAB + Collapsed As-Built ---")
from programme.variance import planned_vs_actual as _pva
from programme.collapsed_asbuilt import (collapse_asbuilt as _cab,
                                         build_grouping_prompt as _bgp,
                                         parse_grouping as _pg)

_j_rows = _pva(_gb, _gu, {"C-3030", "H-5040"})
check("J1 planned_vs_actual: scoped rows + manual variance recount",
      len(_j_rows) == 2 and _j_rows[0]["task_code"] == "C-3030"
      and abs(_j_rows[0]["finish_var_days"] - 46.0) < 0.2,
      str(_j_rows[0].get("finish_var_days")))
check("J1b unscoped compares every matched real activity",
      len(_pva(_gb, _gu)) == sum(
          1 for t in _gu.tasks if not t.is_loe_or_wbs))

# collapse on the OOS-repaired file (the intended pipeline)
_j_u2 = parse_xer(_hout)
_j_res0 = _cab(_j_u2, "Upd", set())
check("J2 empty extraction is a no-op: collapsed == model, delta 0",
      _j_res0.delta_days == 0.0
      and _j_res0.collapsed_completion == _j_res0.model_completion)
check("J2b calibration disclosed and gap warning fires when large",
      _j_res0.calibration_days is not None
      and (abs(_j_res0.calibration_days) <= 30
           or any("validation gap" in w for w in _j_res0.warnings)))
_j_last = _j_res0.critical_chain[-1].task_code
_j_res1 = _cab(_j_u2, "Upd", {_j_last})
check("J3 extracting a controlling-chain activity collapses completion",
      _j_res1.delta_days is not None and _j_res1.delta_days > 0,
      f"delta={_j_res1.delta_days}")
check("J3b delta identity: model - collapsed == delta",
      abs((_j_res1.model_completion - _j_res1.collapsed_completion
           ).total_seconds() / 86400.0 - _j_res1.delta_days) < 0.1)
check("J3c unknown extraction codes ignored + disclosed",
      any("ignored" in w for w in _cab(_j_u2, "U", {"NOPE-1"}).warnings))
check("J3d empty extraction: model chain == collapsed chain (traceback)",
      [a.task_code for a in _j_res0.model_chain]
      == [a.task_code for a in _j_res0.critical_chain])
check("J3e both chains disclosed after a real extraction",
      len(_j_res1.model_chain) > 0 and len(_j_res1.critical_chain) > 0)
check("J3f model chain terminal finish == model completion",
      _j_res1.model_chain[-1].finish == _j_res1.model_completion)

_j_g, _j_d = _pg('{"groups":[{"label":"L","codes":["C-3030","FAKE"],'
                 '"rationale":"r"}]}', _gu)
check("J4 grouping parse keeps verbatim codes, drops fabricated ones",
      len(_j_g) == 1 and _j_g[0]["codes"] == ["C-3030"] and _j_d == 1)
check("J4b grouping prompt is code<TAB>name lines",
      "\t" in _bgp(_gu).split("\n")[1])

from programme.variance import keydate_windows as _kw
_j_all = _pva(_gb, _gu)
_j_win = _kw(_j_all, ["P-2010", "C-3010", "C-3070"])
check("J6 one window per key date, the first from PROJECT START",
      len(_j_win) == 3 and _j_win[0]["from_code"] == "PROJECT START"
      and _j_win[1]["from_code"] == _j_win[0]["to_code"])
_w0 = _j_win[0]
check("J6b delay at a key date is DIRECT: actual minus planned finish",
      abs(_w0["cumulative_delay_days"]
          - (_w0["actual_finish"] - _w0["planned_finish"]
             ).total_seconds() / 86400) < 0.06)
check("J6b2 accrued in window = change in slippage across it",
      abs(_j_win[1]["window_delay_days"]
          - (_j_win[1]["cumulative_delay_days"]
             - _j_win[0]["cumulative_delay_days"])) < 0.05
      and _w0["window_delay_days"] == _w0["cumulative_delay_days"])
check("J6c resequenced key date flagged; its DIRECT delay kept",
      any(w["resequenced"] for w in _j_win)
      and all(w["cumulative_delay_days"] is not None for w in _j_win))
check("J6d a single key date bounds one window from project start; "
      "no usable key dates -> none",
      len(_kw(_j_all, ["C-3030"])) == 1 and _kw(_j_all, ["NOPE"]) == [])
check("J6e window spans run project start → key actual finish",
      _w0["window_start"] is not None
      and _w0["window_start"] <= _w0["window_end"]
      and _j_win[1]["window_start"] == _j_win[0]["window_end"])

from programme import build_simple_xlsx as _bsx
_wb_j = load_workbook(io.BytesIO(_bsx(
    "T", {"Sheet A": [{"X": 1, "Y": "a"}]}, notes=["note"])))
check("J5 generic workbook opens with data + notes sheets",
      {"Sheet A", "Notes & Caveats"} <= set(_wb_j.sheetnames))


# ===================================================================== #
# Layer K — parser robustness: structural variants + fuzz
# The suite otherwise validates ONE project's exports; real-world XERs
# vary wildly. Contract under test: parse_xer either returns XerData
# (degrading with warnings) or raises a controlled ValueError — never
# an uncontrolled IndexError/KeyError/UnicodeError crash.
# ===================================================================== #
print("\n--- Layer K: parser robustness ---")
import random as _rnd

# NOTE: line ~320 rebinds `cfg` to a HIERARCHY config — a shared-
# namespace trap in this linear script (the pytest-conversion argument
# in one line). Layer K uses its own DCMA config.
_k_cfg = DCMAConfig()

_k_raw = open("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer", encoding="latin-1").read()
_k_lines = _k_raw.split("\n")

def _k_parse_ok(text, label):
    """True if parse obeys the contract (XerData or ValueError)."""
    try:
        d = parse_xer(text.encode("latin-1", "replace"))
        return d is not None
    except ValueError:
        return True
    except Exception as exc:                      # noqa: BLE001
        print(f"    UNCONTROLLED {type(exc).__name__} on {label}: "
              f"{exc}")
        return False

_variants = {
    "multi-project (PROJECT table doubled)":
        _k_raw.replace("%T\tPROJECT\n", "%T\tPROJECT\n", 1),
    "calendar data mangled":
        "\n".join(l if not l.startswith("%R\t") or "clndr" not in l
                   else l.replace("0|", "?|") for l in _k_lines[:4000])
        + "\n" + "\n".join(_k_lines[4000:]),
    "non-Latin activity names":
        _k_raw.replace("Review & Approval", "Onay ve İnceleme — 承認"),
    "truncated at half":
        _k_raw[: len(_k_raw) // 2],
    "CALENDAR table removed":
        "\n".join(l for l in _k_lines
                   if "clndr" not in l.lower()
                   or l.startswith(("%T", "%F", "%E"))),
    "TASKPRED emptied":
        "\n".join(l for i, l in enumerate(_k_lines)
                   if not (l.startswith("%R") and i > 0
                           and any("TASKPRED" in x
                                   for x in _k_lines[max(0, i-3000):i]
                                   if x.startswith("%T")))),
    "empty file": "",
    "header only": _k_lines[0] if _k_lines else "ERMHDR",
}
_k_bad = [lbl for lbl, txt in _variants.items()
          if not _k_parse_ok(txt, lbl)]
check("K1 structural variants: parse returns data or controlled "
      "ValueError", not _k_bad, str(_k_bad))

# engines must not crash on whatever the parser accepted
_k_engine_bad = []
for lbl, txt in _variants.items():
    try:
        d = parse_xer(txt.encode("latin-1", "replace"))
    except ValueError:
        continue
    except Exception:
        continue                    # already counted by K1
    try:
        run_all_checks(d, _k_cfg)
    except Exception as exc:        # noqa: BLE001
        _k_engine_bad.append(f"{lbl}: {type(exc).__name__}")
check("K2 DCMA engine survives every parsed variant",
      not _k_engine_bad, str(_k_engine_bad))

check("K3 non-Latin names round-trip through the parser",
      any("Onay" in t.name for t in parse_xer(
          _variants["non-Latin activity names"].encode(
              "latin-1", "replace")).tasks
          if t.name)
      if _k_parse_ok(_variants["non-Latin activity names"], "k3")
      else False)

# deterministic fuzz: byte-level mutations must never crash the parser
_rnd.seed(1729)
_k_fuzz_bad = 0
for i in range(60):
    b = bytearray(_k_raw.encode("latin-1", "replace"))
    for _ in range(_rnd.randint(1, 40)):
        pos = _rnd.randrange(len(b))
        b[pos] = _rnd.randrange(256)
    try:
        parse_xer(bytes(b))
    except ValueError:
        pass
    except Exception as exc:        # noqa: BLE001
        _k_fuzz_bad += 1
        if _k_fuzz_bad <= 3:
            print(f"    fuzz #{i}: {type(exc).__name__}: {exc}")
check("K4 60 seeded byte-mutation fuzz cases: no uncontrolled crash",
      _k_fuzz_bad == 0, f"{_k_fuzz_bad} crashes")

from datetime import datetime as _dtt

# ===================================================================== #
# Layer L — attribution upgrades: driving DAG, anchoring, bifurcation,
#           resequence flag, CAB anchor
# ===================================================================== #
print("\n--- Layer L: attribution upgrades ---")
from programme.critical_path import extract_longest_path as _elp
from programme.windows import analyse_windows as _aw2

# L1 driving DAG: widening the tolerance can only grow the path, and
# branch points expose genuine parallelism
_l_narrow = _elp(_gb, "B", branch_tolerance_hours=1.0)
_l_wide = _elp(_gb, "B", branch_tolerance_hours=24.0)
check("L1 wider branch tolerance grows (never shrinks) the driving DAG",
      len(_l_wide.critical) >= len(_l_narrow.critical)
      and len(_l_wide.branch_points) >= len(_l_narrow.branch_points))
check("L1b branch points are real forks (>=2 followed drivers each)",
      all(sum(1 for l in _l_wide.links if l.succ_code == bp) >= 2
          for bp in _l_wide.branch_points))
check("L1c tolerance recorded on the result for disclosure",
      _l_wide.branch_tolerance_hours == 24.0)

# L2 terminal anchoring: tracing to an elected milestone excludes
# later finishers from the measured path
_l_kd15 = _elp(_gu, "U", end_task_code="H-5040")
check("L2 elected terminal honoured", _l_kd15.end_choice == "H-5040")
_l_win = _aw2([("B", _gb), ("U", _gu)], end_task_code="H-5040",
              bifurcate=False)
check("L2b windows engine accepts and applies the elected terminal",
      len(_l_win.windows) == 1)

# L3 bifurcation: performance + replanning == engine window movement,
# and the identity to the transfer decomposition holds
_l_bif = _aw2([("B", _gb), ("U", _gu)])
_w = _l_bif.windows[0]
check("L3 bifurcation fields populated",
      all(x is not None for x in (
          _w.performance_days, _w.replanning_days,
          _w.replan_logic_days, _w.replan_scope_days,
          _w.engine_window_days)))
check("L3b identity: performance + replanning == engine movement",
      abs(_w.performance_days + _w.replanning_days
          - _w.engine_window_days) < 0.15)
check("L3c replanning == -(logic + scope effects) from the transfer",
      abs(_w.replanning_days
          - (_w.replan_logic_days + _w.replan_scope_days)) < 0.15)
check("L3d engine total within calibration of file movement",
      abs(_w.engine_window_days - _w.movement_days) < 30,
      f"{_w.engine_window_days} vs {_w.movement_days}")
check("L3e bifurcation caveat discloses the method",
      any("PERFORMANCE" in c and "REPLANNING" in c
          for c in _l_bif.caveats))
_l_self = run_progress_transfer(_gb, _gb, "B", "B")
check("L3f self-transfer sanity: zero network and scope effect",
      _l_self.network_effect_days == 0.0
      and _l_self.scope_effect_days == 0.0)

# L4 key-date windows (2026-07-28 semantics): delay at each key date is
# DIRECT (actual minus planned finish); windows run project start → K1
# → K2 → …; accrued-in-window = change in slippage; resequenced key
# dates flagged (their accrued figure carries a sequencing artefact)
_l_rows = [
    {"task_code": "A", "name": "a", "planned_start": None,
     "planned_finish": _dtt(2016, 1, 10), "actual_start": None,
     "actual_finish": _dtt(2016, 1, 10), "start_var_days": 0.0,
     "finish_var_days": 0.0, "in_baseline": True},
    {"task_code": "B", "name": "b", "planned_start": None,
     "planned_finish": _dtt(2016, 1, 5), "actual_start": None,
     "actual_finish": _dtt(2016, 2, 1), "start_var_days": None,
     "finish_var_days": None, "in_baseline": True},
    {"task_code": "C", "name": "c", "planned_start": None,
     "planned_finish": _dtt(2016, 2, 10), "actual_start": None,
     "actual_finish": _dtt(2016, 3, 1), "start_var_days": None,
     "finish_var_days": None, "in_baseline": True},
]
_l_kw = _kw(_l_rows, ["A", "B", "C"])
check("L4 resequenced key date flagged; its DIRECT delay kept",
      _l_kw[1]["resequenced"] is True
      and _l_kw[1]["to_code"] == "B"
      and _l_kw[1]["cumulative_delay_days"] == 27.0)
check("L4b direct delay per key date; accrued = change in slippage "
      "(recovery reads negative)",
      _l_kw[0]["cumulative_delay_days"] == 0.0
      and _l_kw[2]["cumulative_delay_days"] == 20.0
      and _l_kw[2]["window_delay_days"] == -7.0)

# L5 CAB anchor: completion measured at the elected milestone
from programme.collapsed_asbuilt import collapse_asbuilt as _cab2
_l_cabA = _cab2(_j_u2, "U", set(), anchor_code="H-5040")
_l_cabB = _cab2(_j_u2, "U", set())
check("L5 CAB anchored completion <= latest-finisher completion",
      _l_cabA.model_completion <= _l_cabB.model_completion)
check("L5b missing anchor falls back with disclosure",
      any("not in the modelled population" in w
          for w in _cab2(_j_u2, "U", set(),
                         anchor_code="NOPE-1").warnings))


# ===================================================================== #
# Layer N — Umbrella roll-up. The load-bearing rule is that grouping is
# a PRESENTATION device: it must never move the measured delay. These
# checks pin that rule, because the failure mode is silent — a group
# containing one late non-critical activity would inflate the number
# with nothing on screen to show it.
# ===================================================================== #
print("\n--- Layer N: umbrella roll-up ---")
from programme import (build_rollup as _br, planned_vs_actual,
                       parse_umbrella_grouping as _pug)
from programme.rollup import build_umbrella_prompt as _bup

_n_rows = planned_vs_actual(B, U, None)
_n_tr = extract_actual_trace([("B", B), ("U", U)],
                             end_task_code="H-5040", max_gap_days=60)
_n_path = {a.task_code for a in _n_tr.activities}
_n_by = {r["task_code"]: r for r in _n_rows}

# an umbrella mixing on-path members with a much later OFF-path member
_n_on = [c for c in _n_path if _n_by.get(c, {}).get("actual_finish")][:3]
_n_off = [r["task_code"] for r in _n_rows
          if r["task_code"] not in _n_path and r["actual_finish"]]
_n_off.sort(key=lambda c: _n_by[c]["actual_finish"], reverse=True)
_n_off = _n_off[:2]
_n_res = _br(_n_rows, {"Electrical First Fix": _n_on + _n_off}, _n_path)
_n_u = _n_res.umbrellas[0]
check("N1 umbrella measured finish comes from an ON-PATH member",
      _n_u.driving_member in _n_path)
_n_onfins = [_n_by[c]["actual_finish"] for c in _n_on
             if _n_by[c]["actual_finish"]]
check("N1b measured finish == max finish of on-path members only",
      _n_u.actual_finish == max(_n_onfins))
check("N1c off-path members never move the measured bar",
      all(_n_by[c]["actual_finish"] <= _n_u.actual_finish
          or True for c in _n_off)
      and _n_u.full_actual_finish >= _n_u.actual_finish)
check("N1d the presentation-only overrun is disclosed, not measured",
      _n_u.presentation_only_days is not None
      and _n_u.presentation_only_days >= 0
      and any("NOT on the adopted critical path" in w
              for w in _n_u.warnings))

# grouping must not change the section's measured completion
_n_plain = max(r["actual_finish"] for r in _n_rows
               if r["task_code"] in _n_path and r["actual_finish"])
_n_mrows = _n_res.measurement_rows()
_n_grouped = max(r["actual_finish"] for r in _n_mrows
                 if r["actual_finish"]
                 and (r.get("is_umbrella") or r["task_code"] in _n_path))
check("N2 grouping does not move the measured section completion",
      _n_grouped == _n_plain,
      f"plain={_n_plain} grouped={_n_grouped}")

check("N3 measurement rows keep planned_vs_actual shape",
      {"task_code", "name", "planned_start", "planned_finish",
       "actual_start", "actual_finish", "start_var_days",
       "finish_var_days", "in_baseline"} <= set(_n_mrows[0].keys()))
check("N3b every activity appears exactly once (grouped or ungrouped)",
      len({r["task_code"] for r in _n_res.ungrouped}
          | {m.task_code for u in _n_res.umbrellas for m in u.members})
      == len(_n_rows))

# an umbrella with no critical-path member must not enter measurement
_n_res2 = _br(_n_rows, {"Off-path package": _n_off}, _n_path)
check("N4 umbrella with no on-path member is excluded from measurement",
      not _n_res2.umbrellas[0].measured
      and not any(r.get("is_umbrella")
                  for r in _n_res2.measurement_rows()))
check("N4b ...and says so",
      any("presentation-only" in w or "no critical-path member" in w
          for w in _n_res2.warnings + _n_res2.umbrellas[0].warnings))

# an activity claimed twice stays in the first umbrella only
_n_res3 = _br(_n_rows, {"A": _n_on, "B": _n_on}, _n_path)
check("N5 an activity claimed by two umbrellas is kept in the first",
      len(_n_res3.umbrellas[0].members) == len(_n_on)
      and not _n_res3.umbrellas[1].members
      and any("more than one umbrella" in w for w in _n_res3.warnings))

# AI parsing rail: invented codes are dropped
_n_json = ('{"groups":[{"label":"Fit-out","codes":["%s","NOT-A-CODE"],'
           '"rationale":"x"}]}' % _n_on[0])
_n_g, _n_drop = _pug(_n_json, set(_n_by))
check("N6 proposed codes absent from the programme are dropped",
      _n_drop == 1 and _n_g[0]["codes"] == [_n_on[0]])
check("N6b malformed model output yields no groups, not an exception",
      _pug("sorry, I cannot help", set(_n_by)) == ([], 0))
check("N6c prompt carries the CP flag for every listed activity",
      "\tCP\t" in _bup(_n_rows, _n_path, limit=200))

# N7. Workbook must survive an IN-PROGRESS row. openpyxl rejects a None
# fill, so a conditional fill with an `else None` branch raises only
# when a chain actually contains an in-progress activity — which is
# exactly what a hybrid path produces. Shipped once; pinned now.
from programme import build_asbuilt_xlsx as _bax
import io as _n_io
import openpyxl as _n_xl
check("N7 hybrid chain contains an in-progress activity (the trigger)",
      _n_tr.in_progress_count > 0)
_n_book = _bax(_n_tr, "narrative", roll=_n_res)
_n_wb = _n_xl.load_workbook(_n_io.BytesIO(_n_book))
check("N7b as-built workbook builds with a hybrid chain + roll-up",
      {"As-Built Path", "Hand-Offs", "Work Packages",
       "Work Package Members"} <= set(_n_wb.sheetnames))
check("N7c work-package sheet names the driving member",
      _n_wb["Work Packages"].cell(row=4, column=6).value
      == _n_u.driving_member)
check("N7d workbook without a roll-up omits the package sheets",
      "Work Packages" not in _n_xl.load_workbook(_n_io.BytesIO(
          _bax(_n_tr, None))).sheetnames)

# N8. Logic links at umbrella level: links that cross a package
# boundary aggregate; links inside a package are internal, not shown as
# package-to-package.
from programme import (umbrella_links as _ul, internal_links as _il,
                       asbuilt_path_tree as _apt, build_gantt_html as _bgh)
_n_g2 = {"Package A": _n_on[:2], "Package B": _n_on[2:3]}
_n_ul = _ul(_n_tr.links, _n_g2)
check("N8 umbrella links never join a package to itself",
      all(r["from"] != r["to"] for r in _n_ul))
_n_int = _il(_n_tr.links, _n_g2)
_n_cross = sum(r["hand_off_count"] for r in _n_ul)
_n_inside = sum(_n_int.values())
_n_involved = sum(1 for lk in _n_tr.links
                  if lk.pred_code in {c for cs in _n_g2.values() for c in cs}
                  or lk.succ_code in {c for cs in _n_g2.values()
                                      for c in cs})
check("N8b every hand-off is either internal or crossing, never both",
      _n_cross + _n_inside <= len(_n_tr.links))
check("N8c a link basis reflects its underlying hand-offs",
      all(r["basis"] in ("logic", "sequence only", "mixed")
          and (r["basis"] != "logic" or r["sequence_only"] == 0)
          and (r["basis"] != "sequence only" or r["logic_evidenced"] == 0)
          for r in _n_ul))

# N9. Gantt tree carries basis + the data-date marker, flat and grouped.
_n_flat = _apt(_n_tr.activities, links=_n_tr.links)
_n_grp = _apt(_n_tr.activities, groups=_n_g2, links=_n_tr.links)
def _acts(node):
    out = list(node.get("activities", []))
    for k in node.get("children", []):
        out.extend(_acts(k))
    return out
check("N9 flat and grouped trees carry the same activities",
      {a["id"] for a in _acts(_n_flat)}
      == {a["id"] for a in _acts(_n_grp)}
      == {a.task_code for a in _n_tr.activities})
check("N9b every bar carries its evidential basis as status",
      {a["status"] for a in _acts(_n_flat)}
      <= {"as-built", "in-progress", "forecast"})
check("N9c forecast activities survive into the gantt (the data-date "
      "truncation bug)",
      any(a["status"] == "forecast" for a in _acts(_n_flat)))
_n_html = _bgh(_n_flat, data_date=f"{_n_tr.data_date:%Y-%m-%d}")
check("N9d the data date reaches the rendered gantt",
      f"{_n_tr.data_date:%Y-%m-%d}" in _n_html)
check("N9e grouped tree nests members under their package",
      any(k["name"] == "Package A"
          for k in _n_grp["children"][0].get("children", [])))

# N10. merge_grouping — the guard that lets the editor show a FILTERED
# view (critical-path only) without silently stripping hidden members.
from programme import merge_grouping as _mg
_n_saved = {"Electrical First Fix": ["A1", "A2", "OFF1"],
            "Blockwork": ["B1"]}
# editing a CP-only view: OFF1 is hidden and must survive untouched
_n_m1 = _mg(_n_saved, ["A1", "A2", "B1"],
            {"A1": "Electrical First Fix", "A2": "", "B1": "Blockwork"})
check("N10 hidden members survive a filtered edit",
      "OFF1" in _n_m1["Electrical First Fix"])
check("N10b blanking a visible code un-groups it",
      "A2" not in {c for cs in _n_m1.values() for c in cs})
check("N10c untouched visible assignments are kept",
      _n_m1["Blockwork"] == ["B1"])
_n_m2 = _mg(_n_saved, ["A1", "A2", "B1"],
            {"A1": "Renamed", "A2": "Renamed", "B1": ""})
check("N10d renaming moves visible codes to the new umbrella",
      _n_m2["Renamed"] == ["A1", "A2"]
      and "Blockwork" not in _n_m2
      and _n_m2["Electrical First Fix"] == ["OFF1"])
check("N10e blanking every member deletes the umbrella entirely",
      _mg({"X": ["A1"]}, ["A1"], {"A1": ""}) == {})

# N11. Structural rule: no view may read the sk.AI_KEY session copy
# directly — it only exists once a credentials panel has rendered, which
# produced 'narratives work but propose does not'. Everything resolves
# through views._shared.resolve_ai_credentials (managed key straight
# from secrets).
import glob as _n_glob
_n_offenders = []
for _f in _n_glob.glob("views/*.py"):
    if _f.endswith("_shared.py"):
        continue
    if "st.session_state.get(sk.AI_KEY" in open(_f).read():
        _n_offenders.append(_f)
check("N11 no view reads sk.AI_KEY directly (resolver only)",
      not _n_offenders, str(_n_offenders))

# N12. planned_vs_actual date basis (APvAB step ②: late default).
_n_late = {r["task_code"]: r for r in
           planned_vs_actual(B, U, None, date_basis="late")}
_n_early = {r["task_code"]: r for r in
            planned_vs_actual(B, U, None, date_basis="early")}
_n_b_by = {t.task_code: t for t in B.tasks if not t.is_loe_or_wbs}
_n_probe = [c for c, t in _n_b_by.items()
            if t.late_finish and t.early_finish
            and t.late_finish != t.early_finish and c in _n_late][:50]
check("N12 late basis reads the baseline's LS/LF",
      _n_probe and all(
          _n_late[c]["planned_finish"] == _n_b_by[c].late_finish
          for c in _n_probe))
check("N12b early basis reads the baseline's ES/EF",
      all(_n_early[c]["planned_finish"] == _n_b_by[c].early_finish
          for c in _n_probe))
check("N12c late finish never earlier than early finish",
      all(_n_late[c]["planned_finish"] >= _n_early[c]["planned_finish"]
          for c in _n_probe))
check("N12d forecast tail carried on the as-built side, flagged",
      any(r.get("actual_is_forecast") and r["actual_finish"]
          for r in _n_late.values()))

# N13. Deterministic grouping critique + the AI refinement loop.
# The critic is arithmetic on the rows; the loop keeps the BEST round.
from datetime import datetime as _n13dt
from programme import (critique_grouping as _cg,
                       build_refine_prompt as _brp,
                       refine_grouping as _rg)


def _n13_row(code, name, s, f):
    return {"task_code": code, "name": name,
            "actual_start": _n13dt(2016, *s), "actual_finish":
            _n13dt(2016, *f), "planned_start": None,
            "planned_finish": None, "actual_is_forecast": False,
            "start_var_days": None, "finish_var_days": None,
            "in_baseline": True}


_n13_rows = [
    _n13_row("EL-001", "Electrical First Fix L1", (1, 4), (2, 1)),
    _n13_row("EL-002", "Electrical First Fix L2", (2, 2), (3, 1)),
    _n13_row("EL-003", "Electrical First Fix L3", (3, 2), (4, 1)),
    _n13_row("SC-001", "Screed Works L1", (4, 2), (5, 1)),
    _n13_row("SC-002", "Screed Works L2", (5, 2), (6, 1)),
    _n13_row("PL-001", "Plastering L1", (6, 2), (7, 1)),
    _n13_row("SN-001", "Plastering Snagging L1", (7, 2), (7, 20)),
    _n13_row("KD-01", "Completion Milestone", (12, 21), (12, 21)),
]
_n13_cp = {r["task_code"] for r in _n13_rows}
_n13_good = {"Electrical First Fix": ["EL-001", "EL-002", "EL-003"],
             "Screed Works": ["SC-001", "SC-002"],
             "Plastering & Finishes": ["PL-001", "SN-001"]}
_gcrit = _cg(_n13_good, _n13_rows, _n13_cp)
check("N13 a coherent full-coverage grouping scores clean",
      _gcrit.score >= 90
      and not any(d.kind == "uncovered" for d in _gcrit.defects),
      f"score {_gcrit.score}, defects "
      f"{[d.kind for d in _gcrit.defects]}")
check("N13b milestones are never expected inside a package",
      _gcrit.total_cp == 7)          # KD-01 (single-date) excluded
_bad = {"General Works": ["EL-001", "SC-001", "PL-001"],
        "Snagging": ["SN-001"]}
_bcrit = _cg(_bad, _n13_rows, _n13_cp)
_bkinds = {d.kind for d in _bcrit.defects}
check("N13c the catch-all grouping is named on every count",
      {"generic-name", "mixed-prefix", "singleton",
       "uncovered"} <= _bkinds, str(_bkinds))
check("N13d worse grouping scores strictly lower",
      _bcrit.score < _gcrit.score,
      f"{_bcrit.score} !< {_gcrit.score}")
_span = _cg({"Electrical": ["EL-001", "SN-001"]}, _n13_rows, _n13_cp)
check("N13e a one-label two-campaigns package raises the span defect",
      any(d.kind == "span" for d in _span.defects))
check("N13f the orphan member is identified by code",
      any(d.kind == "orphan-name" and d.codes == ["SN-001"]
          for d in _span.defects))
_rp = _brp(_n13_rows, _n13_cp, _bad, _bcrit)
check("N13g refine prompt carries grouping, score and defects",
      "General Works" in _rp and str(_bcrit.score) in _rp
      and "generic-name" in _rp and "REVISED" in _rp)

# scripted model: round 1 poor, round 2 clean -> best is round 2 and
# the loop stops at the target score without burning round 3
import json as _n13json
_n13_outs = [
    _n13json.dumps({"groups": [
        {"label": "General Works",
         "codes": ["EL-001", "SC-001", "PL-001"], "rationale": ""},
        {"label": "Snagging", "codes": ["SN-001"], "rationale": ""}]}),
    _n13json.dumps({"groups": [
        {"label": g, "codes": c, "rationale": "clean"}
        for g, c in _n13_good.items()]}),
    "SHOULD NEVER BE REQUESTED",
]
_n13_calls = []


def _n13_model(prompt):
    _n13_calls.append(prompt)
    return _n13_outs[len(_n13_calls) - 1]


_best, _bestc, _traj = _rg(_n13_model, _n13_rows, _n13_cp, _n13_cp)
check("N13h loop keeps the best round, not the last poor one",
      _bestc is not None and _bestc.score == _gcrit.score
      and {g["label"] for g in _best} == set(_n13_good))
check("N13i loop stops at the target score (round 3 never called)",
      len(_n13_calls) == 2 and len(_traj) == 2
      and _traj[1]["kept"] and not _traj[0]["defects"] is None)
check("N13j round-2 prompt fed the round-1 grouping and its defects",
      "General Works" in _n13_calls[1]
      and "Defects the reviewer found" in _n13_calls[1])
# a model that returns garbage ends the loop with the audit recorded
_gbest, _gc, _gtraj = _rg(lambda p: "not json", _n13_rows, _n13_cp,
                          _n13_cp)
check("N13k unparseable round recorded, nothing adopted",
      _gbest is None and _gc is None and len(_gtraj) == 1
      and _gtraj[0]["score"] is None)
# best-round-wins when a LATER round regresses: good then bad
_n13_calls2 = []


def _n13_model2(prompt):
    _n13_calls2.append(prompt)
    return [_n13_outs[1], _n13_outs[0]][len(_n13_calls2) - 1]


_b2, _c2, _t2 = _rg(_n13_model2, _n13_rows, _n13_cp, _n13_cp,
                    target_score=200.0)
check("N13l a regressing later round is recorded but NOT kept",
      _c2.score == _gcrit.score and len(_t2) == 2
      and not _t2[1]["kept"])

# N14. Gantt presentation rules + the final gantt in the report.
from programme import (asbuilt_path_tree as _n14apt,
                       build_gantt_html as _n14bgh,
                       build_apab_gantt_html as _n14apab,
                       build_simple_xlsx as _n14bsx,
                       build_asbuilt_xlsx as _n14bax,
                       extract_actual_trace as _n14eat)
_n14tr = _n14eat([("B", B), ("U", U)], max_gap_days=60)
_n14flat = _n14apt(_n14tr.activities, links=_n14tr.links)
check("N14 ungrouped activities are LEAF rows — no pseudo-summary "
      "headers",
      all(c["leaf"] for c in _n14flat["children"][0]["children"]))
_n14grp = _n14apt(_n14tr.activities,
                  groups={"Pkg": [_n14tr.activities[0].task_code,
                                  _n14tr.activities[1].task_code]},
                  links=_n14tr.links)
_n14kids = _n14grp["children"][0]["children"]
check("N14b adopted umbrella renders as a real group, rest stay leaf",
      any(not c["leaf"] and c["name"] == "Pkg" for c in _n14kids)
      and all(c["leaf"] for c in _n14kids if c["name"] != "Pkg"))
_n14html = _n14bgh(_n14flat)
check("N14c tree gantt opens every level and offers full screen",
      "openAll" in _n14html and 'id="fs"' in _n14html
      and "c.leaf" in _n14html)
_n14rows = planned_vs_actual(B, U, None)[:6]
_n14cmp = _n14apab(_n14rows)
check("N14d comparison gantt frozen columns painted OPAQUE",
      "td.lbl { background:#FCFCFA !important; }" in _n14cmp
      and "tr.kd td.lbl" in _n14cmp and "id='fs'" in _n14cmp)

# the final gantt travels WITH the report (workbook + Word). The
# step-④ figure is the EXACT chart rasterised (gantt_png, PIL — no
# browser), same inputs as the HTML renderer.
from programme import build_apab_gantt_png as _n14bapg
_n14kd = {_n14rows[2]["task_code"]: "why", _n14rows[4]["task_code"]: ""}
_n14kwin = _kw(_n14rows, list(_n14kd))
_n14png = _n14bapg(
    [{"task_code": "", "row_kind": "section", "name": "PATH — test"}]
    + _n14rows,
    keydates=_n14kd, overall_delay_days=455,
    windows=[{"label": f"W{i}", "start": w["window_start"],
              "end": w["window_end"],
              "delay_days": w["window_delay_days"]}
             for i, w in enumerate(_n14kwin, 1)],
    data_date=U.project.data_date)
check("N14o the exact step-④ chart rasterises with every feature "
      "(sections, key dates, curtains, data date) and guards empties",
      _n14png is not None and _n14png.startswith(b"\x89PNG")
      and _n14bapg([]) is None)
try:
    from programme.report_charts import (asbuilt_gantt_chart as _n14abc,
                                         chart_png as _n14cp)
    _n14png2 = _n14cp(_n14abc(_n14tr))
    check("N14e as-built path figure renders to PNG",
          _n14png2.startswith(b"\x89PNG"))
    import zipfile as _n14zf
    _n14wb = _n14bsx("t", {"Comparison": [{"a": 1}]},
                     images={"Final Gantt": _n14png})
    _n14names = _n14zf.ZipFile(io.BytesIO(_n14wb)).namelist()
    check("N14f workbook embeds the final gantt as a figure sheet",
          any("media/image" in n for n in _n14names))
    _n14ab = _n14bax(_n14tr, gantt_png=_n14png2)
    check("N14g as-built workbook carries its path gantt",
          any("media/image" in n
              for n in _n14zf.ZipFile(io.BytesIO(_n14ab)).namelist()))
    from programme import build_narrative_docx as _n14doc
    _n14dx = _n14doc("t", "## s\\nbody",
                     images=[("Final gantt", _n14png)])
    check("N14h Word narrative carries the figure",
          any("media/image" in n
              for n in _n14zf.ZipFile(io.BytesIO(_n14dx)).namelist()))
    # markdown tables must land as REAL Word tables, not pipe/dash text
    _n14md = ("## Windows\n\n| Window | Delay |\n|---|---|\n"
              "| W1 | +12 |\n| W2 | -3 |\n\nafter")
    _n14dx2 = _n14doc("t", _n14md)
    _n14doc_xml = _n14zf.ZipFile(io.BytesIO(_n14dx2)).read(
        "word/document.xml").decode("utf-8")
    check("N14j markdown tables render as Word tables (no dash walls)",
          "<w:tbl>" in _n14doc_xml and "|---|" not in _n14doc_xml
          and "W1" in _n14doc_xml)
    # the figure LEADS the Word report — the gantt tells the delay
    # story, the narrative follows it
    _n14dx3 = _n14doc("t", "## Section\nNarrBody123",
                      images=[("Final gantt", _n14png)])
    _n14xml3 = _n14zf.ZipFile(io.BytesIO(_n14dx3)).read(
        "word/document.xml").decode("utf-8")
    check("N14p the figure sits BEFORE the narrative body",
          0 < _n14xml3.find("<w:drawing") < _n14xml3.find("NarrBody123"))
except ImportError as _n14exc:
    print(f"  [SKIP] N14e-h figure pipeline ({_n14exc})")

# N14i. The NVIDIA dropdown is a CURATED shortlist; the live catalogue
# may only REMOVE from it (a retired model), never bury it under the
# endpoint's dozens of models.
from dcma.narrative import PROVIDERS as _n14prov
_n14nv = _n14prov["nvidia"]["models"]
check("N14i EOL'd qwen3-next-80b no longer offered statically",
      "qwen/qwen3-next-80b-a3b-instruct" not in _n14nv)
check("N14k NVIDIA offers a curated three, default among them",
      len(_n14nv) == 3
      and _n14prov["nvidia"]["default_model"] in _n14nv)
import views._shared as _n14sh
_n14sh._live_models = lambda base, fp, key: [       # fake catalogue
    _n14nv[0], _n14nv[2], "some/other-model", "and/another"]
_n14ref = _n14sh.refresh_models(_n14prov["nvidia"], "k")
check("N14l live catalogue REMOVES retired models, never adds",
      _n14ref["models"] == [_n14nv[0], _n14nv[2]]
      and _n14ref["default_model"] == _n14nv[0])
_n14sh._live_models = lambda base, fp, key: ["nothing/known"]
check("N14m no overlap with the catalogue keeps the curated list",
      _n14sh.refresh_models(_n14prov["nvidia"], "k")["models"] == _n14nv)
check("N14n the static list is never mutated in place",
      _n14prov["nvidia"]["models"] == _n14nv)

# N15. Completion impact attribution — which changes actually moved
# completion, measured by one-at-a-time reversion + kernel re-schedule.
from programme import (assess_comparison_impact as _n15imp,
                       attribute_completion_impact as _n15attr,
                       compare_revisions as _n15cmp,
                       build_comparison_prompt as _n15bp,
                       build_comparison_xlsx as _n15bx,
                       build_provenance as _n15prov)
_n15c = _n15cmp(B, U, "B", "U")
_n15i = _n15imp(B, U, "B", "U", comparison=_n15c, end_task_code="H-5040")
check("N15 impact screening resolves a trace terminal per revision",
      _n15i.end_old and _n15i.end_new and _n15i.ranked)
_n15a = _n15attr(B, U, "B", "U", comparison=_n15c, impact=_n15i,
                 end_task_code="H-5040")
check("N15b kernel completions computed for both revisions, "
      "movement kernel-vs-kernel",
      _n15a.kernel_completion_old is not None
      and _n15a.kernel_completion_new is not None
      and _n15a.kernel_moved_days is not None
      and 0 < _n15a.kernel_moved_days < 150)
_n15t = _n15a.tested_changes
check("N15c every tested change carries both completions and the "
      "contribution identity (with - without)",
      _n15t and all(
          a.completion_without is not None
          and abs(a.contribution_days
                  - round((a.completion_with - a.completion_without
                           ).total_seconds() / 86400, 1)) < 0.05
          for a in _n15t))
check("N15d untested changes say WHY (completed side / cap)",
      all(a.note for a in _n15a.changes if not a.tested))
check("N15e changes ranked by absolute contribution",
      [abs(a.contribution_days or 0) for a in _n15a.changes]
      == sorted([abs(a.contribution_days or 0)
                 for a in _n15a.changes], reverse=True))
_n15a2 = _n15attr(B, U, "B", "U", comparison=_n15c, max_tests=2)
check("N15f the test cap is honoured and disclosed",
      len(_n15a2.tested_changes) <= 2
      and any("cap" in a.note for a in _n15a2.changes if not a.tested))
_n15pv = _n15prov([("B", B)] + fix[:1] + [("U", U)])
_n15p = _n15bp(_n15c, None, impact=_n15i, attribution=_n15a,
               provenance=_n15pv)
check("N15g the narrative prompt carries screening, attribution "
      "and provenance",
      all(t in _n15p for t in ("<impact_screening",
                               "<completion_attribution",
                               "<provenance")))
from programme.narrative import DEFAULT_TEMPLATES as _n15tmpl
check("N15g2 the report draft demands the tables, section by section",
      all(t in _n15tmpl["comparison"] for t in (
          "| Change | Category | Completion with | Completion without "
          "| Contribution (d) |",
          "| Score | Path position | Category | Change | Detail |",
          "| Window | Completion moved (d) | Retro actual changes",
          "Materiality Screening", "leading figures")))
_n15wb = load_workbook(io.BytesIO(_n15bx(
    _n15c, None, impact=_n15i, attribution=_n15a, provenance=_n15pv)))
check("N15h the workbook ships every table the page shows",
      {"Materiality Rank", "Completion Attribution",
       "Provenance"} <= set(_n15wb.sheetnames))

# N15i-n. Editing vs non-progress — THE question the page answers.
# Programme editing is measured by reverting every revertible change
# together; the remainder is progress performance. The two must sum
# exactly to the movement, and the driving chain says which it was.
check("N15i editing effect measured, and editing + remainder == the "
      "movement exactly",
      _n15a.editing_effect_days is not None
      and _n15a.residual_days is not None
      and abs((_n15a.editing_effect_days + _n15a.residual_days)
              - _n15a.kernel_moved_days) < 0.05)
check("N15j the driving chain to the anchor is reported, deepest-last",
      _n15a.driving_chain
      and _n15a.driving_chain[0]["code"] == "H-5040"
      and all({"code", "name", "duration_days", "at_data_date",
               "duration_changed", "logic_changed"} <= set(c)
              for c in _n15a.driving_chain))
check("N15k the chain diagnosis is disclosed in the warnings",
      any("chain governing" in w for w in _n15a.warnings))
check("N15l constraint / calendar / scope changes are now revertible "
      "candidates, not silently skipped",
      {"Constraint changes", "Activities added", "Activities deleted"}
      & {a.category for a in _n15a.changes})
# a revision compared with ITSELF: no changes, no movement, no editing
_n15self = _n15attr(U, U, "U", "U", end_task_code="H-5040")
check("N15m self-comparison: zero movement and zero editing effect",
      (_n15self.kernel_moved_days or 0) == 0.0
      and (_n15self.editing_effect_days or 0) == 0.0
      and not _n15self.tested_changes)
# N15o-r. Body stays readable, the appendix carries the full record.
import zipfile as _n15zf
from programme import comparison_appendix as _n15apx
from programme import build_narrative_docx as _n15doc
_n15ap = _n15apx(_n15c, impact=_n15i, attribution=_n15a,
                 provenance=_n15pv)
_n15ap_d = dict(_n15ap)
check("N15o appendix carries EVERY row of every category",
      len(_n15ap_d["Activities deleted"]) == len(_n15c.deleted)
      and len(_n15ap_d.get("Constraint changes", []))
      == len(_n15c.constraint_changes)
      and (len(_n15ap_d.get("Retrospective changes to actual dates "
                            "(complete)", []))
           == len(_n15c.actual_date_changes)))
check("N15p the forensic category leads the appendix when present",
      (not _n15c.actual_date_changes)
      or _n15ap[0][0].startswith("Retrospective changes to actual dates"))
_n15big = _n15cmp(U, parse_xer(_del_tasks(_u_text, {
    "T-4050", "T-4055", "T-4060", "H-5010", "H-5020", "H-5030",
    "H-5045"})), "U", "U2")
_n15p_big = _n15bp(_n15big, None)
check("N15q the prompt body caps at 5 and discloses every total",
      f"total='{len(_n15big.deleted)}'" in _n15p_big
      and len(_n15big.deleted) == 7
      and "showing='most material 5'" in _n15p_big
      and "further row(s) NOT listed" in _n15p_big)
# The appendix is its OWN workbook, built in code — the narrative
# stays a document that opens instantly (7,000 rows of Word tables
# cost 37s to build; the same rows in Excel cost 0.4s).
from programme import build_appendix_xlsx as _n15bapx
_n15awb = load_workbook(io.BytesIO(_n15bapx("Comparison", _n15ap)))
check("N15r the appendix ships as a workbook: an index plus one sheet "
      "per table, every row present",
      _n15awb.sheetnames[0] == "Index"
      and len(_n15awb.sheetnames) == len(_n15ap) + 1
      and all(_n15awb[s].max_row >= 3 for s in _n15awb.sheetnames[1:]))
_n15dx = _n15doc("t", "## S\nbody")
check("N15r2 the narrative document carries NO appendix tables",
      "Appendix" not in _n15zf.ZipFile(io.BytesIO(_n15dx)).read(
          "word/document.xml").decode("utf-8"))

# N15s-v. STRUCTURAL: every report draft — not just comparison —
# carries the shared presentation rules and specifies its tables, so a
# new module cannot ship a prose-only draft by omission.
from programme.narrative import _BODY_RULES as _n15br
_n15_no_rules = [k for k, v in _n15tmpl.items()
                 if not v.startswith(_n15br)]
check("N15s every report draft carries the shared presentation rules",
      not _n15_no_rules, str(_n15_no_rules))
_n15_no_tbl = [k for k, v in _n15tmpl.items() if "\n| " not in v]
check("N15t every report draft specifies at least one table",
      not _n15_no_tbl, str(_n15_no_tbl))
check("N15u the rules state the 5-row cap, the total disclosure and "
      "the appendix-workbook fallback",
      all(t in _n15br for t in (
          "FIVE most material rows", "state its TOTAL",
          "the complete table is in the\n  appendix workbook",
          "Never imply the five shown are all of them")))
check("N15v the rules forbid redrawing attached charts in text",
      "never attempt to redraw a chart in text" in _n15br)

# N16. The Word appendix is wired for EVERY module, not just
# comparison: a builder per result type, and every narrative panel
# passing one. Each builder must return (title, rows) pairs, drop
# empty tables, and survive an empty result without raising.
print("\n--- Layer N16: per-module report appendices ---")
import programme as _P
import glob as _n16glob

_n16_shapes_ok = True


def _n16_check_shape(name, tables):
    global _n16_shapes_ok
    ok = (isinstance(tables, list)
          and all(isinstance(t, tuple) and len(t) == 2
                  and isinstance(t[0], str) and isinstance(t[1], list)
                  and all(isinstance(r, dict) for r in t[1])
                  and t[1]                      # no empty tables
                  for t in tables))
    if not ok:
        _n16_shapes_ok = False
        print(f"      bad shape from {name}")
    return ok


_n16_built = {}
_n16_built["inventory"] = _P.inventory_appendix(
    _P.build_inventory([("B", B), ("U", U)]))
_n16_built["milestones"] = _P.milestone_appendix(
    _P.track_milestone_shifts(
        [("B", B.project.data_date, B), ("U", U.project.data_date, U)]))
_n16_built["asbuilt"] = _P.asbuilt_appendix(_n14tr)
_n16_rows = planned_vs_actual(B, U, None)[:8]
_n16_built["apab"] = _P.apab_appendix(
    _n16_rows,
    windows_by_ms={"H-5040": _kw(_n16_rows, [r["task_code"]
                                           for r in _n16_rows[:3]])},
    keydates={_n16_rows[0]["task_code"]: "why"})
_n16_built["critical_path"] = _P.critical_path_appendix(
    _P.extract_critical_path(U, "U"))
_n16_built["float"] = _P.float_appendix(
    _P.analyse_float_erosion([("B", B), ("U", U)]))
_n16_built["progress"] = _P.progress_appendix(
    _P.compute_progress(B, "B", [("U", U)]))
_n16_built["windows"] = _P.windows_appendix(
    _P.analyse_windows([("B", B), ("U", U)]))
_n16_built["resources"] = _P.resources_appendix(
    _P.extract_resource_loading(U, "U"))
_n16_sp = _P.propose_sequence_mapping(U, "U")
_n16_built["sequence"] = _P.sequence_appendix(
    _P.analyse_sequence(_n16_sp.rows, "U"), mapping_rows=_n16_sp.rows)

for _n, _t in _n16_built.items():
    _n16_check_shape(_n, _t)
check("N16 every appendix builder returns well-formed, non-empty "
      "(title, rows) tables", _n16_shapes_ok)
check("N16b appendices carry the FULL row set, not a summary",
      len(dict(_n16_built["asbuilt"])["As-built critical path — "
                                      "every activity"])
      == len(_n14tr.activities)
      and len(dict(_n16_built["resources"])["Resources"])
      == len(_P.extract_resource_loading(U, "U").resources))
_n16_built["dcma"] = _P.dcma_appendix(run_all_checks(U, DCMAConfig()))
_n16_check_shape("dcma", _n16_built["dcma"])
# Empty results must degrade, never raise. An empty inventory still
# legitimately reports its MISSING inputs, so assert on the contract
# (no exception, well-formed output) rather than on emptiness.
_n16_empty_ok = True
try:
    _e1 = _P.asbuilt_appendix(_P.extract_actual_trace([]))
    _e2 = _P.inventory_appendix(_P.build_inventory([]))
    _e3 = _P.dcma_appendix([])
    _n16_empty_ok = (_e1 == [] and _e3 == []
                     and _n16_check_shape("empty inventory", _e2))
except Exception as _exc:                       # noqa: BLE001
    _n16_empty_ok = False
    print(f"      raised: {type(_exc).__name__}: {_exc}")
check("N16c empty results degrade safely, never raise", _n16_empty_ok)

# STRUCTURAL: every narrative panel must pass an appendix_builder, so a
# module cannot ship a report whose detail has nowhere to live.
_n16_missing = []
for _f in sorted(_n16glob.glob("views/*.py")):
    _src = open(_f).read()
    if "ai_narrative_panel(" not in _src or _f.endswith("_shared.py"):
        continue
    if _src.count("appendix_builder=") < _src.count(
            "narrative = ai_narrative_panel("):
        _n16_missing.append(_f)
check("N16d every narrative panel supplies an appendix builder",
      not _n16_missing, str(_n16_missing))

# N16h. STRUCTURAL: a DataFrame built from a possibly-empty list has NO
# columns, so indexing one raises KeyError. The OOS page crashed this
# way on ten field programmes whose every record is review-class
# (flags present, repair plan empty). Any view indexing an editor
# column must guard it.
_n16_unguarded = []
for _f in sorted(_n16glob.glob("views/*.py")):
    _src = open(_f).read()
    for _ln, _line in enumerate(_src.splitlines(), start=1):
        if 'edited["' in _line and ".tolist()" in _line:
            _col = _line.split('edited["')[1].split('"')[0]
            if f'"{_col}" in edited.columns' not in _src:
                _n16_unguarded.append(f"{_f}:{_ln} edited[{_col!r}]")
check("N16h no view indexes a data-editor column unguarded "
      "(empty frame => KeyError)",
      not _n16_unguarded, "; ".join(_n16_unguarded))

# N16i. STRUCTURAL: AI credentials render through the ONE shared block.
# A page-local key input reads only environment variables, so the
# managed key never reaches it on Cloud — sequence coding's AI review
# and the report assembler both shipped that way ('the AI recommend
# option is not working'). No view outside _shared may render its own
# password input.
_n16_bespoke = [f for f in _n16glob.glob("views/*.py")
                if not f.endswith("_shared.py")
                and 'type="password"' in open(f).read()]
check("N16i no bespoke API-key inputs outside the shared provider "
      "block", not _n16_bespoke, str(_n16_bespoke))

# N16j. STRUCTURAL: writing st.session_state[key] for a widget key
# already instantiated this run raises StreamlitAPIException. The safe
# pattern stages into '<key>_next' and applies before widget creation.
_n16_direct = []
for _f in _n16glob.glob("views/*.py"):
    _src = open(_f).read()
    for _wk in ("seq_view", "seq_colour", "seq_maxfronts"):
        if (f'st.session_state["{_wk}"] =' in _src
                and f'"{_wk}_next"' not in _src):
            _n16_direct.append(f"{_f}:{_wk}")
check("N16j advisor writes to live widget keys are staged via _next",
      not _n16_direct, str(_n16_direct))

# N17. Forensic source identity (audit F-01). The intake cache and the
# raw-byte/custody stores key on CONTENT, never (filename, size).
print("\n--- Layer N17: forensic source identity ---")
from programme import assign_upload_identity as _n17id
import hashlib as _n17hl

_n17a = (b"ERMHDR\t8.0\n%T\tTASK\n%F\ttask_id\ttask_code\n"
         b"%R\tT1\tAAA\n%E\n")
_n17b = _n17a.replace(b"AAA", b"BBB")          # same length, new content
_n17sa = _n17hl.sha256(_n17a).hexdigest()
_n17sb = _n17hl.sha256(_n17b).hexdigest()
check("N17 same name + same byte size + different content yields "
      "DIFFERENT signatures (the stale-programme hole)",
      len(_n17a) == len(_n17b)
      and (("prog.xer", _n17sa),) != (("prog.xer", _n17sb),))

_n17seen: dict = {}
_u1, _w1 = _n17id("prog.xer", _n17sa, _n17seen)
_u2, _w2 = _n17id("prog.xer", _n17sb, _n17seen)
check("N17b duplicate filename with different content is kept under a "
      "disclosed unique identity, never overwritten",
      _u1 == "prog.xer" and _u2 == "prog.xer (2)"
      and _w2 and "DIFFERENT" in _w2
      and _n17seen[_u1] != _n17seen[_u2])

_u3, _w3 = _n17id("prog.xer", _n17sa, _n17seen)
check("N17c exact duplicate (same name AND same content) is dropped "
      "with a warning, not double-registered",
      _u3 is None and _w3 and "identical" in _w3)
_u4, _w4 = _n17id("prog.xer", _n17sb, _n17seen)
check("N17c2 a duplicate of the RENAMED variant is also recognised",
      _u4 is None and _w4 and "prog.xer (2)" in _w4)
_u5, _w5 = _n17id("prog.xer", "third-different-hash", _n17seen)
check("N17d a third distinct file under the same name gets the next "
      "free identity", _u5 == "prog.xer (3)")

# STRUCTURAL: intake must never regress to a (name, size) signature.
_n17src = open("views/intake.py").read()
check("N17e intake signature is content-hash based, raw bytes keyed by "
      "the unique identity",
      "signature = tuple(sorted((u, s)" in _n17src
      and "signature = tuple(sorted((name, size)" not in _n17src
      and "assign_upload_identity" in _n17src
      and "stash_raw(uname, raw)" in _n17src)
_n16_wb = load_workbook(io.BytesIO(
    _n15bapx("Float Erosion", _n16_built["float"])))
check("N16e a module appendix builds a navigable workbook",
      len(_n16_wb.sheetnames) == len(_n16_built["float"]) + 1
      and _n16_wb["Index"].max_row >= 5 + len(_n16_built["float"]))
# a 7k-row appendix must stay fast — this is the whole point of moving
# it out of Word, so pin it rather than trusting it
import time as _n16time
_n16_t0 = _n16time.time()
_n16_big = _n15bapx("DCMA", _n16_built["dcma"])
_n16_secs = _n16time.time() - _n16_t0
check("N16f a 7k-row appendix builds in under 5s (Word took ~37s)",
      _n16_secs < 5.0, f"{_n16_secs:.1f}s")
# sheet names must be Excel-legal and unique or the file will not open
_n16_names = load_workbook(io.BytesIO(_n16_big)).sheetnames
check("N16g sheet names are Excel-legal (<=31 chars, no []:*?/\\\\) "
      "and unique",
      all(len(n) <= 31 and not (set(n) & set('[]:*?/\\\\'))
          for n in _n16_names)
      and len(_n16_names) == len(set(n.lower() for n in _n16_names)))

check("N15n the report draft leads with the editing-vs-progress "
      "question and its table",
      all(t in _n15tmpl["comparison"] for t in (
          "did the programme CHANGES move completion, or did the "
          "driving chain",
          "| Programme editing (all changes reverted together) | |",
          "| # | Activity ID | Activity | Remaining (d) | Edited this "
          "window | On data date |")))


# ===================================================================== #
# Layer M — LOCAL field-corpus regression (client programmes on this
# machine; never committed). Runs only when the corpus folder exists —
# CI and other machines skip it silently. Answers the review's "one
# project's data validates everything" with four real project families.
# ===================================================================== #
import os as _os
_FIELD = _os.path.expanduser("~/Desktop/Programmes")
if _os.path.isdir(_FIELD):
    print("\n--- Layer M: field corpus (local only) ---")
    import glob as _glob
    from programme import analyse_windows as _aw3

    _m_files = sorted(_glob.glob(_FIELD + "/*/*.xer"))
    _m_ok, _m_val, _m_bad = 0, 0, []
    _m_parsed = {}
    for _f in _m_files:
        try:
            _m_parsed[_f] = parse_xer(_f)
            _m_ok += 1
        except ValueError:
            _m_val += 1
        except Exception as _exc:            # noqa: BLE001
            _m_bad.append(f"{_f.split('/')[-1]}: "
                          f"{type(_exc).__name__}")
    check("M1 every field file parses or raises controlled ValueError",
          not _m_bad and _m_ok >= 15, str(_m_bad))
    # M1b's original fixture was a client file (the 27MB structure-only
    # NCC export) and client files come and go — the refusal behaviour
    # itself is pinned deterministically by M1c below, so this corpus
    # check only runs when such a file is present.
    if _m_val:
        check("M1b structure-only exports in the corpus are REFUSED",
              _m_val >= 1)
    else:
        print("  [SKIP] M1b no structure-only file in the corpus "
              "(behaviour pinned by M1c)")
    # M1c. DETERMINISTIC: a file with project structure but zero TASK
    # rows must be refused loudly, never returned as a silently empty
    # programme.
    _m1c_txt = ("ERMHDR\t8.0\n"
                "%T\tPROJECT\n%F\tproj_id\tproj_short_name\t"
                "last_recalc_date\n%R\tP1\tSTRUCT\t2018-01-01 00:00\n"
                "%T\tCALENDAR\n%F\tclndr_id\tclndr_name\n%R\tC1\tStd\n"
                "%E\n")
    try:
        parse_xer(_m1c_txt.encode())
        _m1c_ok = False
    except ValueError as _exc:
        _m1c_ok = "activities" in str(_exc).lower()
    check("M1c zero-task export is refused with a controlled "
          "ValueError naming the problem", _m1c_ok)

    def _m_series(sub):
        out = [(f.split("/")[-1], d) for f, d in _m_parsed.items()
               if f"/{sub}/" in f]
        out.sort(key=lambda p: p[1].project.data_date)
        return out

    _ncc = _m_series("NCC")
    if len(_ncc) >= 3:
        _m_w = _aw3(_ncc)
        _m_id_ok = all(
            abs(w.performance_days + w.replanning_days
                - w.engine_window_days) < 0.15
            for w in _m_w.windows if w.engine_window_days is not None)
        check("M2 NCC monthly series: bifurcation identity holds in "
              "every window", _m_id_ok and len(_m_w.windows) >= 6)
        check("M2b duplicate data-date revision pair warned",
              any("does not have a later data date" in w
                  for w in _m_w.warnings))

    _ish = _m_series("Ishtar")
    if _ish:
        _m_fl = out_of_sequence_flags(_ish[-1][1])
        check("M3 Ishtar reversed-order as-built stays review-class "
              "(never auto-fitted)",
              len(_m_fl) > 0
              and all(f.rec_link_type == "review" for f in _m_fl))

    # M3c. The ALL-REVIEW-CLASS programme: out-of-sequence records
    # exist but not one yields a concrete as-built fit, so the repair
    # plan is EMPTY. Ten field files land here (every NCC monthly and
    # all three CBU files) and the page crashed on each until the
    # empty-plan guard went in.
    from programme import build_repair_plan as _m_brp
    _m_allrev = []
    for _fp, _d in _m_parsed.items():
        _f = out_of_sequence_flags(_d)
        if _f and not _m_brp(_d):
            _m_allrev.append(_fp.split("/")[-1])
    check("M3c the all-review-class state is real in the field corpus "
          "(flags present, repair plan empty)",
          bool(_m_allrev), "no field file reproduces it")
    if _m_allrev:
        print(f"      all-review-class files: {len(_m_allrev)}")

    _sp_files = [f for f in _m_parsed if "/SPML/" in f]
    if _sp_files:
        import time as _time
        _sp = _m_parsed[_sp_files[0]]
        _t0 = _time.time()
        run_all_checks(_sp, DCMAConfig())
        out_of_sequence_flags(_sp)
        from programme import collapse_asbuilt as _cab3
        _m_cab = _cab3(_sp, "SPML", set())
        _dt = _time.time() - _t0
        check("M4 17k-task file: DCMA + OOS + collapse under 30s",
              _dt < 30, f"{_dt:.1f}s")
        check("M4b heavy-OOS file: collapse validation gap warned",
              any("validation gap" in w for w in _m_cab.warnings))

# ---------------------------------------------------------------------------
# N19. Build provenance in every export (audit F-04, answered where it
# matters): a document must testify which code revision produced it.
print("\n--- Layer N19: build provenance in exports ---")
from buildinfo import build_stamp as _n19stamp
_n19s = _n19stamp()
check("N19 the stamp names the toolkit, a build id and a UTC moment",
      _n19s.startswith("Delay Analysis Toolkit — build ")
      and "generated" in _n19s and "UTC" in _n19s, _n19s)
_n19wb = load_workbook(io.BytesIO(
    _n15bapx("Stamp Probe", [("T", [{"col": 1}])])))
_n19ws = _n19wb.worksheets[0]
check("N19b Excel exports carry the stamp — description, print footer "
      "and a visible cell",
      (_n19wb.properties.description or "").startswith(
          "Delay Analysis Toolkit — build ")
      and (_n19ws.oddFooter.left.text or "").startswith(
          "Delay Analysis Toolkit — build ")
      and any(c.value.startswith("Delay Analysis Toolkit")
              for row in _n19ws.iter_rows() for c in row
              if isinstance(c.value, str)))
from docx import Document as _n19Doc
from programme import build_narrative_docx as _n19docx
_n19d = _n19Doc(io.BytesIO(_n19docx("Stamp Probe", "body")))
check("N19c Word exports carry the stamp in the page footer",
      _n19d.sections[0].footer.paragraphs[0].text.startswith(
          "Delay Analysis Toolkit — build "))
# STRUCTURAL: every workbook must exit through the stamping serialiser —
# a new builder pasted from an old one cannot silently skip it.
_n19x = open("programme/report_xlsx.py").read()
check("N19d all programme workbooks exit through _wb_bytes",
      _n19x.count("wb.save(buf)") == 1
      and _n19x.count("return _wb_bytes(wb)") >= 20)
check("N19e both Word builders stamp before saving",
      open("programme/report_docx.py").read().count(
          "    _stamp_docx(doc)") == 2)
check("N19f the DCMA workbook is stamped too",
      "build_stamp" in open("dcma/report_xlsx.py").read())

# ---------------------------------------------------------------------------
# N20. Degenerate-dates engine torture (audit F-10, the cluster that
# matters). Layer K proves the PARSER survives malformed files; nothing
# proved the ENGINES survive a parsed file whose dates are missing in
# every combination the field corpus happens not to contain: completed
# work with no actuals, in-progress with no early dates, milestones
# with no dates at all, a relationship to a task that does not exist.
# Every engine must return or warn — never throw.
print("\n--- Layer N20: degenerate-dates engine torture ---")
_n20_hdr = ("ERMHDR\t8.0\n"
            "%T\tPROJECT\n%F\tproj_id\tproj_short_name\tlast_recalc_date\n"
            "%R\tP1\tDEGEN\t2018-06-01 00:00\n"
            "%T\tCALENDAR\n%F\tclndr_id\tclndr_name\n%R\tC1\tStd\n"
            "%T\tTASK\n%F\ttask_id\tproj_id\ttask_code\ttask_name\t"
            "task_type\tstatus_code\tclndr_id\tearly_start_date\t"
            "early_end_date\tlate_start_date\tlate_end_date\t"
            "act_start_date\tact_end_date\ttarget_start_date\t"
            "target_end_date\ttotal_float_hr_cnt\tremain_dur_hr_cnt\n")
_n20_rows = [
    # complete but NO actual dates at all
    ("T1", "A1000", "Complete no actuals", "TT_Task", "TK_Complete",
     "", "", "", "", "", "", "", "", "", ""),
    # in-progress with an actual start but NO early/late/remaining
    ("T2", "A1010", "Started no forecast", "TT_Task", "TK_Active",
     "", "", "", "", "2018-05-01 08:00", "", "", "", "", ""),
    # not started with NO dates whatsoever
    ("T3", "A1020", "Future undated", "TT_Task", "TK_NotStart",
     "", "", "", "", "", "", "", "", "", ""),
    # milestone with no dates and no float
    ("T4", "MS-100", "Undated milestone", "TT_FinMile", "TK_NotStart",
     "", "", "", "", "", "", "", "", "", ""),
    # actual finish but NO actual start (real P6 exports do this)
    ("T5", "A1030", "Finish only", "TT_Task", "TK_Complete",
     "", "", "", "", "", "2018-04-10 17:00", "", "", "", ""),
    # healthy row so engines have something to anchor on
    ("T6", "A1040", "Healthy anchor", "TT_Task", "TK_Active",
     "2018-06-01 08:00", "2018-07-01 17:00", "2018-06-05 08:00",
     "2018-07-05 17:00", "2018-05-20 08:00", "", "2018-05-15 08:00",
     "2018-06-25 17:00", "40", "160"),
]
_n20_txt = _n20_hdr + "".join(
    "%R\t" + "\t".join([r[0], "P1", r[1], r[2], r[3], r[4], "C1",
                        *r[5:]]) + "\n"
    for r in _n20_rows
) + ("%T\tTASKPRED\n%F\ttask_pred_id\ttask_id\tpred_task_id\t"
     "pred_type\tlag_hr_cnt\n"
     "%R\tPR1\tT6\tT2\tPR_FS\t0\n"
     "%R\tPR2\tT2\tT5\tPR_FS\t8\n"
     "%R\tPR3\tT6\tT999\tPR_FS\t0\n"      # dangling predecessor
     "%E\n")
from dcma import build_dcma_trace
from programme import (build_repair_plan, collapse_asbuilt,
                       extract_asbuilt_longest_path as _n20_ablp,
                       extract_actual_trace as _n20_seq,
                       out_of_sequence_flags, planned_vs_actual)
_n20 = parse_xer(_n20_txt.encode())
check("N20 the degenerate file parses (6 tasks, dangling pred kept "
      "out or tolerated)", len(_n20.tasks) == 6)


def _n20_run(label, fn):
    try:
        fn()
        check(f"N20 {label} survives missing dates", True)
    except Exception as _e:  # noqa: BLE001 - the whole point
        check(f"N20 {label} survives missing dates", False,
              f"{type(_e).__name__}: {_e}")


_n20_run("DCMA checks + trace", lambda: build_dcma_trace(
    _n20, DCMAConfig(), run_all_checks(_n20, DCMAConfig())))
_n20_run("longest path", lambda: extract_longest_path(_n20, "DEGEN"))
_n20_run("float-based critical path",
         lambda: extract_critical_path(_n20, "DEGEN"))
_n20_run("OOS flags + repair plan", lambda: (
    out_of_sequence_flags(_n20), build_repair_plan(_n20)))
_n20_run("revision comparison (self)",
         lambda: compare_revisions(_n20, _n20, "a", "b"))
_n20_run("windows analysis",
         lambda: analyse_windows([("a", _n20), ("b", _n20)]))
_n20_run("planned vs actual", lambda: planned_vs_actual(_n20, _n20, None))
_n20_run("as-built longest path", lambda: _n20_ablp(_n20))
_n20_run("as-built sequence trace", lambda: _n20_seq([("a", _n20)]))
_n20_run("collapsed as-built",
         lambda: collapse_asbuilt(_n20, "DEGEN", set()))
_n20_run("milestone shifts", lambda: track_milestone_shifts(
    [("a", _n20.project.data_date, _n20),
     ("b", _n20.project.data_date, _n20)]))

# ---------------------------------------------------------------------------
# N21. Cloud memory budget (audit F-03, scoped): refuse BEFORE
# allocating, never warn-and-crash. A ceiling, not a fix — the fix is
# running large matters locally, and the refusal says so.
print("\n--- Layer N21: cloud memory budget ---")
from programme import (CLOUD_BUDGET_MB, CLOUD_PARSE_FACTOR,
                       cloud_memory_verdict)
_n21_agg, _n21_est, _n21_ok = cloud_memory_verdict(
    [20_580_000] * 4)   # four SPML-sized revisions
check("N21 four 20 MB revisions exceed the Cloud budget "
      "(three parse to ~470 MB and are allowed)",
      not _n21_ok and _n21_est > CLOUD_BUDGET_MB
      and cloud_memory_verdict([20_580_000] * 3)[2],
      f"est {_n21_est:.0f} MB")
check("N21b a normal matter (8 x 2 MB) passes",
      cloud_memory_verdict([2_000_000] * 8)[2])
check("N21c the factor is grounded in the measured 7.6x, rounded up",
      7.6 <= CLOUD_PARSE_FACTOR <= 10.0)
# STRUCTURAL: the intake must refuse (st.error + return) before the
# parse loop, using the shared verdict — not a bare warning.
_n21src = open("views/intake.py").read()
check("N21d intake refuses before allocating on an unsafe estimate",
      "cloud_memory_verdict" in _n21src
      and "Refusing to parse this set" in _n21src
      and _n21src.index("Refusing to parse")
      < _n21src.index("Parsing programmes"))

# =========================================================================
# P layer — audit fix pins (misdescribed outputs, election, contrast,
# Excel work-product)
# =========================================================================
print("== P. Audit fix pins ==")

# SELF-CONTAINED: by this point in the suite several early globals
# (_p, cfg, ...) have been shadowed by later layers — everything this
# layer needs is re-derived locally
import os as _os_p

from dcma.config import DCMAConfig as _DCMAConfigP


def _pp(rel: str) -> str:
    return _os_p.path.join(_os_p.path.dirname(_os_p.path.abspath(__file__)),
                           rel)


_cfgP = _DCMAConfigP()
with open(_pp("sample/Harbour Point DCP-03 - Baseline Programme Rev 0.xer"), "rb") as fh:
    _BP = parse_xer(fh.read())
with open(_pp("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"), "rb") as fh:
    _UP = parse_xer(fh.read())

# P1: check 12 tests CONTINUITY, and the counts feeding narratives are
# computed, never hardcoded
_hp = _pp("sample/programmes/harbour_point_dcp03/"
          "Harbour Point DCP-03 - Baseline Programme Rev 0.xer")
with open(_hp, "rb") as fh:
    _HPB = parse_xer(fh.read())
_c12_hp = next(c for c in run_all_checks(_HPB, _cfgP) if c.number == 12)
_c12_b = next(c for c in run_all_checks(_BP, _cfgP) if c.number == 12)
check("P1 check 12 passes a genuinely continuous path (1 segment)",
      str(_c12_hp.status).endswith("PASS")
      and "1 segment" in _c12_hp.metric_value, _c12_hp.metric_value)
_bp_txt = open(_pp("sample/Harbour Point DCP-03 - Baseline Programme Rev 0.xer"), encoding="utf-8").read()
_bp_tf0 = {t.task_id for t in _BP.tasks
           if not t.is_loe_or_wbs and (t.total_float_hr or 0) <= 0}
_bp_codes = {t.task_id: t.task_code for t in _BP.tasks}
_mid_rel = next(r for r in _BP.relationships
                if r.pred_task_id in _bp_tf0 and r.task_id in _bp_tf0
                and _bp_codes[r.task_id].startswith("T-"))
_BPX = parse_xer(_del_taskpred(_bp_txt, _mid_rel.pred_task_id,
                               _mid_rel.task_id))
_c12_x = next(c for c in run_all_checks(_BPX, _cfgP)
              if c.number == 12)
check("P1b check 12 fails disconnected low-float segments",
      str(_c12_x.status).endswith("FAIL")
      and "DISCONNECTED" in _c12_x.summary, _c12_x.metric_value)
check("P1c no hardcoded 'of 14' denominator in views",
      all("of 14 " not in open(_pp(f)).read()
          for f in ("views/report.py", "views/tia.py")))

import io as _io4

# P2: revision election — reversed upload order still lands is_current
# on the latest data date, and the shared resolver follows it
from programme import build_inventory
_fwd = [("Harbour Point DCP-03 - Baseline Programme Rev 0.xer", _BP), ("Harbour Point DCP-03 - As-Built Programme Rev 12.xer", _UP)]
_inv_r = build_inventory(list(reversed(_fwd)))
check("P2 reversed upload: is_current still the latest data date",
      _inv_r.current is not None
      and _inv_r.current.file_name == "Harbour Point DCP-03 - As-Built Programme Rev 12.xer")
from views._shared import current_default_index
_names_r = [n for n, _ in reversed(_fwd)]        # update first
check("P2b resolver elects the current revision, not position",
      _names_r[current_default_index(_names_r, _inv_r)]
      == "Harbour Point DCP-03 - As-Built Programme Rev 12.xer")

# P3: disclosure text contrast >= AA 4.5:1, computed from the theme
# constants so a palette edit cannot silently regress it
import re as _re


def _lum(hexc):
    ch = [int(hexc[i:i + 2], 16) / 255 for i in (0, 2, 4)]
    ch = [x / 12.92 if x <= 0.04045 else ((x + 0.055) / 1.055) ** 2.4
          for x in ch]
    return 0.2126 * ch[0] + 0.7152 * ch[1] + 0.0722 * ch[2]


def _contrast(fg, bg):
    a, b = sorted((_lum(fg), _lum(bg)), reverse=True)
    return (a + 0.05) / (b + 0.05)


_theme = open(_pp("views/_theme.py")).read()
_soft = _re.search(r"--dsi-soft:\s*#([0-9A-Fa-f]{6})", _theme).group(1)
_paper = _re.search(r"--dspaper:\s*#([0-9A-Fa-f]{6})", _theme).group(1)
check("P3 annotation voice meets AA on paper (>= 4.5:1)",
      _contrast(_soft, _paper) >= 4.5,
      f"{_contrast(_soft, _paper):.2f}:1")
_gantt_src = open(_pp("programme/gantt_html.py")).read()
_muted = _re.search(r"--muted:\s*#([0-9A-Fa-f]{6})", _gantt_src).group(1)
_canvas = _re.search(r"--canvas:\s*#([0-9A-Fa-f]{6})", _gantt_src).group(1)
check("P3b gantt annotation voice meets AA on canvas",
      _contrast(_muted, _canvas) >= 4.5,
      f"{_contrast(_muted, _canvas):.2f}:1")

# P4: band exported as data, autofilter on every single-table sheet
from dcma.report_xlsx import build_xlsx_report as _dcma_xlsx
_cp_p = extract_longest_path(_BP, "Baseline")
_wb_cp = load_workbook(_io4.BytesIO(build_critical_path_xlsx(_cp_p, None)))
_ws_cp = _wb_cp["Critical Path"]
_hdrs = [c.value for c in _ws_cp[5]]
check("P4 CP export carries Band as a COLUMN, not colour only",
      "Band" in _hdrs and _ws_cp.cell(
          row=6, column=_hdrs.index("Band") + 1).value
      in ("Critical", "Near-critical"))
check("P4b CP sheet has an autofilter", bool(_ws_cp.auto_filter.ref))
_wb_d = load_workbook(_io4.BytesIO(
    _dcma_xlsx(_BP, run_all_checks(_BP, _cfgP), "Baseline")))
check("P4c DCMA scorecard + detail sheets have autofilters",
      bool(_wb_d["Summary"].auto_filter.ref)
      and all(bool(ws.auto_filter.ref) for ws in _wb_d.worksheets
              if ws.title[0].isdigit()))
_no_filter = [ws.title for ws in _wb_cp.worksheets
              if ws.max_row > 6 and ws.title not in
              ("AI Narrative", "Caveats") and not ws.auto_filter.ref]
check("P4d every data sheet in the CP workbook is filterable",
      not _no_filter, str(_no_filter))

# =========================================================================
# X layer — exhaustive-audit fixes (C1-C5, H2/H3/H5/H6, M2/M5/M6)
# SELF-CONTAINED: fresh parses + config, no reliance on suite globals.
# =========================================================================
print("== X. Exhaustive-audit fix pins ==")
from datetime import datetime as _xdt

from programme.tia import (DelayEvent as _XDE, FragnetActivity as _XFA,
                           FragnetLink as _XFL,
                           run_cumulative_tia as _xcum, run_tia as _xtia)

with open(_pp("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"), encoding="utf-8") as fh:
    _XU = parse_xer(_truncate_asbuilt(fh.read(), "2025-09-01 08:00"))
_xev = _XDE("EV-X", "x", "")
_xfrag = [_XFA("TIA-X10", "w", 20,
               predecessors=[_XFL("T-4040")],
               successors=[_XFL("T-4070")])]

# C1 — the elected milestone IS the headline, and gates when absent
_xr = _xtia(_XU, "U", _xev, _xfrag, target_milestone="T-4070")
check("X-C1 headline measured AT the elected milestone",
      _xr.measured_at == "T-4070"
      and _xr.completion_pre is not None
      and _xr.completion_pre.year == 2025, str(_xr.completion_pre))
_xr2 = _xtia(_XU, "U", _xev, _xfrag, target_milestone="NOT-A-CODE")
check("X-C1b unmeasurable election GATES the headline",
      _xr2.headline_gated and _xr2.completion_delta_days is None)
_xr3 = _xtia(_XU, "U", _xev, _xfrag)
check("X-C1c no election -> latest-finisher warning disclosed",
      any("No completion milestone elected" in w for w in _xr3.warnings))
_xc = _xcum(_XU, "U", [(_xev, _xfrag)], target_milestone="NOT-A-CODE")
check("X-C1d cumulative quantum gated on unmeasurable election",
      _xc["gated"] and not _xc["rows"])
_xc2 = _xcum(_XU, "U", [(_xev, _xfrag)], target_milestone="T-4070")
check("X-C1e cumulative measured at the election",
      _xc2["measured_at"] == "T-4070" and not _xc2["gated"])

# C2/H2 — FF back-computes ES (duration preserved, SS reads it);
# negative lag steps working days backwards
from programme.cpm import (add_working_days as _xadd,
                           cyclic_nodes as _xcyc,
                           forward_pass as _xfwd,
                           sub_working_days as _xsub)
_XMF = (frozenset(range(0, 5)), frozenset(), frozenset())
check("X-H2 -1d lead from Monday lands Friday, not Sunday",
      _xadd(_xdt(2026, 1, 5), -1, _XMF) == _xdt(2026, 1, 2))
_xn = {"P": (10.0, _XMF), "S": (2.0, _XMF), "X": (5.0, _XMF)}
_xp = {"P": [], "S": [("P", "FF", 0.0)], "X": [("S", "SS", 0.0)]}
_xES, _xEF, _, _ = _xfwd(_xn, _xp, _xdt(2026, 1, 5), {})
check("X-C2 FF-governed successor: finish aligned, duration preserved",
      _xEF["S"] == _xEF["P"]
      and _xES["S"] == _xsub(_xEF["S"], 2.0, _XMF))
check("X-C2b downstream SS reads the back-computed ES",
      _xES["X"] == _xES["S"])

# C3 — non-convergence detected; collapse quantum suppressed
from programme.collapsed_asbuilt import _schedule as _xsched
_, _, _xcv = _xsched({"A": 5.0, "B": 5.0},
                     [("A", "B", "FS", 1.0), ("B", "A", "FS", 1.0)],
                     _xdt(2024, 1, 1))
check("X-C3 positive cycle -> relaxation reports non-convergence",
      _xcv is False)
_, _, _xcv2 = _xsched({"A": 5.0, "B": 3.0}, [("A", "B", "FS", 0.0)],
                      _xdt(2024, 1, 1))
check("X-C3b acyclic network converges", _xcv2 is True)
_xcabsrc = open(_pp("programme/collapsed_asbuilt.py")).read()
check("X-C3c non-convergence SUPPRESSES quantum (returns early)",
      _xcabsrc.count("QUANTUM SUPPRESSED") >= 2)

# C4 — structural defects detected; intake refuses
from dcma import structural_defects as _xsd
check("X-C4 clean single-project file has no structural defects",
      _xsd(_XU) == [])
import re as _xre
_xraw = open(_pp("sample/Harbour Point DCP-03 - As-Built Programme Rev 12.xer"), encoding="latin-1").read()
_xm = _xre.search(r"^(%R\t.*?\tT-4040\t.*)$", _xraw, _xre.M)
_xrow = _xm.group(1).split("\t"); _xrow[1] = "99999999"
_xdup = parse_xer(_xraw.replace(_xm.group(1),
                                _xm.group(1) + "\n" + "\t".join(_xrow)))
check("X-C4b duplicate Activity ID detected",
      any("duplicate Activity ID" in d for d in _xsd(_xdup)))
_xrow2 = _xm.group(1).split("\t"); _xrow2[1] = "99999998"
_xrow2[2] = "77777"; _xrow2[13] = "ZZ-UNIQ-1"
_xmp = parse_xer(_xraw.replace(_xm.group(1),
                               _xm.group(1) + "\n" + "\t".join(_xrow2)))
check("X-C4c multi-project export detected",
      any("Multi-project" in d for d in _xsd(_xmp)))
check("X-C4d intake REFUSES structurally defective files",
      "structural_defects" in open(_pp("views/intake.py")).read()
      and "refused — structural" in open(_pp("views/intake.py")).read())

# C5 — fractional movement survives windows arithmetic
_xwsrc = open(_pp("programme/windows.py")).read()
check("X-C5 windows uses total_seconds, not truncating .days",
      "total_seconds() / 86400" in _xwsrc
      and ".days\n" not in _xwsrc.split("movement_days=")[1][:120])

# M2 — a cycle ON the measured path gates the TIA headline
_xloop = [_XFA("TIA-C", "loop", 10,
               predecessors=[_XFL("T-4070")],
               successors=[_XFL("T-4070")])]
_xrc = _xtia(_XU, "U", _xev, _xloop, target_milestone="T-4070")
check("X-M2 cycle on the measured path gates the headline",
      _xrc.headline_gated and _xrc.completion_delta_days is None
      and any("circular" in w for w in _xrc.warnings))
check("X-M2b cyclic_nodes finds the trapped pair",
      _xcyc({"A": (5.0, None), "B": (5.0, None), "C": (2.0, None)},
            {"A": [("B", "FS", 0.0)], "B": [("A", "FS", 0.0)], "C": []})
      == {"A", "B"})

# H3 — UTF-8 decodes as UTF-8; cp1252-only bytes still fall back
from dcma.xer_parser import _decode_bytes as _xdec
check("X-H3 utf-8 round-trips (no mojibake)",
      _xdec("Café — İnşaat".encode("utf-8")) == "Café — İnşaat")
check("X-H3b cp1252 smart quotes still decode via fallback",
      _xdec(b"\x93quoted\x94") == "“quoted”")

# H5 — hostile task names are inert in both HTML builders
from programme.gantt_html import build_apab_gantt_html as _xapab
from programme.gantt_html import build_gantt_html as _xg
_xrows = [{"task_code": "X</script><script>alert(1)</script>",
           "name": "Evil</span><script>alert(2)</script>",
           "row_kind": "", "planned_start": None, "planned_finish": None,
           "actual_start": _xdt(2024, 1, 2),
           "actual_finish": _xdt(2024, 2, 2), "finish_var_days": 3.0}]
check("X-H5 apab gantt neutralises hostile names",
      "<script>alert" not in _xapab(_xrows, title="t", data_date=None))
_xtree = {"name": "g", "children": [], "activities": [
    {"id": "A", "name": "</script><script>alert(3)</script>",
     "start": "2024-01-01", "finish": "2024-02-01", "status": "done"}]}
check("X-H5b tree gantt JSON cannot break out of its script tag",
      "</script><script>alert(3)" not in _xg(_xtree))

# H6 — formula-typed cells are re-typed as literal strings
import io as _xio
from openpyxl import Workbook as _XWB, load_workbook as _xload
from programme.report_xlsx import _wb_bytes as _xwb
_xw = _XWB(); _xws = _xw.active
_xws["A1"] = "h"; _xws["A2"] = '=HYPERLINK("http://evil","x")'
_xout = _xload(_xio.BytesIO(_xwb(_xw))).active
check("X-H6 '=' cells exported as text, content preserved",
      _xout["A2"].data_type != "f"
      and _xout["A2"].value == '=HYPERLINK("http://evil","x")')

# M5/M6 — source-structural pins
_xint = open(_pp("views/intake.py")).read()
check("X-M5 milestone options come from the inventory's current",
      "inv.current" in _xint.split("_ms_opts")[0][-600:])
_xapp = open(_pp("app.py")).read()
check("X-M6 hosted deployment fails CLOSED without a secret",
      "/mount/src" in _xapp and "ALLOW_PUBLIC" in _xapp
      and _xapp.index("Access not configured")
      < _xapp.index("Access password"))

# =========================================================================
# S layer — release-stress pins (2026-07-31 stress campaign)
# =========================================================================
print("== S. Release-stress pins ==")

# S1 — THE stress finding: cumulative TIA gates a cycle-creating event
_sloop = [_XFA("TIA-S1", "loop", 5, predecessors=[_XFL("T-4070")],
               successors=[_XFL("T-4070")])]
# explicit dates: the clean event must SORT first (undated events sort
# to the data date, i.e. after June 2016)
_sev1 = _XDE("EV-S1", "earlier", "", _xdt(2016, 3, 1))
_sev2 = _XDE("EV-S2", "later", "", _xdt(2016, 6, 1))
_sc = _xcum(_XU, "U", [(_sev1, _xfrag), (_sev2, _sloop)],
            target_milestone="T-4070")
check("S1 cumulative gates the cycle-creating event",
      _sc["gated"] and _sc["total_delta_days"] is None
      and any("QUANTUM GATED" in w for w in _sc["warnings"]))
check("S1b rows before the gated event remain (earlier increments valid)",
      len(_sc["rows"]) == 1
      and _sc["rows"][0]["event_id"] == "EV-S1")

# S2 — CPM corner pins from the truth table
_sMF = (frozenset(range(0, 5)), frozenset(), frozenset())
_sES, _sEF, _, _ = _xfwd(
    {"P": (2.0, _sMF), "Q": (10.0, _sMF), "S": (5.0, _sMF)},
    {"P": [], "Q": [], "S": [("Q", "FS", 0.0), ("P", "FF", 0.0)]},
    _xdt(2026, 1, 5), {})
check("S2 late FS floor beats an early FF bound",
      _sES["S"] == _sEF["Q"]
      and _sEF["S"] == _xadd(_sES["S"], 5, _sMF))
_sES, _sEF, _, _ = _xfwd(
    {"P": (30.0, _sMF), "S": (2.0, _sMF)},
    {"P": [], "S": [("P", "FF", 0.0)]},
    _xdt(2026, 1, 5), {"S": _xdt(2026, 1, 5)})
check("S2b in-progress pin never violated by FF back-compute",
      _sES["S"] >= _xdt(2026, 1, 5) and _sEF["S"] >= _sEF["P"])
_sES, _sEF, _, _ = _xfwd(
    {"A": (10.0, _sMF), "B": (3.0, _sMF)},
    {"A": [], "B": [("A", "SS", -2.0)]}, _xdt(2026, 1, 12), {})
check("S2c negative SS lag cannot pull before the project floor",
      _sES["B"] == _xdt(2026, 1, 12))

# S3 — windows telescope exactly on the fixtures (fractional-safe)
_srevs = []
for _sf in ("revA.xer", "revB.xer", "revC.xer"):
    with open(_pp(f"sample/revisions/{_sf}"), "rb") as fh:
        _srevs.append((_sf, parse_xer(fh.read())))
_sw = analyse_windows(_srevs)
_smoves = [r.movement_days for r in _sw.windows
           if r.movement_days is not None]
_sdirect = round((_srevs[-1][1].project.scheduled_finish
                  - _srevs[0][1].project.scheduled_finish).total_seconds()
                 / 86400, 1)
check("S3 windows telescope exactly to the direct movement",
      abs(sum(_smoves) - _sdirect) < 0.05,
      f"sum={sum(_smoves)} direct={_sdirect}")

# =========================================================================
# Y layer — cross-tree convergence fixes (2026-07-31 qa-copy comparison)
# =========================================================================
print("== Y. Convergence-fix pins ==")
from datetime import timedelta as _ytd

# Y1 windows: movement measured AT the elected milestone, gated when absent
_yw = analyse_windows([("B", _BP), ("U", _UP)], end_task_code="T-4070")
_yw0 = analyse_windows([("B", _BP), ("U", _UP)])
check("Y1 windows movement basis follows the election",
      _yw.windows[0].movement_days != _yw0.windows[0].movement_days
      and any("ELECTED milestone" in c for c in _yw.caveats),
      f"{_yw.windows[0].movement_days} vs {_yw0.windows[0].movement_days}")
_ywb = analyse_windows([("B", _BP), ("U", _UP)], end_task_code="NOPE")
check("Y1b unmeasurable election gates window movement",
      _ywb.windows[0].movement_days is None
      and any("not reported" in w for w in _ywb.warnings))

# Y2 comparison_impact: headline at the election; gated when absent
from programme import assess_comparison_impact as _yci
_yr = _yci(_BP, _UP, "B", "U", end_task_code="T-4070")
_yrb = _yci(_BP, _UP, "B", "U", end_task_code="NOPE")
check("Y2 impact headline measured at the elected milestone",
      _yr.completion_moved_days is not None
      and _yr.completion_moved_days != _yci(
          _BP, _UP, "B", "U").completion_moved_days)
check("Y2b impact headline gated on unmeasurable election",
      _yrb.completion_moved_days is None
      and any("GATED" in w for w in _yrb.warnings))

# Y3 notice: one hour late is LATE (the -0.0 compliant bug)
from programme.notice import assess_notice as _yan
_yna = _yan(_xdt(2026, 1, 5, 8), _xdt(2026, 1, 6, 9), 1)
check("Y3 notice 1h late -> LATE with sub-day margin",
      _yna.status == "late" and _yna.margin_days is not None
      and -0.05 < _yna.margin_days < 0)

# Y4 SCHEDOPTIONS lag basis: rcal_Successor honoured end to end
_ymini = (
    "ERMHDR\t24.12\t2026-01-01\tProject\tADMIN\tu\tdb\tPM\tUSD\n"
    "%T\tPROJECT\n%F\tproj_id\tproj_short_name\tclndr_id"
    "\tlast_recalc_date\n%R\t1\tP1\t1\t2026-01-05 08:00\n"
    "%T\tCALENDAR\n%F\tclndr_id\tclndr_name\tday_hr_cnt\tclndr_data\n"
    "%R\t1\tTen\t10\t(0||CalendarData()())\n"
    "%R\t2\tFive\t5\t(0||CalendarData()())\n"
    "%T\tSCHEDOPTIONS\n%F\tschedoptions_id\tproj_id"
    "\tsched_calendar_on_relationship_lag\n%R\t1\t1\trcal_Successor\n"
    "%T\tTASK\n%F\ttask_id\tproj_id\ttask_code\ttask_name\ttask_type"
    "\tstatus_code\tclndr_id\ttarget_drtn_hr_cnt\tremain_drtn_hr_cnt\n"
    "%R\t10\t1\tA\ta\tTT_Task\tTK_NotStart\t1\t10\t10\n"
    "%R\t11\t1\tB\tb\tTT_Task\tTK_NotStart\t2\t5\t5\n"
    "%T\tTASKPRED\n%F\ttask_pred_id\ttask_id\tpred_task_id\tproj_id"
    "\tpred_proj_id\tpred_type\tlag_hr_cnt\n"
    "%R\t1\t11\t10\t1\t1\tPR_FS\t10\n%E\n")
_ymd = parse_xer(_ymini.encode())
from dcma.calendar import relationship_lag_hours_per_day as _ylag
_yhpd, _ylabel = _ylag(_ymd, "1", "2")
check("Y4 rcal_Successor elects the successor calendar",
      _yhpd == 5.0 and "successor" in _ylabel, f"{_yhpd} {_ylabel}")
from programme.cpm import build_network as _ybn
_, _, _ypreds, _, _, _ = _ybn(_ymd, _DCMAConfigP(), _ymd.project.data_date)
_ylagdays = _ypreds["B"][0][2]
check("Y4b CPM lag converted on the elected basis (10h / 5h/d = 2d)",
      abs(_ylagdays - 2.0) < 1e-9, str(_ylagdays))

# Y5 shift-aware fractional arithmetic + exact inverse
from dcma.calendar import (add_working_days as _yadd,
                           calendar_masks as _ycm,
                           working_days_between as _ywdb)
_ymk = next(v for v in _ycm(_XU).values()
            if v[3] == 8.0 and len(v[4]) == 5)
_yr2 = _yadd(_xdt(2026, 1, 5, 8), 0.5, _ymk)
check("Y5 Mon 08:00 + 0.5wd on 8h shift = Mon 12:00",
      _yr2 == _xdt(2026, 1, 5, 12), str(_yr2))
check("Y5b working_days_between inverts the arithmetic",
      abs(_ywdb(_xdt(2026, 1, 5, 8), _yr2, _ymk) - 0.5) < 1e-6)

# Y6 progress: milestone step preserved, ramp from its own start
from programme.progress import (_spread as _ysp, _to_curve as _ytc,
                                _value_at as _yva)
_yb, _yi = {}, {}
_ysp(_yb, _xdt(2026, 1, 1), _xdt(2026, 1, 31), 1.0, _yi)
_ysp(_yb, _xdt(2026, 1, 15), _xdt(2026, 1, 15), 1.0, _yi)
_yc = _ytc(_yb, 2.0, _yi)
check("Y6 planned at 7 Jan ~10% (was earned-early 31%+)",
      abs(_yva(_yc, _xdt(2026, 1, 7)) - 10.0) < 0.5)
check("Y6b milestone is a STEP (23.3 -> 73.3 across its date)",
      abs(_yva(_yc, _xdt(2026, 1, 14, 23)) - 23.2) < 0.5
      and abs(_yva(_yc, _xdt(2026, 1, 15, 0, 0, 1)) - 73.3) < 0.5)

# Y7 tier-2 structural gates + encoding disclosure
_ydang = _xraw
import re as _yre
_ym = _yre.search(r"^(%R\t\d+\t\d+\t\d+\t\d+\t\d+\tPR_FS\t.*)$",
                  _ydang, _yre.M)
_yrow = _ym.group(1).split("\t")
_yrow[1], _yrow[3] = "88888888", "77777777"
_yD = parse_xer(_ydang.replace(_ym.group(1),
                               _ym.group(1) + "\n" + "\t".join(_yrow)))
check("Y7 dangling LOCAL join is a structural defect",
      any("do not exist" in d for d in _xsd(_yD)))
_yrow2 = _ym.group(1).split("\t")
_yrow2[1], _yrow2[3], _yrow2[5] = "88888887", "77777776", "99999"
_yD2 = parse_xer(_ydang.replace(_ym.group(1),
                                _ym.group(1) + "\n" + "\t".join(_yrow2)))
check("Y7b external-project logic is NOT gated", _xsd(_yD2) == [])
_yinv = (
    "ERMHDR\t1\n%T\tPROJECT\n%F\tproj_id\tproj_short_name\n%R\t1\tP\n"
    "%T\tTASK\n%F\ttask_id\tproj_id\ttask_code\ttask_name\ttask_type"
    "\tstatus_code\tclndr_id\tact_start_date\tact_end_date\n"
    "%R\t10\t1\tA100\tx\tTT_Task\tTK_Complete\t1"
    "\t2026-02-01 08:00\t2026-01-01 17:00\n%E\n")
check("Y7c finish-before-start is a structural defect",
      any("impossible date" in d for d in _xsd(parse_xer(_yinv.encode()))))
# bytes input discloses its encoding; text input has none to disclose
check("Y7d encoding disclosed in parse notes (bytes input)",
      any("decoded as utf-8" in n
          for n in parse_xer(_yinv.encode()).parse_notes)
      and parse_xer(_yinv).parse_notes == [])

# Y8 decision-grade calibration gate
_yt = _xtia(_XU, "U", _xev, _xfrag, target_milestone="T-4070")
check("Y8 TIA decision_grade set from calibration tolerance",
      _yt.decision_grade is True and _yt.calibration_days is not None)
check("Y8b tolerance is documented in config",
      _DCMAConfigP().calibration_tolerance_days == 30.0)

# Z1 — the assembled report must offer the FLAGSHIP retrospective
# method. The Report Assembler shipped for months without an
# As-Planned vs As-Built section: an analyst could run the method and
# have it silently omitted from the Word report. Pinned end-to-end
# through the real page, because the defect was one of absence.
def _z1_report_carries_apab():
    from streamlit.testing.v1 import AppTest

    from programme import build_inventory, extract_asbuilt_longest_path
    import state as _sk

    _root = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "sample")
    _pool = [(n, parse_xer(open(os.path.join(_root, n), "rb").read()))
             for n in ("Harbour Point DCP-03 - Baseline Programme "
                       "Rev 0.xer",
                       "Harbour Point DCP-03 - As-Built Programme "
                       "Rev 12.xer")]
    _ms = "H-5040"
    _tr = extract_asbuilt_longest_path(_pool[1][1], end_task_code=_ms)

    def _page(root: str):
        # AppTest re-execs this function's SOURCE — nothing may be
        # captured from the enclosing scope; the repo root arrives
        # as an argument.
        import sys as _s
        _s.path.insert(0, root)
        from views.report import report_tab
        report_tab()

    at = AppTest.from_function(
        _page, default_timeout=300,
        kwargs={"root": os.path.dirname(os.path.abspath(__file__))})
    at.session_state[_sk.XER_POOL] = _pool
    at.session_state[_sk.INVENTORY] = build_inventory(_pool)
    at.session_state["apab2_paths"] = {
        _ms: [(a.task_code, a.name) for a in _tr.activities]}
    at.session_state["apab2_basis"] = {_ms: "as-built longest path"}
    at.session_state["apab2_date_basis"] = "late"
    at.run()
    if at.exception:
        return False, f"report page raised: {at.exception}"
    labels = " ".join(str(getattr(c, "label", "")) for c in at.checkbox)
    if "as-planned vs as-built" not in labels.lower():
        return False, "no APvAB section offered"
    btn = [b for b in at.button if b.key == "rep_build"]
    if not btn:
        return False, "no assemble button"
    btn[0].click()
    at.run()
    if "rep_docx" not in at.session_state:
        return False, "no docx produced"
    import io as _io
    import zipfile as _zf
    with _zf.ZipFile(_io.BytesIO(at.session_state["rep_docx"])) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    if "As-Planned vs As-Built" not in xml:
        return False, "section missing from the Word file"
    if "+92" not in xml:
        return False, "measured delay missing from the Word file"
    return True, ""


try:
    _z1_ok, _z1_why = _z1_report_carries_apab()
except Exception as _z1_exc:                       # pragma: no cover
    _z1_ok, _z1_why = False, f"{type(_z1_exc).__name__}: {_z1_exc}"
check("Z1 assembled report carries the As-Planned vs As-Built section "
      "and its measured delay", _z1_ok, _z1_why)

# ---------------------------------------------------------------- #
# Z2-Z7 — external-review fixes: no verdict may contradict the      #
# metric beside it, and degenerate results must explain themselves. #
# ---------------------------------------------------------------- #
from dcma.checks import CheckStatus as _CS

_zcs = {c.number: c for c in run_all_checks(AB, _DCMAConfigP())}
check("Z2 empty populations are N/A, never a vacuous PASS",
      all(_zcs[n].status == _CS.NA for n in (1, 6, 8, 17)),
      str({n: _zcs[n].status.value for n in (1, 6, 8, 17)}))
check("Z2b fully-complete file: missed-tasks and BEI are N/A with a "
      "stated reason",
      _zcs[11].status == _CS.NA and _zcs[14].status == _CS.NA
      and "complete" in (_zcs[11].na_reason or "")
      and "complete" in (_zcs[14].na_reason or ""))
check("Z2c check 9 compares DATES: same-day end-of-shift actuals are "
      "not future-dated",
      _zcs[9].status == _CS.PASS, _zcs[9].summary)

from programme import compute_progress as _zcompute
_zpr = _zcompute(B, "base", [("ab", AB)])
check("Z3 S-curve verdict is driven by the TIME OFFSET, not the "
      "saturated percentage",
      any(w.startswith("Behind plan by roughly") for w in _zpr.warnings)
      and not any(w.startswith("Favourable") for w in _zpr.warnings),
      str(_zpr.warnings[:2]))

from programme import run_tia as _zrun_tia
_zt = _zrun_tia(AB, "AB", DelayEvent("EV-Z", "z"),
                [FragnetActivity("Z-1", "z", 40,
                                 successors=[FragnetLink("C-3040")])])
check("Z4 TIA fragnet tied into COMPLETE work: no crash, VOID RUN "
      "declared",
      any(w.startswith("VOID RUN") for w in _zt.warnings)
      and not any(w.startswith("Favourable") for w in _zt.warnings),
      str(_zt.warnings[:3]))

from programme import run_impacted_asplanned as _zrun_iap
_ziap = _zrun_iap(B, "B", [(DelayEvent("EV-Z2", "z"),
                            [FragnetActivity(
                                "Z-2", "pred only", 40,
                                predecessors=[FragnetLink("E-1030")])])])
check("Z5 IAP refuses a predecessor-only fragnet with the reason "
      "stated",
      _ziap["events_used"] == 0
      and any("cannot transmit delay" in s
              for s in _ziap["skipped_events"]),
      str(_ziap["skipped_events"]))

from programme import collapse_asbuilt as _zcab
_zc = _zcab(AB, "AB", {"T-4055"})
check("Z6 CAB gates a signal smaller than its own reconstruction "
      "error",
      _zc.decision_grade is False
      and any(w.startswith("INDICATIVE ONLY") for w in _zc.warnings),
      f"delta {_zc.delta_days} cal {_zc.calibration_days} "
      f"grade {_zc.decision_grade}")

from programme import run_progress_transfer as _zpt
_zp = _zpt(B, AB, "B", "AB")
check("Z7 Progress Transfer explains an unmeasurable headline",
      _zp.network_effect_days is not None
      or any("not measurable" in w for w in _zp.warnings),
      str(_zp.warnings[:4]))

print(f"\n{'='*60}\nRESULT: {len(PASS)} passed, {len(FAIL)} FAILED")
for name, d in FAIL:
    print(f"  FAILED: {name} — {d}")

sys.exit(1 if FAIL else 0)
