"""File upload, listing, deletion, stats, and export endpoints."""

import os
from pathlib import Path
from typing import List

from fastapi import APIRouter, UploadFile, File, BackgroundTasks, Depends, HTTPException
from fastapi.responses import StreamingResponse

from backend.core.security import get_current_user, UserContext
from backend.core.projects import ProjectContext, get_current_project, require_project_editor
from backend.models.responses import FileInfo, UploadResult
from backend.services.file_service import FileService
from backend.tasks.indexing import index_file_background

router = APIRouter()
_file_service = FileService()


def _delete_legacy_project_file(file_id: str, project_id: str) -> dict | None:
    """Delete a project-owned source that predates the document registry.

    Bulk corpora use the exact file name as the Files API id. Membership in the
    project-scoped chunk store or catalog must be established before any write;
    a guessed global filename is never sufficient authorization.
    """
    file_name = Path(file_id).name
    if not file_name or file_name != file_id:
        return None

    from src.chunk_store import get_chunk_store
    from src.catalog import get_catalog
    from src.document_rag import generate_doc_id

    chunk_store = get_chunk_store()
    con = chunk_store.connection()
    chunk_rows = con.execute(
        "SELECT DISTINCT doc_id FROM chunks WHERE project_id=? AND file_name=?",
        [project_id, file_name],
    ).fetchall()
    doc_ids = {str(row[0]) for row in chunk_rows if row and row[0]}

    catalog = get_catalog()
    catalog_entries = [
        entry for entry in catalog.entries.values()
        if (getattr(entry, "project_id", "") or "") == project_id
        and file_id in (Path(entry.source_file).name, generate_doc_id(entry.source_file))
    ]
    if catalog_entries and file_name == file_id:
        # A catalog id can be a generated hash; use its real source name for
        # derivative cleanup and disk resolution.
        file_name = Path(catalog_entries[0].source_file).name
    doc_ids.update(generate_doc_id(entry.source_file) for entry in catalog_entries)

    if not chunk_rows and not catalog_entries:
        return None

    # Check whether a legacy global source is shared before removing it from
    # disk. Project-local sources are isolated by construction.
    other_chunk_ref = bool(con.execute(
        "SELECT 1 FROM chunks WHERE project_id<>? AND file_name=? LIMIT 1",
        [project_id, file_name],
    ).fetchone())
    other_catalog_ref = any(
        (getattr(entry, "project_id", "") or "") != project_id
        and Path(entry.source_file).name == file_name
        for entry in catalog.entries.values()
    )
    try:
        from src.document_registry import get_document_registry
        other_registry_ref = any(
            (getattr(record, "project_id", "") or "") != project_id
            and record.file_name == file_name
            for record in get_document_registry().get_all()
        )
    except Exception:
        other_registry_ref = False
    shared_source = other_chunk_ref or other_catalog_ref or other_registry_ref

    result = {"file_name": file_name, "legacy": True}

    table_names = sorted({
        table.table_name
        for entry in catalog_entries
        for table in entry.tables
        if table.table_name
    })
    if table_names:
        try:
            from src.data_analyzer_sql import get_data_analyzer
            result["tables_dropped"] = get_data_analyzer().drop_tables(table_names)
        except Exception as exc:
            result["table_cleanup_error"] = str(exc)

    source_paths = {Path(entry.source_file) for entry in catalog_entries}
    for entry in catalog_entries:
        catalog.remove_entry(entry.source_file, project_id=project_id)
    result["catalog_entries_deleted"] = len(catalog_entries)

    # clear_file is project-filtered through the request's ProjectContext and
    # removes both Qdrant points and mirrored lexical chunks.
    from src.document_rag import get_document_rag
    get_document_rag().clear_file(file_name, project_id=project_id)
    result["rag_cleaned"] = True

    try:
        from src.event_timeline import get_event_timeline
        result["events_deleted"] = sum(
            get_event_timeline().delete_by_document(doc_id, project_id=project_id)
            for doc_id in doc_ids
        )
    except Exception as exc:
        result["event_cleanup_error"] = str(exc)

    from src.config import DATA_DIR, STORAGE_DIR
    data_root = Path(DATA_DIR).resolve()
    storage_root = Path(STORAGE_DIR).resolve()
    project_roots = (
        (data_root / "projects" / project_id).resolve(),
        (storage_root / "projects" / project_id).resolve(),
    )
    source_paths.update({
        data_root / "edinburgh_pdfs" / file_name,
        data_root / "projects" / project_id / "documents" / file_name,
        data_root / "projects" / project_id / "emails" / file_name,
        data_root / "projects" / project_id / "tables" / file_name,
    })
    deleted_paths = []
    for candidate in source_paths:
        try:
            resolved = candidate.resolve()
            project_local = any(resolved.is_relative_to(root) for root in project_roots)
            inside_managed_data = resolved.is_relative_to(data_root)
            if not resolved.is_file() or not (project_local or inside_managed_data):
                continue
            if not project_local and shared_source:
                continue
            resolved.unlink()
            deleted_paths.append(str(resolved))
        except (OSError, ValueError):
            continue
    result["source_files_deleted"] = len(deleted_paths)
    result["shared_source_retained"] = shared_source
    return result


def _edinburgh_data_table_infos(project_id: str = "") -> List[FileInfo]:
    """FileInfo rows for the edinburgh corpus's SQL data tables (Excel/CSV), read
    from the catalog by corpus tag. These live as DuckDB/parquet tables (not in the
    chunk store), so without this they never reach the 'edinburgh' Spreadsheets list.
    doc_id == generate_doc_id(source_file) so the viewer resolves them via the
    registry → direct file render. Read-only."""
    from src.catalog import get_catalog
    from src.document_rag import generate_doc_id
    out: List[FileInfo] = []
    for entry in get_catalog().entries.values():
        if (getattr(entry, "corpus", "demo") or "demo").lower() != "edinburgh":
            continue
        if project_id and (getattr(entry, "project_id", "") or "") != project_id:
            continue
        if entry.source_type not in ("excel", "csv"):
            continue
        fname = os.path.basename(entry.source_file)
        cols = list(entry.tables[0].columns or [])[:8] if entry.tables else []
        out.append(FileInfo(
            id=generate_doc_id(entry.source_file),
            name=fname,
            file_type="data",
            tables=len(entry.tables),
            rows=sum(t.row_count for t in entry.tables),
            status="completed",
            data_table_status="registered",
            data_tables_count=len(entry.tables),
            columns=cols,
            sheets=len(entry.tables),
        ))
    return out


def _edinburgh_file_infos(project_id: str = "") -> List[FileInfo]:
    """FileInfo rows for the 'edinburgh' account: bulk vectors-only PDFs/emails
    (chunk-store docs not in the registry) PLUS its SQL data tables (Excel/CSV)
    from the catalog, so Documents AND Spreadsheets both reflect its own corpus.
    Read-only."""
    from src.document_registry import get_document_registry
    from src.chunk_store import get_chunk_store
    registry = get_document_registry()
    reg_names = {
        r.file_name for r in (
            registry.get_completed(project_id=project_id)
            if project_id else registry.get_completed()
        )
    }
    con = get_chunk_store().connection()
    sql = ("SELECT file_name, MAX(page_number) FROM chunks "
           "WHERE file_name IS NOT NULL AND file_name <> ''")
    params = []
    if project_id:
        sql += " AND project_id=?"
        params.append(project_id)
    rows = con.execute(sql + " GROUP BY file_name", params).fetchall()
    out: List[FileInfo] = []
    for file_name, max_page in rows:
        if file_name in reg_names:
            continue
        ext = os.path.splitext(file_name)[1].lower()
        ftype = ("email" if ext in (".eml", ".msg")
                 else "data" if ext in (".xlsx", ".xls", ".csv") else "document")
        try:
            pages = int(max_page) if max_page else None
        except (TypeError, ValueError):
            pages = None
        out.append(FileInfo(id=file_name, name=file_name, file_type=ftype,
                            pages=pages, status="completed"))
    # Merge the edinburgh SQL data tables (Excel/CSV) — these are not in the chunk store.
    try:
        out.extend(_edinburgh_data_table_infos(project_id))
    except Exception:
        pass
    return out


@router.post("/upload", response_model=UploadResult)
async def upload_file(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    user: UserContext = Depends(get_current_user),
    project: ProjectContext = Depends(require_project_editor),
):
    saved_path, file_id, is_duplicate = await _file_service.save(
        file, project.project_id, username=user.username,
    )
    if is_duplicate:
        return UploadResult(
            file_id=file_id,
            filename=file.filename,
            status="completed",
        )
    from backend.tasks.ingestion_jobs import get_ingestion_job_store
    from src.project_store import get_project_store
    get_project_store().set_vector_state(project.project_id, "provisioning")
    job = get_ingestion_job_store().enqueue(
        project.project_id, file_id, saved_path, file.filename or "upload",
        requested_by=user.username,
    )
    return UploadResult(
        file_id=file_id,
        filename=file.filename,
        status=job["status"],
    )


@router.get("/files", response_model=List[FileInfo])
async def list_files(
    user: UserContext = Depends(get_current_user),
    project: ProjectContext = Depends(get_current_project),
):
    raw = _file_service.list_files(project_id=project.project_id)
    registered = [
        FileInfo(
            id=f.get("id", ""),
            name=f.get("name", ""),
            file_type=f.get("file_type", ""),
            pages=f.get("pages"),
            ocr_pages=f.get("ocr_pages", 0),
            tables=f.get("tables", 0),
            rows=f.get("rows", 0),
            status=f.get("status", "completed"),
            data_table_status=f.get("data_table_status"),
            data_tables_count=f.get("data_tables_count", 0),
            columns=f.get("columns", []),
            sheets=f.get("sheets", 0),
        )
        for f in raw
    ]
    if str((user.features or {}).get("corpus") or "").lower() == "edinburgh":
        # A migrated bulk corpus and newly uploaded project files coexist. The
        # old conditional returned only registry rows as soon as the first new
        # upload existed, making all 7k legacy sources appear to vanish. Merge
        # the inventories and prefer the richer registry row on name clashes.
        merged = {(item.file_type, item.name): item
                  for item in _edinburgh_file_infos(project.project_id)}
        merged.update({(item.file_type, item.name): item for item in registered})
        return list(merged.values())
    return registered


@router.delete("/files/{file_id}")
async def delete_file(
    file_id: str,
    project: ProjectContext = Depends(require_project_editor),
):
    """Delete one selected-project file and all of its derived evidence."""
    from src.document_registry import get_document_registry
    rec = get_document_registry().get(file_id)
    if rec:
        if (getattr(rec, "project_id", "") or "") != project.project_id:
            raise HTTPException(404, "file_not_found")
        from src.file_router import delete_document
        result = delete_document(file_id)
        if "error" in result:
            raise HTTPException(500, result["error"])
        return {"ok": True, "cleanup": result}

    legacy_result = _delete_legacy_project_file(file_id, project.project_id)
    if legacy_result is None:
        raise HTTPException(404, "file_not_found")
    from src.billing_store import get_billing_store
    get_billing_store().release_storage(project_id=project.project_id, file_id=file_id)
    return {"ok": True, "cleanup": legacy_result}


@router.post("/files/{file_id}/reindex")
async def reindex_file(
    file_id: str,
    background_tasks: BackgroundTasks,
    project: ProjectContext = Depends(require_project_editor),
):
    """Delete and re-index a single file (clean slate)."""
    from pathlib import Path
    from src.document_registry import get_document_registry
    from src.file_router import delete_document
    from src.document_rag import generate_doc_id

    registry = get_document_registry()
    rec = registry.get(file_id)
    if not rec or getattr(rec, "project_id", "") != project.project_id:
        return {"ok": False, "detail": "Document not found"}

    file_path = rec.file_path
    file_name = rec.file_name

    # 1. Full cleanup from all stores
    delete_document(file_id)

    # 2. Check file still on disk (downloaded from GCS at startup)
    if not Path(file_path).exists():
        return {"ok": False, "detail": f"File not on disk: {file_path}"}

    # 3. Re-index in background
    new_doc_id = generate_doc_id(file_path)
    from backend.tasks.ingestion_jobs import get_ingestion_job_store
    get_ingestion_job_store().enqueue(
        project.project_id, new_doc_id, file_path, file_name,
    )

    return {"ok": True, "file_name": file_name, "new_file_id": new_doc_id, "status": "reindexing"}


@router.post("/files/reindex-stuck")
async def reindex_stuck_files(
    background_tasks: BackgroundTasks,
    project: ProjectContext = Depends(require_project_editor),
):
    """Find all processing/error files, clean them up, and re-index."""
    from pathlib import Path
    from src.document_registry import get_document_registry
    from src.file_router import delete_document
    from src.document_rag import generate_doc_id

    registry = get_document_registry()
    stuck = [r for r in registry.get_all(project_id=project.project_id)
             if r.status in ("processing", "error")]

    results = []
    for rec in stuck:
        file_path = rec.file_path
        # Full cleanup
        delete_document(rec.doc_id)

        # Check file on disk
        if not Path(file_path).exists():
            results.append({"name": rec.file_name, "status": "skipped", "reason": "file not on disk"})
            continue

        # Re-index
        new_id = generate_doc_id(file_path)
        from backend.tasks.ingestion_jobs import get_ingestion_job_store
        get_ingestion_job_store().enqueue(
            project.project_id, new_id, file_path, rec.file_name,
        )
        results.append({"name": rec.file_name, "status": "reindexing", "new_id": new_id})

    return {"total": len(stuck), "results": results}


@router.get("/stats")
async def get_stats(project: ProjectContext = Depends(get_current_project)):
    """Return vector count and table count for dashboard metrics."""
    vectors = 0
    tables = 0
    try:
        from src.document_rag import get_document_rag
        rag = get_document_rag()
        vectors = rag.count_project_points(project.project_id)
    except Exception:
        pass
    try:
        from src.data_analyzer_sql import get_data_analyzer
        analyzer = get_data_analyzer()
        tables = len(analyzer.list_tables())
    except Exception:
        pass
    return {"vectors": vectors, "tables": tables}


@router.get("/files/export")
async def export_files_excel(project: ProjectContext = Depends(get_current_project)):
    """Export file list as multi-sheet Excel (.xlsx) grouped by file type."""
    import io
    from datetime import datetime
    try:
        import pandas as pd
        from openpyxl.styles import Font, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "pandas/openpyxl not available")

    raw = _file_service.list_files(project_id=project.project_id)

    # Group files by type
    groups = {
        "Documents": [f for f in raw if f.get("file_type") == "document"],
        "Emails": [f for f in raw if f.get("file_type") == "email"],
        "Data Files": [f for f in raw if f.get("file_type") == "data"],
    }

    def _fmt_date(iso_str):
        if not iso_str:
            return ""
        try:
            return datetime.fromisoformat(iso_str).strftime("%Y-%m-%d")
        except Exception:
            return iso_str[:10] if len(iso_str) >= 10 else iso_str

    # Column definitions per sheet type
    COLS = {
        "Documents": ["File Name", "Upload Date", "Document Date", "Pages", "Tables", "Rows"],
        "Emails": ["File Name", "Upload Date", "Document Date", "Sender", "Receiver", "Pages", "Tables", "Rows"],
        "Data Files": ["File Name", "Upload Date", "Document Date", "Sheets", "Tables", "Rows"],
    }

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, files in groups.items():
            rows = []
            for f in files:
                row = {
                    "File Name": f.get("name", ""),
                    "Upload Date": _fmt_date(f.get("created_at", "")),
                    "Document Date": f.get("document_date", ""),
                }
                if sheet_name == "Emails":
                    row["Sender"] = f.get("sender", "")
                    row["Receiver"] = f.get("recipient", "")
                    row["Pages"] = f.get("pages") or ""
                elif sheet_name == "Data Files":
                    row["Sheets"] = f.get("tables", 0)
                else:
                    row["Pages"] = f.get("pages") or ""
                row["Tables"] = f.get("tables", 0)
                row["Rows"] = f.get("rows", 0)
                rows.append(row)

            cols = COLS[sheet_name]
            df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)

            # Write data starting at row 3 (leave 2 rows for header)
            header_rows = 2
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=header_rows)

            # Add notice/title header using openpyxl
            ws = writer.sheets[sheet_name]
            title_cell = ws.cell(row=1, column=1, value=f"COAir - {sheet_name}")
            title_cell.font = Font(bold=True, size=13)
            date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            date_cell.font = Font(italic=True, size=10, color="666666")

            # Auto-adjust column widths
            for col_idx, col_name in enumerate(cols, 1):
                max_len = len(col_name)
                for row_data in rows:
                    val = str(row_data.get(col_name, ""))
                    max_len = max(max_len, len(val))
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Add Summary sheet with document type breakdown
        summary_rows = []
        for sheet_name, files in groups.items():
            summary_rows.append({"Category": sheet_name, "Count": len(files)})
        # Add doc_type breakdown from notice metadata
        try:
            from src.document_registry import get_document_registry
            from src.notice_extractor import get_notice_extractor
            registry = get_document_registry()
            extractor = get_notice_extractor()
            doc_type_counts: dict[str, int] = {}
            for rec in registry.get_completed():
                if rec.notice_extracted:
                    notice = extractor.load_notice(rec.doc_id)
                    dt = (notice.doc_type if notice and notice.doc_type else "unclassified").title()
                else:
                    dt = rec.file_type.title() if rec.file_type else "Other"
                doc_type_counts[dt] = doc_type_counts.get(dt, 0) + 1
            summary_rows.append({"Category": "", "Count": ""})
            summary_rows.append({"Category": "Document Types", "Count": ""})
            for dt, count in sorted(doc_type_counts.items(), key=lambda x: x[1], reverse=True):
                summary_rows.append({"Category": f"  {dt}", "Count": count})
        except Exception:
            pass

        summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(columns=["Category", "Count"])
        summary_df.to_excel(writer, index=False, sheet_name="Summary", startrow=2)
        ws_summary = writer.sheets["Summary"]
        ws_summary.cell(row=1, column=1, value="Project Document Summary").font = Font(bold=True, size=13)
        ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10, color="666666")
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 12

    buf.seek(0)

    filename = "COAir_Documents.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
