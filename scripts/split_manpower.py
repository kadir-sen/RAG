"""
Split Sample Manpower.xlsx into monthly sheets + yearly totals sheet.

Input:  Single sheet with daily manpower data (SEP-2015 to JAN-2023)
Output: One sheet per month (e.g. "SEP-2015") + "Yearly Totals" summary sheet
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────
INPUT_FILE = Path(__file__).parent.parent / "data" / "tables" / "Sample Manpower.xlsx"
OUTPUT_FILE = Path(__file__).parent.parent / "data" / "tables" / "Sample Manpower - Monthly.xlsx"

MONTH_NAMES = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC",
}

# Worker category columns (col index 2-17 in raw data)
WORKER_COLUMNS = [
    "Group Leader HVAC",
    "Group Leader PL&FF",
    "Group Leader Electrical",
    "Low Voltage Workers",
    "Plumber",
    "Plumbers (SP Manpower)",
    "Electrician",
    "Electrician (SP Manpower)",
    "Welder (Plumbing)",
    "Driver",
    "HVAC",
    "Pipe Fitter (Plumbing)",
    "Helper (Plumbing)",
    "Helper (Plumbing) (SP Manpower)",
    "BMS Subcontractor",
    "Store",
]


def load_daily_data(file_path: Path) -> pd.DataFrame:
    """Load and parse only the daily rows from the Excel file."""
    raw = pd.read_excel(file_path, sheet_name=0, header=None)

    # Header at row 5 (0-indexed), data starts at row 6
    headers = ["Date", "Day"] + WORKER_COLUMNS + ["Total Direct"]
    data_rows = raw.iloc[6:].copy()
    data_rows.columns = range(len(data_rows.columns))

    # Build clean DataFrame with only the columns we need (0-18)
    df = data_rows.iloc[:, :19].copy()
    df.columns = headers

    # Parse dates
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # Daily rows have a Day value (MON, TUE, etc.), monthly/yearly summaries don't
    daily = df[df["Day"].notna() & df["Date"].notna()].copy()

    # Convert numeric columns
    for col in WORKER_COLUMNS + ["Total Direct"]:
        daily[col] = pd.to_numeric(daily[col], errors="coerce").fillna(0).astype(int)

    daily = daily.sort_values("Date").reset_index(drop=True)

    print(f"Loaded {len(daily)} daily rows")
    print(f"Date range: {daily['Date'].min().date()} to {daily['Date'].max().date()}")

    return daily


def split_by_month(daily: pd.DataFrame) -> dict:
    """Split daily data into {sheet_name: DataFrame} by month."""
    daily["Year"] = daily["Date"].dt.year
    daily["Month"] = daily["Date"].dt.month

    monthly_sheets = {}
    for (year, month), group in daily.groupby(["Year", "Month"]):
        sheet_name = f"{MONTH_NAMES[month]}-{year}"
        sheet_df = group.drop(columns=["Year", "Month"]).copy()
        monthly_sheets[sheet_name] = sheet_df

    print(f"Split into {len(monthly_sheets)} monthly sheets")
    return monthly_sheets


def build_yearly_totals(daily: pd.DataFrame) -> pd.DataFrame:
    """Build yearly totals summary from daily data."""
    daily["Year"] = daily["Date"].dt.year

    yearly = daily.groupby("Year")[WORKER_COLUMNS + ["Total Direct"]].sum()
    yearly = yearly.reset_index()

    # Add a grand total row
    grand_total = yearly[WORKER_COLUMNS + ["Total Direct"]].sum()
    grand_total["Year"] = "GRAND TOTAL"
    yearly = pd.concat([yearly, pd.DataFrame([grand_total])], ignore_index=True)

    print(f"Yearly totals: {len(yearly) - 1} years + grand total")
    return yearly


def style_worksheet(ws, is_yearly=False):
    """Apply formatting to a worksheet."""
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    number_align = Alignment(horizontal="right")
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Style header row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        is_last_row = row_idx == ws.max_row and is_yearly
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx >= 3 or (is_yearly and col_idx >= 2):
                cell.alignment = number_align
            if is_last_row:
                cell.font = total_font
                cell.fill = total_fill

    # Monthly sheet: style the monthly total row (last row)
    if not is_yearly:
        last_row = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 25)

    # Freeze header row
    ws.freeze_panes = "A2"


def write_output(monthly_sheets: dict, yearly_totals: pd.DataFrame, output_path: Path):
    """Write all sheets to the output Excel file."""
    print(f"\nWriting to: {output_path.name}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write yearly totals first
        yearly_totals.to_excel(writer, sheet_name="Yearly Totals", index=False)

        # Write monthly sheets in chronological order
        for sheet_name in sorted(monthly_sheets.keys(),
                                  key=lambda s: pd.Timestamp(f"1-{s}")):
            sheet_df = monthly_sheets[sheet_name]

            # Format date column
            export_df = sheet_df.copy()
            export_df["Date"] = export_df["Date"].dt.strftime("%Y-%m-%d")

            # Add monthly total row
            total_row = {"Date": "TOTAL", "Day": ""}
            for col in WORKER_COLUMNS + ["Total Direct"]:
                total_row[col] = export_df[col].sum()
            export_df = pd.concat([export_df, pd.DataFrame([total_row])], ignore_index=True)

            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  {sheet_name}: {len(sheet_df)} days, total direct = {int(total_row['Total Direct'])}")

    # Apply styling
    wb = load_workbook(output_path)

    # Style yearly totals
    style_worksheet(wb["Yearly Totals"], is_yearly=True)

    # Style monthly sheets
    for sheet_name in wb.sheetnames:
        if sheet_name != "Yearly Totals":
            style_worksheet(wb[sheet_name], is_yearly=False)

    wb.save(output_path)
    print(f"\nDone! {len(wb.sheetnames)} sheets written to {output_path.name}")


def main():
    if not INPUT_FILE.exists():
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        sys.exit(1)

    print(f"Input: {INPUT_FILE.name}")
    print("=" * 60)

    daily = load_daily_data(INPUT_FILE)
    monthly_sheets = split_by_month(daily)
    yearly_totals = build_yearly_totals(daily)

    print("\n--- Yearly Totals ---")
    print(yearly_totals[["Year", "Total Direct"]].to_string(index=False))

    write_output(monthly_sheets, yearly_totals, OUTPUT_FILE)


if __name__ == "__main__":
    main()
