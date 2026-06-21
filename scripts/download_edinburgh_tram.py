"""Bulk-download the Edinburgh Tram Inquiry evidence archive and build a manifest.

The inquiry (https://www.edinburghtraminquiry.org) publishes ~7,600 open-source
documents (witness statements, hearing transcripts, supporting documents — mails,
letters, reports). Each file is named only by its reference number
(e.g. WED00000659.pdf), so we also keep the official "Published Evidence Documents
Record" spreadsheet which says what each reference actually is.

How it works
------------
The documents listing is a FacetWP grid: page 1 is server-rendered, but pages
2..766 are only served by FacetWP's JSON refresh API. That API (and the PDFs
themselves on some paths) sits behind a WAF that filters non-browser TLS
fingerprints, so we use `curl_cffi` (Chrome impersonation) as the HTTP client —
plain urllib/requests get a 403 HTML challenge page.

Pipeline
--------
  1. Harvest every reference -> exact PDF url by paging the FacetWP API       (authoritative urls + folders)
  2. Download + parse the Published Evidence Documents Record (.xlsx)         (authoritative descriptions + categories)
  3. Download every PDF (resumable, throttled, retried) into pdfs/
  4. Write a joined manifest (.csv + .json), a download log, and missing.csv

The harvested url map is cached to url_map.json so re-runs skip the crawl.

Run:
    PYTHONPATH=. python scripts/download_edinburgh_tram.py --manifest-only   # harvest + manifest, no PDFs
    PYTHONPATH=. python scripts/download_edinburgh_tram.py --limit 20         # small end-to-end test
    PYTHONPATH=. python scripts/download_edinburgh_tram.py                    # full run (resumable)
    PYTHONPATH=. python scripts/download_edinburgh_tram.py --refresh-urls     # force re-harvest the url map
"""
from __future__ import annotations

import argparse
import csv
import html as html_lib
import json
import re
import sys
import time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

BASE = "https://www.edinburghtraminquiry.org"
DOCUMENTS_PAGE = f"{BASE}/evidence-and-hearings/public-hearing/documents/"
REFRESH_API = f"{BASE}/wp-json/facetwp/v1/refresh"
DOCUMENTS_URI = "evidence-and-hearings/public-hearing/documents"
MANIFEST_XLSX_URL = f"{BASE}/wp-content/uploads/2023/05/Published-Evidence-Documents-Record.xlsx"
ARXIV_PDF_URL = "https://arxiv.org/pdf/2604.14169"

# Inquiry references look like WED00000659, TRI00000001, CEC02086924, etc.
REF_RE = re.compile(r"\b([A-Z]{2,6}\d{4,})\b")
# Captures the YYYY/MM/<stem> portion of an uploads PDF url.
UPLOADS_PDF_RE = re.compile(r"/wp-content/uploads/(\d{4}/\d{2}/[^\"'?#<> ]+?\.pdf)", re.IGNORECASE)
# Per-row markup from the FacetWP template.
WRAPPER_SPLIT = re.compile(r'class="docWrapper"')
TITLE_RE = re.compile(r'docTitle"><span>Title:\s*</span>(.*?)</p>', re.IGNORECASE | re.DOTALL)
DATE_RE = re.compile(r'Hearing date:\s*</span>(.*?)</p>', re.IGNORECASE | re.DOTALL)


# --------------------------------------------------------------------------- #
# HTTP client — curl_cffi (Chrome impersonation) is mandatory to clear the WAF.
# --------------------------------------------------------------------------- #
def make_session():
    try:
        from curl_cffi import requests as cffi_requests
    except ImportError:
        print("ERROR: curl_cffi is required (the site WAF blocks non-browser TLS).\n"
              "  Install it:  uv pip install --python .venv/bin/python curl_cffi",
              file=sys.stderr)
        raise SystemExit(3)
    session = cffi_requests.Session(impersonate="chrome")
    session.headers.update({"Referer": DOCUMENTS_PAGE, "Origin": BASE})
    # Prime cookies / pass any first-visit checks.
    session.get(DOCUMENTS_PAGE, timeout=30)
    return session


def get_bytes(session, url: str, retries: int = 3, backoff: float = 1.5,
              timeout: int = 60) -> tuple[int, bytes]:
    status = 0
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=timeout)
            status = r.status_code
            if status == 200:
                return status, r.content
            if status in (404, 410):
                return status, b""
        except Exception as e:  # noqa: BLE001 — transient network/TLS errors
            status = 0
            if attempt == retries - 1:
                print(f"    network error on {url}: {type(e).__name__}: {e}", file=sys.stderr)
        if attempt < retries - 1:
            time.sleep(backoff ** (attempt + 1))
    return status, b""


# --------------------------------------------------------------------------- #
# 1. Harvest reference -> exact PDF url(s) via the FacetWP refresh API.
# --------------------------------------------------------------------------- #
def _refresh_payload(paged: int) -> dict[str, Any]:
    return {
        "action": "facetwp_refresh",
        "data": {
            "facets": {"main_search": [], "document_published_date": [],
                       "witness_first_name": [], "witness_last_name": [], "document_type": []},
            "frozen_facets": {},
            "http_params": {"get": [], "uri": DOCUMENTS_URI, "url_vars": []},
            "template": "search_filters",
            "extras": {"sort": "default"},
            "soft_refresh": 0, "is_bfcache": 0, "first_load": 0, "paged": paged,
        },
    }


def _parse_rows(template_html: str) -> list[dict[str, str]]:
    """Each docWrapper -> {reference, url, title, date}. Multi-part rows emit one record per PDF."""
    rows: list[dict[str, str]] = []
    chunks = WRAPPER_SPLIT.split(template_html)
    for chunk in chunks[1:]:  # chunks[0] is preamble before the first wrapper
        title_m = TITLE_RE.search(chunk)
        date_m = DATE_RE.search(chunk)
        title = html_lib.unescape(title_m.group(1).strip()) if title_m else ""
        date = html_lib.unescape(date_m.group(1).strip()) if date_m else ""
        for m in UPLOADS_PDF_RE.finditer(chunk):
            rel = m.group(1)  # YYYY/MM/stem.pdf
            stem = rel.rsplit("/", 1)[-1][:-4]  # drop .pdf
            rows.append({"reference": stem.upper(), "url": f"{BASE}/wp-content/uploads/{rel}",
                         "title": title, "date": date})
    return rows


def harvest_urls(session, max_pages: int, throttle: float) -> dict[str, dict[str, str]]:
    """Page the FacetWP API. Returns {reference: {url, title, date}}."""
    out: dict[str, dict[str, str]] = {}
    total_pages = max_pages
    page = 1
    while page <= total_pages:
        try:
            r = session.post(REFRESH_API, json=_refresh_payload(page), timeout=60)
            data = r.json()
        except Exception as e:  # noqa: BLE001
            print(f"  page {page}: request failed ({type(e).__name__}: {e}) — retrying once", file=sys.stderr)
            time.sleep(2.0)
            try:
                r = session.post(REFRESH_API, json=_refresh_payload(page), timeout=60)
                data = r.json()
            except Exception as e2:  # noqa: BLE001
                print(f"  page {page}: failed again ({e2}); stopping harvest", file=sys.stderr)
                break

        if isinstance(data, dict):
            pager = data.get("settings", {}).get("pager", {}) if isinstance(data.get("settings"), dict) else {}
            if pager.get("total_pages"):
                total_pages = min(max_pages, int(pager["total_pages"]))
            new = 0
            for row in _parse_rows(data.get("template", "")):
                if row["reference"] not in out:
                    out[row["reference"]] = {k: row[k] for k in ("url", "title", "date")}
                    new += 1
            if page == 1 or page % 25 == 0 or page == total_pages:
                print(f"  page {page}/{total_pages}: +{new} (total {len(out)})")
        else:
            print(f"  page {page}: unexpected response {str(data)[:80]}", file=sys.stderr)
        page += 1
        time.sleep(throttle)
    return out


def load_or_harvest(session, out_dir: Path, max_pages: int, throttle: float,
                    refresh: bool) -> dict[str, dict[str, str]]:
    cache = out_dir / "url_map.json"
    if cache.exists() and not refresh:
        url_map = json.loads(cache.read_text(encoding="utf-8"))
        print(f"  loaded cached url map: {len(url_map)} references ({cache.name})")
        return url_map
    url_map = harvest_urls(session, max_pages, throttle)
    cache.write_text(json.dumps(url_map, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  ✓ harvested {len(url_map)} references -> {cache.name}")
    return url_map


# --------------------------------------------------------------------------- #
# 2. Published Evidence Documents Record (.xlsx) -> reference -> {category, description, parts}
# --------------------------------------------------------------------------- #
def download_record(session, out_dir: Path) -> Path | None:
    dest = out_dir / "manifest_source.xlsx"
    if dest.exists() and dest.stat().st_size > 0:
        print(f"  record already present: {dest.name}")
        return dest
    status, body = get_bytes(session, MANIFEST_XLSX_URL)
    if status != 200 or not body:
        print(f"  ✗ record download failed (status={status})", file=sys.stderr)
        return None
    dest.write_bytes(body)
    print(f"  ✓ saved {dest.name} ({len(body):,} bytes)")
    return dest


def parse_record(xlsx_path: Path) -> dict[str, dict[str, str]]:
    """Columns: Document Reference | Category | Parts | File Format | Description."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    records: dict[str, dict[str, str]] = {}
    first = True
    for raw in ws.iter_rows(values_only=True):
        values = ["" if c is None else str(c).strip() for c in raw]
        if first:  # header row
            first = False
            continue
        if not any(values):
            continue
        ref = values[0].upper() if values else ""
        if not REF_RE.fullmatch(ref):
            continue
        records[ref] = {
            "category": values[1] if len(values) > 1 else "",
            "parts": values[2] if len(values) > 2 else "",
            "format": values[3] if len(values) > 3 else "",
            "description": values[4] if len(values) > 4 else "",
        }
    wb.close()
    return records


def lookup_record(records: dict[str, dict[str, str]], reference: str) -> dict[str, str]:
    """Exact match, else strip a trailing part suffix (e.g. WED00001_2 -> WED00001)."""
    if reference in records:
        return records[reference]
    base = re.sub(r"[_\-]\d+$", "", reference)
    return records.get(base, {})


# --------------------------------------------------------------------------- #
# 3 + 4. Download + write manifest / logs.
# --------------------------------------------------------------------------- #
def download_pdf(session, url: str, dest: Path, throttle: float) -> tuple[str, int, int]:
    if dest.exists() and dest.stat().st_size > 0:
        return "skipped", 200, dest.stat().st_size
    status, body = get_bytes(session, url)
    if status == 200 and body:
        dest.write_bytes(body)
        time.sleep(throttle)
        return "downloaded", status, len(body)
    time.sleep(throttle)
    return "failed", status, 0


def write_manifest(out_dir: Path, rows: list[dict[str, Any]]) -> None:
    cols = ["reference", "title", "category", "description", "date", "format",
            "parts", "url", "local_path", "status"]
    with (out_dir / "manifest.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    (out_dir / "manifest.json").write_text(
        json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  ✓ manifest.csv + manifest.json ({len(rows)} rows)")


def write_logs(out_dir: Path, log_rows: list[dict[str, Any]], missing: list[dict[str, Any]]) -> None:
    with (out_dir / "download_log.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["reference", "url", "http_status", "bytes", "status"])
        w.writeheader()
        w.writerows(log_rows)
    with (out_dir / "missing.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["reference", "description", "category", "url", "http_status"])
        w.writeheader()
        w.writerows(missing)


def save_reference_paper(session, out_dir: Path) -> None:
    ref_dir = out_dir / "_reference"
    ref_dir.mkdir(parents=True, exist_ok=True)
    dest = ref_dir / "chronological-rag-2604.14169.pdf"
    if dest.exists() and dest.stat().st_size > 0:
        print(f"  reference paper already present: {dest.name}")
        return
    status, body = get_bytes(session, ARXIV_PDF_URL)
    if status == 200 and body:
        dest.write_bytes(body)
        print(f"  ✓ saved reference paper ({len(body):,} bytes)")
    else:
        print(f"  ✗ reference paper download failed (status={status})", file=sys.stderr)


# --------------------------------------------------------------------------- #
def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=str(ROOT / "data" / "edinburgh_tram"))
    parser.add_argument("--limit", type=int, default=0, help="Only process the first N references")
    parser.add_argument("--types", default="", help="Comma-separated reference prefixes to keep (e.g. WED,TRI)")
    parser.add_argument("--manifest-only", action="store_true", help="Harvest + write manifest, download no PDFs")
    parser.add_argument("--refresh-urls", action="store_true", help="Re-harvest the url map (ignore cache)")
    parser.add_argument("--throttle", type=float, default=0.4, help="Seconds between requests (default 0.4)")
    parser.add_argument("--max-pages", type=int, default=800, help="Max FacetWP pages to crawl (default 800)")
    parser.add_argument("--skip-reference-paper", action="store_true")
    args = parser.parse_args()

    out_dir = Path(args.out)
    pdfs_dir = out_dir / "pdfs"
    out_dir.mkdir(parents=True, exist_ok=True)
    pdfs_dir.mkdir(parents=True, exist_ok=True)
    prefixes = [p.strip().upper() for p in args.types.split(",") if p.strip()]

    session = make_session()

    print("1. Harvesting reference -> PDF url via FacetWP API")
    url_map = load_or_harvest(session, out_dir, args.max_pages, args.throttle, args.refresh_urls)

    print("\n2. Published Evidence Documents Record (.xlsx)")
    xlsx_path = download_record(session, out_dir)
    records = parse_record(xlsx_path) if xlsx_path else {}
    print(f"  parsed {len(records)} references from record")

    refs = sorted(url_map)
    if prefixes:
        refs = [r for r in refs if any(r.startswith(p) for p in prefixes)]
    if args.limit:
        refs = refs[: args.limit]
    print(f"\n3. {len(refs)} references to process"
          + (f" (filtered to {prefixes})" if prefixes else "")
          + (f", limited to {args.limit}" if args.limit else "") + "\n")

    manifest_rows: list[dict[str, Any]] = []
    log_rows: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    counts: dict[str, int] = {}

    for i, ref in enumerate(refs, 1):
        harvested = url_map[ref]
        rec = lookup_record(records, ref)
        url = harvested["url"]
        dest = pdfs_dir / f"{ref}.pdf"
        row = {
            "reference": ref,
            "title": harvested.get("title", ""),
            "category": rec.get("category", ""),
            "description": rec.get("description", ""),
            "date": harvested.get("date", ""),
            "format": rec.get("format", ""),
            "parts": rec.get("parts", ""),
            "url": url,
            "local_path": str(dest.relative_to(out_dir)),
            "status": "",
        }

        if args.manifest_only:
            row["status"] = "manifest-only"
        else:
            label, http_status, nbytes = download_pdf(session, url, dest, args.throttle)
            row["status"] = label
            log_rows.append({"reference": ref, "url": url, "http_status": http_status,
                             "bytes": nbytes, "status": label})
            if label == "failed":
                row["local_path"] = ""
                missing.append({"reference": ref, "description": rec.get("description", ""),
                                "category": rec.get("category", ""), "url": url, "http_status": http_status})
            if i % 50 == 0:
                done = counts.get("downloaded", 0) + counts.get("skipped", 0)
                print(f"  [{i}/{len(refs)}] {counts.get('downloaded',0)} dl, "
                      f"{counts.get('skipped',0)} skip, {counts.get('failed',0)} fail")
        counts[row["status"]] = counts.get(row["status"], 0) + 1
        manifest_rows.append(row)

    print("\n4. Writing manifest + logs")
    write_manifest(out_dir, manifest_rows)
    if not args.manifest_only:
        write_logs(out_dir, log_rows, missing)
        print(f"  ✓ download_log.csv ({len(log_rows)}), missing.csv ({len(missing)})")

    if not args.skip_reference_paper:
        print("\n5. Saving arXiv reference paper")
        save_reference_paper(session, out_dir)

    print("\n" + "=" * 60)
    print(f"Output:      {out_dir}")
    print(f"References:  {len(refs)}")
    for k in ("downloaded", "skipped", "failed", "manifest-only"):
        if counts.get(k):
            print(f"  {k:<14} {counts[k]}")
    print("=" * 60)
    return 0 if counts.get("failed", 0) == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
