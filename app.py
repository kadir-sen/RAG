"""
Agentic RAG Chatbot - Modern Dark UI
Clean, minimal, professional interface with dark theme.
"""
import os
import hashlib
from datetime import datetime
from pathlib import Path

# Load env first
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import streamlit as st
import pandas as pd

# Page config - MUST be first
st.set_page_config(
    page_title="Document Analysis Platform",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Dark theme CSS
st.markdown("""
<style>
    /* Force dark background everywhere */
    .stApp, .main, [data-testid="stAppViewContainer"], [data-testid="stHeader"] {
        background-color: #1a1a2e !important;
    }

    /* Main content area */
    .main .block-container {
        background-color: #1a1a2e !important;
        padding: 1rem 2rem 6rem 2rem;
        max-width: 1400px;
        margin: 0 auto;
    }

    /* Hide default elements (keep header for sidebar toggle) */
    #MainMenu, footer {visibility: hidden;}
    .stDeployButton {display: none;}
    [data-testid="stHeader"] {background-color: #1a1a2e !important;}

    /* Sidebar - darker */
    [data-testid="stSidebar"] {
        background-color: #0f0f1a !important;
    }
    [data-testid="stSidebar"] > div:first-child {
        background-color: #0f0f1a !important;
    }
    [data-testid="stSidebar"] * {
        color: #e0e0e0 !important;
    }
    [data-testid="stSidebar"] .stButton button {
        background-color: #2d2d44 !important;
        border: 1px solid #404060 !important;
        color: #e0e0e0 !important;
        border-radius: 8px;
    }
    [data-testid="stSidebar"] .stButton button:hover {
        background-color: #3d3d5c !important;
        border-color: #10a37f !important;
    }

    /* File uploader in sidebar */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] {
        background-color: #1a1a2e !important;
        border: 1px dashed #404060 !important;
        border-radius: 8px;
        padding: 0.5rem;
    }

    /* All text white */
    p, span, label, .stMarkdown, h1, h2, h3, h4, h5, h6 {
        color: #e0e0e0 !important;
    }

    /* Chat input */
    .stChatInput > div {
        background-color: #2d2d44 !important;
        border: 1px solid #404060 !important;
        border-radius: 12px !important;
    }
    .stChatInput input {
        color: #e0e0e0 !important;
        background-color: transparent !important;
    }
    .stChatInput input::placeholder {
        color: #808080 !important;
    }

    /* Expander */
    .streamlit-expanderHeader {
        background-color: #2d2d44 !important;
        color: #e0e0e0 !important;
        border-radius: 8px;
    }
    .streamlit-expanderContent {
        background-color: #1f1f35 !important;
        border: 1px solid #404060 !important;
    }

    /* Metrics */
    [data-testid="stMetric"] {
        background-color: #2d2d44 !important;
        padding: 0.75rem;
        border-radius: 8px;
    }
    [data-testid="stMetricValue"] {
        color: #10a37f !important;
    }

    /* Dataframe */
    .stDataFrame {
        background-color: #2d2d44 !important;
    }

    /* Code blocks */
    .stCodeBlock {
        background-color: #0d0d15 !important;
    }
    code {
        background-color: #2d2d44 !important;
        color: #10a37f !important;
    }

    /* Buttons */
    .stButton > button {
        background-color: #10a37f !important;
        color: white !important;
        border: none !important;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 500;
    }
    .stButton > button:hover {
        background-color: #0d8a6a !important;
    }

    /* Success/Error messages */
    .stSuccess, .stError, .stWarning, .stInfo {
        background-color: #2d2d44 !important;
        color: #e0e0e0 !important;
    }

    /* Spinner */
    .stSpinner > div {
        border-color: #10a37f !important;
    }

    /* Tabs for dual LLM answers */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #2d2d44 !important;
        border-radius: 8px;
        padding: 0.25rem;
        gap: 0.25rem;
    }
    .stTabs [data-baseweb="tab"] {
        color: #e0e0e0 !important;
        background-color: transparent !important;
        border-radius: 6px;
        padding: 0.5rem 1rem;
    }
    .stTabs [aria-selected="true"] {
        background-color: #404060 !important;
        color: #10a37f !important;
    }
    .stTabs [data-baseweb="tab-panel"] {
        background-color: #1f1f35 !important;
        border: 1px solid #404060 !important;
        border-radius: 0 0 8px 8px;
        padding: 1rem;
    }

    /* Citation panel (right column) */
    [data-testid="stVerticalBlockBorderWrapper"]:has(> div > [data-testid="stVerticalBlock"] > [data-testid="stMarkdown"] > div > h3:first-child) {
        border-left: 1px solid #404060;
        padding-left: 1rem;
    }

    /* Bottom document drawer */
    /* Document drawer cards use Streamlit native columns + buttons */

    /* Dashboard file cards */
    .file-card {
        background: #2d2d44;
        border: 1px solid #404060;
        border-radius: 8px;
        padding: 12px;
        margin-bottom: 8px;
        transition: border-color 0.2s;
    }
    .file-card:hover { border-color: #10a37f; }
    .stat-card {
        background: #2d2d44;
        border-radius: 10px;
        padding: 16px;
        text-align: center;
    }
</style>
""", unsafe_allow_html=True)


def init_session():
    """Initialize session state."""
    defaults = {
        "messages": [],
        "docs_count": 0,
        "tables_count": 0,
        "initialized": False,
        "processed_files": set(),
        "uploaded_files_log": [],  # [{name, type, time, info}]
        "show_pdf_viewer": False,
        "pdf_path": None,
        "pdf_page": 1,
        "pdf_highlight": None,
        # OCR settings (fixed - English only, auto mode)
        "ocr_mode": "auto",
        "ocr_language": "eng",
        # Conversation management
        "active_conversation_id": None,
        # Inline source features
        "citation_verifications": {},  # {msg_idx: [CitationVerification...]}
        "expanded_pdf_sources": set(),  # {"msgIdx_srcIdx"} - inline PDF viewer (legacy)
        "expanded_excel_sources": set(),  # {"msgIdx_srcIdx"} - inline Excel viewer (legacy)
        "selected_pdf_source": None,  # {file_path, page_number, highlight_text, file_name, is_ocr}
        "selected_excel_source": None,  # {table_name, source_file, total_rows, result_columns}
        "notice_cache": {},  # {doc_id: NoticeMetadata or None}
        # Platform view mode
        "view_mode": "chat",  # always chat
        # Contextual document drawer
        "drawer_documents": [],  # Sources for bottom drawer
        "drawer_title": "",  # Drawer header text
        # Active mode: "chat" | "mail" | "doc_analysis"
        "active_mode": "chat",
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

    # Ensure user has an active conversation
    if st.session_state.get("active_conversation_id") is None:
        store = _get_conversation_store()
        conversations = store.list_conversations()

        # Clean up empty "New Chat" entries (keep at most 1)
        empty_chats = [c for c in conversations if c.message_count == 0]
        for ec in empty_chats[1:]:  # keep first empty, delete rest
            store.delete_conversation(ec.conversation_id)
        if len(empty_chats) > 1:
            conversations = store.list_conversations()

        if conversations:
            _switch_conversation(conversations[0].conversation_id)
        else:
            meta = store.create_conversation()
            st.session_state["active_conversation_id"] = meta.conversation_id


def check_api_keys():
    """Validate API keys at startup."""
    from src.config import validate_config

    is_valid, errors = validate_config()
    if not is_valid:
        st.error("⚠️ Configuration Error")
        for e in errors:
            st.warning(e)
        st.code("# Create .env file:\nGOOGLE_API_KEY=your_key\nPINECONE_API_KEY=your_key\nOPENAI_API_KEY=your_key\nANTHROPIC_API_KEY=your_key")
        st.stop()


def save_uploaded_file(uploaded_file, target_dir: Path) -> str:
    """Save uploaded file and return path."""
    target_dir.mkdir(parents=True, exist_ok=True)
    file_path = target_dir / uploaded_file.name
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return str(file_path)


def process_unified_upload(uploaded_file):
    """Process any file type through unified file router."""
    from src.file_router import route_file, ProcessingResult
    from src.config import DOCUMENTS_DIR, TABLES_DIR, EMAILS_DIR

    ext = Path(uploaded_file.name).suffix.lower()

    # Dedup check — session level
    key = f"{uploaded_file.name}_{uploaded_file.size}"
    if key in st.session_state.processed_files:
        return ProcessingResult(success=False, file_path="", file_type="duplicate")

    # Dedup check — persistent (fast registry lookup, no full scan)
    from src.document_registry import get_document_registry
    registry = get_document_registry()
    if registry.find_duplicate(uploaded_file.name, uploaded_file.size // 1024):
        return ProcessingResult(success=False, file_path="", file_type="duplicate")

    # Save to appropriate directory
    if ext in (".pdf", ".docx", ".doc", ".txt"):
        file_path = save_uploaded_file(uploaded_file, DOCUMENTS_DIR)
    elif ext in (".eml", ".msg"):
        file_path = save_uploaded_file(uploaded_file, EMAILS_DIR)
    else:
        file_path = save_uploaded_file(uploaded_file, TABLES_DIR)

    # Sync source file to GCS for persistence across Cloud Run redeploys
    try:
        from src.gcs_storage import sync_uploaded_file_to_gcs
        sync_uploaded_file_to_gcs(file_path)
    except Exception:
        pass

    result = route_file(file_path)

    if result.success:
        st.session_state.processed_files.add(key)

        # Log uploaded file for sidebar display
        file_info = {
            "name": uploaded_file.name,
            "type": result.file_type,
            "time": datetime.now().strftime("%H:%M"),
        }
        if result.tables_extracted:
            file_info["tables"] = result.tables_extracted
        if result.total_rows:
            file_info["rows"] = result.total_rows
        if result.ocr_pages:
            file_info["ocr"] = result.ocr_pages
        if result.notice_extracted:
            file_info["notice"] = True
        if result.attachments_processed:
            file_info["attachments"] = result.attachments_processed
        st.session_state.uploaded_files_log.append(file_info)

        # Update counts (local — no Pinecone API call)
        if result.file_type in ("document", "email"):
            st.session_state.docs_count += result.ocr_pages or 1
            if result.file_type == "email" and result.tables_extracted:
                st.session_state.tables_count += result.tables_extracted
        elif result.file_type == "data":
            st.session_state.tables_count += result.tables_extracted or 1

    return result


def _render_upload_result(filename: str, result):
    """Render upload result feedback in sidebar."""
    info_parts = []

    if result.file_type == "document":
        if result.ocr_pages > 0:
            info_parts.append(f"{result.ocr_pages} OCR")
        if result.tables_extracted > 0:
            info_parts.append(f"{result.tables_extracted} tables")
        if result.notice_extracted:
            info_parts.append("notice")

    elif result.file_type == "email":
        if result.attachments_processed > 0:
            info_parts.append(f"{result.attachments_processed} attachments")
        if result.notice_extracted:
            info_parts.append("notice")

    elif result.file_type == "data":
        if result.tables_extracted > 0:
            info_parts.append(f"{result.tables_extracted} tables")
        if result.total_rows > 0:
            info_parts.append(f"{result.total_rows} rows")
        if result.converter_generated:
            info_parts.append("converter generated")
        elif result.converter_used:
            info_parts.append("converter used")
        if result.target_schema:
            info_parts.append(f"schema: {result.target_schema}")

    info_str = f" ({', '.join(info_parts)})" if info_parts else ""
    st.success(f"✓ {filename}{info_str}")

    # Show notice summary
    if result.notice_summary:
        ns = result.notice_summary
        with st.expander(f"Notice: {filename}", expanded=False):
            if ns.get('date'):
                st.write(f"**Date:** {ns['date']}")
            if ns.get('sender'):
                st.write(f"**From:** {ns['sender'][:50]}")
            if ns.get('recipient'):
                st.write(f"**To:** {ns['recipient'][:50]}")
            if ns.get('subject'):
                st.write(f"**Subject:** {ns['subject'][:80]}")
            if ns.get('actions'):
                st.write(f"**Actions:** {', '.join(ns['actions'][:3])}")

    # Offer template save for data files
    if result.file_type == "data" and not result.converter_used:
        _render_save_template_button(filename)




def load_project_files():
    """Load files from project folders."""
    from src.document_rag import get_document_rag
    from src.data_analyzer_sql import get_data_analyzer

    rag = get_document_rag()
    analyzer = get_data_analyzer()
    base = Path(__file__).parent

    doc_count = 0
    data_count = 0

    for folder in base.iterdir():
        if folder.is_dir() and not folder.name.startswith((".", "src", "storage", "__")):
            doc_count += rag.add_documents_from_folder(str(folder))
            data_count += analyzer.load_files_from_folder(str(folder))

    if doc_count > 0:
        rag.build_index()

    st.session_state.docs_count = len(rag.file_registry)
    st.session_state.tables_count = len(analyzer.list_tables())
    st.session_state.initialized = True

    return doc_count, data_count


def _render_save_template_button(file_name: str):
    """Show 'Save as Template' option after successful data file upload."""
    btn_key = f"save_tmpl_{file_name}"
    form_key = f"tmpl_form_{file_name}"

    if st.button("Save as Template", key=btn_key, use_container_width=True):
        st.session_state[f"_show_tmpl_form_{file_name}"] = True

    if st.session_state.get(f"_show_tmpl_form_{file_name}"):
        with st.form(form_key):
            tmpl_name = st.text_input("Template Name", value=f"{Path(file_name).stem} Format")
            tmpl_category = st.selectbox(
                "Category",
                ["dpr", "invoice", "manpower", "progress", "mep", "custom"],
                index=5,
            )
            submitted = st.form_submit_button("Create Template")

        if submitted:
            try:
                from src.config import TABLES_DIR
                from src.excel_table_extractor import ExcelTableExtractor
                from src.template_engine import get_template_store

                file_path = str(TABLES_DIR / file_name)
                extractor = ExcelTableExtractor()
                tables = extractor.extract_tables(file_path)

                if tables:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=False, data_only=True)
                    template = extractor.create_template_from_extraction(
                        tables=tables,
                        file_path=file_path,
                        sheet_names=wb.sheetnames,
                        template_name=tmpl_name,
                        category=tmpl_category,
                    )
                    wb.close()

                    store = get_template_store()
                    store.add_template(template)
                    st.success(f"Template '{tmpl_name}' created!")
                    st.session_state.pop(f"_show_tmpl_form_{file_name}", None)
                else:
                    st.warning("No tables found to create template from.")
            except Exception as e:
                st.error(f"Template creation failed: {e}")
                print(f"[Template] Error creating template from {file_name}: {e}")


def _render_template_management():
    """Render template management section in sidebar (admin only)."""
    try:
        from src.template_engine import get_template_store
        store = get_template_store()
    except Exception:
        return

    templates = store.list_templates()
    if not templates:
        return

    st.markdown("---")
    st.markdown("### Templates")

    for t in templates:
        tid = t["template_id"]
        with st.expander(f"{t['name']} ({t['category']})", expanded=False):
            st.caption(f"Source: {t['source_file']}")
            st.caption(f"Sheets: {t['sheet_count']} | Matches: {t['match_count']} | v{t['version']}")
            if st.button("Delete", key=f"del_{tid}"):
                store.remove_template(tid)
                st.rerun()



def _gather_uploaded_files() -> list:
    """Gather ALL uploaded files from DocumentRegistry (primary), plus RAG/catalog/disk fallbacks."""
    from src.config import DOCUMENTS_DIR, TABLES_DIR, EMAILS_DIR

    files = []
    seen = set()

    # Extension → type mapping
    EXT_TYPE = {
        ".pdf": "document", ".docx": "document", ".doc": "document", ".txt": "document",
        ".eml": "email", ".msg": "email",
        ".xlsx": "data", ".xls": "data", ".csv": "data",
    }

    # 1) Primary: DocumentRegistry (unified source of truth)
    try:
        from src.document_registry import get_document_registry
        registry = get_document_registry()

        # Hydrate from existing sources if registry is empty
        all_docs = registry.get_all()
        if not all_docs:
            rag_reg = {}
            try:
                from src.document_rag import get_document_rag
                rag_reg = get_document_rag().file_registry
            except Exception:
                pass
            catalog_entries = {}
            try:
                from src.catalog import get_catalog
                catalog_entries = get_catalog().entries
            except Exception:
                pass
            if rag_reg or catalog_entries:
                registry.hydrate_from_existing(rag_reg, catalog_entries)
                all_docs = registry.get_all()

        for rec in all_docs:
            if rec.status != "completed" or rec.file_name in seen:
                continue
            seen.add(rec.file_name)
            entry = {
                "name": rec.file_name,
                "type": rec.file_type,
                "doc_id": rec.doc_id,
                "time": rec.completed_at[:16].replace("T", " ") if rec.completed_at else "",
            }
            if rec.table_names:
                entry["tables"] = len(rec.table_names)
            if rec.notice_extracted:
                entry["notice"] = True
            if rec.file_size_kb:
                entry["size_kb"] = rec.file_size_kb
            files.append(entry)
    except Exception:
        pass

    # 2) Fallback: RAG file_registry for OCR/page info not in registry
    try:
        from src.document_rag import get_document_rag
        rag = get_document_rag()
        for fname, info in rag.file_registry.items():
            if fname in seen:
                # Enrich existing entry with OCR info
                for f in files:
                    if f["name"] == fname and info.get("ocr_pages", 0) > 0:
                        f["ocr"] = info["ocr_pages"]
                continue
            seen.add(fname)
            ext = info.get("file_type", Path(fname).suffix.lower())
            entry = {
                "name": fname,
                "type": EXT_TYPE.get(ext, "document"),
                "time": "",
            }
            if info.get("ocr_pages", 0) > 0:
                entry["ocr"] = info["ocr_pages"]
            files.append(entry)
    except Exception:
        pass

    # 3) Fallback: catalog for row counts not in registry
    try:
        from src.catalog import get_catalog
        catalog = get_catalog()
        for cat_entry in catalog.entries.values():
            fname = Path(cat_entry.source_file).name
            if fname in seen:
                # Enrich existing entry with row counts
                for f in files:
                    if f["name"] == fname:
                        total_rows = sum(t.row_count for t in cat_entry.tables)
                        if total_rows:
                            f["rows"] = total_rows
                        summary = ""
                        for t in cat_entry.tables:
                            if getattr(t, "summary", ""):
                                summary = t.summary
                                break
                        if summary:
                            f["summary"] = summary
                continue
            seen.add(fname)
            total_rows = sum(t.row_count for t in cat_entry.tables)
            summary = ""
            for t in cat_entry.tables:
                if getattr(t, "summary", ""):
                    summary = t.summary
                    break
            files.append({
                "name": fname,
                "type": "data",
                "time": "",
                "tables": len(cat_entry.tables),
                "rows": total_rows,
                "summary": summary,
            })
    except Exception:
        pass

    # 4) Scan disk directories for any files not yet registered
    for directory, fallback_type in [
        (DOCUMENTS_DIR, "document"),
        (EMAILS_DIR, "email"),
        (TABLES_DIR, "data"),
    ]:
        try:
            if not directory.exists():
                continue
            for fp in directory.iterdir():
                if fp.is_file() and fp.name not in seen:
                    seen.add(fp.name)
                    ext = fp.suffix.lower()
                    files.append({
                        "name": fp.name,
                        "type": EXT_TYPE.get(ext, fallback_type),
                        "time": "",
                    })
        except Exception:
            pass

    return files


def _render_uploaded_files():
    """Render scrollable uploaded files list + single-click Excel download in sidebar."""
    import io

    # On first call, gather existing files from all sources
    if not st.session_state.uploaded_files_log:
        existing = _gather_uploaded_files()
        if existing:
            st.session_state.uploaded_files_log = existing

    files_log = st.session_state.uploaded_files_log
    if not files_log:
        return

    st.markdown("---")
    st.markdown(f"### Uploaded Files ({len(files_log)})")

    # Handle pending delete action
    if st.session_state.get("_pending_delete_doc_id"):
        _doc_id = st.session_state.pop("_pending_delete_doc_id")
        _doc_name = st.session_state.pop("_pending_delete_name", "")
        try:
            from src.file_router import delete_document
            delete_document(_doc_id)
            st.session_state.uploaded_files_log = [
                f for f in st.session_state.uploaded_files_log
                if f.get("doc_id") != _doc_id
            ]
            st.toast(f"Deleted: {_doc_name}")
            st.rerun()
        except Exception as e:
            st.error(f"Delete failed: {e}")

    # Scrollable file list with delete buttons
    with st.container(height=220):
        for i, f in enumerate(files_log):
            icon = {"document": "📄", "email": "📧", "data": "📊"}.get(f.get("type", ""), "📁")

            chips = []
            if f.get("ocr"):
                chips.append(f'{f["ocr"]} OCR')
            if f.get("tables"):
                chips.append(f'{f["tables"]} tbl')
            if f.get("rows"):
                chips.append(f'{f["rows"]} rows')
            if f.get("notice"):
                chips.append("notice")
            if f.get("attachments"):
                chips.append(f'{f["attachments"]} att')
            chip_str = f" ({', '.join(chips)})" if chips else ""

            col_name, col_del = st.columns([0.85, 0.15])
            with col_name:
                st.markdown(
                    f'<div style="font-size:0.8rem;color:#e0e0e0;overflow:hidden;'
                    f'text-overflow:ellipsis;white-space:nowrap;padding:2px 0;">'
                    f'{icon} {f["name"]}'
                    f'<span style="color:#10a37f;font-size:0.7rem;">{chip_str}</span></div>',
                    unsafe_allow_html=True,
                )
            with col_del:
                if f.get("doc_id") and st.button(
                    "🗑", key=f"del_file_{i}",
                    help=f"Delete {f['name']}",
                ):
                    st.session_state["_pending_delete_doc_id"] = f["doc_id"]
                    st.session_state["_pending_delete_name"] = f["name"]
                    st.rerun()

    # Single-click Excel download (always ready)
    rows = []
    for f in files_log:
        rows.append({
            "File Name": f["name"],
            "Type": f.get("type", ""),
            "Upload Time": f.get("time", ""),
            "Size (KB)": f.get("size_kb", ""),
            "Tables": f.get("tables", ""),
            "Rows": f.get("rows", ""),
            "OCR Pages": f.get("ocr", ""),
            "Notice": "Yes" if f.get("notice") else "",
            "Attachments": f.get("attachments", ""),
            "Summary": f.get("summary", ""),
        })
    df_export = pd.DataFrame(rows)
    buf = io.BytesIO()
    df_export.to_excel(buf, index=False, sheet_name="Uploaded Files")
    buf.seek(0)
    st.download_button(
        label=f"Download File List ({len(files_log)})",
        data=buf,
        file_name=f"uploaded_files_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_files_excel",
        use_container_width=True,
    )


def _get_conversation_store():
    """Get ConversationStore for default user (cached in session_state)."""
    key = "_conversation_store"
    if key not in st.session_state:
        from src.conversation_store import ConversationStore
        st.session_state[key] = ConversationStore("default")
    return st.session_state[key]


def _switch_conversation(conv_id: str):
    """Load a conversation into session state."""
    store = _get_conversation_store()
    conv = store.get_conversation(conv_id)
    if conv:
        st.session_state["active_conversation_id"] = conv_id
        st.session_state["messages"] = [
            {
                "role": m.role,
                "content": m.content,
                "query_type": m.query_type or "",
                "sources": m.sources or [],
                "sql": m.sql,
                "result_data": m.result_data,
                "dual_answers": m.dual_answers,
            }
            for m in conv.messages
        ]


def render_conversation_sidebar():
    """Render conversation list and management in sidebar."""
    store = _get_conversation_store()

    # New Chat button — skip if current chat is already empty
    if st.button("New Chat", use_container_width=True, key="new_chat"):
        if st.session_state.get("messages"):
            # Current chat has messages → create new one
            meta = store.create_conversation()
            st.session_state["active_conversation_id"] = meta.conversation_id
            st.session_state["messages"] = []
            st.rerun()
        # else: current chat is already empty, no-op

    conversations = store.list_conversations()
    active_id = st.session_state.get("active_conversation_id")

    # Hide empty chats from sidebar (except the active one)
    conversations = [
        c for c in conversations
        if c.message_count > 0 or c.conversation_id == active_id
    ]

    # Conversation search
    search_term = st.text_input(
        "Search conversations",
        key="conv_search",
        placeholder="Search...",
        label_visibility="collapsed",
    )
    if search_term:
        search_lower = search_term.lower()
        conversations = [c for c in conversations if search_lower in c.title.lower()]

    for conv in conversations:
        is_active = conv.conversation_id == active_id
        col1, col2, col3 = st.columns([0.75, 0.12, 0.13])

        with col1:
            label = conv.title[:35]
            if len(conv.title) > 35:
                label += "..."
            btn_type = "primary" if is_active else "secondary"
            if st.button(
                label,
                key=f"conv_{conv.conversation_id}",
                use_container_width=True,
                type=btn_type,
            ):
                if not is_active:
                    _switch_conversation(conv.conversation_id)
                    st.rerun()

        with col2:
            if st.button("R", key=f"ren_{conv.conversation_id}", help="Rename"):
                st.session_state[f"_renaming_{conv.conversation_id}"] = True
                st.rerun()

        with col3:
            if st.button("X", key=f"del_{conv.conversation_id}"):
                store.delete_conversation(conv.conversation_id)
                if active_id == conv.conversation_id:
                    remaining = store.list_conversations()
                    if remaining:
                        _switch_conversation(remaining[0].conversation_id)
                    else:
                        new_meta = store.create_conversation()
                        st.session_state["active_conversation_id"] = new_meta.conversation_id
                        st.session_state["messages"] = []
                st.rerun()

        # Inline rename (admin/user can double-click title)
        if st.session_state.get(f"_renaming_{conv.conversation_id}"):
            new_name = st.text_input(
                "Rename",
                value=conv.title,
                key=f"rename_input_{conv.conversation_id}",
                label_visibility="collapsed",
            )
            if st.button("Save", key=f"save_rename_{conv.conversation_id}"):
                store.rename_conversation(conv.conversation_id, new_name)
                st.session_state.pop(f"_renaming_{conv.conversation_id}", None)
                st.rerun()


def _handle_mail_draft():
    """Build thread from selected emails and generate draft reply."""
    from src.light_graph import get_light_graph
    from src.thread_builder import get_thread_builder, ThreadMessage, CorrespondenceThread
    from src.content_generator import draft_reply

    graph = get_light_graph()
    builder = get_thread_builder()

    # 1. Collect selected doc_ids
    selected_ids = [
        key.replace("mail_sel_", "")
        for key in st.session_state
        if key.startswith("mail_sel_") and st.session_state[key]
    ]
    if not selected_ids:
        return

    # 2. Get notice data for selected emails (chronological order)
    try:
        placeholders = ",".join(["?"] * len(selected_ids))
        rows = graph._db.execute(
            f"SELECT doc_id, date, sender, recipient, subject, doc_type, file_name, actions "
            f"FROM notices WHERE doc_id IN ({placeholders}) ORDER BY date ASC",
            selected_ids,
        ).fetchall()
    except Exception as e:
        st.session_state.messages.append({
            "role": "assistant",
            "content": f"Error querying email metadata: {e}",
            "query_type": "draft",
        })
        return

    if not rows:
        st.session_state.messages.append({
            "role": "assistant",
            "content": "No notice metadata found for selected emails. Cannot generate draft.",
            "query_type": "draft",
        })
        return

    # 3. Build ThreadMessage list
    messages = []
    for row in rows:
        body_preview = builder._get_body_preview(row[6])  # file_name
        actions = [a.strip() for a in (row[7] or "").split(",") if a.strip()]
        messages.append(ThreadMessage(
            doc_id=row[0], date=row[1] or "", sender=row[2] or "",
            recipient=row[3] or "", subject=row[4] or "",
            body_preview=body_preview, actions=actions,
            file_name=row[6] or "", doc_type=row[5] or "",
        ))

    # 4. Build CorrespondenceThread (we reply as the last message's recipient)
    last = messages[-1]
    thread = CorrespondenceThread(
        party_a=last.recipient,
        party_b=last.sender,
        messages=messages,
        topic=last.subject,
    )

    # 5. Generate draft
    instruction = st.session_state.get("mail_draft_instruction", "")
    our_company = st.session_state.get("mail_our_company", "")
    draft_text = draft_reply(thread, instruction=instruction, our_company=our_company)

    # 6. Add to chat
    draft_msg = {
        "role": "assistant",
        "content": (
            f"**Draft Reply** ({thread.party_a} \u2192 {thread.party_b})\n\n"
            f"**Subject:** {last.subject}\n\n"
            f"---\n\n{draft_text}\n\n---\n\n"
            f"*Generated from {len(selected_ids)} selected email(s). "
            f"Please review and edit before sending.*"
        ),
        "query_type": "draft",
    }
    st.session_state.messages.append(draft_msg)

    # Save to conversation store
    store = _get_conversation_store()
    conv_id = st.session_state.get("active_conversation_id")
    if conv_id:
        from src.conversation_store import Message
        from datetime import datetime
        store.add_message(conv_id, Message(
            role="assistant", content=draft_msg["content"],
            timestamp=datetime.now().isoformat(), query_type="draft",
        ))


def _handle_mail_summarize():
    """Build thread from selected emails and generate summary."""
    from src.light_graph import get_light_graph
    from src.thread_builder import get_thread_builder, ThreadMessage, CorrespondenceThread
    from src.content_generator import summarize_thread

    graph = get_light_graph()
    builder = get_thread_builder()

    # 1. Collect selected doc_ids
    selected_ids = [
        key.replace("mail_sel_", "")
        for key in st.session_state
        if key.startswith("mail_sel_") and st.session_state[key]
    ]
    if not selected_ids:
        return

    # 2. Get notice data for selected emails (chronological order)
    try:
        placeholders = ",".join(["?"] * len(selected_ids))
        rows = graph._db.execute(
            f"SELECT doc_id, date, sender, recipient, subject, doc_type, file_name, actions "
            f"FROM notices WHERE doc_id IN ({placeholders}) ORDER BY date ASC",
            selected_ids,
        ).fetchall()
    except Exception as e:
        st.session_state.messages.append({
            "role": "assistant",
            "content": f"Error querying email metadata: {e}",
            "query_type": "document",
        })
        return

    if not rows:
        st.session_state.messages.append({
            "role": "assistant",
            "content": "No notice metadata found for selected emails.",
            "query_type": "document",
        })
        return

    # 3. Build ThreadMessage list
    messages = []
    for row in rows:
        body_preview = builder._get_body_preview(row[6])
        actions = [a.strip() for a in (row[7] or "").split(",") if a.strip()]
        messages.append(ThreadMessage(
            doc_id=row[0], date=row[1] or "", sender=row[2] or "",
            recipient=row[3] or "", subject=row[4] or "",
            body_preview=body_preview, actions=actions,
            file_name=row[6] or "", doc_type=row[5] or "",
        ))

    # 4. Build CorrespondenceThread
    senders = {m.sender for m in messages if m.sender}
    recipients = {m.recipient for m in messages if m.recipient}
    parties = list(senders | recipients)
    thread = CorrespondenceThread(
        party_a=parties[0] if parties else "Unknown",
        party_b=parties[1] if len(parties) > 1 else parties[0] if parties else "Unknown",
        messages=messages,
        topic=messages[-1].subject if messages else "",
    )

    # 5. Generate summary
    summary_text = summarize_thread(thread)

    # 6. Add to chat
    summary_msg = {
        "role": "assistant",
        "content": (
            f"**Thread Summary** ({thread.party_a} \u2194 {thread.party_b})\n\n"
            f"---\n\n{summary_text}\n\n---\n\n"
            f"*Summary of {len(selected_ids)} selected email(s).*"
        ),
        "query_type": "document",
    }
    st.session_state.messages.append(summary_msg)

    # Save to conversation store
    store = _get_conversation_store()
    conv_id = st.session_state.get("active_conversation_id")
    if conv_id:
        from src.conversation_store import Message
        from datetime import datetime
        store.add_message(conv_id, Message(
            role="assistant", content=summary_msg["content"],
            timestamp=datetime.now().isoformat(), query_type="document",
        ))


def render_doc_analysis_sidebar():
    """Document analysis mode — jargon-aware content search."""
    from src.jargon_manager import get_jargon_manager
    from src.light_graph import get_light_graph
    from src.document_rag import get_document_rag

    st.markdown("### \U0001F50D Document Analysis")

    search_term = st.text_input(
        "Search term",
        key="doc_analysis_input",
        placeholder="e.g., EOT, delay, payment, NCR...",
    )

    if not search_term:
        st.caption("Enter a keyword or abbreviation to search document contents.")
        return

    jargon = get_jargon_manager()

    # Jargon check — expand abbreviation
    expanded = jargon.expand(search_term.upper())
    search_terms = [search_term.lower()]
    if expanded:
        st.info(f"**{search_term.upper()}** = {expanded}")
        search_terms.append(expanded.lower())
    else:
        # Reverse lookup: does this text map to an abbreviation?
        abbr = jargon.abbreviate(search_term)
        if abbr:
            st.info(f"**{search_term}** \u2192 {abbr}")
            search_terms.append(abbr.lower())

    # DuckDB notices search (subject, topics, actions)
    graph = get_light_graph()
    like_patterns = []
    params = []
    for term in search_terms:
        for field in ["subject", "topics", "actions"]:
            like_patterns.append(f"LOWER({field}) LIKE ?")
            params.append(f"%{term}%")

    where_clause = " OR ".join(like_patterns)
    rows = []
    try:
        rows = graph._db.execute(
            f"SELECT doc_id, file_name, date, sender, recipient, subject, doc_type, topics "
            f"FROM notices WHERE {where_clause} "
            f"ORDER BY date ASC",
            params,
        ).fetchall()
    except Exception:
        pass

    # Pinecone RAG content search if few notice results
    rag_results = []
    if len(rows) < 5:
        try:
            rag = get_document_rag()
            expanded_query = " ".join(search_terms)
            rag_resp = rag.query(expanded_query, top_k=10)
            if rag_resp and rag_resp.get("sources"):
                found_ids = {r[0] for r in rows}
                for src in rag_resp["sources"]:
                    if src.get("doc_id") not in found_ids:
                        rag_results.append(src)
        except Exception:
            pass

    total = len(rows) + len(rag_results)
    st.caption(f"{total} document(s) found")

    if not total:
        st.warning("No documents found matching this term.")
        return

    # Results — chronological order
    with st.container(height=450):
        for row in rows:
            doc_id, file_name, date, sender, recipient, subject, doc_type, topics = row
            date_str = date or "\u2014"
            st.markdown(
                f"**{file_name}**\n\n"
                f"{date_str} | {doc_type or 'document'}\n\n"
                f"*{subject or 'No subject'}*"
            )
            if sender or recipient:
                st.caption(f"{sender or '?'} \u2192 {recipient or '?'}")
            st.markdown("---")

        # RAG results (documents not in notices)
        if rag_results:
            st.markdown("**Additional matches (content search):**")
            for src in rag_results:
                st.markdown(
                    f"**{src.get('file_name', 'Unknown')}**\n\n"
                    f"Score: {src.get('score', 0):.0%} | "
                    f"Page {src.get('page_number', '?')}"
                )
                if src.get("highlight_text"):
                    st.caption(src["highlight_text"][:100] + "...")
                st.markdown("---")


def render_mail_sidebar():
    """Render email list in sidebar for Correspondence Mode."""
    from src.document_registry import get_document_registry
    from src.light_graph import get_light_graph

    registry = get_document_registry()
    graph = get_light_graph()

    # Get email files from registry
    all_records = registry.get_completed()
    email_records = [r for r in all_records if r.file_type == "email"]

    if not email_records:
        st.info("No email files uploaded yet. Upload .eml or .msg files to use Correspondence Mode.")
        return

    st.markdown("### \u2709 Correspondence")
    st.caption(f"{len(email_records)} emails available")

    # Enrich with notice metadata from DuckDB
    enriched = {}
    try:
        email_filenames = [r.file_name for r in email_records]
        placeholders = ",".join(["?"] * len(email_filenames))
        rows = graph._db.execute(
            f"SELECT doc_id, file_name, date, sender, recipient, subject, doc_type "
            f"FROM notices WHERE file_name IN ({placeholders}) "
            f"ORDER BY date ASC",
            email_filenames,
        ).fetchall()
        for row in rows:
            enriched[row[1]] = {
                "doc_id": row[0], "date": row[2], "sender": row[3],
                "recipient": row[4], "subject": row[5], "doc_type": row[6],
            }
    except Exception:
        pass

    # Scrollable email list with checkboxes
    with st.container(height=400):
        for rec in email_records:
            meta = enriched.get(rec.file_name, {})
            sender = meta.get("sender", "Unknown")
            recipient = meta.get("recipient", "Unknown")
            date = meta.get("date", "\u2014")
            subject = meta.get("subject", rec.file_name)
            doc_id = meta.get("doc_id", rec.doc_id)

            date_display = str(date)[:10] if date and date != "\u2014" else ""
            st.checkbox(
                f"{date_display} \u2014 {subject[:50]}",
                key=f"mail_sel_{doc_id}",
            )
            st.caption(f"{sender} \u2192 {recipient}")

    # Collect selected doc_ids
    selected_ids = [
        key.replace("mail_sel_", "")
        for key in st.session_state
        if key.startswith("mail_sel_") and st.session_state[key]
    ]

    st.markdown("---")
    if selected_ids:
        st.success(f"{len(selected_ids)} email(s) selected")
        st.text_area(
            "Draft instruction (optional)",
            placeholder="e.g., Accept the extension request and confirm the new deadline...",
            key="mail_draft_instruction",
        )
        st.text_input(
            "Our company",
            key="mail_our_company",
            placeholder="e.g., TABH Construction",
        )
        if st.button("Generate Draft Reply", type="primary", use_container_width=True):
            with st.spinner("Generating..."):
                _handle_mail_draft()
            st.rerun()

        if st.button("Summarize Thread", use_container_width=True):
            with st.spinner("Summarizing..."):
                _handle_mail_summarize()
            st.rerun()

        if st.button("Clear Selection", use_container_width=True):
            for key in list(st.session_state.keys()):
                if key.startswith("mail_sel_"):
                    st.session_state[key] = False
            st.rerun()
    else:
        st.caption("Select emails above to generate a draft reply.")


def render_sidebar():
    """Render sidebar."""
    with st.sidebar:
        st.markdown("## Document Analysis Platform")
        st.caption("Documents, data & correspondence")

        st.markdown("---")

        _mode = st.session_state.get("active_mode", "chat")

        # Mail mode: skip conversation list, show email list directly
        if _mode == "mail":
            render_mail_sidebar()
            return

        # Other modes: normal conversation + upload sidebar
        render_conversation_sidebar()

        st.markdown("---")

        # File upload
        st.markdown("### Upload Files")
        all_files = st.file_uploader(
            "All file types",
            type=["pdf", "docx", "doc", "txt", "xlsx", "xls", "csv", "eml", "msg"],
            accept_multiple_files=True,
            key="unified_upload",
            label_visibility="collapsed",
        )

        if all_files:
            # Track already-processed files to avoid re-processing on Streamlit rerun
            if "processed_uploads" not in st.session_state:
                st.session_state.processed_uploads = set()

            new_files = [
                f for f in all_files
                if (f.name, f.size) not in st.session_state.processed_uploads
            ]

            if new_files:
                progress = st.progress(0, text="Processing files...")
                for i, f in enumerate(new_files):
                    progress.progress(
                        i / len(new_files),
                        text=f"Processing {f.name} ({i+1}/{len(new_files)})...",
                    )
                    result = process_unified_upload(f)
                    st.session_state.processed_uploads.add((f.name, f.size))
                    if result.success:
                        _render_upload_result(f.name, result)
                    elif result.file_type == "duplicate":
                        st.warning(f"Document already exists: {f.name}")
                    elif result.error:
                        st.warning(f"Could not process {f.name}: {result.error}")
                progress.progress(1.0, text="All files processed.")
                progress.empty()

        st.markdown("---")

        # Stats - use local counts (no Pinecone API call on every render)
        vec_count = st.session_state.docs_count

        col1, col2 = st.columns(2)
        with col1:
            st.metric("Vectors", vec_count)
        with col2:
            st.metric("Tables", st.session_state.tables_count)

        # Uploaded files list (all users can see)
        _render_uploaded_files()

        # Template management
        _render_template_management()

        # Clear all data (Pinecone + notices + graph)
        st.markdown("---")
        if st.button("Clear All Data", use_container_width=True):
            from src.document_rag import get_document_rag
            from src.notice_extractor import NOTICES_DIR
            from src.light_graph import GRAPH_DIR, get_light_graph

            rag = get_document_rag()
            rag.clear_index()

            # Clear notice JSON files
            for f in NOTICES_DIR.glob("*.json"):
                f.unlink()

            # Clear graph JSON files
            for f in GRAPH_DIR.glob("*.json"):
                f.unlink()

            # Reset light graph singleton
            graph = get_light_graph()
            graph.graph.nodes = {}
            graph.graph.edges = []
            graph._sync_notices_to_duckdb()

            # Clear table catalog (local + GCS)
            try:
                from src.catalog import get_catalog
                get_catalog().clear_all()
            except Exception:
                pass

            # Clear DuckDB table registry
            try:
                from src.data_analyzer_sql import get_data_analyzer
                analyzer = get_data_analyzer()
                for tbl in list(analyzer.tables.keys()):
                    try:
                        analyzer.conn.execute(f'DROP TABLE IF EXISTS "{tbl}"')
                    except Exception:
                        pass
                analyzer.tables.clear()
                analyzer.file_paths.clear()
            except Exception:
                pass

            # Clear review session files
            try:
                from src.config import REVIEW_SESSIONS_DIR
                for f in REVIEW_SESSIONS_DIR.glob("*.json"):
                    f.unlink()
            except Exception:
                pass

            # Clear converter registry
            try:
                from src.config import CONVERTER_REGISTRY_FILE
                if CONVERTER_REGISTRY_FILE.exists():
                    CONVERTER_REGISTRY_FILE.unlink()
            except Exception:
                pass

            st.session_state.docs_count = 0
            st.session_state.tables_count = 0
            st.session_state.processed_files = set()
            st.session_state.uploaded_files_log = []
            st.session_state.messages = []
            st.session_state["_startup_synced"] = False

            # Reset inline source state
            st.session_state.citation_verifications = {}
            st.session_state["expanded_pdf_sources"] = set()
            st.session_state["expanded_excel_sources"] = set()
            st.session_state["selected_pdf_source"] = None
            st.session_state["selected_excel_source"] = None
            st.session_state["notice_cache"] = {}

            # Reset current conversation
            conv_id = st.session_state.get("active_conversation_id")
            if conv_id:
                conv_store = _get_conversation_store()
                conv_store.delete_conversation(conv_id)
                new_meta = conv_store.create_conversation()
                st.session_state["active_conversation_id"] = new_meta.conversation_id

                st.rerun()


def _render_pdf_page(file_path: str, page_num: int, highlight: str = "", show_download: bool = False):
    """Render a single PDF page with optional highlight annotations."""
    try:
        page_num = int(page_num)
    except (ValueError, TypeError):
        page_num = 1

    if not file_path or not os.path.exists(file_path):
        st.error("PDF file not found")
        return

    try:
        import fitz

        doc = fitz.open(file_path)
        total_pages = len(doc)

        if page_num < 1 or page_num > total_pages:
            page_num = 1

        page = doc[page_num - 1]

        # Highlight the reference text with light blue
        highlight_found = False
        if highlight and len(highlight) > 5:
            clean_highlight = highlight.replace('\n', ' ').replace('  ', ' ').strip()

            search_attempts = []
            sentences = [s.strip() for s in clean_highlight.split('. ') if len(s.strip()) > 10]
            for sent in sentences[:3]:
                search_attempts.append(sent[:80])
                search_attempts.append(sent[:50])

            search_attempts.extend([
                clean_highlight[:100],
                clean_highlight[:70],
                clean_highlight[:50],
                clean_highlight[:30],
            ])

            seen = set()
            unique_attempts = []
            for s in search_attempts:
                if s and s not in seen and len(s) > 5:
                    seen.add(s)
                    unique_attempts.append(s)

            for term in unique_attempts:
                instances = page.search_for(term)
                if instances:
                    for inst in instances:
                        annot = page.add_highlight_annot(inst)
                        annot.set_colors(stroke=(1.0, 0.95, 0.0))
                        annot.set_opacity(0.35)
                        annot.update()
                    highlight_found = True
                    break

        mat = fitz.Matrix(2.0, 2.0)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        doc.close()

        st.image(img_bytes, use_container_width=True)
        st.caption(f"Page {page_num} / {total_pages}")

        if highlight and not highlight_found:
            st.warning("Could not locate the exact text on this page.")

        if show_download:
            with open(file_path, "rb") as f:
                st.download_button("Download PDF", f, Path(file_path).name, "application/pdf",
                                   key=f"dl_pdf_{page_num}")

    except Exception as e:
        st.error(f"Error rendering PDF: {e}")


def render_user_message(content: str):
    """Render user message bubble."""
    import html as _html
    safe = _html.escape(str(content or "").strip())
    if not safe:
        return
    st.markdown(f"""
    <div style="display: flex; justify-content: flex-end; margin: 1rem 0;">
        <div style="background: linear-gradient(135deg, #10a37f, #0d8a6a);
                    color: white;
                    padding: 1rem 1.25rem;
                    border-radius: 1.25rem 1.25rem 0.25rem 1.25rem;
                    max-width: 80%;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.3);">
            {safe}
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_assistant_message(content: str, query_type: str = ""):
    """Render assistant message bubble."""
    # Badge colors
    badge_styles = {
        "document": "background: #1e40af; color: #93c5fd;",
        "data": "background: #166534; color: #86efac;",
        "hybrid": "background: #92400e; color: #fcd34d;",
    }
    badge_style = badge_styles.get(query_type, "")
    badge_html = f'<span style="{badge_style} padding: 0.2rem 0.6rem; border-radius: 1rem; font-size: 0.7rem; font-weight: 600; text-transform: uppercase; margin-bottom: 0.5rem; display: inline-block;">{query_type}</span><br>' if query_type else ""

    st.markdown(f"""
    <div style="display: flex; justify-content: flex-start; margin: 1rem 0;">
        <div style="background: #2d2d44;
                    color: #e0e0e0;
                    padding: 1rem 1.25rem;
                    border-radius: 1.25rem 1.25rem 1.25rem 0.25rem;
                    max-width: 80%;
                    border: 1px solid #404060;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.3);">
            {badge_html}
            {content}
        </div>
    </div>
    """, unsafe_allow_html=True)


def score_sources_for_relevance(sources: list, user_question: str) -> None:
    """
    Enrich RAG source dicts with relevance_score, relevance_label, and issue_tags.
    Uses keyword scoring (Tier 1 only, no LLM) for speed.
    """
    import json as _json
    from src.notice_extractor import NoticeMetadata, NOTICES_DIR
    from src.document_reviewer import DocumentReviewer
    from src.config import REVIEW_HIGH_THRESHOLD, REVIEW_LOW_THRESHOLD

    reviewer = DocumentReviewer()
    question_lower = user_question.lower()
    cache = st.session_state.get("notice_cache", {})

    for src in sources:
        if src.get("type") == "structured_data" or not src.get("file_path"):
            continue

        doc_id = hashlib.md5(src["file_path"].encode()).hexdigest()[:16]

        if doc_id not in cache:
            notice_path = NOTICES_DIR / f"{doc_id}.json"
            if notice_path.exists():
                try:
                    with open(notice_path, 'r', encoding='utf-8') as f:
                        cache[doc_id] = NoticeMetadata(**_json.load(f))
                except Exception:
                    cache[doc_id] = None
            else:
                cache[doc_id] = None

        notice = cache.get(doc_id)
        if notice is None:
            src["relevance_score"] = None
            src["relevance_label"] = "unscored"
            src["issue_tags"] = []
            continue

        score, matched_tags, _ = reviewer._keyword_score(notice, question_lower)

        if score >= REVIEW_HIGH_THRESHOLD:
            label = "relevant"
        elif score <= REVIEW_LOW_THRESHOLD:
            label = "not_relevant"
        else:
            label = "borderline"

        src["relevance_score"] = round(score, 2)
        src["relevance_label"] = label
        src["issue_tags"] = matched_tags

    st.session_state["notice_cache"] = cache


# Relevance badge colors
_BADGE_COLORS = {
    "relevant": "#22c55e",
    "borderline": "#eab308",
    "not_relevant": "#ef4444",
    "unscored": "#666666",
}

# Issue tag colors
_ISSUE_COLORS = {
    "delay": "#ef4444",
    "payment_dispute": "#f59e0b",
    "quality_concern": "#8b5cf6",
    "safety_issue": "#ec4899",
    "scope_change": "#06b6d4",
    "communication_gap": "#64748b",
    "contractual_dispute": "#dc2626",
}


def _render_download_button(file_path: str, file_name: str, key: str):
    """Render a download button for non-PDF files."""
    _MIME_MAP = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".doc": "application/msword",
        ".txt": "text/plain",
        ".eml": "message/rfc822",
        ".msg": "application/vnd.ms-outlook",
    }
    ext = Path(file_path).suffix.lower()
    mime = _MIME_MAP.get(ext, "application/octet-stream")
    label = f"Download {ext.upper().lstrip('.')}"
    try:
        with open(file_path, "rb") as f:
            st.download_button(label, f, file_name, mime, key=key)
    except Exception:
        st.caption(f"File not available: {file_name}")


def render_sources(sources: list, msg_idx: int):
    """Render sources with relevance badges, issue tags, inline PDF, and citation verification."""
    if not sources:
        return

    verifications = st.session_state.get("citation_verifications", {}).get(str(msg_idx))

    # Auto-expand sources if first source has highlight text
    has_highlight = any(s.get("highlight_text") for s in sources if s.get("type") != "structured_data")
    with st.expander(f"Sources ({len(sources)})", expanded=has_highlight):
        # Verify Citations button
        doc_sources = [s for s in sources if s.get("type") != "structured_data"]
        if doc_sources and not verifications:
            if st.button("Verify Citations", key=f"verify_{msg_idx}"):
                from src.document_reviewer import get_citation_verifier
                verifier = get_citation_verifier()
                msg = st.session_state.messages[msg_idx] if msg_idx < len(st.session_state.messages) else {}
                answer_text = msg.get("content", "")
                v_results = verifier.verify_citations(answer_text, sources)
                if "citation_verifications" not in st.session_state:
                    st.session_state.citation_verifications = {}
                st.session_state.citation_verifications[str(msg_idx)] = [
                    {"source_idx": v.source_idx, "text_found": v.text_found_on_page,
                     "supported": v.claim_supported, "note": v.verification_note,
                     "confidence": v.confidence}
                    for v in v_results
                ]
                st.rerun()

        if verifications:
            verified_count = sum(1 for v in verifications if v.get("text_found"))
            total_v = len(verifications)
            v_color = "#22c55e" if verified_count == total_v else "#eab308" if verified_count > 0 else "#ef4444"
            st.markdown(
                f'<span style="color:{v_color};font-size:0.75rem;font-weight:600;">'
                f'Citations: {verified_count}/{total_v} verified</span>',
                unsafe_allow_html=True,
            )

        for i, src in enumerate(sources):
            src_type = src.get("type", "")

            if src_type == "structured_data":
                table_name = src.get("table_name", src.get("file_name", "Unknown"))
                rows = src.get("row_count_returned", 0)
                total = src.get("total_rows", 0)
                sql = src.get("sql_query", "")
                narrative = src.get("table_narrative", "")
                schema = src.get("target_schema", "")
                date_range = src.get("date_range", "")

                # Source file reference (original Excel name)
                source_file = src.get("source_file", "")
                display_file = ""
                if source_file:
                    display_file = Path(source_file).name
                if not display_file:
                    fn = src.get("file_name", "")
                    if fn and not fn.startswith("t_"):
                        display_file = fn

                # Header: Excel file + table info
                if display_file:
                    st.markdown(
                        f'<div style="background:rgba(34,197,94,0.1);border-left:3px solid #10a37f;'
                        f'padding:6px 10px;border-radius:4px;margin-bottom:4px;">'
                        f'<span style="color:#10a37f;font-weight:600;">Excel</span> '
                        f'<span style="color:#e0e0e0;font-weight:500;">{display_file}</span>'
                        f'<br/><span style="color:#888;font-size:0.8rem;">'
                        f'{table_name} | {rows}/{total} rows</span></div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(f"**{table_name}** - Rows: {rows}/{total}")

                if narrative:
                    st.markdown(f"> {narrative}")

                # Metadata chips
                meta_parts = []
                if schema:
                    meta_parts.append(f"Schema: `{schema}`")
                if date_range:
                    meta_parts.append(f"Period: `{date_range}`")
                if meta_parts:
                    st.markdown(" | ".join(meta_parts))

                # Inline result table (SQL query result) with highlighting
                result_preview = src.get("result_preview")
                if result_preview:
                    st.caption("Query Result:")
                    highlight_cols = src.get("highlight_columns", [])
                    df = pd.DataFrame(result_preview)

                    # Format date columns to YYYY-MM-DD
                    for col in df.columns:
                        if "date" in col.lower() or "tarih" in col.lower():
                            try:
                                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")
                            except Exception:
                                pass

                    # Format numeric columns to 2 decimal places
                    for col in df.select_dtypes(include=["float64", "float32"]).columns:
                        df[col] = df[col].round(2)

                    if highlight_cols:
                        def _highlight_key(col):
                            col_name = col.name.lower().replace('"', '')
                            is_key = any(
                                h.lower() in col_name or col_name in h.lower()
                                for h in highlight_cols
                            )
                            if is_key:
                                return [
                                    'background-color: rgba(34, 197, 94, 0.15)'
                                ] * len(col)
                            return [''] * len(col)

                        styled = df.style.apply(_highlight_key)
                        st.dataframe(styled, use_container_width=True,
                                     hide_index=True)
                    else:
                        st.dataframe(df, use_container_width=True,
                                     hide_index=True)

                # View Full Table — open in right citation panel
                table_name_db = src.get("table_name", "")
                if table_name_db:
                    if st.button(f"View Full Data ({total} rows)", key=f"view_excel_{msg_idx}_{i}"):
                        st.session_state["selected_excel_source"] = {
                            "table_name": table_name_db,
                            "source_file": display_file or table_name,
                            "total_rows": total,
                            "result_columns": src.get("result_columns", []),
                        }
                        st.rerun()

                # Expandable SQL details
                if sql:
                    with st.expander("SQL Query"):
                        st.code(sql, language="sql")
            elif src_type in ("thread_message", "notice"):
                # Chronological document source — clickable
                date = src.get("date", "")
                sender = src.get("sender", "")
                recipient = src.get("recipient", "")
                subject = src.get("subject", "")
                highlight = src.get("highlight_text", "")
                file_name = src.get("file_name", "")
                file_path = src.get("file_path", "")
                evidence = src.get("evidence", [])

                # Header: date + sender → recipient
                st.markdown(
                    f'<div style="background:rgba(59,130,246,0.1);border-left:3px solid #3b82f6;'
                    f'padding:6px 10px;border-radius:4px;margin-bottom:4px;">'
                    f'<span style="color:#60a5fa;font-weight:600;">{date}</span> '
                    f'<span style="color:#e0e0e0;">{sender} → {recipient}</span>'
                    f'<br/><span style="color:#93c5fd;font-size:0.85rem;">'
                    f'{subject}</span>'
                    f'<br/><span style="color:#888;font-size:0.75rem;">{file_name}</span></div>',
                    unsafe_allow_html=True,
                )

                # Evidence / highlight preview
                if highlight:
                    st.markdown(
                        f'<div style="background:rgba(255,213,79,0.1);border-left:3px solid #ffd54f;'
                        f'padding:6px 10px;margin:4px 0;border-radius:4px;font-size:0.82rem;'
                        f'color:#bbb;font-style:italic;">{highlight[:300]}</div>',
                        unsafe_allow_html=True,
                    )
                elif evidence:
                    # Show first evidence span
                    ev_text = evidence[0].get('text', str(evidence[0])) if isinstance(evidence[0], dict) else str(evidence[0])
                    st.markdown(
                        f'<div style="background:rgba(255,213,79,0.1);border-left:3px solid #ffd54f;'
                        f'padding:6px 10px;margin:4px 0;border-radius:4px;font-size:0.82rem;'
                        f'color:#bbb;font-style:italic;">{ev_text[:300]}</div>',
                        unsafe_allow_html=True,
                    )

                # View Document button — open PDF in right panel
                if file_path and os.path.exists(file_path):
                    if file_path.lower().endswith('.pdf'):
                        # Detect OCR
                        is_ocr = False
                        try:
                            from src.document_rag import get_document_rag
                            rag = get_document_rag()
                            reg = rag.file_registry.get(file_name, {})
                            is_ocr = (reg.get("ocr_pages", 0) or 0) > 0
                        except Exception:
                            pass

                        if st.button(f"View Document", key=f"notice_doc_{msg_idx}_{i}"):
                            st.session_state["selected_pdf_source"] = {
                                "file_path": file_path,
                                "page_number": src.get("page_number", 1),
                                "highlight_text": highlight if not is_ocr else "",
                                "file_name": file_name,
                                "is_ocr": is_ocr,
                                "total_pages": src.get("total_pages") or reg.get("page_count", 1),
                            }
                            st.rerun()
                    else:
                        # Download button for non-PDF files
                        _render_download_button(file_path, file_name, f"notice_dl_{msg_idx}_{i}")

            elif src_type == "search_result":
                # Unified search result — clickable with download/view
                file_name = src.get("file_name", "")
                file_path = src.get("file_path", "")
                file_type = src.get("file_type", "")
                ext = src.get("extension", "")
                date = src.get("date", "")
                sender = src.get("sender", "")
                subject = src.get("subject", "")
                description = src.get("description", "")

                ext_icons = {
                    ".pdf": "📄", ".xlsx": "📊", ".xls": "📊",
                    ".csv": "📊", ".docx": "📝", ".doc": "📝",
                    ".txt": "📃", ".eml": "📧", ".msg": "📧",
                }
                icon = ext_icons.get(ext, {"document": "📄", "data": "📊", "email": "📧"}.get(file_type, "📁"))

                # Header card
                meta_parts = []
                if date:
                    meta_parts.append(f'<span style="color:#60a5fa;font-weight:600;">{date}</span>')
                if sender:
                    meta_parts.append(f'<span style="color:#e0e0e0;">From: {sender}</span>')
                meta_html = " — ".join(meta_parts) if meta_parts else ""

                desc_html = ""
                if subject:
                    desc_html = f'<br/><span style="color:#93c5fd;font-size:0.85rem;">{subject}</span>'
                elif description:
                    desc_html = f'<br/><span style="color:#93c5fd;font-size:0.85rem;">{description}</span>'

                st.markdown(
                    f'<div style="background:rgba(16,163,127,0.1);border-left:3px solid #10a37f;'
                    f'padding:6px 10px;border-radius:4px;margin-bottom:4px;">'
                    f'{icon} <span style="color:#e0e0e0;font-weight:600;">{file_name}</span>'
                    f'{"<br/>" + meta_html if meta_html else ""}'
                    f'{desc_html}</div>',
                    unsafe_allow_html=True,
                )

                # Action button: View PDF or Download file
                if file_path and os.path.exists(file_path):
                    if ext == '.pdf':
                        if st.button("View Document", key=f"search_view_{msg_idx}_{i}"):
                            st.session_state["selected_pdf_source"] = {
                                "file_path": file_path,
                                "page_number": 1,
                                "highlight_text": "",
                                "file_name": file_name,
                                "is_ocr": False,
                            }
                            st.rerun()
                    else:
                        _render_download_button(file_path, file_name, f"search_dl_{msg_idx}_{i}")

            else:
                # Document source
                file_name = src.get("file_name", "Unknown")
                page = src.get("page_number", 1)
                total_pages = src.get("total_pages", 1)
                highlight = src.get("highlight_text", "")
                file_path = src.get("file_path", "")

                try:
                    page = int(page)
                except (ValueError, TypeError):
                    page = 1

                # Relevance badge
                rel_label = src.get("relevance_label", "unscored")
                rel_score = src.get("relevance_score")
                badge_color = _BADGE_COLORS.get(rel_label, "#666")
                score_str = f" {rel_score:.0%}" if rel_score is not None else ""
                rel_display = rel_label.replace("_", " ").title()

                # Citation verification badge
                v_badge = ""
                if verifications:
                    v = next((v for v in verifications if v.get("source_idx") == i), None)
                    if v:
                        if v.get("text_found") and v.get("supported"):
                            v_badge = ' <span style="color:#22c55e;font-size:0.7rem;">✓</span>'
                        elif v.get("text_found"):
                            v_badge = ' <span style="color:#eab308;font-size:0.7rem;">⚠</span>'
                        else:
                            v_badge = ' <span style="color:#ef4444;font-size:0.7rem;">✗</span>'

                # Header with relevance badge
                st.markdown(
                    f'**{file_name}** - p.{page}/{total_pages} '
                    f'<span style="background:{badge_color};color:white;padding:1px 7px;'
                    f'border-radius:10px;font-size:0.65rem;font-weight:600;">'
                    f'{rel_display}{score_str}</span>{v_badge}',
                    unsafe_allow_html=True,
                )

                # Issue tag chips
                issue_tags = src.get("issue_tags", [])
                if issue_tags:
                    chips = " ".join(
                        f'<span style="background:{_ISSUE_COLORS.get(t, "#555")};color:white;'
                        f'padding:1px 6px;border-radius:10px;font-size:0.6rem;">'
                        f'{t.replace("_", " ")}</span>'
                        for t in issue_tags
                    )
                    st.markdown(chips, unsafe_allow_html=True)

                # Highlight text
                if highlight:
                    hl_text = highlight[:300]
                    if len(highlight) > 300:
                        hl_text += "..."
                    st.markdown(
                        f'<div style="background:rgba(255,213,79,0.15);border-left:3px solid #ffd54f;'
                        f'padding:8px 12px;margin:4px 0;border-radius:4px;font-size:0.85rem;'
                        f'color:#e0e0e0;font-style:italic;">{hl_text}</div>',
                        unsafe_allow_html=True,
                    )

                # PDF viewer — open in right citation panel
                if file_path and os.path.exists(file_path) and file_path.lower().endswith('.pdf'):
                    # Detect OCR document
                    is_ocr = False
                    try:
                        from src.document_rag import get_document_rag
                        rag = get_document_rag()
                        reg = rag.file_registry.get(file_name, {})
                        is_ocr = (reg.get("ocr_pages", 0) or 0) > 0
                    except Exception:
                        pass

                    if st.button(f"View Page {page}", key=f"view_{msg_idx}_{i}"):
                        st.session_state["selected_pdf_source"] = {
                            "file_path": file_path,
                            "page_number": page,
                            "highlight_text": highlight if not is_ocr else "",
                            "file_name": file_name,
                            "is_ocr": is_ocr,
                            "total_pages": src.get("total_pages") or reg.get("page_count", 1),
                        }
                        st.rerun()

            st.markdown("---")


def render_provider_answer(answer: dict, provider: str, msg_idx: int, tab_idx: int):
    """Render a single provider's answer inside a tab."""
    if not answer:
        st.warning(f"No response from {provider}")
        return

    # Provider badge
    badge_colors = {
        "gemini": ("background: #4285f4; color: white;", "gemini"),
        "openai": ("background: #10a37f; color: white;", "OpenAI"),
        "claude": ("background: #d97706; color: white;", "Claude"),
    }
    style, label = badge_colors.get(provider, ("background: #555; color: white;", provider))
    st.markdown(
        f'<span style="{style} padding: 0.25rem 0.75rem; border-radius: 1rem; '
        f'font-size: 0.75rem; font-weight: 600;">{label}</span>',
        unsafe_allow_html=True,
    )

    # Answer text
    answer_text = answer.get("answer", answer.get("error", "No answer"))
    st.markdown(answer_text)

    # Sources
    sources = answer.get("sources", [])
    if sources:
        render_sources(sources, msg_idx * 100 + tab_idx)

    # SQL query
    sql = answer.get("sql")
    if sql:
        with st.expander("🔍 SQL Query"):
            st.code(sql, language="sql")

    # Result data
    result_data = answer.get("result_data")
    if result_data:
        with st.expander("📊 Data Results"):
            df = pd.DataFrame(result_data)
            st.dataframe(df, use_container_width=True, hide_index=True)

    # Error
    error = answer.get("error")
    if error and not answer.get("answer"):
        st.error(f"Error: {error}")


def render_dual_answers(msg: dict, idx: int):
    """Render dual-provider answers in tabs."""
    answers = msg.get("dual_answers", {})
    query_type = msg.get("query_type", "")

    # Query type badge
    badge_styles = {
        "document": "background: #1e40af; color: #93c5fd;",
        "data": "background: #166534; color: #86efac;",
        "hybrid": "background: #92400e; color: #fcd34d;",
    }
    badge_style = badge_styles.get(query_type, "")
    if badge_style:
        st.markdown(
            f'<span style="{badge_style} padding: 0.2rem 0.6rem; border-radius: 1rem; '
            f'font-size: 0.7rem; font-weight: 600; text-transform: uppercase;">'
            f'{query_type}</span>',
            unsafe_allow_html=True,
        )

    # Build tabs dynamically from available providers
    from src.config import LLM_PROVIDERS, GEMINI_MODEL, OPENAI_MODEL, ANTHROPIC_MODEL

    provider_labels = {
        "gemini": f"Gemini ({GEMINI_MODEL})",
        "openai": f"OpenAI ({OPENAI_MODEL})",
        "claude": f"Claude ({ANTHROPIC_MODEL})",
    }
    active_providers = [p for p in LLM_PROVIDERS if p in answers]
    if not active_providers:
        active_providers = list(answers.keys())

    tab_names = [provider_labels.get(p, p) for p in active_providers]
    tabs = st.tabs(tab_names)

    for tab, prov in zip(tabs, active_providers):
        with tab:
            render_provider_answer(answers.get(prov, {}), prov, idx, active_providers.index(prov))


def _render_doc_sources_table(sources: list, msg_idx: int):
    """Render document sources as inline clickable table in chat."""
    for i, src in enumerate(sources):
        file_name = src.get("file_name", "Unknown")
        file_path = src.get("file_path", "")
        page = src.get("page_number", 1)
        total_pages = src.get("total_pages", 1)
        date_raw = src.get("date", "")

        # Format date
        date_display = "\u2014"
        if date_raw:
            try:
                from datetime import datetime as _dt
                _parsed = _dt.strptime(str(date_raw)[:10], "%Y-%m-%d")
                date_display = _parsed.strftime("%d %b %Y")
            except Exception:
                date_display = str(date_raw)[:10]

        # 3-column row: Date | File Name | View button
        col_date, col_name, col_btn = st.columns([0.18, 0.62, 0.20])
        with col_date:
            st.caption(date_display)
        with col_name:
            st.markdown(f"**{file_name}**", help=src.get("subject", ""))
            sender = src.get("sender", "")
            recipient = src.get("recipient", "")
            if sender and recipient and sender != "Unknown":
                st.caption(f"{sender} \u2192 {recipient}")
        with col_btn:
            if file_path and os.path.exists(file_path) and file_path.lower().endswith('.pdf'):
                if st.button("View", key=f"doc_view_{msg_idx}_{i}", use_container_width=True):
                    is_ocr = False
                    _tp = total_pages
                    try:
                        from src.document_rag import get_document_rag
                        reg = get_document_rag().file_registry.get(file_name, {})
                        is_ocr = (reg.get("ocr_pages", 0) or 0) > 0
                        if not _tp or _tp <= 1:
                            _tp = reg.get("page_count", 1)
                    except Exception:
                        pass
                    st.session_state["selected_pdf_source"] = {
                        "file_path": file_path,
                        "page_number": page,
                        "highlight_text": src.get("highlight_text", ""),
                        "file_name": file_name,
                        "is_ocr": is_ocr,
                        "total_pages": _tp,
                    }
                    st.rerun()


def render_message(msg: dict, idx: int):
    """Render a chat message."""
    if msg.get("_thinking"):
        # Animated thinking indicator
        st.markdown(
            "<div style='display:flex;align-items:center;gap:8px;margin:1rem 0;'>"
            "<div style='width:8px;height:8px;border-radius:50%;background:#10a37f;"
            "animation:pulse 1.2s infinite;'></div>"
            "<span style='color:#888;font-style:italic;'>Thinking...</span>"
            "</div>"
            "<style>@keyframes pulse{0%,100%{opacity:.3}50%{opacity:1}}</style>",
            unsafe_allow_html=True,
        )
        return
    if msg["role"] == "user":
        render_user_message(msg["content"])
    elif msg.get("dual_answers"):
        render_dual_answers(msg, idx)
    else:
        render_assistant_message(msg["content"], msg.get("query_type", ""))

        # Draft email — show copyable code block
        if msg.get("query_type") == "draft":
            content = msg.get("content", "")
            # Extract draft text between --- markers if present
            if "---" in content:
                parts = content.split("---")
                if len(parts) >= 3:
                    draft_text = parts[1].strip()
                else:
                    draft_text = parts[-1].strip()
            else:
                draft_text = content
            with st.expander("Draft Email (Copy)", expanded=False):
                st.code(draft_text, language=None)

        # Sources badge — clicking selects this message's sources for the right panel
        sources = msg.get("sources", [])
        if not sources and msg.get("dual_answers"):
            for prov_ans in msg["dual_answers"].values():
                if isinstance(prov_ans, dict):
                    sources = prov_ans.get("sources", [])
                    if sources:
                        break
        if sources and msg.get("query_type") == "document":
            _render_doc_sources_table(sources, idx)
        elif sources:
            with st.expander(f"Sources ({len(sources)})"):
                for si, src in enumerate(sources):
                    fname = src.get("file_name", src.get("table_name", "Unknown"))
                    stype = src.get("type", "")
                    if stype == "structured_data":
                        tname = src.get("table_name", "")
                        cols = src.get("columns_used", [])
                        rows = src.get("row_count_returned", 0)
                        st.markdown(f"**{fname}** \u2014 `{tname}` ({rows} rows)")
                        if cols:
                            st.caption(f"Columns: {', '.join(cols)}")
                    else:
                        page = src.get("page_number", "")
                        st.markdown(f"**{fname}**" + (f" \u2014 p.{page}" if page else ""))

        # SQL query
        sql = msg.get("sql")
        if sql:
            with st.expander("SQL Query"):
                st.code(sql, language="sql")

        # Result data
        result_data = msg.get("result_data")
        if result_data:
            with st.expander("Data Results"):
                df = pd.DataFrame(result_data)
                st.dataframe(df, use_container_width=True, hide_index=True)


def render_citation_panel():
    """Render citation/source panel on the right side of the chat."""
    st.markdown("### Sources")

    messages = st.session_state.messages
    selected_idx = st.session_state.get("selected_citation_msg")

    # Default to latest assistant message with sources
    if selected_idx is None:
        for i in range(len(messages) - 1, -1, -1):
            msg = messages[i]
            if msg.get("role") == "assistant":
                sources = msg.get("sources", [])
                if not sources and msg.get("dual_answers"):
                    for prov_ans in msg["dual_answers"].values():
                        if isinstance(prov_ans, dict) and prov_ans.get("sources"):
                            sources = prov_ans["sources"]
                            break
                if sources:
                    selected_idx = i
                    break

    if selected_idx is None or selected_idx >= len(messages):
        st.caption("Ask a question to see sources here.")
        return

    msg = messages[selected_idx]
    sources = msg.get("sources", [])

    # For dual answers, get sources from first provider
    if not sources and msg.get("dual_answers"):
        for prov_ans in msg["dual_answers"].values():
            if isinstance(prov_ans, dict) and prov_ans.get("sources"):
                sources = prov_ans["sources"]
                break

    if not sources:
        st.caption("No sources for this answer.")
        return

    render_sources(sources, selected_idx)

    # PDF viewer in citation panel with page navigation
    pdf_src = st.session_state.get("selected_pdf_source")
    if pdf_src:
        st.markdown("---")
        col_h, col_close = st.columns([0.85, 0.15])
        with col_h:
            st.markdown(f"**{pdf_src['file_name']}**")
        with col_close:
            if st.button("✕", key="close_pdf_panel"):
                st.session_state["selected_pdf_source"] = None
                st.rerun()

        # Page navigation — lazy-load total_pages from fitz if missing
        page = pdf_src.get("page_number", 1)
        total = pdf_src.get("total_pages")
        if not total:
            try:
                import fitz
                _doc = fitz.open(pdf_src["file_path"])
                total = len(_doc)
                _doc.close()
                pdf_src["total_pages"] = total
            except Exception:
                total = "?"
        nav1, nav2, nav3 = st.columns([0.3, 0.4, 0.3])
        with nav1:
            if page > 1:
                if st.button("< Prev", key="pdf_prev"):
                    pdf_src["page_number"] = page - 1
                    pdf_src["highlight_text"] = ""
                    st.rerun()
        with nav2:
            st.caption(f"Page {page} / {total}")
        with nav3:
            can_next = not isinstance(total, int) or page < total
            if can_next:
                if st.button("Next >", key="pdf_next"):
                    pdf_src["page_number"] = page + 1
                    pdf_src["highlight_text"] = ""
                    st.rerun()

        _render_pdf_page(
            pdf_src["file_path"],
            pdf_src["page_number"],
            pdf_src.get("highlight_text", ""),
            show_download=True,
        )
        if pdf_src.get("is_ocr"):
            st.info("OCR document — text highlighting not available")

    # Excel viewer in citation panel
    excel_src = st.session_state.get("selected_excel_source")
    if excel_src:
        st.markdown("---")
        col_h, col_close = st.columns([0.85, 0.15])
        with col_h:
            st.markdown(f"**{excel_src['source_file']}**")
        with col_close:
            if st.button("✕", key="close_excel_panel"):
                st.session_state["selected_excel_source"] = None
                st.rerun()

        try:
            from src.data_analyzer_sql import get_data_analyzer
            analyzer = get_data_analyzer()
            table = excel_src["table_name"]
            limit = min(excel_src.get("total_rows") or 500, 500)
            full_df = analyzer.conn.execute(
                f"SELECT * FROM {table} LIMIT {limit}"
            ).fetchdf()

            # Format dates
            for col in full_df.columns:
                if "date" in col.lower() or "tarih" in col.lower():
                    try:
                        full_df[col] = pd.to_datetime(
                            full_df[col], errors="coerce"
                        ).dt.strftime("%Y-%m-%d")
                    except Exception:
                        pass

            # Format numerics
            for col in full_df.select_dtypes(include=["float64", "float32"]).columns:
                full_df[col] = full_df[col].round(2)

            # Highlight used columns
            result_cols = excel_src.get("result_columns", [])
            if result_cols:
                def _hl_used(c):
                    cn = c.name.lower().replace('"', '')
                    hit = any(rc.lower() in cn or cn in rc.lower() for rc in result_cols)
                    return ['background-color: rgba(34,197,94,0.1)'] * len(c) if hit else [''] * len(c)

                st.dataframe(full_df.style.apply(_hl_used),
                             use_container_width=True, hide_index=True, height=400)
            else:
                st.dataframe(full_df, use_container_width=True, hide_index=True, height=400)

            total = excel_src.get("total_rows") or 0
            if total and len(full_df) < total:
                st.caption(f"Showing {len(full_df)} of {total} rows")
        except Exception as e:
            st.error(f"Could not load table: {e}")


def _gather_dashboard_data() -> dict:
    """Collect all data needed for the dashboard."""
    from src.document_rag import get_document_rag
    from src.data_analyzer_sql import get_data_analyzer
    from src.light_graph import get_light_graph

    rag = get_document_rag()
    analyzer = get_data_analyzer()
    graph = get_light_graph()

    files = []
    for fname, info in rag.file_registry.items():
        files.append({
            "name": fname,
            "type": info.get("file_type", "document"),
            "pages": info.get("page_count", 0),
            "ocr": info.get("ocr_pages", 0),
            "file_path": info.get("file_path", ""),
        })

    # Add data files from catalog
    seen = {f["name"] for f in files}
    try:
        from src.catalog import get_catalog
        for entry in get_catalog().entries.values():
            fname = Path(entry.source_file).name
            if fname not in seen:
                table_names = [t.table_name for t in entry.tables]
                files.append({
                    "name": fname, "type": "excel",
                    "tables": len(entry.tables),
                    "rows": sum(t.row_count for t in entry.tables),
                    "table_names": table_names,
                })
                seen.add(fname)
    except Exception:
        pass

    stats = graph.get_statistics()
    recent = graph.timeline()[:5] if graph.graph.nodes else []
    parties = graph.get_all_parties()[:8] if graph.graph.nodes else []

    return {
        "files": files,
        "docs": len(rag.file_registry),
        "tables": len(analyzer.list_tables()),
        "notices": stats.get("node_count", 0),
        "edges": stats.get("edge_count", 0),
        "recent": recent,
        "parties": parties,
    }


@st.dialog("Table Viewer", width="large")
def _show_excel_dialog(table_name: str, source_file: str, total_rows: int):
    """Modal dialog for viewing full Excel table data."""
    st.markdown(f"**{source_file}** — `{table_name}`")
    try:
        from src.data_analyzer_sql import get_data_analyzer
        analyzer = get_data_analyzer()
        limit = min(total_rows or 500, 500)
        df = analyzer.conn.execute(
            f'SELECT * FROM "{table_name}" LIMIT {limit}'
        ).fetchdf()

        # Format dates
        for col in df.columns:
            if "date" in col.lower() or "tarih" in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
        # Round numerics
        for col in df.select_dtypes(include=["float64", "float32"]).columns:
            df[col] = df[col].round(2)

        st.dataframe(df, use_container_width=True, hide_index=True, height=500)

        if total_rows and limit < total_rows:
            st.caption(f"Showing {limit} of {total_rows} rows")
    except Exception as e:
        st.error(f"Could not load table: {e}")


def _apply_result_to_session(result: dict, query: str):
    """Apply a router result to session state as an assistant message."""
    from src.types import QueryType
    assistant_dict = {
        "role": "assistant",
        "content": result.get("answer", ""),
        "query_type": result.get("query_type", ""),
        "sources": result.get("sources", []),
        "sql": result.get("sql"),
        "result_data": result.get("result_data"),
    }
    st.session_state.messages.append(assistant_dict)

    # Populate drawer for timeline/thread queries
    sources = result.get("sources", [])
    qtype = result.get("query_type", "")
    doc_sources = [s for s in sources if s.get("type") in ("notice", "thread_message", None)
                   and s.get("file_name")]
    if qtype in (QueryType.TIMELINE.value, QueryType.THREAD.value) and doc_sources:
        st.session_state["drawer_documents"] = doc_sources
        st.session_state["drawer_title"] = f"Related Documents ({len(doc_sources)})"
        st.session_state["drawer_sort"] = "chronological"
    elif qtype == QueryType.DOCUMENT.value and len(doc_sources) >= 3:
        st.session_state["drawer_documents"] = doc_sources
        st.session_state["drawer_title"] = "Related Documents"
        st.session_state["drawer_sort"] = "relevance"
    else:
        st.session_state["drawer_documents"] = []


def render_dashboard():
    """Render the interactive dashboard — file inventory, stats, quick actions."""
    st.markdown("# Document Analysis Platform")
    st.caption("Examine documents, tables, relationships, and timelines")

    data = _gather_dashboard_data()

    # ── Stats row ──
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Documents", data["docs"])
    with c2:
        st.metric("Tables", data["tables"])
    with c3:
        st.metric("Notices", data["notices"])
    with c4:
        st.metric("Relationships", data["edges"])

    st.markdown("---")

    # ── Quick actions ──
    a1, a2, a3 = st.columns(3)
    with a1:
        if st.button("Ask a Question", use_container_width=True, type="primary"):
            st.session_state["view_mode"] = "chat"
            st.rerun()
    with a2:
        if st.button("View Timeline", use_container_width=True):
            st.session_state["view_mode"] = "chat"
            st.session_state.messages.append({"role": "user", "content": "Show all notices chronologically"})
            from src.router import get_router
            router = get_router()
            result = router.route_and_execute("Show all notices chronologically")
            _apply_result_to_session(result, "Show all notices chronologically")
            st.rerun()
    with a3:
        if st.button("Browse Relationships", use_container_width=True):
            st.session_state["view_mode"] = "chat"
            st.session_state.messages.append({"role": "user", "content": "Show communication flow"})
            from src.router import get_router
            router = get_router()
            result = router.route_and_execute("Show communication flow")
            _apply_result_to_session(result, "Show communication flow")
            st.rerun()

    st.markdown("---")

    # ── Tabs ──
    tab_files, tab_recent, tab_parties = st.tabs(["All Files", "Recent Activity", "Parties"])

    with tab_files:
        _render_file_inventory(data["files"])

    with tab_recent:
        _render_recent_activity(data["recent"])

    with tab_parties:
        _render_parties(data["parties"])


def _render_file_inventory(files: list):
    """Render file inventory grid."""
    if not files:
        st.info("No documents loaded. Upload files using the sidebar.")
        return

    # Filter row
    fc1, fc2 = st.columns([0.7, 0.3])
    with fc1:
        search = st.text_input("Search files", placeholder="Filter by name...",
                               label_visibility="collapsed", key="dash_file_search")
    with fc2:
        type_filter = st.selectbox("Type", ["All", "PDF", "Excel", "Email"],
                                   label_visibility="collapsed", key="dash_type_filter")

    filtered = files
    if search:
        filtered = [f for f in filtered if search.lower() in f["name"].lower()]
    if type_filter != "All":
        type_map = {"PDF": "pdf", "Excel": "excel", "Email": "email", "PDF": "document"}
        tval = type_filter.lower()
        filtered = [f for f in filtered if tval in (f.get("type") or "").lower()]

    st.caption(f"{len(filtered)} file(s)")

    # Render as rows with 3 columns
    for row_start in range(0, len(filtered), 3):
        cols = st.columns(3)
        for col_idx, f in enumerate(filtered[row_start:row_start + 3]):
            with cols[col_idx]:
                ftype = (f.get("type") or "document").lower()
                badge_colors = {"pdf": "#ef4444", "document": "#ef4444", "excel": "#22c55e",
                                "email": "#3b82f6", "data": "#22c55e"}
                bc = badge_colors.get(ftype, "#6b7280")
                type_label = {"pdf": "PDF", "document": "PDF", "excel": "Excel",
                              "email": "Email", "data": "Excel"}.get(ftype, ftype.upper())

                # Card HTML
                chips = []
                if f.get("pages"):
                    chips.append(f'{f["pages"]} pg')
                if f.get("ocr"):
                    chips.append(f'{f["ocr"]} OCR')
                if f.get("tables"):
                    chips.append(f'{f["tables"]} tbl')
                if f.get("rows"):
                    chips.append(f'{f["rows"]} rows')
                chip_str = f' <span style="color:#888;font-size:0.7rem;">{", ".join(chips)}</span>' if chips else ""

                st.markdown(
                    f'<div class="file-card">'
                    f'<span style="background:{bc};padding:2px 8px;border-radius:10px;'
                    f'color:white;font-size:0.65rem;font-weight:600;">{type_label}</span>'
                    f'{chip_str}'
                    f'<br/><span style="color:#e0e0e0;font-size:0.85rem;">'
                    f'{f["name"][:40]}</span></div>',
                    unsafe_allow_html=True,
                )

                # Open button
                btn_key = f"open_dash_{row_start}_{col_idx}"
                if ftype in ("pdf", "document"):
                    fp = f.get("file_path", "")
                    if fp and os.path.exists(fp):
                        if st.button("Open", key=btn_key, use_container_width=True):
                            is_ocr = (f.get("ocr", 0) or 0) > 0
                            st.session_state["selected_pdf_source"] = {
                                "file_path": fp, "page_number": 1,
                                "highlight_text": "",
                                "file_name": f["name"],
                                "is_ocr": is_ocr,
                            }
                            st.rerun()
                elif ftype in ("excel", "data") and f.get("table_names"):
                    if st.button("Open", key=btn_key, use_container_width=True):
                        tname = f["table_names"][0]
                        _show_excel_dialog(tname, f["name"], f.get("rows", 0))


def _render_recent_activity(recent: list):
    """Render recent documents from graph timeline."""
    if not recent:
        st.caption("No document activity yet. Upload and process documents first.")
        return

    for i, node in enumerate(recent):
        date = node.get("date", "—")
        sender = (node.get("sender") or "Unknown")[:30]
        recipient = (node.get("recipient") or "Unknown")[:30]
        subject = (node.get("subject") or "")[:80]
        file_name = node.get("file_name", "")
        doc_type = node.get("doc_type", "")

        st.markdown(
            f'<div class="file-card">'
            f'<span style="color:#10a37f;font-weight:600;font-size:0.8rem;">{date}</span>'
            f' <span style="color:#888;font-size:0.7rem;">[{doc_type}]</span>'
            f'<br/><span style="color:#93c5fd;font-size:0.82rem;">'
            f'{sender} → {recipient}</span>'
            f'<br/><span style="color:#e0e0e0;font-size:0.85rem;">{subject}</span>'
            f'<br/><span style="color:#666;font-size:0.72rem;">{file_name}</span></div>',
            unsafe_allow_html=True,
        )


def _render_parties(parties: list):
    """Render party list with communication stats."""
    if not parties:
        st.caption("No party data available. Upload email/notice documents first.")
        return

    for p in parties:
        party_name = p.get("party", "Unknown")
        sent = p.get("sent_count", 0)
        received = p.get("received_count", 0)
        st.markdown(
            f'<div class="file-card">'
            f'<span style="color:#e0e0e0;font-weight:600;">{party_name}</span>'
            f'<br/><span style="color:#22c55e;font-size:0.8rem;">'
            f'Sent: {sent}</span>'
            f' | <span style="color:#3b82f6;font-size:0.8rem;">'
            f'Received: {received}</span></div>',
            unsafe_allow_html=True,
        )


def _open_drawer_document(node):
    """Open a document from the Related Documents drawer."""
    fname = node.get("file_name", "")
    file_path = node.get("file_path", "")

    if not file_path:
        try:
            from src.document_rag import get_document_rag
            reg = get_document_rag().file_registry.get(fname, {})
            file_path = reg.get("file_path", "")
        except Exception:
            pass

    if file_path and os.path.exists(file_path) and file_path.lower().endswith(".pdf"):
        is_ocr = False
        try:
            from src.document_rag import get_document_rag
            reg = get_document_rag().file_registry.get(fname, {})
            is_ocr = (reg.get("ocr_pages", 0) or 0) > 0
        except Exception:
            pass

        _total_pages = reg.get("page_count", 1) if reg else 1
        st.session_state["selected_pdf_source"] = {
            "file_path": file_path,
            "page_number": node.get("page_number", 1),
            "highlight_text": node.get("highlight_text", ""),
            "file_name": fname,
            "is_ocr": is_ocr,
            "total_pages": _total_pages,
        }
        st.rerun()


def render_document_drawer():
    """Render the bottom contextual document drawer.

    Opens only when a query returns related documents about a topic/event/party.
    Does NOT open for direct Q&A or data/SQL answers.
    """
    data = st.session_state.get("drawer_documents", [])
    title = st.session_state.get("drawer_title", "Related Documents")

    if not data:
        return

    # Filter: only real documents (not Excel/structured_data)
    doc_nodes = [n for n in data if n.get("type") != "structured_data"]
    if not doc_nodes:
        return

    st.markdown("---")

    # Header bar with close button
    h1, h2 = st.columns([0.88, 0.12])
    with h1:
        st.markdown(f"**{title}** ({len(doc_nodes)} documents)")
    with h2:
        if st.button("Close", key="drawer_close"):
            st.session_state["drawer_documents"] = []
            st.rerun()

    # Sort documents based on query context
    sort_mode = st.session_state.get("drawer_sort", "relevance")
    if sort_mode == "chronological":
        doc_nodes.sort(key=lambda n: n.get("date", ""), reverse=False)
    else:
        doc_nodes.sort(key=lambda n: n.get("score") or 0, reverse=True)

    # Clickable card grid (5 columns per row)
    MAX_COLS = 5
    badge_colors = {
        "notice": "#ef4444", "email": "#3b82f6", "letter": "#8b5cf6",
        "thread_message": "#3b82f6",
    }
    type_icons = {
        "email": "\u2709\ufe0f", "letter": "\U0001F4E8",
        "notice": "\U0001F4CB", "thread_message": "\u2709\ufe0f",
    }

    for row_start in range(0, min(len(doc_nodes), 30), MAX_COLS):
        batch = doc_nodes[row_start:row_start + MAX_COLS]
        cols = st.columns(MAX_COLS)
        for i, node in enumerate(batch):
            with cols[i]:
                file_name = node.get("file_name", "Unknown")
                doc_type = node.get("doc_type", node.get("type", "document"))
                score = node.get("score")  # Pinecone cosine similarity
                bc = badge_colors.get(doc_type, "#6b7280")
                icon = type_icons.get(doc_type, "\U0001F4C4")
                type_label = "email" if doc_type == "thread_message" else doc_type

                # Short filename (without extension)
                short = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
                if len(short) > 25:
                    short = short[:22] + "..."

                # Score display with color coding
                score_html = ""
                if score is not None:
                    if score >= 0.5:
                        s_color, s_label = "#22c55e", "High"
                    elif score >= 0.3:
                        s_color, s_label = "#eab308", "Mid"
                    else:
                        s_color, s_label = "#f97316", "Low"
                    score_html = (
                        f'<span style="color:{s_color};font-size:0.68rem;'
                        f'font-weight:600;"> {s_label} ({score:.2f})</span>'
                    )
                elif sort_mode == "chronological":
                    # Show date for chronological mode
                    date = node.get("date", "")
                    if date:
                        score_html = (
                            f'<span style="color:#10a37f;font-size:0.68rem;'
                            f'font-weight:600;"> {date}</span>'
                        )

                # Card info above button
                st.markdown(
                    f'<div style="background:#2d2d44;border:1px solid #404060;'
                    f'border-radius:8px;padding:10px;text-align:center;margin-bottom:4px;">'
                    f'<div style="margin-bottom:6px;">'
                    f'<span style="background:{bc};padding:2px 8px;border-radius:10px;'
                    f'color:white;font-size:0.65rem;">{icon} {type_label}</span>'
                    f'{score_html}</div>'
                    f'<div style="font-size:0.82rem;color:#e0e0e0;word-break:break-word;'
                    f'line-height:1.3;">{short}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if st.button("Open", key=f"drawer_doc_{row_start + i}",
                             use_container_width=True):
                    _open_drawer_document(node)


def _save_user_message(content: str):
    """Save user message to session state and conversation store."""
    from src.conversation_store import Message
    st.session_state.messages.append({"role": "user", "content": content})
    store = _get_conversation_store()
    conv_id = st.session_state.get("active_conversation_id")
    if conv_id:
        store.add_message(conv_id, Message(
            role="user", content=content, timestamp=datetime.now().isoformat(),
        ))


def _save_assistant_message(content: str, query_type: str = "document",
                            sources=None, sql=None, result_data=None):
    """Save assistant message to session state and conversation store."""
    from src.conversation_store import Message
    msg = {"role": "assistant", "content": content, "query_type": query_type}
    if sources:
        msg["sources"] = sources
    if sql:
        msg["sql"] = sql
    if result_data is not None:
        msg["result_data"] = result_data
    st.session_state.messages.append(msg)
    store = _get_conversation_store()
    conv_id = st.session_state.get("active_conversation_id")
    if conv_id:
        store.add_message(conv_id, Message(
            role="assistant", content=content,
            timestamp=datetime.now().isoformat(), query_type=query_type,
        ))
        if len(st.session_state.messages) <= 2:
            first_msg = st.session_state.messages[0].get("content", "") if st.session_state.messages else content
            store.auto_title(conv_id, first_msg)


def handle_mail_input(user_input: str, skip_user_save: bool = False):
    """Handle chat input in Correspondence Mode — draft/reply or email Q&A."""
    from src.light_graph import get_light_graph
    from src.thread_builder import get_thread_builder, ThreadMessage, CorrespondenceThread
    from src.content_generator import draft_reply

    if not skip_user_save:
        _save_user_message(user_input)

    # Check for selected emails from sidebar
    selected_ids = [
        key.replace("mail_sel_", "")
        for key in st.session_state
        if key.startswith("mail_sel_") and st.session_state[key]
    ]

    graph = get_light_graph()
    builder = get_thread_builder()

    if selected_ids:
        # Build thread from selected emails
        try:
            placeholders = ",".join(["?"] * len(selected_ids))
            rows = graph._db.execute(
                f"SELECT doc_id, date, sender, recipient, subject, doc_type, file_name, actions "
                f"FROM notices WHERE doc_id IN ({placeholders}) ORDER BY date ASC",
                selected_ids,
            ).fetchall()
        except Exception:
            rows = []

        if not rows:
            _save_assistant_message(
                "No notice metadata found for selected emails. Cannot generate draft.",
                query_type="draft",
            )
            return

        messages = []
        for row in rows:
            body_preview = builder._get_body_preview(row[6])
            actions = [a.strip() for a in (row[7] or "").split(",") if a.strip()]
            messages.append(ThreadMessage(
                doc_id=row[0], date=row[1] or "", sender=row[2] or "",
                recipient=row[3] or "", subject=row[4] or "",
                body_preview=body_preview, actions=actions,
                file_name=row[6] or "", doc_type=row[5] or "",
            ))

        last = messages[-1]
        thread = CorrespondenceThread(
            party_a=last.recipient or "Unknown", party_b=last.sender or "Unknown",
            messages=messages, topic=last.subject,
        )

        # Intent detection: draft vs. explain/summarize
        import re as _re
        draft_patterns = r'\b(?:draft|reply|cevap|taslak|yaz|hazırla|oluştur|generate\s+(?:draft|reply)|respond)\b'
        is_draft = bool(_re.search(draft_patterns, user_input.lower()))

        if is_draft:
            our_company = st.session_state.get("mail_our_company", "")
            draft_text = draft_reply(thread, instruction=user_input, our_company=our_company)

            response = (
                f"**Draft Reply** ({thread.party_a} \u2192 {thread.party_b})\n\n"
                f"**Subject:** {last.subject}\n\n"
                f"---\n\n{draft_text}\n\n---\n\n"
                f"*Generated from {len(selected_ids)} selected email(s). "
                f"Please review and edit before sending.*"
            )
            _save_assistant_message(response, query_type="draft")
        else:
            # Summarize/explain mode — answer questions about selected emails
            from src.llm_client import get_llm_client
            llm = get_llm_client()

            context_parts = []
            for m in messages:
                context_parts.append(
                    f"[{m.date}] {m.sender} \u2192 {m.recipient}\n"
                    f"Subject: {m.subject}\n"
                    f"Content: {m.body_preview[:500]}\n"
                    f"Actions: {', '.join(m.actions) if m.actions else 'None'}"
                )
            email_context = "\n\n---\n\n".join(context_parts)

            prompt = (
                f"Below are {len(messages)} email(s) in chronological order:\n\n"
                f"{email_context}\n\n"
                f"User question: {user_input}\n\n"
                f"Please answer the user's question based on these emails."
            )
            answer = llm.generate_text(prompt, max_tokens=1024)
            _save_assistant_message(answer, query_type="document")
    else:
        # No emails selected — use normal router with mail-aware context
        from src.router import get_router
        router = get_router()
        result = router.route_and_execute(user_input)
        _save_assistant_message(
            result.get("answer", "No answer generated."),
            query_type=result.get("query_type", "document"),
            sources=result.get("sources"),
            sql=result.get("sql"),
            result_data=result.get("result_data"),
        )


def handle_doc_analysis_input(user_input: str, skip_user_save: bool = False):
    """Handle chat input in Document Analysis mode — jargon-aware search."""
    from src.jargon_manager import get_jargon_manager
    from src.light_graph import get_light_graph
    from src.document_rag import get_document_rag

    if not skip_user_save:
        _save_user_message(user_input)

    jargon = get_jargon_manager()

    # 1. Jargon expansion — expand abbreviations within sentences
    import re as _re
    expanded_query = jargon.expand_query(user_input.strip())
    search_terms = [user_input.lower()]
    jargon_info = ""

    if expanded_query != user_input.strip():
        # Jargon expansion found abbreviations — add expanded form as search term
        search_terms.append(expanded_query.lower())
        # Show expanded abbreviations to user
        found_abbrs = _re.findall(r'\b([A-Z][A-Z0-9&/]{1,10})\s*\(([^)]+)\)', expanded_query)
        if found_abbrs:
            jargon_parts = [f"**{ab}** = {meaning}" for ab, meaning in found_abbrs]
            jargon_info = " | ".join(jargon_parts) + "\n\n"
    else:
        # Single word — try reverse lookup
        abbr = jargon.abbreviate(user_input.strip())
        if abbr:
            search_terms.append(abbr.lower())
            jargon_info = f"**{user_input.strip()}** \u2192 {abbr}\n\n"

    # 2. DuckDB notices search
    graph = get_light_graph()
    like_patterns = []
    params = []
    for term in search_terms:
        for field in ["subject", "topics", "actions"]:
            like_patterns.append(f"LOWER({field}) LIKE ?")
            params.append(f"%{term}%")

    where_clause = " OR ".join(like_patterns)
    rows = []
    try:
        rows = graph._db.execute(
            f"SELECT doc_id, file_name, date, sender, recipient, subject, doc_type, topics "
            f"FROM notices WHERE {where_clause} "
            f"ORDER BY date ASC",
            params,
        ).fetchall()
    except Exception as e:
        logger.warning(f"[DocAnalysis] DuckDB search failed: {e}")

    # 3. RAG fallback if few results
    rag_results = []
    if len(rows) < 5:
        try:
            rag = get_document_rag()
            rag_query = expanded_query if expanded_query != user_input.strip() else user_input
            rag_resp = rag.query(rag_query, top_k=10)
            if rag_resp and rag_resp.get("sources"):
                found_ids = {r[0] for r in rows}
                for src in rag_resp["sources"]:
                    if src.get("doc_id") not in found_ids:
                        rag_results.append(src)
        except Exception as e:
            logger.warning(f"[DocAnalysis] RAG fallback failed: {e}")

    # 4. LLM topic extraction fallback — if very few results, extract keywords via LLM
    total = len(rows) + len(rag_results)
    if total < 3 and len(user_input.split()) > 2:
        try:
            from src import llm_client
            topic_prompt = (
                "Extract the main topic/keyword from this query for document search. "
                "Return ONLY the key topic words in the original language, nothing else.\n\n"
                f"Query: {user_input}"
            )
            topic_resp = llm_client.generate_text(topic_prompt, max_tokens=50)
            if topic_resp and topic_resp.text:
                extracted_topic = topic_resp.text.strip().lower()
                if extracted_topic and extracted_topic not in search_terms:
                    search_terms.append(extracted_topic)
                    # Re-run DuckDB search with new topic
                    extra_patterns = []
                    extra_params = []
                    for field in ["subject", "topics", "actions"]:
                        extra_patterns.append(f"LOWER({field}) LIKE ?")
                        extra_params.append(f"%{extracted_topic}%")
                    extra_where = " OR ".join(extra_patterns)
                    try:
                        extra_rows = graph._db.execute(
                            f"SELECT doc_id, file_name, date, sender, recipient, subject, doc_type, topics "
                            f"FROM notices WHERE {extra_where} ORDER BY date ASC",
                            extra_params,
                        ).fetchall()
                        existing_ids = {r[0] for r in rows}
                        for r in extra_rows:
                            if r[0] not in existing_ids:
                                rows.append(r)
                                existing_ids.add(r[0])
                    except Exception as e:
                        logger.warning(f"[DocAnalysis] LLM topic re-search failed: {e}")
        except Exception as e:
            logger.warning(f"[DocAnalysis] LLM topic extraction failed: {e}")

    # 5. Build response
    total = len(rows) + len(rag_results)
    if not total:
        _save_assistant_message(
            jargon_info + "No documents found matching this term.",
            query_type="document",
        )
        return

    # Build sources list for clickable document buttons
    doc_sources = []
    rag_obj = None
    try:
        from src.document_rag import get_document_rag
        rag_obj = get_document_rag()
    except Exception:
        pass

    # Build sources for inline table (text kept short — table shows details)
    for row in rows:
        _doc_id, file_name, date, sender, recipient, subject, doc_type, _topics = row
        _reg = {}
        _file_path = ""
        if rag_obj:
            _reg = rag_obj.file_registry.get(file_name, {})
            _file_path = _reg.get("file_path", "")
        doc_sources.append({
            "file_name": file_name,
            "file_path": _file_path,
            "page_number": 1,
            "total_pages": _reg.get("page_count", 1),
            "date": date or "",
            "sender": sender or "",
            "recipient": recipient or "",
            "subject": subject or "",
            "doc_type": doc_type or "document",
            "type": "notice",
        })

    for src in rag_results:
        doc_sources.append({
            **src,
            "type": "document",
            "total_pages": src.get("total_pages", 1),
            "date": src.get("date", ""),
        })

    summary = jargon_info + f"**{total}** related document(s) found:"
    _save_assistant_message(summary, query_type="document", sources=doc_sources)


def handle_input(user_input: str, skip_user_save: bool = False):
    """Process user input with conversation persistence and chat memory."""
    from src.router import get_router
    from src.config import LLM_PROVIDERS, CHAT_MEMORY_MESSAGES, CHAT_MEMORY_MAX_CHARS
    from src.conversation_store import Message, format_chat_context

    store = _get_conversation_store()
    conv_id = st.session_state.get("active_conversation_id")

    if not skip_user_save:
        # 1. Save user message to store
        now = datetime.now().isoformat()
        user_msg = Message(role="user", content=user_input, timestamp=now)
        if conv_id:
            store.add_message(conv_id, user_msg)

        # 2. Add to session state for immediate display
        st.session_state.messages.append({
            "role": "user",
            "content": user_input,
        })

    # 3. Build context-augmented query from recent messages
    if conv_id:
        recent = store.get_recent_messages(conv_id, CHAT_MEMORY_MESSAGES)
        # Exclude the just-added user message from context
        context_msgs = recent[:-1] if recent else []
        context = format_chat_context(context_msgs, CHAT_MEMORY_MESSAGES, CHAT_MEMORY_MAX_CHARS)
    else:
        context = ""
    augmented_query = f"{context}\n\nCurrent question: {user_input}" if context else user_input

    # 4. Route and execute
    router = get_router()

    if len(LLM_PROVIDERS) >= 2:
        # Dual-LLM mode
        result = router.route_and_execute_dual(augmented_query)
        answers = result.get("answers", {})
        assistant_dict = {
            "role": "assistant",
            "content": "",
            "dual_answers": answers,
            "query_type": result.get("query_type", ""),
        }
    else:
        # Single-LLM mode
        result = router.route_and_execute(augmented_query)
        assistant_dict = {
            "role": "assistant",
            "content": result.get("answer", "No answer"),
            "query_type": result.get("query_type", ""),
            "sources": result.get("sources", []),
            "sql": result.get("sql"),
            "result_data": result.get("result_data"),
        }

    st.session_state.messages.append(assistant_dict)

    # Auto-score sources for relevance (keyword-based, no LLM)
    src_list = assistant_dict.get("sources", [])
    if src_list:
        score_sources_for_relevance(src_list, user_input)

    # Also score sources in dual-LLM answers
    dual = assistant_dict.get("dual_answers", {})
    if dual:
        for prov_ans in dual.values():
            if isinstance(prov_ans, dict):
                prov_sources = prov_ans.get("sources", [])
                if prov_sources:
                    score_sources_for_relevance(prov_sources, user_input)

    # 5. Populate document drawer for relevant query types
    from src.types import QueryType
    qtype = assistant_dict.get("query_type", "")
    all_sources = assistant_dict.get("sources", [])
    if not all_sources and dual:
        # Get sources from first dual provider
        for prov_ans in dual.values():
            if isinstance(prov_ans, dict) and prov_ans.get("sources"):
                all_sources = prov_ans["sources"]
                break

    doc_sources = [s for s in all_sources
                   if s.get("type") in ("notice", "thread_message", None)
                   and s.get("file_name")]

    if qtype in (QueryType.TIMELINE.value, QueryType.THREAD.value) and doc_sources:
        st.session_state["drawer_documents"] = doc_sources
        st.session_state["drawer_title"] = f"Related Documents ({len(doc_sources)})"
        st.session_state["drawer_sort"] = "chronological"
    elif qtype == QueryType.DOCUMENT.value and len(doc_sources) >= 3:
        st.session_state["drawer_documents"] = doc_sources
        st.session_state["drawer_title"] = "Related Documents"
        st.session_state["drawer_sort"] = "relevance"
    else:
        st.session_state["drawer_documents"] = []

    # 6. Save assistant message to store
    if conv_id:
        assistant_msg = Message(
            role="assistant",
            content=assistant_dict.get("content", ""),
            timestamp=datetime.now().isoformat(),
            query_type=assistant_dict.get("query_type"),
            sources=assistant_dict.get("sources"),
            sql=assistant_dict.get("sql"),
            result_data=assistant_dict.get("result_data"),
            dual_answers=assistant_dict.get("dual_answers"),
        )
        store.add_message(conv_id, assistant_msg)

        # 6. Auto-title on first message
        conv_meta = next(
            (c for c in store.list_conversations()
             if c.conversation_id == conv_id), None
        )
        if conv_meta and conv_meta.title == "New Chat":
            store.auto_title(conv_id, user_input)


def _inject_scroll_restore():
    """Inject JS to save/restore scroll position across st.rerun() calls."""
    import streamlit.components.v1 as components
    components.html("""
    <script>
    (function() {
        const key = 'stScrollPos';
        const saved = sessionStorage.getItem(key);
        if (saved) {
            const pos = parseInt(saved);
            sessionStorage.removeItem(key);
            setTimeout(function() {
                var el = window.parent.document.querySelector('[data-testid="stAppViewContainer"]')
                      || window.parent.document.querySelector('section.main');
                if (el) el.scrollTop = pos;
            }, 150);
        }
        function save() {
            var el = window.parent.document.querySelector('[data-testid="stAppViewContainer"]')
                  || window.parent.document.querySelector('section.main');
            if (el && el.scrollTop > 50) {
                sessionStorage.setItem(key, el.scrollTop.toString());
            }
        }
        setInterval(save, 500);
    })();
    </script>
    """, height=0)


def main():
    """Main application entry point (single-user, no auth)."""
    init_session()
    check_api_keys()
    _inject_scroll_restore()

    # Auto-sync shared state on first session
    if not st.session_state.get("_startup_synced"):
        # 1. Pinecone vectors
        try:
            from src.document_rag import get_document_rag
            rag = get_document_rag()
            stats = rag.pinecone_index.describe_index_stats()
            vec_count = stats.get("total_vector_count", 0)
            if vec_count > 0:
                if not rag.index:
                    rag.load_index()
                st.session_state.docs_count = vec_count
                print(f"[Startup] Loaded {vec_count} vectors from Pinecone")
        except Exception as e:
            print(f"[Startup] Pinecone sync error: {e}")

        # 2. GCS sync (must run before graph rebuild so notice files are available)
        try:
            from src.gcs_storage import (
                sync_catalog_from_gcs, sync_all_parquets_from_gcs,
                sync_user_conversations_from_gcs,
                sync_review_sessions_from_gcs,
                sync_document_registry_from_gcs,
                sync_all_uploads_from_gcs,
            )
            sync_catalog_from_gcs()
            sync_all_parquets_from_gcs()
            sync_document_registry_from_gcs()
            sync_all_uploads_from_gcs()
            sync_user_conversations_from_gcs("default")
            sync_review_sessions_from_gcs()
        except Exception as gcs_err:
            print(f"[Startup] GCS sync: {gcs_err}")

        # 3. Light graph (after GCS sync so notice JSONs are downloaded)
        try:
            from src.light_graph import get_light_graph
            graph = get_light_graph()
            if not graph.graph.nodes:
                graph.rebuild_from_notices()
                if graph.graph.nodes:
                    print(f"[Startup] Rebuilt graph: {len(graph.graph.nodes)} notices")
        except Exception as e:
            print(f"[Startup] Graph rebuild error: {e}")

        # 4. Reload tables from catalog (ALWAYS attempt, independent of above)
        try:
            from src.data_analyzer_sql import get_data_analyzer
            analyzer = get_data_analyzer()
            if not analyzer.list_tables():
                loaded = analyzer.load_from_catalog()
                if loaded > 0:
                    print(f"[Startup] Reloaded {loaded} tables from catalog")
            st.session_state.tables_count = len(analyzer.list_tables())
        except Exception as e:
            print(f"[Startup] Table load error: {e}")

        st.session_state["_startup_synced"] = True

    render_sidebar()

    # Chat view (always)
    has_viewer = (st.session_state.get("selected_pdf_source") is not None
                  or st.session_state.get("selected_excel_source") is not None)

    if st.session_state.messages:
        if has_viewer:
            chat_col, citation_col = st.columns([0.65, 0.35])
            with chat_col:
                for i, msg in enumerate(st.session_state.messages):
                    render_message(msg, i)
            with citation_col:
                render_citation_panel()
        else:
            for i, msg in enumerate(st.session_state.messages):
                render_message(msg, i)
    else:
        # Welcome screen -- matching reference design
        mode = st.session_state.get("active_mode", "chat")
        if mode == "chat":
            # Spacer to push content toward vertical center
            st.markdown("<div style='margin-top:80px'></div>", unsafe_allow_html=True)

            # Centered lightning icon + greeting
            st.markdown(
                "<div style='text-align:center;'>"
                "<div style='display:inline-flex;align-items:center;justify-content:center;"
                "width:56px;height:56px;border-radius:16px;"
                "background:linear-gradient(135deg,#4F8EF7,#6C63FF);margin-bottom:16px;'>"
                "<span style='font-size:28px;'>&#9889;</span></div>"
                "<h2 style='margin:0 0 4px 0;'>How can I help you today?</h2>"
                "<p style='color:#888;font-size:0.95rem;margin:0;'>"
                "Ask me anything from technical problems to creative writing or daily tasks.</p>"
                "</div>",
                unsafe_allow_html=True,
            )

            st.markdown("")

            # Centered chat input bar
            spacer_l, input_col, spacer_r = st.columns([0.15, 0.70, 0.15])
            with input_col:
                icol1, icol2 = st.columns([0.88, 0.12])
                with icol1:
                    st.text_input(
                        "Message",
                        placeholder="Message AI Assistant...",
                        key="welcome_chat_input",
                        label_visibility="collapsed",
                    )
                with icol2:
                    st.button("➤", key="welcome_send", type="primary",
                              use_container_width=True)

            st.markdown("")

            # Mode cards -- Correspondence Mode & Document Analysis
            spacer_l2, card_col, spacer_r2 = st.columns([0.20, 0.60, 0.20])
            with card_col:
                c1, c2 = st.columns(2, gap="medium")
                with c1:
                    with st.container(border=True):
                        st.markdown("##### \u2709\uFE0F Correspondence Mode")
                        st.caption("Draft emails, organize inbox, and manage communications.")
                        if st.button("Open Correspondence Mode", key="btn_mail", use_container_width=True):
                            st.session_state.active_mode = "mail"
                            st.rerun()
                with c2:
                    with st.container(border=True):
                        st.markdown("##### \U0001F4C4 Document Analysis")
                        st.caption("Extract insights, summarize, and query PDFs or text files.")
                        if st.button("Open Document Analysis", key="btn_doc", use_container_width=True):
                            st.session_state.active_mode = "doc_analysis"
                            st.rerun()

    # Mode indicator + back button
    mode = st.session_state.get("active_mode", "chat")
    if mode != "chat":
        col_back, col_label = st.columns([1, 5])
        with col_back:
            if st.button("← Back", key="btn_back_to_chat"):
                st.session_state.active_mode = "chat"
                st.rerun()
        with col_label:
            label = "✉ Correspondence Mode" if mode == "mail" else "🔍 Document Analysis"
            st.markdown(f"**{label}**")

    # Handle welcome screen input (Enter key or send button)
    if not st.session_state.messages and st.session_state.get("active_mode", "chat") == "chat":
        wi = st.session_state.get("welcome_chat_input", "")
        if wi and wi.strip():
            _save_user_message(wi.strip())
            st.session_state["_pending_query"] = wi.strip()
            st.session_state["_pending_mode"] = "chat"
            st.rerun()

    # Phase 2: Show thinking indicator for pending query
    pending = st.session_state.pop("_pending_query", None)
    pending_mode = st.session_state.pop("_pending_mode", "chat")
    if pending:
        # Add temporary thinking message and rerun to display it
        st.session_state.messages.append({
            "role": "assistant", "content": "Thinking...", "_thinking": True,
        })
        st.session_state["_thinking_query"] = pending
        st.session_state["_thinking_mode"] = pending_mode
        st.rerun()

    # Phase 3: Execute query (thinking bubble already visible)
    thinking_query = st.session_state.pop("_thinking_query", None)
    thinking_mode = st.session_state.pop("_thinking_mode", "chat")
    if thinking_query:
        # Remove thinking indicator
        st.session_state.messages = [
            m for m in st.session_state.messages if not m.get("_thinking")
        ]
        # Process the query
        if thinking_mode == "mail":
            handle_mail_input(thinking_query, skip_user_save=True)
        elif thinking_mode == "doc_analysis":
            handle_doc_analysis_input(thinking_query, skip_user_save=True)
        else:
            handle_input(thinking_query, skip_user_save=True)
        st.rerun()

    # Chat input -- always visible in doc_analysis/mail modes, otherwise when conversation is active
    if st.session_state.messages or st.session_state.get("active_mode") in ("doc_analysis", "mail"):
        placeholder = {
            "chat": "Ask about your documents or data...",
            "mail": "Describe what you'd like to know or draft...",
            "doc_analysis": "Search documents by keyword or topic...",
        }.get(mode, "Ask about your documents or data...")

        # Phase 1: Save user message and rerun to show bubble immediately
        user_input = st.chat_input(placeholder)
        if user_input and user_input.strip():
            user_input = user_input.strip()
            _save_user_message(user_input)
            st.session_state["_pending_query"] = user_input
            st.session_state["_pending_mode"] = mode
            st.rerun()

    # Contextual document drawer (bottom, always last)
    render_document_drawer()


if __name__ == "__main__":
    main()
