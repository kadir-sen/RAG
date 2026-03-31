"""
Generate variant Excel files from template formats.
Produces 10 copies of each template with shifted dates and randomized numerical values.
"""

import copy
import random
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
NUM_COPIES = 10  # copies 2..11
MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


def shift_date(val, delta_days: int):
    """Shift a date value by delta_days. Handles str and datetime."""
    if isinstance(val, datetime):
        return val + timedelta(days=delta_days)
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(val, fmt)
                new_dt = dt + timedelta(days=delta_days)
                return new_dt.strftime(fmt)
            except ValueError:
                continue
    return val


def vary(value, pct=0.20, minimum=None):
    """Apply random variation to a numeric value."""
    if not isinstance(value, (int, float)):
        return value
    factor = 1 + random.uniform(-pct, pct)
    result = value * factor
    if minimum is not None:
        result = max(minimum, result)
    if isinstance(value, int):
        return max(1, round(result))
    return round(result, 2)


# ---------------------------------------------------------------------------
# Equipment Log Generator
# ---------------------------------------------------------------------------
def generate_equipment_logs():
    src = SCRIPT_DIR / "Equipment Log.xlsx"
    wb_src = openpyxl.load_workbook(src, data_only=True)
    ws_src = wb_src["Sheet1"]

    headers = [cell.value for cell in ws_src[1]]
    rows = []
    for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
        rows.append(list(row))
    wb_src.close()

    for copy_idx in range(2, 2 + NUM_COPIES):
        delta_days = (copy_idx - 1) * 90  # +3 months per copy
        random.seed(copy_idx * 1000 + 1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        for r_idx, row_data in enumerate(rows, 2):
            # Date (col 0)
            ws.cell(row=r_idx, column=1, value=shift_date(row_data[0], delta_days))
            # Block (col 1) - unchanged
            ws.cell(row=r_idx, column=2, value=row_data[1])
            # Floor (col 2) - unchanged
            ws.cell(row=r_idx, column=3, value=row_data[2])
            # Machinery Name (col 3) - unchanged
            ws.cell(row=r_idx, column=4, value=row_data[3])
            # Hours (col 4) - vary +/-20%
            ws.cell(row=r_idx, column=5, value=vary(row_data[4], 0.20, minimum=1))

        out = SCRIPT_DIR / f"Equipment Log {copy_idx}.xlsx"
        wb.save(out)
        wb.close()
        print(f"  Created: {out.name}")


# ---------------------------------------------------------------------------
# Manpower Production Log Generator
# ---------------------------------------------------------------------------
def generate_manpower_logs():
    src = SCRIPT_DIR / "Manpower Production Log.xlsx"
    wb_src = openpyxl.load_workbook(src, data_only=True)
    ws_src = wb_src["Sheet1"]

    headers = [cell.value for cell in ws_src[1]]
    rows = []
    for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=True):
        rows.append(list(row))
    wb_src.close()

    for copy_idx in range(2, 2 + NUM_COPIES):
        delta_days = (copy_idx - 1) * 90
        random.seed(copy_idx * 2000 + 2)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        for r_idx, row_data in enumerate(rows, 2):
            # Date (col 0)
            ws.cell(row=r_idx, column=1, value=shift_date(row_data[0], delta_days))
            # Block (col 1)
            ws.cell(row=r_idx, column=2, value=row_data[1])
            # Floor (col 2)
            ws.cell(row=r_idx, column=3, value=row_data[2])
            # Activity Description (col 3)
            ws.cell(row=r_idx, column=4, value=row_data[3])
            # Job Description (col 4)
            ws.cell(row=r_idx, column=5, value=row_data[4])
            # Number of Workers (col 5) - vary +/-2
            workers = row_data[5]
            if isinstance(workers, (int, float)):
                workers = max(1, round(workers + random.randint(-2, 2)))
            ws.cell(row=r_idx, column=6, value=workers)
            # Quantification (col 6) - vary +/-15%
            ws.cell(row=r_idx, column=7, value=vary(row_data[6], 0.15, minimum=1))
            # Unit of Measure (col 7)
            ws.cell(row=r_idx, column=8, value=row_data[7])

        out = SCRIPT_DIR / f"Manpower Production Log {copy_idx}.xlsx"
        wb.save(out)
        wb.close()
        print(f"  Created: {out.name}")


# ---------------------------------------------------------------------------
# IPC Sample Generator
# ---------------------------------------------------------------------------
def generate_ipc_samples():
    src = SCRIPT_DIR / "IPC Sample.xlsx"
    wb_src = openpyxl.load_workbook(src, data_only=True)

    # Read first sheet as template for structure
    template_sheet = wb_src[wb_src.sheetnames[0]]
    headers = [cell.value for cell in template_sheet[1]]
    template_rows = []
    for row in template_sheet.iter_rows(min_row=2, max_row=template_sheet.max_row, values_only=True):
        template_rows.append(list(row))

    # Identify which rows are "section headers" (no unit rate / BOQ Qty = 'LS')
    section_header_indices = set()
    for i, row in enumerate(template_rows):
        if row[4] == '-' or (row[3] == 'LS' and row[4] == '-'):
            section_header_indices.add(i)

    wb_src.close()

    for copy_idx in range(2, 2 + NUM_COPIES):
        random.seed(copy_idx * 3000 + 3)
        month_offset = (copy_idx - 1) * 3  # shift start month by 3 per copy

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Generate 4 monthly sheets
        for sheet_num in range(4):
            abs_month = (0 + month_offset + sheet_num) % 12  # 0-indexed month
            year = 2025 + (0 + month_offset + sheet_num) // 12
            month_name = MONTH_NAMES[abs_month]
            sheet_title = f"IPC_{month_name}_{year}"

            ws = wb.create_sheet(title=sheet_title)

            # Write headers
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)

            # Progress speed factor for this copy (0.7 to 1.3)
            progress_speed = 0.7 + random.random() * 0.6

            for r_idx, row_data in enumerate(template_rows):
                out_row = r_idx + 2  # 1-indexed, after header

                # Activity Code (A), Activity Name (B), Unit (C) - unchanged
                ws.cell(row=out_row, column=1, value=row_data[0])
                ws.cell(row=out_row, column=2, value=row_data[1])
                ws.cell(row=out_row, column=3, value=row_data[2])

                if r_idx in section_header_indices:
                    # Section header row - keep structure
                    ws.cell(row=out_row, column=4, value=row_data[3])  # BOQ Qty
                    ws.cell(row=out_row, column=5, value=row_data[4])  # Unit Rate
                    # Total BOQ Amount
                    boq_amount = row_data[5]
                    if isinstance(boq_amount, (int, float)):
                        boq_amount = vary(boq_amount, 0.10)
                    ws.cell(row=out_row, column=6, value=boq_amount)
                    # Percentages
                    ws.cell(row=out_row, column=7, value=row_data[6])  # Previous %
                    ws.cell(row=out_row, column=8, value=row_data[7] if row_data[7] else 0.0)
                    ws.cell(row=out_row, column=9, value=row_data[8])  # Current %
                    ws.cell(row=out_row, column=10, value=row_data[9] if row_data[9] else 0.0)
                    ws.cell(row=out_row, column=11, value=row_data[10])  # Cumulative %
                    ws.cell(row=out_row, column=12, value=row_data[11] if row_data[11] else 0.0)
                    continue

                # Detail row - apply variations
                boq_qty = row_data[3]  # Keep BOQ Qty same (contract quantity)
                ws.cell(row=out_row, column=4, value=boq_qty)

                unit_rate = row_data[4]
                if isinstance(unit_rate, (int, float)):
                    unit_rate = vary(unit_rate, 0.10)
                ws.cell(row=out_row, column=5, value=unit_rate)

                # Total BOQ Amount = BOQ Qty * Unit Rate
                if isinstance(boq_qty, (int, float)) and isinstance(unit_rate, (int, float)):
                    total_boq = round(boq_qty * unit_rate, 2)
                else:
                    total_boq = row_data[5]
                ws.cell(row=out_row, column=6, value=total_boq)

                # Generate progress percentages
                # Each sheet represents progressive months
                # Cumulative progress increases over sheets
                base_progress = min(1.0, (sheet_num + 1) * 0.25 * progress_speed)
                item_variation = random.uniform(-0.15, 0.15)
                cumulative_pct = max(0.0, min(1.0, base_progress + item_variation))

                if sheet_num == 0:
                    prev_pct = 0.0
                else:
                    prev_pct = max(0.0, cumulative_pct - random.uniform(0.05, 0.35))

                current_pct = cumulative_pct - prev_pct

                # Previous Amount
                prev_amount = round(total_boq * prev_pct, 2) if isinstance(total_boq, (int, float)) else 0.0
                ws.cell(row=out_row, column=7, value=f"{prev_pct:.2%}" if prev_pct > 0 else None)
                ws.cell(row=out_row, column=8, value=prev_amount)

                # Current Amount
                curr_amount = round(total_boq * current_pct, 2) if isinstance(total_boq, (int, float)) else 0.0
                ws.cell(row=out_row, column=9, value=f"{current_pct:.2%}" if current_pct > 0 else None)
                ws.cell(row=out_row, column=10, value=curr_amount)

                # Cumulative
                cum_amount = round(prev_amount + curr_amount, 2)
                ws.cell(row=out_row, column=11, value=f"{cumulative_pct:.2%}")
                ws.cell(row=out_row, column=12, value=cum_amount)

                # Quantities
                if isinstance(boq_qty, (int, float)):
                    prev_qty = round(boq_qty * prev_pct, 2)
                    curr_qty = round(boq_qty * current_pct, 2)
                    cum_qty = round(prev_qty + curr_qty, 2)
                else:
                    prev_qty = curr_qty = cum_qty = 0.0

                ws.cell(row=out_row, column=13, value=prev_qty)
                ws.cell(row=out_row, column=14, value=curr_qty)
                ws.cell(row=out_row, column=15, value=cum_qty)

        out = SCRIPT_DIR / f"IPC Sample {copy_idx}.xlsx"
        wb.save(out)
        wb.close()
        print(f"  Created: {out.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Generating Equipment Log variants...")
    generate_equipment_logs()

    print("\nGenerating Manpower Production Log variants...")
    generate_manpower_logs()

    print("\nGenerating IPC Sample variants...")
    generate_ipc_samples()

    print("\nDone! 30 files generated.")
