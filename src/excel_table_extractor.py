"""
Excel Table Extractor with native table detection, block detection,
and invoice/form-aware extraction for construction domain documents.

Handles:
- Native Excel tables (ListObjects)
- Block detection for clean tabular regions
- Invoice/form layouts with merged cells, header metadata, and data tables
- Full sheet fallback with smart header detection

Supports jargon-aware column normalization and merged cell handling.
"""
import re
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
from dataclasses import dataclass, field

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .logger import logger
from .catalog import get_catalog, TableMetadata


@dataclass
class ExtractedTable:
    """Represents an extracted table from Excel."""
    df: pd.DataFrame
    sheet_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    extraction_method: str  # "native_table" | "block_detect" | "invoice_detect" | "full_sheet"
    table_name: Optional[str] = None  # Excel native table name if available
    column_jargon: Dict[str, str] = field(default_factory=dict)  # col -> expanded meaning
    header_metadata: Dict[str, str] = field(default_factory=dict)  # label -> value from header section


# ── Label patterns for detecting form/invoice metadata rows ──────────
METADATA_LABEL_PATTERNS = re.compile(
    r'(?:name\s*of\s*project|project\s*name|customer\s*name|client'
    r'|sub\s*contractor|contractor\s*name|contract\s*value'
    r'|commencement\s*date|completion\s*date|start\s*date|end\s*date'
    r'|order\s*ref|reference|invoice\s*(?:no|number|date)'
    r'|date\s*of|attention|subject|provisional\s*sum'
    r'|variation|rev(?:ision)?|description\s*of\s*work)',
    re.IGNORECASE,
)

# Patterns that indicate a serial-number / line-item column
SERIAL_COL_PATTERNS = re.compile(
    r'^(?:s\.?\s*no\.?|sr\.?\s*no\.?|sl\.?\s*no\.?|item|no\.?|#|line)$',
    re.IGNORECASE,
)


class ExcelTableExtractor:
    """
    Extracts tables from Excel files using multiple strategies:
    0. Template-based extraction (user-confirmed format reuse)
    1. Native Excel tables (ListObjects)
    2. Dense table with multi-row headers (DPR, manpower reports)
    3. Block detection for clean tabular regions
    4. Invoice/form-aware extraction (merged cells, header metadata)
    5. Full sheet fallback

    Supports jargon-aware column normalization and merged cell handling.
    """

    # Block detection thresholds
    MIN_BLOCK_ROWS = 3
    MIN_BLOCK_COLS = 2
    MAX_EMPTY_ROWS_IN_BLOCK = 2
    MAX_EMPTY_COLS_IN_BLOCK = 1

    # Summary row keywords to filter out
    SUMMARY_KEYWORDS = {
        'total', 'grand total',
        'sum', 'subtotal', 'average',
        'net amount', 'net total', 'remaining total',
    }

    # Broader partial-match tokens for summary detection in any column
    SUMMARY_TOKENS = [
        'total', 'subtotal', 'sub total', 'net amount',
        'grand total', 'amount due',
    ]

    def __init__(self):
        """Initialize Excel table extractor."""
        self.catalog = get_catalog()
        self._jargon = None

    @property
    def jargon(self):
        """Lazy-load jargon manager."""
        if self._jargon is None:
            from .jargon_manager import get_jargon_manager
            self._jargon = get_jargon_manager()
        return self._jargon

    # ── Main entry point ─────────────────────────────────────────

    def extract_tables(self, file_path: str) -> List[ExtractedTable]:
        """
        Extract all tables from an Excel file.
        Tries strategies in order: template → native → dense → block → invoice → full-sheet.
        """
        path = Path(file_path)
        logger.info(f"[ExcelExtractor] Processing: {path.name}")

        tables: List[ExtractedTable] = []

        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)

            # ── Strategy 0: Template-based extraction ──
            template_result = self._extract_via_template(file_path, wb)
            if template_result is not None:
                tables, template_sheets = template_result
                if template_sheets and len(template_sheets) == len(wb.sheetnames):
                    # All sheets matched by template - done
                    wb.close()
                    logger.info(f"[ExcelExtractor] All sheets matched by template ({len(tables)} tables)")
                    return tables
                # Some sheets matched - continue with heuristics for unmatched
                matched_names = template_sheets or set()
            else:
                matched_names = set()

            for sheet_name in wb.sheetnames:
                if sheet_name in matched_names:
                    continue  # Already extracted via template
                ws = wb[sheet_name]

                if ws.max_row is None or ws.max_row < 2:
                    continue

                logger.info(
                    f"[ExcelExtractor] Sheet: {sheet_name} "
                    f"({ws.max_row} rows, {ws.max_column} cols, "
                    f"{len(ws.merged_cells.ranges)} merged ranges)"
                )

                # Build filled matrix (unmerge cells)
                matrix = self._unmerge_and_fill(ws)

                # Strategy 1: Native Excel tables
                native_tables = self._extract_native_tables(ws, sheet_name)
                if native_tables:
                    tables.extend(native_tables)
                    logger.info(f"[ExcelExtractor] Found {len(native_tables)} native table(s)")

                # Strategy 2: Dense table with multi-row headers
                dense_table = self._extract_dense_table(matrix, sheet_name)
                if dense_table:
                    tables.append(dense_table)
                    logger.info(
                        f"[ExcelExtractor] Dense table: {len(dense_table.df)} rows, "
                        f"{len(dense_table.df.columns)} cols"
                    )
                    continue  # Skip other strategies for this sheet

                # Strategy 3: Block detection on filled matrix
                block_tables = self._extract_block_tables_from_matrix(
                    matrix, sheet_name, native_tables
                )
                if block_tables:
                    tables.extend(block_tables)
                    logger.info(f"[ExcelExtractor] Found {len(block_tables)} block table(s)")

                # Strategy 4: Invoice/form-aware extraction
                if not native_tables and not block_tables:
                    invoice_tables = self._extract_invoice_tables(matrix, sheet_name)
                    if invoice_tables:
                        tables.extend(invoice_tables)
                        logger.info(f"[ExcelExtractor] Found {len(invoice_tables)} invoice table(s)")

                # Strategy 5: Full sheet fallback
                if not any(t.sheet_name == sheet_name for t in tables):
                    full_sheet = self._extract_full_sheet_from_matrix(matrix, sheet_name)
                    if full_sheet:
                        tables.append(full_sheet)
                        logger.info("[ExcelExtractor] Using full sheet as table")

            wb.close()

        except Exception as e:
            logger.warning(f"[ExcelExtractor] openpyxl failed: {e}")
            logger.info("[ExcelExtractor] Falling back to pandas/xlrd for .xls support")
            try:
                tables = self._extract_via_pandas(file_path)
            except Exception as e2:
                logger.error(f"[ExcelExtractor] Pandas fallback also failed: {e2}")

        logger.info(f"[ExcelExtractor] Total tables extracted: {len(tables)}")
        return tables

    # ── Pandas/xlrd fallback for .xls files ─────────────────────

    def _extract_via_pandas(self, file_path: str) -> List[ExtractedTable]:
        """
        Fallback extraction using pandas (supports .xls via xlrd).
        Builds a matrix from each sheet and runs invoice + block detection.
        """
        tables: List[ExtractedTable] = []
        xls = pd.ExcelFile(file_path)

        for sheet_name in xls.sheet_names:
            try:
                df_raw = pd.read_excel(
                    xls, sheet_name=sheet_name, header=None, dtype=object,
                )
                if df_raw.empty or len(df_raw) < 2:
                    continue

                logger.info(
                    f"[ExcelExtractor] Sheet (pandas): {sheet_name} "
                    f"({len(df_raw)} rows, {len(df_raw.columns)} cols)"
                )

                # Build matrix (replace NaN with None for consistency)
                matrix: List[List] = []
                for _, row in df_raw.iterrows():
                    matrix_row = []
                    for v in row:
                        if pd.isna(v):
                            matrix_row.append(None)
                        else:
                            # Try to preserve numeric types
                            try:
                                fv = float(v)
                                if fv == int(fv) and '.' not in str(v):
                                    matrix_row.append(int(fv))
                                else:
                                    matrix_row.append(fv)
                            except (ValueError, TypeError):
                                matrix_row.append(v)
                    matrix.append(matrix_row)

                # Try dense table first (multi-row headers)
                dense_table = self._extract_dense_table(matrix, sheet_name)
                if dense_table:
                    tables.append(dense_table)
                    logger.info(f"[ExcelExtractor] Dense table (pandas): {len(dense_table.df)} rows x {len(dense_table.df.columns)} cols")
                    continue

                # Run invoice detection (most common for .xls invoices)
                invoice_tables = self._extract_invoice_tables(matrix, sheet_name)
                if invoice_tables:
                    tables.extend(invoice_tables)
                    logger.info(f"[ExcelExtractor] Found {len(invoice_tables)} invoice table(s) (pandas)")
                    continue

                # Block detection
                block_tables = self._extract_block_tables_from_matrix(
                    matrix, sheet_name, [],
                )
                if block_tables:
                    tables.extend(block_tables)
                    logger.info(f"[ExcelExtractor] Found {len(block_tables)} block table(s) (pandas)")
                    continue

                # Full sheet fallback
                full_sheet = self._extract_full_sheet_from_matrix(matrix, sheet_name)
                if full_sheet:
                    tables.append(full_sheet)
                    logger.info("[ExcelExtractor] Using full sheet as table (pandas)")

            except Exception as sheet_err:
                logger.warning(f"[ExcelExtractor] Sheet '{sheet_name}' failed: {sheet_err}")

        return tables

    # ── Merged cell handling ─────────────────────────────────────

    def _unmerge_and_fill(self, ws: Worksheet) -> List[List]:
        """
        Build a matrix with merged cells expanded so every constituent
        cell holds the top-left value.  This is the foundation for all
        non-native extraction strategies.
        """
        max_row = min(ws.max_row or 1, 10000)
        max_col = min(ws.max_column or 1, 500)

        # Build initial matrix from cell values
        matrix: List[List] = [[None] * max_col for _ in range(max_row)]
        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            for cell in row:
                r, c = cell.row - 1, cell.column - 1
                if r < max_row and c < max_col:
                    matrix[r][c] = cell.value

        # Fill merged ranges
        for merge_range in ws.merged_cells.ranges:
            min_r = merge_range.min_row - 1
            max_r = min(merge_range.max_row - 1, max_row - 1)
            min_c = merge_range.min_col - 1
            max_c = min(merge_range.max_col - 1, max_col - 1)
            value = matrix[min_r][min_c]

            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    matrix[r][c] = value

        filled = sum(
            1 for r in range(max_row) for c in range(max_col) if matrix[r][c] is not None
        )
        logger.info(f"[ExcelExtractor] Matrix {max_row}x{max_col}, {filled} filled cells")
        return matrix

    # ── Strategy 0: Template-based extraction ─────────────────────

    def _extract_via_template(
        self, file_path: str, wb,
    ) -> Optional[Tuple[List['ExtractedTable'], set]]:
        """
        Attempt template-based extraction.
        Returns (tables, matched_sheet_names) if a template matches, else None.
        """
        from .template_engine import get_template_store, TemplateMatcher

        store = get_template_store()
        if not store.templates:
            return None

        matcher = TemplateMatcher(store)

        # Build matrices for sheets (needed for scoring stages 2-3)
        sheet_matrices: Dict[str, List[List]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is not None and ws.max_row >= 2:
                sheet_matrices[sheet_name] = self._unmerge_and_fill(ws)

        match = matcher.find_best_template(
            file_path, wb.sheetnames, sheet_matrices,
        )
        if match is None:
            return None

        template, score = match

        if not matcher.is_auto_match(template, score):
            logger.info(
                f"[ExcelExtractor] Template '{template.name}' matched at "
                f"{score:.1f} (below auto-threshold {template.confidence_threshold * 100:.0f}), skipping"
            )
            return None

        logger.info(f"[ExcelExtractor] Template match: '{template.name}' (score={score:.1f})")

        tables: List[ExtractedTable] = []
        matched_sheets: set = set()

        for sheet_name in wb.sheetnames:
            st = matcher.find_sheet_template(template, sheet_name)
            if st is None or sheet_name not in sheet_matrices:
                continue

            matrix = sheet_matrices[sheet_name]
            table = self._apply_sheet_template(st, matrix, sheet_name)
            if table:
                table.extraction_method = f"template:{template.template_id}"
                tables.append(table)
                matched_sheets.add(sheet_name)

        if tables:
            store.record_match(template.template_id)
            return tables, matched_sheets

        return None

    def _apply_sheet_template(
        self, st: 'SheetTemplate', matrix: List[List], sheet_name: str,
    ) -> Optional['ExtractedTable']:
        """Apply a sheet template to extract a table from the matrix."""
        num_rows = len(matrix)

        if not st.header_rows or max(st.header_rows) >= num_rows:
            return None

        # 1. Get headers from known rows
        if st.is_multi_row_header and len(st.header_rows) > 1:
            headers = self._merge_header_rows(
                matrix,
                st.header_rows[0], st.header_rows[-1],
                st.col_start, st.col_end,
            )
        else:
            header_row = st.header_rows[0]
            end_col = min(st.col_end + 1, len(matrix[header_row]) if header_row < num_rows else 0)
            headers = [
                str(matrix[header_row][c]).strip() if matrix[header_row][c] is not None else f"col_{c}"
                for c in range(st.col_start, end_col)
            ]

        if not headers:
            return None

        # 2. Find data end using existing logic
        data_end = self._find_dense_data_end(
            matrix, st.data_start_row, st.col_start, st.col_end,
        )

        if data_end < st.data_start_row:
            return None

        # 3. Extract data rows
        body = []
        for r in range(st.data_start_row, data_end + 1):
            end_col = min(st.col_end + 1, len(matrix[r]))
            row_data = [
                matrix[r][c] if c < len(matrix[r]) else None
                for c in range(st.col_start, st.col_end + 1)
            ]
            body.append(row_data)

        if not body:
            return None

        # Ensure headers match body width
        body_width = len(body[0]) if body else 0
        if len(headers) < body_width:
            headers.extend([f"col_{i}" for i in range(len(headers), body_width)])
        elif len(headers) > body_width:
            headers = headers[:body_width]

        # 4. Build and clean DataFrame
        df = pd.DataFrame(body, columns=headers)
        df = self._clean_dataframe(df)

        if df.empty:
            return None

        # 5. Extract metadata if template expects it
        header_meta = {}
        if st.has_metadata_header and st.header_rows:
            header_meta = self._extract_header_metadata(matrix, st.header_rows[0])

        return ExtractedTable(
            df=df,
            sheet_name=sheet_name,
            start_row=st.header_rows[0] + 1,
            start_col=st.col_start + 1,
            end_row=data_end + 1,
            end_col=st.col_end + 1,
            extraction_method="template",
            header_metadata=header_meta,
        )

    def create_template_from_extraction(
        self,
        tables: List['ExtractedTable'],
        file_path: str,
        sheet_names: List[str],
        template_name: str,
        category: str = "custom",
    ) -> 'FileTemplate':
        """
        Generate a FileTemplate from confirmed extraction results.
        Call this after user confirms the extraction is correct.
        """
        from .template_engine import (
            FileTemplate, SheetTemplate, _generate_template_id,
        )

        path = Path(file_path)

        # Build file name pattern from actual filename
        # e.g. "DPR 180207.xlsx" -> "DPR\s*\d+"
        # Replace digits first, then escape non-digit parts, then rejoin
        stem = path.stem
        parts = re.split(r'(\d+)', stem)
        pattern_parts = []
        for part in parts:
            if re.match(r'^\d+$', part):
                pattern_parts.append(r'\d+')
            elif part:
                pattern_parts.append(re.escape(part))
        file_name_pattern = "".join(pattern_parts)

        sheet_templates: Dict[str, SheetTemplate] = {}

        for table in tables:
            sn = table.sheet_name

            # Infer column types from DataFrame
            col_types = {}
            for col in table.df.columns:
                dtype = table.df[col].dtype
                if np.issubdtype(dtype, np.number):
                    col_types[col] = "numeric"
                else:
                    col_types[col] = "string"

            # Determine header rows (0-indexed)
            header_start = table.start_row - 1  # convert from 1-indexed
            data_start = header_start + 1  # default: single header row

            is_multi = table.extraction_method in ("dense_table", "template")
            header_rows = [header_start]

            # For dense tables, we may have multi-row headers
            # Detect by checking if extraction_method is dense_table
            if "dense" in table.extraction_method:
                # Heuristic: count non-data rows between header_start and data
                # For DPR files this is typically 2-3 rows
                for r_offset in range(1, 5):
                    candidate_row = header_start + r_offset
                    if candidate_row < table.start_row + len(table.df):
                        header_rows.append(candidate_row)
                    else:
                        break
                    # Check if this row is still header-like (mostly strings)
                    # Simple heuristic: the data_start_row is start_row + len(header_rows) - 1
                data_start = header_rows[-1] + 1
                is_multi = True

            # Check for serial number column
            has_serial = False
            if len(table.df.columns) > 0:
                first_col = table.df.columns[0]
                first_vals = table.df[first_col].dropna().head(5)
                try:
                    nums = [int(float(v)) for v in first_vals]
                    if nums == list(range(1, len(nums) + 1)):
                        has_serial = True
                except (ValueError, TypeError):
                    pass

            st = SheetTemplate(
                sheet_name_pattern=sn,
                header_rows=header_rows,
                data_start_row=data_start,
                col_start=table.start_col - 1,  # 0-indexed
                col_end=table.end_col - 1,       # 0-indexed
                column_names=list(table.df.columns),
                column_count=len(table.df.columns),
                column_types=col_types,
                is_multi_row_header=is_multi,
                has_metadata_header=bool(table.header_metadata),
                metadata_labels=list(table.header_metadata.keys()) if table.header_metadata else [],
                has_serial_column=has_serial,
                has_summary_rows=False,  # already cleaned
                extraction_method=table.extraction_method,
            )
            sheet_templates[sn] = st

        ft = FileTemplate(
            template_id=_generate_template_id(template_name, file_path),
            name=template_name,
            category=category,
            file_name_pattern=file_name_pattern,
            sheet_name_patterns=list(sheet_templates.keys()),
            sheet_templates=sheet_templates,
            source_file=path.name,
        )

        logger.info(
            f"[ExcelExtractor] Created template: {template_name} "
            f"({len(sheet_templates)} sheets)"
        )
        return ft

    # ── Strategy 1: Native tables ────────────────────────────────

    def _extract_native_tables(self, ws: Worksheet, sheet_name: str) -> List[ExtractedTable]:
        """Extract native Excel tables (ListObjects)."""
        tables = []
        try:
            if hasattr(ws, 'tables') and ws.tables:
                for table_name, table_obj in ws.tables.items():
                    try:
                        ref = table_obj.ref
                        start, end = ref.split(":")
                        start_col, start_row = self._parse_cell_ref(start)
                        end_col, end_row = self._parse_cell_ref(end)

                        data = []
                        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                                min_col=start_col, max_col=end_col):
                            data.append([cell.value for cell in row])

                        if len(data) < 2:
                            continue

                        df = pd.DataFrame(data[1:], columns=data[0])
                        df = self._clean_dataframe(df)

                        if not df.empty:
                            tables.append(ExtractedTable(
                                df=df,
                                sheet_name=sheet_name,
                                start_row=start_row,
                                start_col=start_col,
                                end_row=end_row,
                                end_col=end_col,
                                extraction_method="native_table",
                                table_name=table_name,
                            ))
                    except Exception as e:
                        logger.warning(f"[ExcelExtractor] Error parsing table {table_name}: {e}")
        except Exception as e:
            logger.warning(f"[ExcelExtractor] Error accessing tables: {e}")
        return tables

    # ── Strategy 2: Dense table with multi-row headers ──────────

    # Values to skip when merging header rows (separators/placeholders)
    HEADER_SKIP_VALUES = {'-', 'x', '#', '', 'nan', 'none'}

    def _extract_dense_table(
        self, matrix: List[List], sheet_name: str,
    ) -> Optional[ExtractedTable]:
        """
        Detect and extract dense tables with multi-row headers.
        Common in DPR manpower reports, equipment logs, etc.

        Multi-row header pattern:
        - Row A: Zone/category names (many duplicates from merged cells)
        - Row B: Sub-area/subcategory names
        - Row C: Day/Night or other repeating sub-headers (small unique set)
        - Row D+: Numeric data
        """
        detection = self._detect_multi_row_header(matrix)
        if detection is None:
            return None

        header_start, header_end, col_start, col_end = detection

        # Merge header rows into composite column names
        merged_headers = self._merge_header_rows(
            matrix, header_start, header_end, col_start, col_end
        )

        if not merged_headers or len(merged_headers) < 3:
            return None

        # Validate: reject if column names are excessively long (not a real header)
        avg_len = sum(len(h) for h in merged_headers) / len(merged_headers)
        if avg_len > 60:
            logger.info(f"[DenseTable] Rejected: avg column name too long ({avg_len:.0f} chars)")
            return None

        # Extract header metadata from rows above the header
        header_meta = self._extract_header_metadata(matrix, header_start)

        # Extract data rows
        data_start = header_end + 1
        data_end = self._find_dense_data_end(matrix, data_start, col_start, col_end)

        if data_end <= data_start:
            return None

        body = [
            matrix[r][col_start:col_end + 1]
            for r in range(data_start, data_end + 1)
        ]

        df = pd.DataFrame(body, columns=merged_headers)
        df = self._clean_dataframe(df)

        if df.empty:
            return None

        logger.info(
            f"[DenseTable] Sheet '{sheet_name}': "
            f"headers=rows {header_start + 1}-{header_end + 1}, "
            f"data={len(df)} rows x {len(df.columns)} cols"
        )

        return ExtractedTable(
            df=df,
            sheet_name=sheet_name,
            start_row=header_start + 1,
            start_col=col_start + 1,
            end_row=data_end + 1,
            end_col=col_end + 1,
            extraction_method="dense_table",
            header_metadata=header_meta,
        )

    def _detect_multi_row_header(
        self, matrix: List[List], max_scan: int = 15,
    ) -> Optional[Tuple[int, int, int, int]]:
        """
        Detect multi-row header pattern in the matrix.

        Returns (header_start, header_end, col_start, col_end) or None.

        Detection signals:
        1. 2+ consecutive string-heavy rows
        2. High value repetition (merged cells duplicate values in neighbors)
        3. Wide span (>10 columns)
        4. Last header row has very few unique values (Day/Night pattern)
        5. Row after headers has numeric data
        """
        num_rows = len(matrix)
        if num_rows < 4:
            return None
        num_cols = len(matrix[0]) if matrix else 0
        if num_cols < 10:
            return None

        # Analyze each row in the scan region
        row_info = []
        for r in range(min(num_rows, max_scan)):
            row = matrix[r]
            non_null = []
            strings = []
            numbers = []
            neighbor_dupes = 0

            for c in range(num_cols):
                v = row[c]
                if v is not None and str(v).strip() and str(v).strip().lower() not in self.HEADER_SKIP_VALUES:
                    non_null.append(c)
                    if isinstance(v, str):
                        strings.append((c, v.strip()))
                    elif isinstance(v, (int, float)) and not isinstance(v, bool):
                        numbers.append((c, v))

                # Check if cell duplicates its left neighbor (merge signal)
                if c > 0 and v is not None and str(v).strip():
                    left = row[c - 1]
                    if left is not None and str(v).strip() == str(left).strip():
                        neighbor_dupes += 1

            col_start = non_null[0] if non_null else 0
            col_end = non_null[-1] if non_null else 0
            width = col_end - col_start + 1 if non_null else 0
            unique_vals = len(set(str(v).strip().lower() for _, v in strings)) if strings else 0

            row_info.append({
                'r': r,
                'non_null': len(non_null),
                'strings': len(strings),
                'numbers': len(numbers),
                'neighbor_dupes': neighbor_dupes,
                'col_start': col_start,
                'col_end': col_end,
                'width': width,
                'unique_vals': unique_vals,
            })

        # Find candidate header groups: consecutive string-heavy wide rows
        best_group = None
        best_score = 0

        for start_idx in range(len(row_info)):
            info = row_info[start_idx]
            # Header row should have many strings and be wide
            if info['strings'] < 5 or info['width'] < 10:
                continue

            # Build group of consecutive header rows
            group = [start_idx]
            for next_idx in range(start_idx + 1, min(start_idx + 5, len(row_info))):
                next_info = row_info[next_idx]
                # Next row should also be string-heavy or have the repeating pattern
                if next_info['strings'] >= 3 and next_info['numbers'] <= next_info['strings']:
                    group.append(next_idx)
                else:
                    break

            if len(group) < 2:
                continue

            # Validate: row after group should have numeric data
            data_row_idx = group[-1] + 1
            if data_row_idx >= len(row_info):
                continue
            data_info = row_info[data_row_idx]
            if data_info['numbers'] < 1:
                continue

            # Score the group
            # High neighbor dupes = merged cells = zone-style headers
            total_dupes = sum(row_info[i]['neighbor_dupes'] for i in group)
            # Last row should have few unique values (Day/Night pattern)
            last_unique = row_info[group[-1]]['unique_vals']
            # Width should be large
            max_width = max(row_info[i]['width'] for i in group)

            score = (
                total_dupes * 3 +     # Merged cells signal
                max_width * 2 +        # Wide tables
                len(group) * 5 +       # More header rows = more complex
                (10 if last_unique <= 5 else 0)  # Day/Night pattern bonus
            )

            if score > best_score:
                best_score = score
                # Determine overall column extent
                g_col_start = min(row_info[i]['col_start'] for i in group)
                g_col_end = max(row_info[i]['col_end'] for i in group)
                best_group = (group[0], group[-1], g_col_start, g_col_end)

        if best_group is None:
            return None

        header_start, header_end, col_start, col_end = best_group

        # Final validation: minimum score threshold
        if best_score < 50:
            return None

        # Require at least one header row with significant merged-cell dupes
        max_dupes = max(row_info[i]['neighbor_dupes'] for i in range(header_start, header_end + 1))
        width = col_end - col_start + 1
        if max_dupes < 5 or (width > 0 and max_dupes / width < 0.15):
            return None

        # Refine column extent: trim to contiguous dense region
        # Use the last header row (most granular, e.g. Day/Night) to find where data stops
        last_header = matrix[header_end]
        refined_end = col_start
        gap = 0
        for c in range(col_start, col_end + 1):
            v = last_header[c] if c < len(last_header) else None
            if v is not None and str(v).strip() and str(v).strip().lower() not in self.HEADER_SKIP_VALUES:
                refined_end = c
                gap = 0
            else:
                gap += 1
                if gap > 5:  # Large gap = end of table
                    break

        # Also check the first header row for Total/Grand Total as natural end
        first_header = matrix[header_start]
        for c in range(col_start, refined_end + 1):
            v = first_header[c] if c < len(first_header) else None
            if v is not None and str(v).strip().lower() == 'grand total':
                refined_end = c  # Include Grand Total but stop after
                break

        if refined_end > col_end:
            refined_end = col_end
        col_end = refined_end

        logger.info(
            f"[DenseTable] Multi-row header detected: "
            f"rows {header_start + 1}-{header_end + 1}, "
            f"cols {col_start + 1}-{col_end + 1}, score={best_score}"
        )

        return (header_start, header_end, col_start, col_end)

    def _merge_header_rows(
        self, matrix: List[List],
        header_start: int, header_end: int,
        col_start: int, col_end: int,
    ) -> List[str]:
        """
        Merge multi-row header into single composite column names.
        For each column, concatenate unique non-empty values from each header row,
        skipping separator values like '-' and 'x'.

        Example:  Row 5: Zone # 1A  |  Row 6: Road &Utility  |  Row 7: Day
        Result:   zone_1a_road_utility_day
        """
        merged = []
        skip_lower = self.HEADER_SKIP_VALUES

        for c in range(col_start, col_end + 1):
            parts = []
            seen = set()

            for r in range(header_start, header_end + 1):
                if c < len(matrix[r]):
                    v = matrix[r][c]
                else:
                    v = None

                if v is None:
                    continue

                text = str(v).strip()
                if not text or text.lower() in skip_lower:
                    continue

                # Skip if this is the same as a value already collected
                # (e.g., merged zones that repeat across rows)
                text_norm = text.lower()
                if text_norm not in seen:
                    seen.add(text_norm)
                    parts.append(text)

            if parts:
                # Join parts with underscore, then normalize
                raw = "_".join(parts)
                # Clean special characters, normalize
                clean = re.sub(r'[#&/\\()]+', '_', raw)
                clean = re.sub(r'\s+', '_', clean)
                clean = re.sub(r'[^a-zA-Z0-9_]', '', clean)
                clean = re.sub(r'_+', '_', clean).strip('_').lower()
                merged.append(clean if clean else f"col_{c}")
            else:
                merged.append(f"col_{c}")

        # Handle duplicates (same composite name for different columns)
        seen_names: Dict[str, int] = {}
        final = []
        for name in merged:
            if name in seen_names:
                seen_names[name] += 1
                final.append(f"{name}_{seen_names[name]}")
            else:
                seen_names[name] = 0
                final.append(name)

        return final

    def _find_dense_data_end(
        self, matrix: List[List],
        data_start: int, col_start: int, col_end: int,
    ) -> int:
        """Find the last data row of a dense table."""
        num_rows = len(matrix)
        end_row = data_start
        empty_streak = 0

        # For wide tables (DPR), use absolute minimum (3 cells)
        # For narrow tables, use percentage
        col_span = col_end - col_start + 1
        min_cells = max(3, int(col_span * 0.03))  # 3% or at least 3

        for r in range(data_start, num_rows):
            non_null = 0
            for c in range(col_start, min(col_end + 1, len(matrix[r]))):
                v = matrix[r][c]
                if v is not None and str(v).strip():
                    non_null += 1

            if non_null >= min_cells:
                end_row = r
                empty_streak = 0
            else:
                empty_streak += 1
                if empty_streak > 3:
                    break

        return end_row

    # ── Strategy 3: Block detection (on filled matrix) ───────────

    def _extract_block_tables_from_matrix(
        self, matrix: List[List], sheet_name: str,
        existing_tables: List[ExtractedTable],
    ) -> List[ExtractedTable]:
        """Block detection on the unmerged/filled matrix."""
        tables = []
        try:
            arr = np.array(matrix, dtype=object)
            non_empty = pd.DataFrame(arr).notna() & (pd.DataFrame(arr) != "")
            non_empty = non_empty.values

            blocks = self._find_table_blocks(non_empty, existing_tables)

            for block in blocks:
                start_row, start_col, end_row, end_col = block
                data = arr[start_row:end_row + 1, start_col:end_col + 1].tolist()

                if len(data) < self.MIN_BLOCK_ROWS:
                    continue

                header_idx = self._detect_header_row(data)
                if header_idx >= len(data) - 1:
                    continue

                headers = data[header_idx]
                body = data[header_idx + 1:]

                df = pd.DataFrame(body, columns=headers)
                df = self._clean_dataframe(df)

                if df.empty or len(df.columns) < self.MIN_BLOCK_COLS:
                    continue

                tables.append(ExtractedTable(
                    df=df,
                    sheet_name=sheet_name,
                    start_row=start_row + 1,
                    start_col=start_col + 1,
                    end_row=end_row + 1,
                    end_col=end_col + 1,
                    extraction_method="block_detect",
                ))
        except Exception as e:
            logger.warning(f"[ExcelExtractor] Block detection error: {e}")
        return tables

    # ── Strategy 3: Invoice / form-aware extraction ──────────────

    def _extract_invoice_tables(
        self, matrix: List[List], sheet_name: str,
    ) -> List[ExtractedTable]:
        """
        Detect invoice/form layouts:
        1. Skip metadata header rows  (label : value)
        2. Find the real column-header row  (3+ strings, data below has numbers)
        3. Extract the data table
        4. Optionally capture header metadata for context
        """
        tables: List[ExtractedTable] = []
        num_rows = len(matrix)
        if num_rows < 3:
            return tables
        num_cols = len(matrix[0]) if matrix else 0

        # ── Phase 1: find candidate header rows ──────────────────
        candidates = self._find_header_candidates(matrix)
        if not candidates:
            logger.info("[InvoiceDetect] No header candidates found")
            return tables

        logger.info(f"[InvoiceDetect] {len(candidates)} header candidate(s)")

        for header_row, col_start, col_end, score in candidates:
            # ── Phase 2: determine data extent below header ──────
            data_end = self._find_data_end(matrix, header_row, col_start, col_end)

            data_rows = data_end - header_row
            if data_rows < 1:
                continue

            # ── Phase 3: extract DataFrame ───────────────────────
            headers = matrix[header_row][col_start:col_end + 1]
            body = [
                matrix[r][col_start:col_end + 1]
                for r in range(header_row + 1, data_end + 1)
            ]

            df = pd.DataFrame(body, columns=headers)
            df = self._clean_dataframe(df)

            if df.empty or len(df) < 1:
                continue

            # ── Phase 4: extract header metadata ─────────────────
            header_meta = self._extract_header_metadata(matrix, header_row)

            tables.append(ExtractedTable(
                df=df,
                sheet_name=sheet_name,
                start_row=header_row + 1,
                start_col=col_start + 1,
                end_row=data_end + 1,
                end_col=col_end + 1,
                extraction_method="invoice_detect",
                header_metadata=header_meta,
            ))

            logger.info(
                f"[InvoiceDetect] Extracted table: "
                f"header=row{header_row + 1}, cols={col_start + 1}-{col_end + 1}, "
                f"{len(df)} data rows, {len(df.columns)} cols"
            )
            # Only extract one table per invoice layout (the primary data table)
            break

        return tables

    def _find_header_candidates(self, matrix: List[List]) -> List[Tuple[int, int, int, float]]:
        """
        Score each row as a potential column-header row.
        Returns list of (row_idx, col_start, col_end, score) sorted by score desc.

        A good header row:
        - Has 3+ non-null unique string values
        - Is NOT a metadata label row (label : value)
        - Rows below it have numeric data
        """
        num_rows = len(matrix)
        num_cols = len(matrix[0]) if matrix else 0
        candidates = []

        for i in range(min(num_rows - 1, 30)):  # scan first 30 rows
            row = matrix[i]

            # Count non-null cells and their positions
            non_null_positions = []
            string_vals = []
            for c, v in enumerate(row):
                if v is not None and str(v).strip():
                    non_null_positions.append(c)
                    if isinstance(v, str) and len(str(v).strip()) > 0:
                        string_vals.append((c, str(v).strip()))

            if len(string_vals) < 3:
                continue

            # Skip if this looks like a metadata/label row (has ":" separator)
            if self._is_metadata_row(row):
                continue

            # Skip if all strings are the same value (merged cell artifact)
            unique_strings = set(s for _, s in string_vals)
            if len(unique_strings) < 3:
                continue

            # Determine column extent from this row
            col_start = non_null_positions[0]
            col_end = non_null_positions[-1]

            # Check rows below: do they have numeric data?
            numeric_rows_below = 0
            for j in range(i + 1, min(i + 6, num_rows)):
                nums_in_row = 0
                for c in range(col_start, col_end + 1):
                    v = matrix[j][c]
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        nums_in_row += 1
                if nums_in_row >= 2:
                    numeric_rows_below += 1

            if numeric_rows_below < 1:
                continue

            # Score: unique strings + numeric rows below + bonus for serial-number column
            score = len(unique_strings) + numeric_rows_below * 2

            # Bonus if first column looks like a serial number header
            first_str = string_vals[0][1] if string_vals else ""
            if SERIAL_COL_PATTERNS.match(first_str):
                score += 5

            # Bonus if next row has a numeric value in the first data column
            if i + 1 < num_rows:
                next_first = matrix[i + 1][col_start]
                if isinstance(next_first, (int, float)) and not isinstance(next_first, bool):
                    score += 3

            candidates.append((i, col_start, col_end, score))

        candidates.sort(key=lambda x: x[3], reverse=True)
        return candidates

    def _find_data_end(
        self, matrix: List[List], header_row: int, col_start: int, col_end: int,
    ) -> int:
        """Find the last row of data below the header."""
        num_rows = len(matrix)
        end_row = header_row
        empty_streak = 0
        col_span = col_end - col_start + 1

        for r in range(header_row + 1, num_rows):
            non_null = sum(
                1 for c in range(col_start, col_end + 1)
                if matrix[r][c] is not None and str(matrix[r][c]).strip()
            )

            # Check if this row is a footer/signature indicator
            row_text = " ".join(
                str(matrix[r][c]) for c in range(col_start, col_end + 1)
                if matrix[r][c] is not None
            ).lower()

            if any(kw in row_text for kw in [
                'for ', 'signature', 'manager', 'director', 'authorized',
                'regards', 'sincerely', 'approved by',
            ]) and non_null <= 2:
                break

            # Row needs >=20% non-empty cells to be considered data
            if col_span > 0 and non_null / col_span >= 0.2:
                end_row = r
                empty_streak = 0
            else:
                empty_streak += 1
                if empty_streak > 3:
                    break

        return end_row

    def _is_metadata_row(self, row: List) -> bool:
        """Check if a row is a metadata label row like 'Name of Project : ...'."""
        non_null = [v for v in row if v is not None and str(v).strip()]
        if not non_null:
            return False

        text = " ".join(str(v) for v in non_null).strip()

        # Check for "label : value" pattern
        if ':' in text and METADATA_LABEL_PATTERNS.search(text):
            return True

        # Single long text string in a row → likely a title/header
        if len(non_null) == 1 and isinstance(non_null[0], str) and len(str(non_null[0])) > 20:
            text_lower = str(non_null[0]).lower()
            if any(kw in text_lower for kw in ['invoice', 'certificate', 'report', 'statement']):
                return True

        return False

    def _extract_header_metadata(
        self, matrix: List[List], header_row: int,
    ) -> Dict[str, str]:
        """
        Extract label:value metadata from rows above the data table header.
        Returns dict like {"project_name": "...", "customer": "...", etc.}
        """
        metadata: Dict[str, str] = {}

        for r in range(0, header_row):
            row = matrix[r]
            non_null = [v for v in row if v is not None and str(v).strip()]
            if not non_null:
                continue

            text = " ".join(str(v) for v in non_null).strip()

            # Try to parse "Label : Value" patterns
            parts = text.split(':', 1)
            if len(parts) == 2:
                label = parts[0].strip().lower()
                value = parts[1].strip()
                if value and len(label) < 50:
                    # Normalize common labels
                    if any(kw in label for kw in ['project']):
                        metadata['project_name'] = value
                    elif any(kw in label for kw in ['customer', 'client']):
                        metadata['customer_name'] = value
                    elif any(kw in label for kw in ['contractor', 'sub con']):
                        metadata['contractor_name'] = value
                    elif any(kw in label for kw in ['contract value']):
                        metadata['contract_value'] = value
                    elif any(kw in label for kw in ['invoice no']):
                        metadata['invoice_number'] = value
                    elif any(kw in label for kw in ['order ref', 'reference', 'ref']):
                        metadata['reference'] = value
                    elif any(kw in label for kw in ['date']):
                        metadata['date'] = value
                    else:
                        # Store with cleaned label
                        clean_label = re.sub(r'[^a-z0-9_]', '_', label)
                        clean_label = re.sub(r'_+', '_', clean_label).strip('_')
                        if clean_label:
                            metadata[clean_label] = value

            # Also check for "INVOICE NO. X" pattern
            m = re.search(r'invoice\s*(?:no\.?\s*)?(\d+)', text, re.IGNORECASE)
            if m and 'invoice_number' not in metadata:
                metadata['invoice_number'] = m.group(1)

        if metadata:
            logger.info(f"[InvoiceDetect] Header metadata: {list(metadata.keys())}")

        return metadata

    # ── Strategy 4: Full sheet fallback (matrix-based) ───────────

    def _extract_full_sheet_from_matrix(
        self, matrix: List[List], sheet_name: str,
    ) -> Optional[ExtractedTable]:
        """Full sheet extraction using the filled matrix."""
        if not matrix or len(matrix) < 2:
            return None

        try:
            # Find the best header row
            header_row = self._detect_header_row_in_matrix(matrix)

            # Determine used column range
            col_start, col_end = self._find_used_columns(matrix, header_row)
            if col_end - col_start < 1:
                return None

            headers = matrix[header_row][col_start:col_end + 1]
            body = [
                matrix[r][col_start:col_end + 1]
                for r in range(header_row + 1, len(matrix))
            ]

            if not body:
                return None

            df = pd.DataFrame(body, columns=headers)
            df = self._clean_dataframe(df)

            if df.empty:
                return None

            return ExtractedTable(
                df=df,
                sheet_name=sheet_name,
                start_row=header_row + 1,
                start_col=col_start + 1,
                end_row=header_row + len(df) + 1,
                end_col=col_end + 1,
                extraction_method="full_sheet",
            )
        except Exception as e:
            logger.warning(f"[ExcelExtractor] Full sheet (matrix) failed: {e}")
            return None

    def _extract_full_sheet(self, file_path: str, sheet_name: str) -> Optional[ExtractedTable]:
        """Extract entire sheet as a single table (pandas fallback)."""
        try:
            df_preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)

            header_row = 0
            for i in range(min(10, len(df_preview))):
                row = df_preview.iloc[i]
                string_count = sum(1 for v in row if isinstance(v, str) and v)
                if string_count >= len(row) * 0.3:
                    header_row = i
                    break

            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            df = self._clean_dataframe(df)

            if df.empty:
                return None

            return ExtractedTable(
                df=df,
                sheet_name=sheet_name,
                start_row=header_row + 1,
                start_col=1,
                end_row=header_row + len(df) + 1,
                end_col=len(df.columns),
                extraction_method="full_sheet",
            )
        except Exception as e:
            logger.warning(f"[ExcelExtractor] Full sheet extraction failed: {e}")
            return None

    # ── Block detection helpers ───────────────────────────────────

    def _find_table_blocks(
        self, non_empty: np.ndarray, existing_tables: List[ExtractedTable],
    ) -> List[Tuple[int, int, int, int]]:
        """Find table-like rectangular blocks in the non-empty mask."""
        blocks = []
        visited = np.zeros_like(non_empty, dtype=bool)

        for table in existing_tables:
            r1, c1 = table.start_row - 1, table.start_col - 1
            r2, c2 = table.end_row - 1, table.end_col - 1
            r1, r2 = max(0, r1), min(non_empty.shape[0] - 1, r2)
            c1, c2 = max(0, c1), min(non_empty.shape[1] - 1, c2)
            visited[r1:r2 + 1, c1:c2 + 1] = True

        rows, cols = non_empty.shape
        for start_row in range(rows):
            for start_col in range(cols):
                if visited[start_row, start_col] or not non_empty[start_row, start_col]:
                    continue

                end_row, end_col = self._expand_block(
                    non_empty, visited, start_row, start_col
                )

                block_rows = end_row - start_row + 1
                block_cols = end_col - start_col + 1

                if block_rows >= self.MIN_BLOCK_ROWS and block_cols >= self.MIN_BLOCK_COLS:
                    blocks.append((start_row, start_col, end_row, end_col))

                visited[start_row:end_row + 1, start_col:end_col + 1] = True

        return blocks

    def _expand_block(
        self, non_empty: np.ndarray, visited: np.ndarray,
        start_row: int, start_col: int,
    ) -> Tuple[int, int]:
        """
        Expand a block from (start_row, start_col).
        Improved: uses the densest row in the first 5 rows for column extent.
        """
        rows, cols = non_empty.shape

        # Determine column extent using the densest of first 5 rows
        best_end_col = start_col
        for probe_r in range(start_row, min(start_row + 5, rows)):
            end_c = start_col
            gap = 0
            for c in range(start_col, cols):
                if visited[probe_r, c]:
                    break
                if non_empty[probe_r, c]:
                    end_c = c
                    gap = 0
                else:
                    gap += 1
                    if gap > self.MAX_EMPTY_COLS_IN_BLOCK:
                        break
            if end_c > best_end_col:
                best_end_col = end_c

        end_col = best_end_col
        col_span = end_col - start_col + 1

        # Find row extent
        end_row = start_row
        empty_row_count = 0

        for r in range(start_row, rows):
            row_non_empty = sum(
                1 for c in range(start_col, end_col + 1)
                if non_empty[r, c]
            )

            if col_span > 0 and row_non_empty / col_span >= 0.25:
                end_row = r
                empty_row_count = 0
            else:
                empty_row_count += 1
                if empty_row_count > self.MAX_EMPTY_ROWS_IN_BLOCK:
                    break

        return end_row, end_col

    # ── Header detection helpers ─────────────────────────────────

    def _detect_header_row(self, data: List[List]) -> int:
        """Detect which row is the header based on heuristics."""
        for i, row in enumerate(data[:8]):  # Check first 8 rows (was 5)
            if row is None:
                continue

            string_count = sum(1 for v in row if isinstance(v, str) and v and len(str(v)) > 0)
            non_null_count = sum(1 for v in row if v is not None)

            if non_null_count < 3:
                continue

            # Skip metadata rows
            if self._is_metadata_row(row):
                continue

            # Header should have mostly strings (>=50%)
            if string_count / non_null_count >= 0.5:
                # Check that next row has different pattern (more numbers)
                if i + 1 < len(data):
                    next_row = data[i + 1]
                    next_num_count = sum(
                        1 for v in next_row
                        if isinstance(v, (int, float)) and not isinstance(v, bool)
                    )
                    next_string_count = sum(1 for v in next_row if isinstance(v, str))
                    if next_num_count > 0 or next_string_count < string_count:
                        return i

        return 0

    def _detect_header_row_in_matrix(self, matrix: List[List]) -> int:
        """Find the best header row in a full matrix, skipping metadata rows."""
        best_row = 0
        best_score = 0

        for i in range(min(len(matrix) - 1, 20)):
            row = matrix[i]
            non_null = [v for v in row if v is not None and str(v).strip()]
            strings = [v for v in non_null if isinstance(v, str) and len(str(v).strip()) > 0]

            if len(strings) < 2:
                continue

            if self._is_metadata_row(row):
                continue

            unique = len(set(str(s).strip().lower() for s in strings))
            if unique < 2:
                continue

            # Score: unique string count + bonus for data below
            score = unique
            if i + 1 < len(matrix):
                next_nums = sum(
                    1 for v in matrix[i + 1]
                    if isinstance(v, (int, float)) and not isinstance(v, bool)
                )
                score += next_nums * 2

            if score > best_score:
                best_score = score
                best_row = i

        return best_row

    def _find_used_columns(
        self, matrix: List[List], start_row: int,
    ) -> Tuple[int, int]:
        """Find the range of columns that contain data."""
        num_cols = len(matrix[0]) if matrix else 0
        col_start = num_cols
        col_end = 0

        for r in range(start_row, min(start_row + 20, len(matrix))):
            for c, v in enumerate(matrix[r]):
                if v is not None and str(v).strip():
                    col_start = min(col_start, c)
                    col_end = max(col_end, c)

        if col_start > col_end:
            return 0, num_cols - 1
        return col_start, col_end

    # ── DataFrame cleaning ───────────────────────────────────────

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and normalize DataFrame with jargon-aware column names."""
        if df.empty:
            return df

        # Drop completely empty rows/columns
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')

        if df.empty:
            return df

        column_jargon = {}
        new_columns = []

        for i, col in enumerate(df.columns):
            if col is None or pd.isna(col) or str(col).strip() == '':
                clean = f"col_{i}"
            else:
                raw = str(col).strip()
                _, meaning = self.jargon.normalize_column_name(raw)
                if meaning:
                    column_jargon[raw.lower()] = meaning

                clean = re.sub(r'[^a-zA-Z0-9_]', '_', raw)
                clean = re.sub(r'_+', '_', clean).strip('_').lower()
                if not clean:
                    clean = f"col_{i}"

                if meaning:
                    column_jargon[clean] = meaning
            new_columns.append(clean)

        # Handle duplicates
        seen: Dict[str, int] = {}
        final_columns = []
        for col in new_columns:
            if col in seen:
                seen[col] += 1
                final_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                final_columns.append(col)

        df.columns = final_columns
        df.attrs['column_jargon'] = column_jargon

        # Remove summary/total rows
        df = self._remove_summary_rows(df)

        return df

    def _remove_summary_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove summary/total rows checking all columns."""
        if df.empty or len(df) < 2:
            return df

        mask = pd.Series(True, index=df.index)

        # Check ALL string columns for summary keywords (not just first 3)
        for col in df.columns:
            if df[col].dtype == 'object' or str(df[col].dtype) == 'string':
                values_lower = df[col].astype(str).str.lower().str.strip()
                # Exact match
                is_summary = values_lower.isin(self.SUMMARY_KEYWORDS)
                # Partial match for tokens like "Net Amount Due"
                for token in self.SUMMARY_TOKENS:
                    is_summary |= values_lower.str.contains(token, na=False, regex=False)
                mask &= ~is_summary

        removed = (~mask).sum()
        if removed > 0:
            logger.info(f"[ExcelExtractor] Removed {removed} summary row(s)")

        return df[mask].reset_index(drop=True)

    # ── Utility ──────────────────────────────────────────────────

    @staticmethod
    def _parse_cell_ref(ref: str) -> Tuple[int, int]:
        """Parse cell reference like 'A1' to (col, row) 1-indexed."""
        match = re.match(r'([A-Z]+)(\d+)', ref.upper())
        if not match:
            return 1, 1

        col_str, row_str = match.groups()
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char) - ord('A') + 1)

        return col, int(row_str)

    @staticmethod
    def _sanitize_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
        """
        Fix mixed-type columns that would break parquet serialization.
        Handles: mixed int/float/str, datetime objects, etc.
        """
        from datetime import datetime as dt_type, date as date_type

        df = df.copy()
        for col in df.columns:
            if df[col].dtype == object:
                # Check types present in the column
                types_present = set()
                has_datetime = False
                for v in df[col].dropna().head(50):
                    if isinstance(v, (dt_type, date_type)):
                        has_datetime = True
                        types_present.add('datetime')
                    elif isinstance(v, (int, float)) and not isinstance(v, bool):
                        types_present.add('numeric')
                    elif isinstance(v, str):
                        types_present.add('string')

                if has_datetime or len(types_present) > 1:
                    # Mixed types or datetime: convert all to string
                    df[col] = df[col].apply(
                        lambda v: str(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None
                    )
                elif types_present == {'numeric'}:
                    # All numeric but stored as object: convert to float
                    df[col] = pd.to_numeric(df[col], errors='coerce')

        return df

    def _generate_table_metadata(
        self, table: ExtractedTable, source_file: str,
    ) -> Tuple[str, List[str], Dict[str, str]]:
        """
        Auto-generate description, semantic tags, and header metadata for a table.
        Uses file name, sheet name, column names, and extraction context.
        """
        tags: List[str] = []
        desc_parts: List[str] = []
        file_stem = Path(source_file).stem.lower()

        # --- Tags from file name ---
        if 'dpr' in file_stem or 'daily' in file_stem:
            tags.extend(['dpr', 'daily_report'])
        if 'manpower' in file_stem or 'man power' in file_stem:
            tags.append('manpower')
        if 'invoice' in file_stem:
            tags.append('invoice')
        if 'equipment' in file_stem:
            tags.append('equipment')
        if 'progress' in file_stem:
            tags.append('progress')
        if 'cost' in file_stem or 'payment' in file_stem:
            tags.extend(['cost', 'payment'])

        # Extract date from file name (e.g., "DPR 180209" → "2018-02-09")
        date_match = re.search(r'(\d{6})', file_stem)
        if date_match:
            ds = date_match.group(1)
            tags.append(f"date_{ds}")

        # --- Tags from sheet name ---
        sheet = (table.sheet_name or "").lower()
        if 'man power' in sheet or 'manpower' in sheet:
            tags.extend(['manpower', 'labor', 'workers', 'headcount'])
            desc_parts.append("Manpower tracking by zone and shift")
        elif 'equipment' in sheet:
            tags.extend(['equipment', 'machinery'])
            desc_parts.append("Equipment deployment")
        elif 'progress' in sheet:
            tags.extend(['progress', 'completion'])
            desc_parts.append("Work progress tracking")
        elif 'road' in sheet:
            tags.extend(['road', 'infrastructure'])
            desc_parts.append("Road works data")

        # --- Tags from extraction method ---
        if table.extraction_method == 'dense_table':
            tags.append('detailed')

        # --- Tags from column names ---
        cols_lower = [c.lower() for c in table.df.columns]
        col_text = " ".join(cols_lower)

        if any('zone' in c for c in cols_lower):
            tags.append('zone_breakdown')
        if any('day' in c or 'night' in c for c in cols_lower):
            tags.append('shift_data')
        if any('cost' in c or 'amount' in c or 'payment' in c for c in cols_lower):
            tags.extend(['financial', 'cost'])
        if any('qty' in c or 'quantity' in c for c in cols_lower):
            tags.append('quantities')
        if any('contractor' in c for c in cols_lower):
            tags.append('contractor')
        if any('total' in c or 'grand_total' in c for c in cols_lower):
            tags.append('summary')

        # --- Build description ---
        file_name = Path(source_file).name
        if not desc_parts:
            desc_parts.append(f"Data from {table.sheet_name or file_name}")

        desc = f"{', '.join(desc_parts)} ({file_name}"
        if table.sheet_name:
            desc += f" / {table.sheet_name}"
        desc += f", {len(table.df)} rows)"

        # --- Header metadata from extraction ---
        hdr_meta = dict(table.header_metadata) if table.header_metadata else {}
        # Add file-level context
        hdr_meta['source_file'] = file_name
        if table.sheet_name:
            hdr_meta['sheet_name'] = table.sheet_name
        if date_match:
            hdr_meta['report_date'] = date_match.group(1)

        # Deduplicate tags
        tags = list(dict.fromkeys(tags))

        return desc, tags, hdr_meta

    def save_to_parquet(self, table: ExtractedTable, source_file: str) -> Optional[TableMetadata]:
        """Save extracted table to Parquet and register in catalog."""
        try:
            table_id = self.catalog.generate_table_id(
                source_file,
                sheet_name=table.sheet_name,
                table_index=0,
            )

            parquet_path = self.catalog.generate_parquet_path(table_id)

            # Sanitize mixed-type columns before saving to parquet
            df = self._sanitize_for_parquet(table.df)
            # Preserve sheet name as a column for period filtering
            if table.sheet_name and '_sheet_name' not in df.columns:
                df['_sheet_name'] = table.sheet_name
            df.to_parquet(parquet_path, index=False)

            logger.info(f"[ExcelExtractor] Saved parquet: {parquet_path.name}")

            col_jargon = getattr(table.df, 'attrs', {}).get('column_jargon', {})
            if table.column_jargon:
                col_jargon.update(table.column_jargon)

            # Generate table-level metadata for query routing
            desc, tags, hdr_meta = self._generate_table_metadata(table, source_file)

            meta = TableMetadata(
                table_id=table_id,
                source_file=source_file,
                source_type="excel",
                table_name=table_id,
                parquet_path=str(parquet_path),
                sheet_name=table.sheet_name,
                row_count=len(df),
                column_count=len(df.columns),
                columns=list(df.columns),
                extraction_method=table.extraction_method,
                file_hash=self.catalog.compute_file_hash(source_file),
                column_jargon=col_jargon,
                description=desc,
                semantic_tags=tags,
                header_metadata=hdr_meta,
            )

            return meta

        except Exception as e:
            logger.error(f"[ExcelExtractor] Error saving parquet: {e}")
            return None


# ── Convenience function ─────────────────────────────────────────

def extract_excel_tables(file_path: str, save_parquet: bool = True) -> List[TableMetadata]:
    """
    Extract tables from Excel file and optionally save to parquet.

    Returns:
        List of TableMetadata for extracted tables
    """
    extractor = ExcelTableExtractor()
    tables = extractor.extract_tables(file_path)

    if not save_parquet:
        return []

    catalog = get_catalog()
    entry = catalog.add_entry(file_path, "excel")

    metadata_list = []
    for table in tables:
        meta = extractor.save_to_parquet(table, file_path)
        if meta:
            catalog.add_table(entry, meta)
            metadata_list.append(meta)

    return metadata_list
